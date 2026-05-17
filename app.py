"""
SOCOM GMAO — Backend v3.2
Audit hardening: bcrypt, rate limiting, CORS restreint, logging structuré,
token avec expiration, gestion connexions DB, taille payload limitée.
"""
import hashlib, json, os, sqlite3, smtplib, io, logging, secrets, time, re

# Pillow pour redimensionnement d'images (v217 — bibliothèque d'images)
try:
    from PIL import Image as PILImage
    HAS_PILLOW = True
except Exception:
    HAS_PILLOW = False

# Web Push (VAPID) — imports optionnels : si la lib n'est pas installée, les push sont désactivés
try:
    from pywebpush import webpush, WebPushException
    HAS_PYWEBPUSH = True
except Exception:
    HAS_PYWEBPUSH = False
from datetime import datetime, date, timedelta
from functools import wraps
from pathlib import Path
from logging.handlers import RotatingFileHandler
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from flask import Flask, jsonify, request, send_file, Response, g

# bcrypt en optionnel (fallback SHA-256 pour la migration progressive)
try:
    import bcrypt
    HAS_BCRYPT = True
except ImportError:
    HAS_BCRYPT = False

# Rate limiting en optionnel (warning si absent)
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    HAS_LIMITER = True
except ImportError:
    HAS_LIMITER = False

from flask_cors import CORS

BASE_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
DB_PATH  = os.environ.get("GMAO_DB", str(BASE_DIR / "gmao.db"))

# ══════════════════════════════════════════════════════════════════════
# WEB PUSH (notifications PWA) — Configuration VAPID
# ══════════════════════════════════════════════════════════════════════
# Clés VAPID : à générer UNE FOIS via `python3 generate_vapid.py` sur le serveur.
# Les clés ici sont des PLACEHOLDERS — remplacer par les vraies valeurs générées,
# OU définir les variables d'environnement GMAO_VAPID_PUBLIC / GMAO_VAPID_PRIVATE / GMAO_VAPID_EMAIL.
VAPID_PUBLIC_KEY = os.environ.get("GMAO_VAPID_PUBLIC", "")
VAPID_PRIVATE_KEY = os.environ.get("GMAO_VAPID_PRIVATE", "")
VAPID_CLAIM_EMAIL = os.environ.get("GMAO_VAPID_EMAIL", "mailto:contact@socom.lu")

# ── Logging structuré avec rotation ──
LOG_DIR = BASE_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)
logger = logging.getLogger("gmao")
logger.setLevel(logging.INFO)
_h = RotatingFileHandler(str(LOG_DIR / "gmao.log"), maxBytes=5*1024*1024, backupCount=5, encoding="utf-8")
_h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
logger.addHandler(_h)
# Aussi en console (stdout gunicorn)
_sh = logging.StreamHandler()
_sh.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
logger.addHandler(_sh)

app = Flask(__name__)
# Limite de taille de requête : 25 Mo (permet upload de plusieurs photos compressées)
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024
app.config["JSON_SORT_KEYS"] = False

# ── CORS : restreindre aux origines légitimes via env, wildcard en fallback ──
_cors_env = os.environ.get("GMAO_CORS_ORIGINS", "")
if _cors_env.strip():
    _cors_origins = [o.strip() for o in _cors_env.split(",") if o.strip()]
else:
    # Fallback: origines locales + wildcard en dev. En prod, DÉFINIR GMAO_CORS_ORIGINS.
    _cors_origins = ["http://localhost", "http://127.0.0.1", "http://localhost:5173",
                     "http://localhost:8000", "http://187.127.68.131"]
    logger.warning("GMAO_CORS_ORIGINS non defini — utilisation de la liste par defaut. "
                   "DEFINIR cette variable en production avec le(s) domaine(s) autorise(s).")
CORS(app, origins=_cors_origins, supports_credentials=False)

# ── Rate limiting (login, reset password) ──
if HAS_LIMITER:
    limiter = Limiter(get_remote_address, app=app,
                      default_limits=[], storage_uri="memory://")
else:
    # Stub qui no-op si la lib n'est pas là
    class _NoopLimiter:
        def limit(self, *a, **k):
            def deco(fn): return fn
            return deco
    limiter = _NoopLimiter()
    logger.warning("flask-limiter non installe — rate limiting desactive. "
                   "Installer avec: pip install flask-limiter")

# ── Token d'auth : durée de vie (secondes), configurable ──
TOKEN_TTL = int(os.environ.get("GMAO_TOKEN_TTL", 60 * 60 * 24 * 7))  # 7 jours par defaut
# Secret serveur utilisé pour signer les tokens (persistant entre redémarrages)
# Stocké dans parametres_app à la première utilisation
_SERVER_SECRET_CACHE = None

# Hook global : gestion d'erreur JSON uniforme + payload trop gros
@app.errorhandler(413)
def _too_large(e):
    return jsonify({"error": "Payload trop volumineux (max 10 Mo)"}), 413

@app.errorhandler(429)
def _too_many(e):
    return jsonify({"error": "Trop de requetes, veuillez reessayer plus tard"}), 429

@app.route("/")
def index():
    f = BASE_DIR / "index.html"
    return send_file(str(f)) if f.exists() else ("index.html introuvable", 404)

@app.route("/mobile")
def mobile():
    f = BASE_DIR / "gmao_mobile.html"
    if not f.exists(): return ("gmao_mobile.html introuvable", 404)
    resp = send_file(str(f))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp

@app.route("/sw.js")
def sw():
    f = BASE_DIR / "sw.js"
    if not f.exists(): return ("", 404)
    # v218.38 : injection dynamique de la version pour invalider le cache automatiquement
    content = f.read_text()
    # Remplacer les occurrences hardcodées par la version courante du backend
    content = content.replace("'gmao-v218.38'", f"'gmao-v{APP_VERSION}'")
    return Response(content, mimetype="application/javascript",
                    headers={
                        "Service-Worker-Allowed": "/",
                        "Cache-Control": "no-cache, no-store, must-revalidate"
                    })

@app.route("/push.js")
def push_js():
    f = BASE_DIR / "push.js"
    if not f.exists(): return ("", 404)
    return Response(f.read_text(), mimetype="application/javascript",
                    headers={"Cache-Control": "no-cache, no-store, must-revalidate"})

@app.route("/sw_web.js")
def sw_web():
    f = BASE_DIR / "sw_web.js"
    if not f.exists(): return ("", 404)
    return Response(f.read_text(), mimetype="application/javascript",
                    headers={
                        "Service-Worker-Allowed": "/",
                        "Cache-Control": "no-cache, no-store, must-revalidate"
                    })

@app.route("/manifest.json")
def manifest():
    return jsonify({
        "name": "SOCOM GMAO",
        "short_name": "GMAO",
        "description": "GMAO Terrain SOCOM",
        "start_url": "/mobile",
        "scope": "/",
        "display": "standalone",
        "orientation": "portrait",
        "background_color": "#0F1E3D",
        "theme_color": "#0F1E3D",
        "icons": [
            {"src": "/icons/socom-icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any"},
            {"src": "/icons/socom-icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any"},
            {"src": "/icons/socom-icon-maskable-512.png", "sizes": "512x512", "type": "image/png", "purpose": "maskable"}
        ]
    })

# v218.38 : version de l'application (à incrémenter à chaque déploiement)
# Le mobile vérifie cette valeur et propose une mise à jour si différente du cache local
APP_VERSION = "218.173"

@app.route("/api/version")
def api_version():
    """Retourne la version actuelle du backend. Permet au SW mobile de détecter une nouvelle version."""
    return jsonify({"version": APP_VERSION})

@app.route("/icons/<path:filename>")
def icons(filename):
    """Sert les icônes PWA depuis le dossier icons/."""
    # Sécurité : empêcher path traversal
    if ".." in filename or filename.startswith("/"):
        return "", 404
    f = BASE_DIR / "icons" / filename
    if not f.exists() or not f.is_file():
        return "", 404
    # Cache long côté navigateur (les icônes ne changent pas)
    resp = send_file(str(f))
    resp.headers["Cache-Control"] = "public, max-age=31536000, immutable"
    return resp

@app.route("/assets/<path:filename>")
def assets(filename):
    """Sert les assets (logos, images) depuis le dossier assets/."""
    if ".." in filename or filename.startswith("/"):
        return "", 404
    f = BASE_DIR / "assets" / filename
    if not f.exists() or not f.is_file():
        return "", 404
    resp = send_file(str(f))
    resp.headers["Cache-Control"] = "public, max-age=31536000, immutable"
    return resp

@app.route("/health")
def health():
    try: get_db().execute("SELECT 1"); return jsonify({"status":"ok","time":now()})
    except Exception as e: return jsonify({"status":"error","error":str(e)}),500

def hp(pw):
    """Hash SHA-256 conservé pour rétro-compatibilité (vérification legacy)."""
    return hashlib.sha256(pw.encode()).hexdigest()

def hash_password(pw):
    """Hash moderne avec bcrypt si dispo, fallback SHA-256 sinon.
    Retourne une chaîne préfixée par 'bcrypt$' ou 'sha256$' pour discriminer."""
    if HAS_BCRYPT:
        salt = bcrypt.gensalt(rounds=12)
        return "bcrypt$" + bcrypt.hashpw(pw.encode(), salt).decode()
    return "sha256$" + hp(pw)

def verify_password(pw, stored):
    """Vérifie un mot de passe contre un hash stocké.
    Supporte 3 formats: bcrypt$..., sha256$..., ou legacy SHA-256 nu (64 hex)."""
    if not stored:
        return False
    if stored.startswith("bcrypt$"):
        if not HAS_BCRYPT:
            return False
        try:
            return bcrypt.checkpw(pw.encode(), stored[7:].encode())
        except Exception:
            return False
    if stored.startswith("sha256$"):
        return stored[7:] == hp(pw)
    # Legacy : hash SHA-256 brut de 64 caractères hex
    if len(stored) == 64 and all(c in "0123456789abcdef" for c in stored.lower()):
        return stored == hp(pw)
    return False

def needs_rehash(stored):
    """True si le hash stocké n'est pas au format bcrypt (et bcrypt dispo)."""
    return HAS_BCRYPT and not stored.startswith("bcrypt$")

def now(): return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
def today(): return date.today().isoformat()
def rows(cur): return [dict(r) for r in cur.fetchall()]
def one(cur): r = cur.fetchone(); return dict(r) if r else None
def to_int(v):
    try: return int(v) if v not in (None,"","null") else None
    except (ValueError, TypeError): return None

# ── Whitelist des colonnes modifiables par table (remplace la confiance aveugle dans les f-string UPDATE) ──
ALLOWED_UPDATE_COLS = {
    "techniciens_astreinte": {"nom","gsm","couleur","actif"},
    "astreinte_specialites": {"nom","description","ordre"},
    "compteur_types":        {"nom","description"},
    "compteurs":             {"type_id","nom","numero","unite","localisation"},
    "graphiques_config":     {"titre","ordre","series","comparaison"},
    "parametres_rapport":    {"parametre","valeur","ordre"},
    "utilisateurs":          {"nom","email","role","actif","password","token_version","matricule","manager_id","techniques","poste_id","superieur_id"},
    "clients":               {"societe","nom","prenom","email","telephone","notes"},
    "projets":               {"numero_projet","nom","client_id","manager_id","description",
                              "date_debut","date_fin","statut","deplacement_km","nb_deplacements"},
    "equipements":           {"designation","type_technique","localisation","marque","modele",
                              "puissance","numero_serie","in_out","date_mise_en_service","statut",
                              "gamme_id","technique_id","notes","planning_id","semaine_planif",
                              "jour_semaine_planif","intervention_samedi","intervention_dimanche",
                              "trafo_marque","trafo_annee","trafo_numero_serie","trafo_puissance_kva",
                              "trafo_refroidissement","trafo_poids_kg","trafo_tension_entree_v",
                              "trafo_courant_a","trafo_norme","trafo_couplage",
                              "trafo_tension_service_v","trafo_reglage_tension_kv"},
    "pieces":                {"type_piece","date_installation","duree_vie_estimee",
                              "date_fin_de_vie","statut","quantite","numero_serie",
                              "reference","commentaire"},
    "equipes":               {"nom","manager_id"},
    "interventions":         {"equipement_id","technicien_id","equipe_id","type","statut",
                              "date_prevue","date_realisation","description","rapport","heure_prevue"},
    "comptes_rendus":        {"date_intervention","observations","actions_realisees",
                              "mesures","recommandations","conclusion"},
    "techniques":            {"nom","description"},
    "gammes":                {"nom","periodicite","temps"},
    "smtp_config":           {"host","port","username","password","sender_email",
                              "sender_name","use_tls","enabled"},
    "plannings":             {"nom","equipe_id","heures_par_jour","heure_debut"},
}

def safe_update(db, table, id_value, fields, id_col="id"):
    """UPDATE sécurisé avec whitelist de colonnes. Retourne le nb de colonnes modifiées."""
    allowed = ALLOWED_UPDATE_COLS.get(table, set())
    sets, params = [], []
    for col, val in fields.items():
        if col in allowed:
            sets.append(f"{col}=?")
            params.append(val)
    if not sets:
        return 0
    params.append(id_value)
    db.execute(f"UPDATE {table} SET {', '.join(sets)} WHERE {id_col}=?", params)
    return len(sets)

def get_db():
    """Récupère/crée la connexion SQLite liée au contexte de requête.
    Fermeture automatique via teardown_appcontext."""
    if "db" not in g:
        conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=30000")
        conn.execute("PRAGMA foreign_keys=ON")
        g.db = conn
    return g.db

@app.teardown_appcontext
def _close_db(exception):
    db = g.pop("db", None)
    if db is not None:
        try: db.close()
        except Exception: pass

def _get_server_secret():
    """Secret HMAC serveur (généré au premier démarrage, persisté en BDD)."""
    global _SERVER_SECRET_CACHE
    if _SERVER_SECRET_CACHE:
        return _SERVER_SECRET_CACHE
    db = get_db()
    r = one(db.execute("SELECT valeur FROM parametres_app WHERE cle='server_secret'"))
    if r and r["valeur"]:
        _SERVER_SECRET_CACHE = r["valeur"]
        return _SERVER_SECRET_CACHE
    secret = secrets.token_hex(32)
    db.execute("INSERT OR REPLACE INTO parametres_app (cle,valeur) VALUES ('server_secret',?)", (secret,))
    db.commit()
    _SERVER_SECRET_CACHE = secret
    return secret

def _sign(data):
    """HMAC-SHA256 hex court pour signer un token."""
    import hmac
    return hmac.new(_get_server_secret().encode(), data.encode(), hashlib.sha256).hexdigest()

def make_token(user, societe_id=None):
    """Token signé: uid.iat.tv.sid.signature — expirable + révocable.

    v218.64 : ajout du societe_id (sid) pour le multi-tenant. Si non fourni,
    on prend la première société active de l'utilisateur via
    utilisateur_societes (par défaut SOCOM = 1 pour les anciens users).
    """
    iat = int(time.time())
    tv = user.get("token_version") or 0
    sid = societe_id
    if sid is None:
        # Trouver la 1re société active de l'user
        try:
            db = get_db()
            r = one(db.execute(
                "SELECT societe_id FROM utilisateur_societes WHERE utilisateur_id=? AND actif=1 ORDER BY societe_id LIMIT 1",
                (user["id"],)
            ))
            sid = (r and r["societe_id"]) or 1
        except Exception:
            sid = 1
    payload = f"{user['id']}.{iat}.{tv}.{sid}"
    sig = _sign(payload)
    return f"{payload}.{sig}"

def verify_token(token):
    """Retourne l'utilisateur si le token est valide, sinon None.

    v218.64 : extrait societe_id du token et l'attache à l'objet user.
    Tolère aussi l'ancien format à 4 segments (uid.iat.tv.sig) pour la
    rétrocompatibilité — on assigne alors la 1re société de l'user.
    """
    if not token or "." not in token:
        return None
    parts = token.split(".")
    # v218.64 : format moderne 5 segments (avec sid), legacy 4 segments
    sid = None
    try:
        if len(parts) == 5:
            uid_s, iat_s, tv_s, sid_s, sig = parts
            uid = int(uid_s); iat = int(iat_s); tv = int(tv_s); sid = int(sid_s)
            payload = f"{uid}.{iat}.{tv}.{sid}"
        elif len(parts) == 4:
            uid_s, iat_s, tv_s, sig = parts
            uid = int(uid_s); iat = int(iat_s); tv = int(tv_s)
            payload = f"{uid}.{iat}.{tv}"
        else:
            return None
    except (ValueError, TypeError):
        return None
    # Vérif signature
    import hmac
    if not hmac.compare_digest(sig, _sign(payload)):
        return None
    # Vérif expiration
    if time.time() - iat > TOKEN_TTL:
        return None
    # Vérif utilisateur + token_version
    db = get_db()
    u = one(db.execute("SELECT * FROM utilisateurs WHERE id=? AND actif=1", (uid,)))
    if not u:
        return None
    current_tv = u.get("token_version") or 0
    if current_tv != tv:
        return None
    # v218.64 : résoudre societe_id (token moderne) ou fallback (token legacy)
    if sid is None:
        try:
            r = one(db.execute(
                "SELECT societe_id FROM utilisateur_societes WHERE utilisateur_id=? AND actif=1 ORDER BY societe_id LIMIT 1",
                (uid,)
            ))
            sid = (r and r["societe_id"]) or 1
        except Exception:
            sid = 1
    # Vérifier que l'user a effectivement accès à cette société
    try:
        access = one(db.execute(
            "SELECT role FROM utilisateur_societes WHERE utilisateur_id=? AND societe_id=? AND actif=1",
            (uid, sid)
        ))
        if not access:
            return None
        # v218.64 : exposer le rôle DANS LA SOCIÉTÉ ACTIVE (peut différer de utilisateurs.role)
        # v218.66 : sauf si utilisateurs.role = 'superadmin' (rôle transversal prioritaire)
        u["societe_id"] = sid
        global_role = u.get("role")
        if global_role == "superadmin":
            # Superadmin : on conserve role='superadmin' partout (utilisé par is_superadmin)
            u["societe_role"] = "superadmin"
            # u["role"] reste à 'superadmin' (déjà)
        else:
            u["societe_role"] = access["role"]
            u["role"] = access["role"]  # request.user.role utilisé partout dans le code
    except Exception:
        u["societe_id"] = sid
        u["societe_role"] = u.get("role")
    return u

def _authenticate():
    """Lit l'Authorization header, valide token moderne OU legacy.
    Retourne l'utilisateur ou None.
    v218.99 : accepte aussi ?token=... en query string (pour <img src=...> qui
    ne peut pas envoyer de header Authorization)."""
    auth = request.headers.get("Authorization", "")
    token = None
    if auth.startswith("Bearer "):
        token = auth[7:]
    else:
        # Fallback : token en query string (pour images embed)
        token = request.args.get("token", "")
    if not token:
        return None
    # v218.64 : format moderne 5 segments (avec societe_id), legacy 4 segments
    if token.count(".") in (3, 4):
        return verify_token(token)
    # Legacy : ancien format "uid:hash" — accepté pour compat ascendante
    if ":" in token:
        parts = token.split(":")
        if len(parts) == 2:
            uid = to_int(parts[0])
            if uid:
                db = get_db()
                u = one(db.execute("SELECT * FROM utilisateurs WHERE id=? AND actif=1",(uid,)))
                if u and hp(f"{uid}:{u['email']}:{u['password']}") == parts[1]:
                    # v218.64 : résoudre la 1re société de l'user pour le legacy
                    try:
                        r = one(db.execute(
                            "SELECT societe_id, role FROM utilisateur_societes WHERE utilisateur_id=? AND actif=1 ORDER BY societe_id LIMIT 1",
                            (uid,)
                        ))
                        u["societe_id"] = (r and r["societe_id"]) or 1
                        # v218.66 : superadmin global prioritaire
                        if u.get("role") == "superadmin":
                            u["societe_role"] = "superadmin"
                        else:
                            societe_role = (r and r["role"]) or u.get("role")
                            u["societe_role"] = societe_role
                            u["role"] = societe_role
                    except Exception:
                        u["societe_id"] = 1
                        u["societe_role"] = u.get("role")
                    return u
    return None

def require_auth(fn):
    @wraps(fn)
    def wrapper(*a, **k):
        u = _authenticate()
        if not u:
            return jsonify({"error":"Non authentifie"}), 401
        request.user = u
        return fn(*a, **k)
    return wrapper

def require_role(*roles):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*a, **k):
            u = _authenticate()
            if not u:
                return jsonify({"error":"Non authentifie"}), 401
            # v218.80 : superadmin a tous les droits (transversal)
            if u["role"] != "superadmin" and u["role"] not in roles:
                return jsonify({"error":"Acces refuse"}), 403
            request.user = u
            return fn(*a, **k)
        return wrapper
    return decorator

def criticite(eid,db):
    p=one(db.execute("SELECT COUNT(*) AS n FROM pieces WHERE equipement_id=? AND statut='A_REMPLACER'",(eid,)))
    if p and p["n"]>0: return "CRITIQUE"
    p2=one(db.execute("SELECT COUNT(*) AS n FROM pieces WHERE equipement_id=? AND statut='A_SURVEILLER'",(eid,)))
    if p2 and p2["n"]>0: return "A_SURVEILLER"
    return "OK"

def statut_piece(fdv_):
    if not fdv_: return "OK"
    try:
        d=(datetime.strptime(fdv_,"%Y-%m-%d").date()-date.today()).days
        return "A_REMPLACER" if d<0 else "A_SURVEILLER" if d<=365 else "OK"
    except: return "OK"

# v217.10 : helper pour ajouter N années à une date.
# Calcul simple : on prend l'année et on ajoute N. Le mois et jour deviennent 01-01.
# 2020-* + 7 ans = 2027-01-01
def _add_years(dt, n):
    """Ajoute n années à la date dt. Renvoie une date au 1er janvier de l'année calculée."""
    return dt.replace(year=dt.year + n, month=1, day=1)

def next_numero(db,prefix,table,col):
    r=one(db.execute(f"SELECT MAX(CAST(SUBSTR({col},LENGTH(?)+1) AS INTEGER)) AS m FROM {table} WHERE {col} LIKE ?",(prefix,prefix+"%")))
    return f"{prefix}{((r['m'] or 0)+1):05d}"

def send_mail(to,subj,body,attachments=None,html=False):
    try:
        db=get_db(); cfg=one(db.execute("SELECT * FROM smtp_config WHERE id=1"))
        if not cfg or not cfg["enabled"]: return
        msg=MIMEMultipart(); msg["From"]=f"{cfg['sender_name']} <{cfg['sender_email']}>"
        msg["To"]=to; msg["Subject"]=subj
        # v218.39 : support HTML (pour le récap causerie avec couleurs rouge/vert)
        msg.attach(MIMEText(body,"html" if html else "plain","utf-8"))
        if attachments:
            from email.mime.base import MIMEBase; from email import encoders
            for fname,fdata in attachments:
                part=MIMEBase('application','octet-stream'); part.set_payload(fdata)
                encoders.encode_base64(part)
                part.add_header('Content-Disposition',f'attachment; filename="{fname}"')
                msg.attach(part)
        with smtplib.SMTP(cfg["host"],cfg["port"],timeout=10) as s:
            if cfg["use_tls"]: s.starttls()
            if cfg["username"]: s.login(cfg["username"],cfg["password"])
            s.sendmail(cfg["sender_email"],to,msg.as_string())
        try:
            db.execute("INSERT INTO mail_log (destinataire,sujet,statut) VALUES (?,?,?)",(to,subj,"ENVOYE"))
            db.commit()
        except Exception as e_log:
            logger.warning(f"mail_log insert failed: {e_log}")
    except Exception as e:
        logger.error(f"[MAIL] echec envoi vers {to}: {e}")
        try:
            db2=get_db()
            db2.execute("INSERT INTO mail_log (destinataire,sujet,statut,erreur) VALUES (?,?,?,?)",(to,subj,"ERREUR",str(e)))
            db2.commit()
        except Exception as e_log:
            logger.warning(f"mail_log error insert failed: {e_log}")

def log_action(user, action, entity_type="", entity_id=None, entity_label="", details=""):
    """Enregistre une action utilisateur dans la table logs pour audit.
    
    Args:
        user: dict utilisateur (request.user) ou None pour actions anonymes
        action: code action (LOGIN, CREATE, UPDATE, DELETE, etc.)
        entity_type: type d'entité concernée (intervention, utilisateur, projet, etc.)
        entity_id: id de l'entité
        entity_label: libellé lisible de l'entité (ex: nom, numéro)
        details: texte libre ou JSON pour info complémentaire
    """
    try:
        db = get_db()
        uid = (user or {}).get("id")
        unom = (user or {}).get("nom", "") or ""
        db.execute(
            "INSERT INTO logs (user_id, user_nom, action, entity_type, entity_id, entity_label, details) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (uid, unom, action, entity_type or "", entity_id, entity_label or "", details or "")
        )
        db.commit()
    except Exception as e:
        logger.warning(f"[log_action] échec: {e}")

def notify(iid,action):
    try:
        db=get_db()
        r=one(db.execute("""SELECT i.numero,i.type,i.statut,e.designation,p.nom AS projet,
               p.numero_projet,p.id AS projet_id,p.logo_filename,
               u.email AS mgr_email,u.nom AS mgr_nom,i.date_realisation,
               i.technicien_id, t.email AS tech_email, t.nom AS tech_nom
               FROM interventions i JOIN equipements e ON i.equipement_id=e.id
               JOIN projets p ON e.projet_id=p.id
               LEFT JOIN utilisateurs u ON p.manager_id=u.id
               LEFT JOIN utilisateurs t ON i.technicien_id=t.id
               WHERE i.id=?""",(iid,)))
        if not r: return
        # Construire la liste des destinataires (manager projet + technicien principal + intervenants des CR)
        recipients = []  # liste de tuples (email, nom)
        seen_emails = set()
        def _add(email, nom):
            if not email: return
            e_low = email.lower().strip()
            if e_low in seen_emails: return
            seen_emails.add(e_low)
            recipients.append((email, nom or ""))
        # Manager projet (destinataire principal historique)
        _add(r.get("mgr_email"), r.get("mgr_nom"))
        # Technicien principal de l'intervention
        _add(r.get("tech_email"), r.get("tech_nom"))
        # Intervenants présents dans les CR
        try:
            ivs_cr = rows(db.execute("""
                SELECT DISTINCT u.email, COALESCE(u.nom, ci.nom, '') AS nom
                FROM cr_intervenants ci
                JOIN comptes_rendus cr ON ci.cr_id=cr.id
                LEFT JOIN utilisateurs u ON ci.utilisateur_id=u.id
                WHERE cr.intervention_id=? AND u.email IS NOT NULL AND u.email!=''
            """, (iid,)))
            for ic in ivs_cr:
                _add(ic.get("email"), ic.get("nom"))
        except Exception as ce:
            logger.warning(f"[NOTIFY] récupération intervenants CR échouée: {ce}")
        if not recipients: return
        attachments=[]
        if r["statut"]=="TERMINEE":
            try:
                from rapport_pdf import generate_rapport
                # Récupère intervention complète (description, localisation, technique)
                iv_full = one(db.execute("""SELECT i.*, e.localisation, e.marque, e.modele, e.technique_id, e.type_technique,
                    e.trafo_marque, e.trafo_annee, e.trafo_numero_serie, e.trafo_puissance_kva,
                    e.trafo_refroidissement, e.trafo_poids_kg, e.trafo_tension_entree_v,
                    e.trafo_courant_a, e.trafo_norme, e.trafo_couplage,
                    e.trafo_tension_service_v, e.trafo_reglage_tension_kv,
                    et.nom AS tableau_nom, et.localisation AS tableau_localisation
                    FROM interventions i JOIN equipements e ON e.id=i.equipement_id
                    LEFT JOIN equipement_tableaux et ON i.tableau_id=et.id
                    WHERE i.id=?""", (iid,)))
                iv_full = iv_full or {}
                crs=rows(db.execute("SELECT * FROM comptes_rendus WHERE intervention_id=? ORDER BY created_at",(iid,)))
                for cr in crs:
                    cr["intervenants"]=rows(db.execute("""
                        SELECT ci.*,COALESCE(u2.nom,ci.nom,'') AS nom
                        FROM cr_intervenants ci LEFT JOIN utilisateurs u2 ON ci.utilisateur_id=u2.id
                        WHERE ci.cr_id=? ORDER BY ci.id""",(cr["id"],)))
                # Construction de la liste structurée des CRs
                comptes_rendus_pdf=[]
                for cr in crs:
                    ivs=[]
                    for iv in cr.get("intervenants",[]):
                        nom2=iv.get("nom","") or ""
                        if not nom2 and iv.get("utilisateur_id"):
                            u2=one(db.execute("SELECT nom FROM utilisateurs WHERE id=?",(iv["utilisateur_id"],)))
                            if u2: nom2=u2["nom"]
                        ivs.append({
                            "nom": nom2 or "—",
                            "date": iv.get("date","") or cr.get("date_intervention",""),
                            "heure_debut": iv.get("heure_debut",""),
                            "heure_fin": iv.get("heure_fin",""),
                            "total_heures": iv.get("total_heures",0)
                        })
                    comptes_rendus_pdf.append({
                        "date": cr.get("date_intervention","") or "—",
                        "numero": cr.get("numero","") or "",
                        "observations": cr.get("observations","") or "",
                        "intervenants": ivs,
                        "photos": [bytes(p["data"]) for p in rows(db.execute("SELECT data FROM cr_photos WHERE cr_id=? ORDER BY id",(cr["id"],)))]
                    })
                # Technicien principal (fallback)
                tech_nom_r=one(db.execute("SELECT nom FROM utilisateurs WHERE id=?",(r.get("technicien_id","0"),)))
                tech_nom=tech_nom_r["nom"] if tech_nom_r else "—"
                # Liste des intervenants uniques
                noms_set=set(); noms_list=[]
                for cr in comptes_rendus_pdf:
                    for iv in cr.get("intervenants",[]):
                        n=iv.get("nom","")
                        if n and n!="—" and n not in noms_set:
                            noms_set.add(n); noms_list.append(n)
                if not noms_list: noms_list=[tech_nom]
                # Date
                date_iv = (comptes_rendus_pdf[0]["date"] if comptes_rendus_pdf else (r.get("date_realisation") or today()))
                date_iv = (date_iv or "").split("\n")[0][:10]
                # Équipement enrichi
                marque_modele = f"{iv_full.get('marque','')} {iv_full.get('modele','')}".strip() or "—"
                # Technique nom
                technique_nom = "—"
                if iv_full.get("technique_id"):
                    tr_row = one(db.execute("SELECT nom FROM techniques WHERE id=?", (iv_full["technique_id"],)))
                    if tr_row: technique_nom = tr_row["nom"]
                # Gamme de maintenance : pour BP uniquement, liste des sous-eq avec date de leur CR
                gamme_maintenance = []
                if r.get("type") == "MAINTENANCE":
                    eq_id_bp = iv_full.get("equipement_id")
                    if eq_id_bp:
                        sub_list = rows(db.execute(
                            "SELECT id, nom, localisation FROM equipement_tableaux WHERE equipement_id=? ORDER BY ordre, id",
                            (eq_id_bp,)
                        ))
                        for s in sub_list:
                            done_row = one(db.execute("""
                                SELECT cr.numero, cr.date_intervention FROM cr_tableaux ct
                                JOIN comptes_rendus cr ON ct.cr_id=cr.id
                                WHERE ct.tableau_id=? AND cr.intervention_id=?
                                ORDER BY cr.date_intervention, cr.id LIMIT 1
                            """, (s["id"], iid)))
                            gamme_maintenance.append({
                                "nom": s["nom"],
                                "localisation": s.get("localisation","") or "",
                                "date_realisation": (done_row.get("date_intervention") if done_row else "") or "",
                                "cr_numero": (done_row.get("numero") if done_row else "") or "",
                            })
                # Liste des opérations de la gamme (pour MAINTENANCE) — affichée en checklist dans le PDF
                gamme_operations = []
                if r.get("type") == "MAINTENANCE":
                    eq_id_bp = iv_full.get("equipement_id")
                    if eq_id_bp:
                        # Récupérer toutes les gammes liées à l'équipement (table equipement_gammes)
                        gids_rows = rows(db.execute(
                            "SELECT gamme_id FROM equipement_gammes WHERE equipement_id=?", (eq_id_bp,)
                        ))
                        gids = [g["gamme_id"] for g in gids_rows]
                        # Fallback : ancienne liaison via equipements.gamme_id
                        eq_row = one(db.execute("SELECT gamme_id FROM equipements WHERE id=?", (eq_id_bp,)))
                        if eq_row and eq_row.get("gamme_id") and eq_row["gamme_id"] not in gids:
                            gids.append(eq_row["gamme_id"])
                        for gid in gids:
                            g_row = one(db.execute("SELECT nom FROM gammes WHERE id=?", (gid,)))
                            ops = rows(db.execute(
                                "SELECT id, description FROM gamme_operations WHERE gamme_id=? ORDER BY ordre, id", (gid,)
                            ))
                            ops_with_status = []
                            for op in ops:
                                done_row = one(db.execute("""
                                    SELECT io.date_realisation, u.nom AS tech_nom, cr.numero AS cr_num
                                    FROM intervention_operations io
                                    LEFT JOIN utilisateurs u ON io.technicien_id = u.id
                                    LEFT JOIN comptes_rendus cr ON io.cr_id = cr.id
                                    WHERE io.intervention_id=? AND io.gamme_operation_id=?
                                """, (iid, op["id"])))
                                ops_with_status.append({
                                    "description": op["description"],
                                    "done": bool(done_row),
                                    "date_realisation": (done_row.get("date_realisation") if done_row else "") or "",
                                    "technicien_nom": (done_row.get("tech_nom") if done_row else "") or "",
                                    "cr_numero": (done_row.get("cr_num") if done_row else "") or "",
                                })
                            if ops_with_status:
                                gamme_operations.append({
                                    "gamme_nom": (g_row.get("nom") if g_row else "") or "",
                                    "operations": ops_with_status
                                })
                # Chemin absolu du logo projet s'il existe
                _projet_logo_path = ""
                _logo_fname = r.get("logo_filename") or ""
                if _logo_fname:
                    _candidate = BASE_DIR / "uploads" / "projet_logos" / _logo_fname
                    if _candidate.exists():
                        _projet_logo_path = str(_candidate)
                # Sous-type maintenance déduit des gammes du bon
                _sous_type_label = _bon_sous_type_label(iid)
                # Mesures techniques (UPS×Entretien, etc.)
                _mesures_techniques = _intervention_mesures_data(iid)
                pdf_bytes = generate_rapport({
                    "type_label": r["type"],
                    "sous_type": _sous_type_label,
                    "mesures_techniques": _mesures_techniques,
                    "numero_iv": r["numero"],
                    "date": date_iv,
                    "client": r["projet"],
                    "numero_projet": r.get("numero_projet","") or "—",
                    "projet_nom": r.get("projet","") or "",
                    "projet_logo_path": _projet_logo_path,
                    "equipement": r["designation"],
                    "marque_modele": marque_modele,
                    "localisation": iv_full.get("localisation","") or "—",
                    "tableau": (iv_full.get("tableau_nom","") + (" (" + iv_full.get("tableau_localisation","") + ")" if iv_full.get("tableau_localisation") else "")) if iv_full.get("tableau_nom") else "",
                    "tableau_label": ("Borne" if (iv_full.get("type_technique") or "").lower()=="borne de charge" else "Tableau"),
                    "trafo": ({
                        "marque": iv_full.get("trafo_marque","") or "",
                        "annee": iv_full.get("trafo_annee","") or "",
                        "numero_serie": iv_full.get("trafo_numero_serie","") or "",
                        "puissance_kva": iv_full.get("trafo_puissance_kva","") or "",
                        "refroidissement": iv_full.get("trafo_refroidissement","") or "",
                        "poids_kg": iv_full.get("trafo_poids_kg","") or "",
                        "tension_entree_v": iv_full.get("trafo_tension_entree_v","") or "",
                        "courant_a": iv_full.get("trafo_courant_a","") or "",
                        "norme": iv_full.get("trafo_norme","") or "",
                        "couplage": iv_full.get("trafo_couplage","") or "",
                        "tension_service_v": iv_full.get("trafo_tension_service_v","") or "",
                        "reglage_tension_kv": iv_full.get("trafo_reglage_tension_kv","") or "",
                    }) if (iv_full.get("type_technique") or "").lower()=="haute tension" else None,
                    "cellules": [],  # v218.160 : table supprimée, utiliser sous_equipements
                    # v218.158 : sous-équipements universels (toutes techniques)
                    "sous_equipements": (rows(db.execute("SELECT designation, marque, type, nombre FROM equipement_sous_equipements WHERE equipement_id=? ORDER BY ordre, id", (iv_full.get("equipement_id"),))) if iv_full.get("equipement_id") else []),
                    "securite_items": (rows(db.execute("""
                        SELECT se.id, se.libelle, se.photo_data, se.photo_mime,
                               isec.present AS present,
                               isec.conforme AS conforme
                        FROM securite_equipements se
                        LEFT JOIN intervention_securite isec ON isec.securite_equipement_id=se.id AND isec.intervention_id=?
                        WHERE se.actif=1
                        ORDER BY se.ordre, se.id
                    """, (iid,))) if (iv_full.get("type_technique") or "").lower()=="haute tension" else []),
                    "technique": technique_nom,
                    "intervenants": ", ".join(noms_list),
                    "description": iv_full.get("description","") or "",
                    "comptes_rendus": comptes_rendus_pdf,
                    "gamme_maintenance": gamme_maintenance,
                    "gamme_operations": gamme_operations,
                    "is_maintenance": (r.get("type") == "MAINTENANCE"),
                })
                attachments.append((f"Rapport_{r['numero']}.pdf",pdf_bytes))
            except Exception as pe: logger.warning(f"[PDF] generation rapport echouee: {pe}")
        # Envoi à tous les destinataires (manager + tech + intervenants)
        subject = f"[GMAO] {action} - {r['numero']}"
        for (email, nom) in recipients:
            try:
                body = f"Bonjour {nom or ''},\n\n{action}\n\nRef: {r['numero']}\nType: {r['type']}\nStatut: {r['statut']}\nEquipement: {r['designation']}\nProjet: {r['projet']}\n\nCordialement,\nSOCOM GMAO"
                send_mail(email, subject, body, attachments=attachments or None)
            except Exception as se:
                logger.warning(f"[NOTIFY] envoi à {email} échoué: {se}")
    except Exception as e: logger.error(f"[NOTIFY] echec iid={iid}: {e}")

# ══════════════════════════════════════════════════════════════════════
# RBAC — Modules, actions, et grille par défaut
# ══════════════════════════════════════════════════════════════════════
RBAC_MODULES = [
    ("projets",       "Projets",                ["read", "create", "update", "delete"]),
    ("equipements",   "Équipements",            ["read", "create", "update", "delete"]),
    ("interventions", "Interventions / Bons",   ["read", "create", "update", "delete"]),
    ("pieces",        "Pièces / Stock",         ["read", "create", "update", "delete"]),
    ("mouvements",    "Mouvements de stock",    ["read", "create", "update", "delete"]),
    ("gammes",        "Gammes de maintenance",  ["read", "create", "update", "delete"]),
    ("planning",      "Plannings & Occupations", ["read", "create", "update", "delete"]),
    ("astreintes",    "Astreintes",             ["read", "create", "update", "delete"]),
    ("conges",        "Demandes de congés",     ["read", "create", "update", "delete"]),
    ("pointage",      "Pointage",               ["read", "create", "update", "delete"]),
    ("utilisateurs",  "Utilisateurs",           ["read", "create", "update", "delete"]),
    ("clients",       "Clients",                ["read", "create", "update", "delete"]),
    ("statistiques",  "Statistiques",           ["read"]),
    ("parametres",    "Paramètres système",     ["read", "create", "update", "delete"]),
]
# Comportement par défaut (basé sur l'état actuel du code)
# Format : { role_code: { module: [actions autorisées] } }
RBAC_DEFAULTS = {
    "admin": "*",  # tout autorisé (compute via RBAC_MODULES)
    "manager": {
        "projets":       ["read", "create", "update"],
        "equipements":   ["read", "create", "update"],
        "interventions": ["read", "create", "update"],
        "pieces":        ["read", "create", "update"],
        "mouvements":    ["read", "create", "update"],
        "gammes":        ["read", "create", "update"],
        "planning":      ["read", "create", "update", "delete"],
        "astreintes":    ["read", "create", "update", "delete"],
        "conges":        ["read", "create", "update"],
        "pointage":      ["read"],
        "utilisateurs":  [],
        "clients":       ["read", "create", "update"],
        "statistiques":  ["read"],
        "parametres":    [],
    },
    "technicien": {
        "projets":       ["read"],
        "equipements":   ["read"],
        "interventions": ["read", "create", "update"],
        "pieces":        ["read"],
        "mouvements":    ["read", "create"],
        "gammes":        ["read"],
        "planning":      ["read"],
        "astreintes":    ["read"],
        "conges":        ["read", "create"],
        "pointage":      ["read"],
        "utilisateurs":  [],
        "clients":       ["read"],
        "statistiques":  [],
        "parametres":    [],
    },
    "acl": {  # Lecture seule (client en consultation)
        m[0]: ["read"] for m in []  # rempli ci-dessous
    },
}
# Compléter "acl" : tout en lecture seule
RBAC_DEFAULTS["acl"] = {m[0]: ["read"] for m in RBAC_MODULES}

def _init_roles_and_permissions(db):
    """Crée les rôles built-in et leur grille de permissions par défaut.
    Idempotent : si un rôle existe déjà, on ne touche pas à ses permissions."""
    builtin_roles = [
        ("superadmin", "Super-administrateur", "superadmin"),  # v218.65 multi-tenant
        ("admin",      "Administrateur",  "admin"),
        ("manager",    "Manager",         "manager"),
        ("technicien", "Technicien",      "technicien"),
        ("acl",        "Client (lecture)", "acl"),
    ]
    for code, label, parent in builtin_roles:
        existing = one(db.execute("SELECT code FROM roles WHERE code=?", (code,)))
        if not existing:
            db.execute(
                "INSERT INTO roles (code, label, parent_role, builtin) VALUES (?, ?, ?, 1)",
                (code, label, parent)
            )
            # Insérer les permissions par défaut pour ce rôle
            defaults = RBAC_DEFAULTS.get(code, {})
            for module_code, _label, actions in RBAC_MODULES:
                for action in actions:
                    if defaults == "*":
                        allowed = 1
                    else:
                        allowed = 1 if action in (defaults.get(module_code) or []) else 0
                    db.execute(
                        "INSERT OR IGNORE INTO permissions (role_code, module, action, allowed) VALUES (?, ?, ?, ?)",
                        (code, module_code, action, allowed)
                    )
            logger.info(f"[RBAC] rôle '{code}' initialisé avec permissions par défaut")
    db.commit()

def _user_role_code(u):
    """Retourne le code du rôle d'un utilisateur. Le code peut être un built-in
    (admin/manager/technicien/acl) ou un rôle personnalisé."""
    return (u or {}).get("role") or "technicien"

def _resolve_parent_role(role_code):
    """Retourne le rôle parent (admin/manager/technicien/acl) d'un rôle donné.
    Pour les built-in, retourne le code tel quel.
    Pour les rôles personnalisés, lit la table roles."""
    if role_code in ("admin", "manager", "technicien", "acl"):
        return role_code
    db = get_db()
    r = one(db.execute("SELECT parent_role FROM roles WHERE code=?", (role_code,)))
    return (r or {}).get("parent_role") or "technicien"

def _user_effective_role(u):
    """Pour le code legacy qui fait `if u['role'] == 'manager'`, on retourne
    le rôle parent. Permet aux rôles personnalisés de fonctionner sans casser
    le code existant qui repose sur les checks d'ownership."""
    return _resolve_parent_role(_user_role_code(u))

def _perms_for_role(role_code):
    """Retourne la grille de permissions pour un rôle donné.
    Format : { module: { action: bool } }"""
    db = get_db()
    rows = list(db.execute(
        "SELECT module, action, allowed FROM permissions WHERE role_code=?",
        (role_code,)
    ))
    grid = {}
    for r in rows:
        m = r["module"]; a = r["action"]; v = bool(r["allowed"])
        grid.setdefault(m, {})[a] = v
    return grid

def _perm_check(user, module, action):
    """Vérifie si l'utilisateur a la permission pour module.action.
    Admin a toujours tout. Pour les autres rôles, lit la grille de permissions."""
    if not user:
        return False
    role_code = _user_role_code(user)
    # L'admin built-in a toujours tout (sécurité par défaut)
    if role_code == "admin":
        return True
    # Pour les autres rôles, lire la grille
    db = get_db()
    r = one(db.execute(
        "SELECT allowed FROM permissions WHERE role_code=? AND module=? AND action=?",
        (role_code, module, action)
    ))
    return bool(r and r.get("allowed"))


# ═══════════════════════════════════════════════════════════════════════════
# MULTI-TENANT PHASE 1 — Fondations DB (v218.62)
# ═══════════════════════════════════════════════════════════════════════════
# Tables métier directes qui reçoivent une colonne `societe_id`.
# Les tables filles (cr_intervenants, intervention_creneaux, equipement_gammes,
# etc.) héritent de leur parent via JOIN — pas de societe_id direct.
_MULTITENANT_TABLES = [
    "clients", "projets", "equipements", "interventions", "comptes_rendus",
    "pieces", "gammes", "techniques", "equipes", "plannings",
    "compteur_types", "compteurs", "occupation_types", "occupations",
    "causeries", "causeries_questions",
    "mesure_blocs", "equipement_statuts", "image_library",
    "maintenance_sous_types", "postes", "securite_equipements",
    "astreinte_specialites", "astreinte_planning", "techniciens_astreinte",
    "parametres_rapport", "analyses_rapport", "graphiques_config",
    "mail_log", "demandes_conges",
    "smtp_config", "parametres_app",
]


def _column_exists(db, table, column):
    """Vérifie si une colonne existe dans une table SQLite."""
    try:
        cols = rows(db.execute(f"PRAGMA table_info({table})"))
        return any(c["name"] == column for c in cols)
    except Exception:
        return False


def _table_exists(db, table):
    try:
        r = one(db.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)
        ))
        return bool(r)
    except Exception:
        return False


def _migrate_multitenant_phase1(db):
    """Migration multi-tenant phase 1.

    1. Crée la table `societes` et y insère SOCOM (id=1) si elle n'existe pas
    2. Crée la table de liaison `utilisateur_societes` et y migre tous les users existants
    3. Ajoute la colonne `societe_id` (NOT NULL DEFAULT 1) sur toutes les tables métier

    Cette migration est entièrement idempotente : peut tourner plusieurs fois.
    Aucun changement fonctionnel : ces colonnes ne sont pas encore lues par le reste
    de l'application. La phase 2 activera le filtrage côté API.
    """
    try:
        # ─── 1. Table societes ────────────────────────────────────────────
        db.execute("""
            CREATE TABLE IF NOT EXISTS societes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nom TEXT NOT NULL,
                code TEXT UNIQUE,
                logo_path TEXT DEFAULT '',
                couleur_primaire TEXT DEFAULT '#0F1E3D',
                adresse TEXT DEFAULT '',
                telephone TEXT DEFAULT '',
                email TEXT DEFAULT '',
                site_web TEXT DEFAULT '',
                config_json TEXT DEFAULT '{}',
                actif INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Insertion SOCOM si la table est vide
        r = one(db.execute("SELECT COUNT(*) AS n FROM societes"))
        if r and r["n"] == 0:
            db.execute("""
                INSERT INTO societes (id, nom, code, couleur_primaire, actif)
                VALUES (1, 'SOCOM', 'socom', '#0F1E3D', 1)
            """)
            logger.info("[multitenant-phase1] societes : SOCOM créée (id=1)")

        # ─── 2. Table utilisateur_societes (liaison N:N user ↔ societe) ──
        db.execute("""
            CREATE TABLE IF NOT EXISTS utilisateur_societes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                utilisateur_id INTEGER NOT NULL REFERENCES utilisateurs(id) ON DELETE CASCADE,
                societe_id INTEGER NOT NULL REFERENCES societes(id) ON DELETE CASCADE,
                role TEXT NOT NULL DEFAULT 'technicien',
                manager_id INTEGER REFERENCES utilisateurs(id) ON DELETE SET NULL,
                techniques TEXT DEFAULT '',
                actif INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(utilisateur_id, societe_id)
            )
        """)
        db.execute("CREATE INDEX IF NOT EXISTS idx_us_user ON utilisateur_societes(utilisateur_id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_us_societe ON utilisateur_societes(societe_id)")

        # Migration des utilisateurs existants vers la liaison SOCOM
        # Chaque user qui n'a pas encore de ligne dans utilisateur_societes est ajouté à SOCOM
        # avec son rôle, manager_id et techniques actuels (copiés depuis utilisateurs).
        existing_users = rows(db.execute("""
            SELECT u.id, u.role, u.manager_id, u.techniques
            FROM utilisateurs u
            WHERE NOT EXISTS (
                SELECT 1 FROM utilisateur_societes us
                WHERE us.utilisateur_id = u.id AND us.societe_id = 1
            )
        """))
        if existing_users:
            for u in existing_users:
                db.execute("""
                    INSERT INTO utilisateur_societes
                    (utilisateur_id, societe_id, role, manager_id, techniques, actif)
                    VALUES (?, 1, ?, ?, ?, 1)
                """, (
                    u["id"],
                    u.get("role") or "technicien",
                    u.get("manager_id"),
                    u.get("techniques") or ""
                ))
            logger.info(f"[multitenant-phase1] utilisateur_societes : {len(existing_users)} user(s) liés à SOCOM")

        # ─── 3. Ajout societe_id sur les tables métier ────────────────────
        # NOTE SQLite : ALTER TABLE ADD COLUMN n'accepte pas REFERENCES sur
        # une table non vide ("Cannot add a REFERENCES column with non-NULL
        # default value"). On ajoute donc la colonne sans contrainte FK
        # explicite — la cohérence sera assurée applicativement (et SQLite
        # ne vérifie pas les FK par défaut de toute façon).
        added = []
        for table in _MULTITENANT_TABLES:
            if not _table_exists(db, table):
                # Table pas encore créée à ce stade : on saute, init_db() la créera plus haut
                continue
            if _column_exists(db, table, "societe_id"):
                continue
            try:
                db.execute(f"ALTER TABLE {table} ADD COLUMN societe_id INTEGER NOT NULL DEFAULT 1")
                # Index pour les futures requêtes filtrées
                db.execute(f"CREATE INDEX IF NOT EXISTS idx_{table}_societe ON {table}(societe_id)")
                added.append(table)
            except Exception as e:
                logger.warning(f"[multitenant-phase1] impossible d'ajouter societe_id à {table}: {e}")

        if added:
            logger.info(f"[multitenant-phase1] societe_id ajouté à {len(added)} table(s) : {', '.join(added)}")

        db.commit()
    except Exception as e:
        logger.error(f"[multitenant-phase1] échec migration : {e}", exc_info=True)


# ═══════════════════════════════════════════════════════════════════════════
# MULTI-TENANT PHASE 2 — Permissions de menus par société (v218.72)
# ═══════════════════════════════════════════════════════════════════════════
# Liste des codes de menus configurables par société.
# Dashboard et Paramètres sont TOUJOURS actifs (indispensables).
_MENUS_CONFIGURABLES = [
    "planning", "conges", "pointage", "bilan", "equipe", "qse",
    "clients", "projets", "equipements", "pieces", "interventions",
    "rapport-energie", "astreinte",
]
# Menus essentiels actifs par défaut à la création d'une société
_MENUS_DEFAULT_NEW_SOCIETE = ["interventions", "equipements", "projets"]


def _migrate_multitenant_phase2(db):
    """Migration phase 2 : ajout du champ menus_actifs (JSON) dans societes.
    SOCOM (id=1) : tous les menus activés (rétro-compat).
    Autres sociétés existantes : tous les menus activés également (le superadmin
    peut désactiver via l'UI).
    """
    try:
        if not _column_exists(db, "societes", "menus_actifs"):
            db.execute("ALTER TABLE societes ADD COLUMN menus_actifs TEXT DEFAULT '[]'")
            # Toutes les sociétés existantes : activer tous les menus configurables
            all_menus = json.dumps(_MENUS_CONFIGURABLES)
            db.execute("UPDATE societes SET menus_actifs=? WHERE menus_actifs IS NULL OR menus_actifs='[]'", (all_menus,))
            db.commit()
            logger.info(f"[multitenant-phase2] menus_actifs ajouté avec {len(_MENUS_CONFIGURABLES)} menus actifs par défaut")
    except Exception as e:
        logger.error(f"[multitenant-phase2] échec migration : {e}", exc_info=True)


# ═══════════════════════════════════════════════════════════════════════════
# COVER PAGES (page de garde personnalisée par société) — v218.81
# ═══════════════════════════════════════════════════════════════════════════
# Blocs disponibles pour composer la page de garde :
# - 'title'      : titre principal (texte, taille, alignement, couleur)
# - 'subtitle'   : sous-titre (texte, taille, alignement, couleur)
# - 'image'      : image téléversée (chemin relatif, largeur en cm, alignement)
# - 'info_box'   : encadré d'infos avec champs dynamiques cochés
# - 'text'       : paragraphe de texte libre
# - 'spacer'     : saut de ligne / espace vertical (taille en cm)
# - 'separator'  : ligne horizontale (couleur, épaisseur)
#
# Champs dynamiques pour info_box :
_COVER_DYNAMIC_FIELDS = [
    'client', 'projet', 'numero_iv', 'date', 'technicien',
    'equipement', 'localisation', 'technique', 'sous_type'
]

# v218.85 : champs additionnels disponibles dans header/footer
_HF_DYNAMIC_FIELDS = _COVER_DYNAMIC_FIELDS + [
    'nom_societe', 'page', 'total_pages', 'date_generation'
]

# Page de garde par défaut (utilisée pour SOCOM lors de la migration)
_COVER_DEFAULT_BLOCKS = [
    {"type": "spacer", "height_cm": 1.5},
    {"type": "title", "text": "RAPPORT DE MAINTENANCE",
     "font_size": 24, "color": "#1E3A8A", "align": "center", "bold": True},
    {"type": "subtitle", "text": "{technique}",
     "font_size": 18, "color": "#1E3A8A", "align": "center", "bold": True},
    {"type": "spacer", "height_cm": 1.2},
    {"type": "info_box", "fields": ["client", "projet"], "title": "INFORMATIONS GÉNÉRALES"},
    {"type": "spacer", "height_cm": 1.0},
    {"type": "image", "path": "", "width_cm": 8.0, "align": "center", "use_projet_logo": True},
]

# v218.109 : Page de garde par défaut pour les bons de dépannage (BC)
_COVER_BC_DEFAULT_BLOCKS = [
    {"type": "spacer", "height_cm": 1.5},
    {"type": "title", "text": "RAPPORT D'INTERVENTION",
     "font_size": 24, "color": "#1E3A8A", "align": "center", "bold": True},
    {"type": "subtitle", "text": "Dépannage",
     "font_size": 18, "color": "#1E3A8A", "align": "center", "bold": True},
    {"type": "spacer", "height_cm": 1.2},
    {"type": "info_box", "fields": ["client", "projet", "date", "numero_iv"], "title": "INFORMATIONS GÉNÉRALES"},
    {"type": "spacer", "height_cm": 1.0},
    {"type": "image", "path": "", "width_cm": 8.0, "align": "center", "use_projet_logo": True},
]

# v218.85 : header par défaut (logo société + n° bon/date à droite + trait bleu)
_HEADER_DEFAULT_BLOCKS = [
    {"type": "row", "items": [
        {"col": "left",   "type": "image",  "use_societe_logo": True, "width_cm": 4.5},
        {"col": "right",  "type": "text",   "text": "{numero_iv}  —  {date}",
         "font_size": 11, "color": "#1E3A8A", "bold": True}
    ]},
    {"type": "separator", "color": "#1E3A8A", "thickness": 1.5}
]

# v218.85 : footer par défaut (adresse à gauche + numéro de page à droite)
_FOOTER_DEFAULT_BLOCKS = [
    {"type": "row", "items": [
        {"col": "left",  "type": "text", "text": "{nom_societe}",
         "font_size": 8, "color": "#64748B", "bold": False},
        {"col": "right", "type": "text", "text": "Page {page}",
         "font_size": 8, "color": "#64748B", "bold": False}
    ]}
]


def _migrate_cover_pages(db):
    """Migration : crée la table cover_pages (1 page de garde par société)
    et initialise SOCOM avec un layout par défaut basé sur l'existant.
    v218.85 : ajoute colonnes header_blocks_json et footer_blocks_json.
    v218.109 : crée aussi la table cover_pages_bc pour les bons de dépannage."""
    try:
        db.execute("""
            CREATE TABLE IF NOT EXISTS cover_pages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                societe_id INTEGER NOT NULL UNIQUE,
                blocks_json TEXT NOT NULL DEFAULT '[]',
                updated_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (societe_id) REFERENCES societes(id) ON DELETE CASCADE
            )
        """)
        # v218.85 : ajouter colonnes header_blocks_json et footer_blocks_json si absentes
        cols = [r["name"] for r in db.execute("PRAGMA table_info(cover_pages)").fetchall()]
        if "header_blocks_json" not in cols:
            db.execute("ALTER TABLE cover_pages ADD COLUMN header_blocks_json TEXT DEFAULT '[]'")
            logger.info("[cover-pages] Colonne header_blocks_json ajoutée")
        if "footer_blocks_json" not in cols:
            db.execute("ALTER TABLE cover_pages ADD COLUMN footer_blocks_json TEXT DEFAULT '[]'")
            logger.info("[cover-pages] Colonne footer_blocks_json ajoutée")
        # Initialiser SOCOM (id=1) si elle n'a pas encore de cover_page
        existing = one(db.execute("SELECT id FROM cover_pages WHERE societe_id=1"))
        if not existing:
            db.execute("""INSERT INTO cover_pages
                (societe_id, blocks_json, header_blocks_json, footer_blocks_json)
                VALUES (1, ?, ?, ?)""",
                (json.dumps(_COVER_DEFAULT_BLOCKS),
                 json.dumps(_HEADER_DEFAULT_BLOCKS),
                 json.dumps(_FOOTER_DEFAULT_BLOCKS)))
            logger.info("[cover-pages] SOCOM (id=1) initialisée avec layouts par défaut")
        else:
            # Vérifier si header/footer sont déjà initialisés pour SOCOM
            row = one(db.execute("""SELECT header_blocks_json, footer_blocks_json
                                    FROM cover_pages WHERE societe_id=1"""))
            updates = {}
            if not row.get("header_blocks_json") or row["header_blocks_json"] in ('[]', '', None):
                updates["header_blocks_json"] = json.dumps(_HEADER_DEFAULT_BLOCKS)
            if not row.get("footer_blocks_json") or row["footer_blocks_json"] in ('[]', '', None):
                updates["footer_blocks_json"] = json.dumps(_FOOTER_DEFAULT_BLOCKS)
            if updates:
                set_clause = ", ".join(f"{k}=?" for k in updates)
                db.execute(f"UPDATE cover_pages SET {set_clause} WHERE societe_id=1",
                           tuple(updates.values()))
                logger.info(f"[cover-pages] SOCOM mise à jour : {list(updates.keys())}")

        # v218.109 : Table cover_pages_bc pour les bons de dépannage
        db.execute("""
            CREATE TABLE IF NOT EXISTS cover_pages_bc (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                societe_id INTEGER NOT NULL UNIQUE,
                blocks_json TEXT NOT NULL DEFAULT '[]',
                header_blocks_json TEXT DEFAULT '[]',
                footer_blocks_json TEXT DEFAULT '[]',
                updated_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (societe_id) REFERENCES societes(id) ON DELETE CASCADE
            )
        """)
        # Initialiser SOCOM avec un layout par défaut BC
        existing_bc = one(db.execute("SELECT id FROM cover_pages_bc WHERE societe_id=1"))
        if not existing_bc:
            db.execute("""INSERT INTO cover_pages_bc
                (societe_id, blocks_json, header_blocks_json, footer_blocks_json)
                VALUES (1, ?, ?, ?)""",
                (json.dumps(_COVER_BC_DEFAULT_BLOCKS),
                 json.dumps(_HEADER_DEFAULT_BLOCKS),
                 json.dumps(_FOOTER_DEFAULT_BLOCKS)))
            logger.info("[cover-pages-bc] SOCOM (id=1) initialisée avec layouts BC par défaut")

        db.commit()
    except Exception as e:
        logger.error(f"[cover-pages] échec migration : {e}", exc_info=True)


def init_db():
    db=get_db()
    db.executescript("""
    CREATE TABLE IF NOT EXISTS utilisateurs (
        id INTEGER PRIMARY KEY AUTOINCREMENT, nom TEXT NOT NULL, email TEXT NOT NULL UNIQUE,
        password TEXT NOT NULL, role TEXT NOT NULL DEFAULT 'technicien'
            CHECK(role IN ('admin','manager','technicien','acl')),
        actif INTEGER DEFAULT 1, matricule TEXT DEFAULT '',
        manager_id INTEGER REFERENCES utilisateurs(id) ON DELETE SET NULL,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS clients (
        id INTEGER PRIMARY KEY AUTOINCREMENT, societe TEXT NOT NULL,
        nom TEXT DEFAULT '', prenom TEXT DEFAULT '', email TEXT DEFAULT '',
        telephone TEXT DEFAULT '', notes TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS projets (
        id INTEGER PRIMARY KEY AUTOINCREMENT, numero_projet TEXT DEFAULT '',
        nom TEXT NOT NULL, client_id INTEGER REFERENCES clients(id),
        manager_id INTEGER REFERENCES utilisateurs(id),
        description TEXT DEFAULT '', date_debut TEXT, date_fin TEXT,
        statut TEXT DEFAULT 'EN_COURS', created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS equipements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        projet_id INTEGER NOT NULL REFERENCES projets(id),
        designation TEXT NOT NULL, type_technique TEXT NOT NULL,
        localisation TEXT DEFAULT '', marque TEXT DEFAULT '', modele TEXT DEFAULT '',
        puissance TEXT DEFAULT '', numero_serie TEXT DEFAULT '', in_out TEXT DEFAULT '',
        date_mise_en_service TEXT, statut TEXT NOT NULL DEFAULT 'EN_SERVICE',
        gamme_id INTEGER, technique_id INTEGER,
        notes TEXT, created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS equipement_gammes (
        equipement_id INTEGER NOT NULL, gamme_id INTEGER NOT NULL,
        PRIMARY KEY (equipement_id, gamme_id)
    );
    -- v218.160 : table equipement_cellules SUPPRIMÉE (DROP au démarrage)
    -- equipement_tableaux et cr_tableaux conservées (FK vers interventions/CRs) mais VIDÉES
    -- au démarrage. Aucune nouvelle saisie possible côté UI (v218.159).
    CREATE TABLE IF NOT EXISTS equipement_tableaux (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        equipement_id INTEGER NOT NULL REFERENCES equipements(id) ON DELETE CASCADE,
        nom TEXT NOT NULL,
        localisation TEXT DEFAULT '',
        ordre INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_eq_tableaux_eq ON equipement_tableaux(equipement_id);
    -- v218.157 : Sous-équipements universels (toutes techniques)
    -- Désignation, Marque, Type, Nombre
    CREATE TABLE IF NOT EXISTS equipement_sous_equipements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        equipement_id INTEGER NOT NULL REFERENCES equipements(id) ON DELETE CASCADE,
        designation TEXT NOT NULL,
        marque TEXT DEFAULT '',
        type TEXT DEFAULT '',
        nombre INTEGER DEFAULT 1,
        ordre INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_eq_souseq_eq ON equipement_sous_equipements(equipement_id);
    -- Catalogue d'équipements de sécurité (admin) pour les bons Haute tension
    CREATE TABLE IF NOT EXISTS securite_equipements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        libelle TEXT NOT NULL,
        photo_data TEXT DEFAULT '',
        photo_mime TEXT DEFAULT '',
        ordre INTEGER DEFAULT 0,
        actif INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_sec_eq_ordre ON securite_equipements(ordre);
    -- Statut de chaque équipement de sécurité sur un bon (HT)
    CREATE TABLE IF NOT EXISTS intervention_securite (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        intervention_id INTEGER NOT NULL REFERENCES interventions(id) ON DELETE CASCADE,
        securite_equipement_id INTEGER NOT NULL REFERENCES securite_equipements(id) ON DELETE CASCADE,
        present INTEGER,
        conforme INTEGER,
        updated_at TEXT DEFAULT (datetime('now')),
        UNIQUE(intervention_id, securite_equipement_id)
    );
    CREATE INDEX IF NOT EXISTS idx_iv_sec_iv ON intervention_securite(intervention_id);
    -- v218.160 : table cr_tableaux conservée (FK depuis CRs) mais VIDÉE au démarrage
    CREATE TABLE IF NOT EXISTS cr_tableaux (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cr_id INTEGER NOT NULL REFERENCES comptes_rendus(id) ON DELETE CASCADE,
        tableau_id INTEGER NOT NULL REFERENCES equipement_tableaux(id) ON DELETE CASCADE,
        created_at TEXT DEFAULT (datetime('now')),
        UNIQUE(cr_id, tableau_id)
    );
    CREATE INDEX IF NOT EXISTS idx_cr_tableaux_cr ON cr_tableaux(cr_id);
    CREATE INDEX IF NOT EXISTS idx_cr_tableaux_tab ON cr_tableaux(tableau_id);
    -- Opérations de gamme cochées sur une intervention
    CREATE TABLE IF NOT EXISTS intervention_operations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        intervention_id INTEGER NOT NULL REFERENCES interventions(id) ON DELETE CASCADE,
        gamme_operation_id INTEGER NOT NULL REFERENCES gamme_operations(id) ON DELETE CASCADE,
        cr_id INTEGER REFERENCES comptes_rendus(id) ON DELETE SET NULL,
        technicien_id INTEGER REFERENCES utilisateurs(id) ON DELETE SET NULL,
        date_realisation TEXT DEFAULT (date('now')),
        created_at TEXT DEFAULT (datetime('now')),
        UNIQUE(intervention_id, gamme_operation_id)
    );
    CREATE INDEX IF NOT EXISTS idx_iv_ops_iv ON intervention_operations(intervention_id);
    CREATE TABLE IF NOT EXISTS techniques (
        id INTEGER PRIMARY KEY AUTOINCREMENT, nom TEXT NOT NULL UNIQUE,
        description TEXT DEFAULT '', created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS gammes (
        id INTEGER PRIMARY KEY AUTOINCREMENT, nom TEXT NOT NULL,
        periodicite TEXT NOT NULL, temps TEXT DEFAULT '00h00', created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS gamme_operations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        gamme_id INTEGER NOT NULL REFERENCES gammes(id) ON DELETE CASCADE,
        ordre INTEGER DEFAULT 0, description TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS pieces (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        equipement_id INTEGER NOT NULL REFERENCES equipements(id),
        type_piece TEXT NOT NULL, date_installation TEXT,
        duree_vie_estimee INTEGER, date_fin_de_vie TEXT,
        statut TEXT NOT NULL DEFAULT 'OK' CHECK(statut IN ('OK','A_SURVEILLER','A_REMPLACER')),
        quantite INTEGER DEFAULT 1, numero_serie TEXT DEFAULT '',
        reference TEXT DEFAULT '', commentaire TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    -- v218.27 : historique des remplacements de pièces
    -- Chaque remplacement copie l'état précédent ici, la pièce courante reste dans `pieces`
    CREATE TABLE IF NOT EXISTS pieces_historique (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        piece_id INTEGER NOT NULL REFERENCES pieces(id) ON DELETE CASCADE,
        type_piece TEXT NOT NULL,
        date_installation TEXT,
        duree_vie_estimee INTEGER,
        date_fin_de_vie TEXT,
        statut TEXT,
        quantite INTEGER DEFAULT 1,
        numero_serie TEXT DEFAULT '',
        reference TEXT DEFAULT '',
        commentaire TEXT,
        date_remplacement TEXT NOT NULL,
        motif_categorie TEXT NOT NULL DEFAULT 'AUTRE'
            CHECK(motif_categorie IN ('PANNE','FIN_DE_VIE','PREVENTIF','AUTRE')),
        motif_detail TEXT DEFAULT '',
        remplace_par INTEGER REFERENCES utilisateurs(id),
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_pieces_hist ON pieces_historique(piece_id, date_remplacement DESC);
    -- v218.169 : types de pièces critiques configurables par société
    CREATE TABLE IF NOT EXISTS types_pieces (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        societe_id INTEGER NOT NULL DEFAULT 1,
        nom TEXT NOT NULL,
        is_batterie INTEGER NOT NULL DEFAULT 0,
        ordre INTEGER DEFAULT 0,
        actif INTEGER NOT NULL DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now')),
        UNIQUE(societe_id, nom),
        FOREIGN KEY (societe_id) REFERENCES societes(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_types_pieces_sid ON types_pieces(societe_id, ordre);
    CREATE TABLE IF NOT EXISTS equipes (
        id INTEGER PRIMARY KEY AUTOINCREMENT, nom TEXT NOT NULL,
        manager_id INTEGER REFERENCES utilisateurs(id),
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS equipe_membres (
        equipe_id INTEGER NOT NULL REFERENCES equipes(id),
        technicien_id INTEGER NOT NULL REFERENCES utilisateurs(id),
        PRIMARY KEY (equipe_id, technicien_id)
    );
    CREATE TABLE IF NOT EXISTS interventions (
        id INTEGER PRIMARY KEY AUTOINCREMENT, numero TEXT UNIQUE,
        equipement_id INTEGER NOT NULL REFERENCES equipements(id),
        technicien_id INTEGER REFERENCES utilisateurs(id),
        equipe_id INTEGER REFERENCES equipes(id),
        type TEXT NOT NULL DEFAULT 'MAINTENANCE' CHECK(type IN ('MAINTENANCE','DEPANNAGE')),
        statut TEXT NOT NULL DEFAULT 'PLANIFIEE'
            CHECK(statut IN ('A_PLANIFIER','PLANIFIEE','EN_COURS','TERMINEE','ANNULEE')),
        date_prevue TEXT, date_realisation TEXT,
        description TEXT DEFAULT '', rapport TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS comptes_rendus (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        intervention_id INTEGER NOT NULL REFERENCES interventions(id),
        date_intervention TEXT, observations TEXT DEFAULT '',
        actions_realisees TEXT DEFAULT '', mesures TEXT DEFAULT '',
        recommandations TEXT DEFAULT '', conclusion TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS cr_intervenants (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cr_id INTEGER NOT NULL REFERENCES comptes_rendus(id) ON DELETE CASCADE,
        utilisateur_id INTEGER REFERENCES utilisateurs(id),
        nom TEXT DEFAULT '', date TEXT DEFAULT '',
        heure_debut TEXT DEFAULT '', heure_fin TEXT DEFAULT '',
        total_heures REAL DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS cr_materiels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cr_id INTEGER NOT NULL REFERENCES comptes_rendus(id) ON DELETE CASCADE,
        designation TEXT NOT NULL,
        quantite REAL DEFAULT 1,
        ordre INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_cr_materiels_cr ON cr_materiels(cr_id);
    CREATE TABLE IF NOT EXISTS cr_photos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cr_id INTEGER NOT NULL REFERENCES comptes_rendus(id) ON DELETE CASCADE,
        filename TEXT DEFAULT '',
        mime_type TEXT DEFAULT 'image/jpeg',
        data BLOB NOT NULL,
        size INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_cr_photos_cr ON cr_photos(cr_id);
    CREATE TABLE IF NOT EXISTS push_subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        utilisateur_id INTEGER NOT NULL REFERENCES utilisateurs(id) ON DELETE CASCADE,
        endpoint TEXT NOT NULL UNIQUE,
        p256dh TEXT NOT NULL,
        auth TEXT NOT NULL,
        user_agent TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_push_user ON push_subscriptions(utilisateur_id);
    CREATE TABLE IF NOT EXISTS intervention_creneaux (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        intervention_id INTEGER NOT NULL REFERENCES interventions(id) ON DELETE CASCADE,
        date TEXT NOT NULL,
        heure_debut TEXT DEFAULT '',
        heure_fin TEXT DEFAULT '',
        technicien_id INTEGER REFERENCES utilisateurs(id),
        notes TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_creneaux_intervention ON intervention_creneaux(intervention_id);
    CREATE INDEX IF NOT EXISTS idx_creneaux_date ON intervention_creneaux(date);
    CREATE INDEX IF NOT EXISTS idx_creneaux_tech ON intervention_creneaux(technicien_id);
    CREATE TABLE IF NOT EXISTS intervention_techniciens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        intervention_id INTEGER NOT NULL REFERENCES interventions(id) ON DELETE CASCADE,
        utilisateur_id INTEGER NOT NULL REFERENCES utilisateurs(id),
        UNIQUE(intervention_id, utilisateur_id)
    );
    CREATE INDEX IF NOT EXISTS idx_ivtech_intervention ON intervention_techniciens(intervention_id);
    CREATE INDEX IF NOT EXISTS idx_ivtech_user ON intervention_techniciens(utilisateur_id);
    CREATE TABLE IF NOT EXISTS occupation_types (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL UNIQUE,
        couleur TEXT DEFAULT '#64748b',
        actif INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS occupations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        technicien_id INTEGER NOT NULL REFERENCES utilisateurs(id) ON DELETE CASCADE,
        type_id INTEGER REFERENCES occupation_types(id),
        date TEXT NOT NULL,
        heure_debut TEXT DEFAULT '',
        heure_fin TEXT DEFAULT '',
        total_heures REAL DEFAULT 0,
        notes TEXT DEFAULT '',
        numero_projet TEXT DEFAULT '',
        nom_chantier TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_occ_tech ON occupations(technicien_id);
    CREATE INDEX IF NOT EXISTS idx_occ_date ON occupations(date);
    CREATE INDEX IF NOT EXISTS idx_occ_type ON occupations(type_id);
    -- ═══════════════════════════════════════════════════════════════════
    -- DEMANDES DE CONGÉS (workflow tech → manager)
    -- statut : EN_ATTENTE, APPROUVEE, REFUSEE, ANNULEE
    -- demi_journee : '' (journée), 'MATIN' (08-12), 'APRESMIDI' (12-16)
    -- motif : 'LEGAL' ou 'EXTRAORDINAIRE'
    -- ═══════════════════════════════════════════════════════════════════
    CREATE TABLE IF NOT EXISTS demandes_conges (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        technicien_id INTEGER NOT NULL REFERENCES utilisateurs(id) ON DELETE CASCADE,
        date_debut TEXT NOT NULL,
        date_fin TEXT NOT NULL,
        demi_journee TEXT DEFAULT '',
        motif TEXT DEFAULT 'LEGAL',
        commentaire TEXT DEFAULT '',
        statut TEXT DEFAULT 'EN_ATTENTE',
        manager_id INTEGER REFERENCES utilisateurs(id),
        decision_par INTEGER REFERENCES utilisateurs(id),
        decision_at TEXT DEFAULT '',
        decision_commentaire TEXT DEFAULT '',
        occupation_ids TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_dc_tech ON demandes_conges(technicien_id);
    CREATE INDEX IF NOT EXISTS idx_dc_manager ON demandes_conges(manager_id);
    CREATE INDEX IF NOT EXISTS idx_dc_statut ON demandes_conges(statut);
    CREATE TABLE IF NOT EXISTS smtp_config (
        id INTEGER PRIMARY KEY DEFAULT 1, host TEXT DEFAULT '',
        port INTEGER DEFAULT 587, username TEXT DEFAULT '', password TEXT DEFAULT '',
        sender_email TEXT DEFAULT '', sender_name TEXT DEFAULT 'SOCOM GMAO',
        use_tls INTEGER DEFAULT 1, enabled INTEGER DEFAULT 0
    );
    INSERT OR IGNORE INTO smtp_config (id) VALUES (1);
    CREATE TABLE IF NOT EXISTS plannings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL,
        equipe_id INTEGER REFERENCES equipes(id),
        heures_par_jour REAL DEFAULT 8.0,
        heure_debut TEXT DEFAULT '08:00',
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS parametres_rapport (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        projet_id INTEGER NOT NULL REFERENCES projets(id) ON DELETE CASCADE,
        annee INTEGER NOT NULL,
        chapitre TEXT NOT NULL,
        parametre TEXT DEFAULT '',
        valeur TEXT DEFAULT '',
        ordre INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS analyses_rapport (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        projet_id INTEGER NOT NULL REFERENCES projets(id) ON DELETE CASCADE,
        annee INTEGER NOT NULL,
        type_analyse TEXT NOT NULL,
        texte TEXT DEFAULT '',
        updated_at TEXT,
        UNIQUE(projet_id, annee, type_analyse)
    );
    CREATE TABLE IF NOT EXISTS graphiques_config (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        projet_id INTEGER NOT NULL REFERENCES projets(id) ON DELETE CASCADE,
        titre TEXT NOT NULL DEFAULT 'Graphique',
        ordre INTEGER DEFAULT 0,
        series TEXT DEFAULT '[]',
        comparaison INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS parametres_app (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cle TEXT NOT NULL UNIQUE,
        valeur TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS releves_compteurs (id INTEGER PRIMARY KEY AUTOINCREMENT, cr_id INTEGER NOT NULL REFERENCES comptes_rendus(id) ON DELETE CASCADE, compteur_id INTEGER NOT NULL REFERENCES compteurs(id), valeur REAL, date_releve TEXT, notes TEXT DEFAULT '');
    CREATE TABLE IF NOT EXISTS compteur_types (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL UNIQUE,
        description TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS compteurs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        equipement_id INTEGER NOT NULL REFERENCES equipements(id) ON DELETE CASCADE,
        type_id INTEGER REFERENCES compteur_types(id),
        nom TEXT NOT NULL,
        numero TEXT DEFAULT '',
        unite TEXT DEFAULT '',
        localisation TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS techniciens_astreinte (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL UNIQUE,
        gsm TEXT DEFAULT '',
        couleur TEXT DEFAULT '#3b82f6',
        actif INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS astreinte_specialites (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL UNIQUE,
        description TEXT DEFAULT '',
        ordre INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS astreinte_planning (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL,
        specialite_id INTEGER NOT NULL REFERENCES astreinte_specialites(id),
        technicien TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')),
        UNIQUE(date, specialite_id)
    );
    -- v218.18 : Déclenchements d'astreinte (1-1 avec une intervention de type ASTREINTE)
    CREATE TABLE IF NOT EXISTS astreinte_declenchements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        intervention_id INTEGER NOT NULL UNIQUE REFERENCES interventions(id) ON DELETE CASCADE,
        demandeur_nom TEXT DEFAULT '',
        demandeur_tel TEXT DEFAULT '',
        demandeur_email TEXT DEFAULT '',
        criticite TEXT DEFAULT 'CRITIQUE',
        heure_appel TEXT DEFAULT '',
        heure_arrivee TEXT DEFAULT '',
        heure_fin TEXT DEFAULT '',
        kilometrage REAL DEFAULT 0,
        specialite_id INTEGER REFERENCES astreinte_specialites(id),
        technicien_astreinte TEXT DEFAULT '',
        declenche_par INTEGER REFERENCES utilisateurs(id),
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_astdec_iv ON astreinte_declenchements(intervention_id);
    -- v218.30 : QSE/RSE — Causeries avec QCM
    CREATE TABLE IF NOT EXISTS causeries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        titre TEXT NOT NULL,
        pdf_filename TEXT DEFAULT '',
        frequence TEXT DEFAULT 'MENSUEL'
            CHECK(frequence IN ('HEBDO','BI_MENSUEL','MENSUEL','TRIMESTRIEL','ANNUEL')),
        date_publication TEXT,
        date_cloture TEXT,
        mail_destinataire TEXT DEFAULT '',
        actif INTEGER DEFAULT 1,
        cree_par INTEGER REFERENCES utilisateurs(id),
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS causeries_questions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        causerie_id INTEGER NOT NULL REFERENCES causeries(id) ON DELETE CASCADE,
        ordre INTEGER DEFAULT 0,
        texte TEXT NOT NULL,
        opt_a TEXT DEFAULT '',
        opt_b TEXT DEFAULT '',
        opt_c TEXT DEFAULT '',
        opt_d TEXT DEFAULT '',
        correct_idx INTEGER NOT NULL CHECK(correct_idx IN (0,1,2,3))
    );
    CREATE INDEX IF NOT EXISTS idx_caus_q ON causeries_questions(causerie_id, ordre);
    CREATE TABLE IF NOT EXISTS causeries_reponses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        causerie_id INTEGER NOT NULL REFERENCES causeries(id) ON DELETE CASCADE,
        utilisateur_id INTEGER NOT NULL REFERENCES utilisateurs(id),
        score INTEGER DEFAULT 0,
        total_questions INTEGER DEFAULT 0,
        completed_at TEXT DEFAULT (datetime('now')),
        UNIQUE(causerie_id, utilisateur_id)
    );
    CREATE INDEX IF NOT EXISTS idx_caus_rep ON causeries_reponses(causerie_id, utilisateur_id);
    CREATE TABLE IF NOT EXISTS causeries_reponses_detail (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        reponse_id INTEGER NOT NULL REFERENCES causeries_reponses(id) ON DELETE CASCADE,
        question_id INTEGER NOT NULL REFERENCES causeries_questions(id),
        choix_index INTEGER NOT NULL,
        est_correct INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS mail_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        destinataire TEXT NOT NULL,
        sujet TEXT NOT NULL,
        statut TEXT DEFAULT 'ENVOYE',
        erreur TEXT DEFAULT '',
        intervention_id INTEGER,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS postes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL UNIQUE,
        ordre INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        created_at TEXT DEFAULT (datetime('now')),
        user_id INTEGER,
        user_nom TEXT DEFAULT '',
        action TEXT NOT NULL,
        entity_type TEXT DEFAULT '',
        entity_id INTEGER,
        entity_label TEXT DEFAULT '',
        details TEXT DEFAULT ''
    );
    CREATE INDEX IF NOT EXISTS idx_logs_created ON logs(created_at);
    CREATE INDEX IF NOT EXISTS idx_logs_user ON logs(user_id);
    CREATE INDEX IF NOT EXISTS idx_logs_entity ON logs(entity_type, entity_id);
    """)
    db.commit()

    # ── Migrations idempotentes (remplace les ALTER TABLE éparpillés) ──
    def _add_col(table, col, ddl):
        cols = {r[1] for r in db.execute(f"PRAGMA table_info({table})").fetchall()}
        if col not in cols:
            try:
                db.execute(f"ALTER TABLE {table} ADD COLUMN {col} {ddl}")
                logger.info(f"[MIGRATION] ajout {table}.{col}")
            except Exception as e:
                logger.error(f"[MIGRATION] echec {table}.{col}: {e}")

    # utilisateurs : token_version (revocation) + reset password
    _add_col("utilisateurs", "token_version",  "INTEGER DEFAULT 0")
    _add_col("utilisateurs", "reset_token",    "TEXT")
    _add_col("utilisateurs", "reset_expires",  "TEXT")

    # plannings (session 7-8)
    for col, ddl in [("equipe_id","INTEGER"), ("heures_par_jour","REAL DEFAULT 8.0"),
                     ("heure_debut","TEXT DEFAULT '08:00'")]:
        _add_col("plannings", col, ddl)
    # v218.126 : capacité simultanée par jour (nombre de bons de maintenance qu'on peut
    # planifier sur la même journée pour ce planning). Défaut = 1.
    _add_col("plannings", "nb_simultane", "INTEGER DEFAULT 1")

    # equipements (session 7-8)
    for col, ddl in [("planning_id","INTEGER"),("semaine_planif","INTEGER"),
                     ("jour_semaine_planif","INTEGER"),("intervention_samedi","INTEGER DEFAULT 0"),
                     ("intervention_dimanche","INTEGER DEFAULT 0")]:
        _add_col("equipements", col, ddl)

    # equipements : champs Transformateur (Haute tension uniquement)
    for col, ddl in [
        ("trafo_marque", "TEXT DEFAULT ''"),
        ("trafo_annee", "TEXT DEFAULT ''"),
        ("trafo_numero_serie", "TEXT DEFAULT ''"),
        ("trafo_puissance_kva", "TEXT DEFAULT ''"),
        ("trafo_refroidissement", "TEXT DEFAULT ''"),
        ("trafo_poids_kg", "TEXT DEFAULT ''"),
        ("trafo_tension_entree_v", "TEXT DEFAULT ''"),
        ("trafo_courant_a", "TEXT DEFAULT ''"),
        ("trafo_norme", "TEXT DEFAULT ''"),
        ("trafo_couplage", "TEXT DEFAULT ''"),
        ("trafo_tension_service_v", "TEXT DEFAULT ''"),
        ("trafo_reglage_tension_kv", "TEXT DEFAULT ''"),
    ]:
        _add_col("equipements", col, ddl)

    # equipement_gammes : planif par gamme (session 19+)
    # Chaque gamme associée à un équipement peut avoir sa propre planif.
    # Les colonnes globales sur equipements restent pour compatibilité (ancienne planif).
    for col, ddl in [("planning_id","INTEGER"),("semaine_planif","INTEGER"),
                     ("jour_semaine_planif","INTEGER"),("intervention_samedi","INTEGER DEFAULT 0"),
                     ("intervention_dimanche","INTEGER DEFAULT 0")]:
        _add_col("equipement_gammes", col, ddl)
    # Mode de planification : SEMAINE (semaine ISO fixe, défaut) ou NTH_JOUR_MOIS (ex: 2ème jeudi du mois)
    # nth_semaine_mois : 1..5 (ou -1 pour "dernier") = rang dans le mois
    _add_col("equipement_gammes", "planif_mode", "TEXT DEFAULT 'SEMAINE'")
    _add_col("equipement_gammes", "nth_semaine_mois", "INTEGER")
    # v218.137 : force_planif = 1 → la date prévue n'est PAS décalée automatiquement
    # (ni weekend, ni férié, ni capacité). Permet de planifier sur samedi/dimanche volontairement.
    _add_col("equipement_gammes", "force_planif", "INTEGER DEFAULT 0")

    # interventions.heure_prevue
    _add_col("interventions", "heure_prevue", "TEXT DEFAULT '08:00'")
    # interventions.tableau_id : sous-équipement (tableau électrique) optionnel pour BT
    _add_col("interventions", "tableau_id", "INTEGER")
    # v218.116 : interventions.gamme_id : référence à la gamme planifiée pour ce bon.
    # Permet de calculer précisément la durée + sous-type sans heuristique.
    # NULL pour les bons legacy / dépannage / astreinte.
    _add_col("interventions", "gamme_id", "INTEGER")
    # Cleanup : suppression des tables astreinte_bis (toute la fonctionnalité retirée)
    try:
        db = get_db()
        db.execute("DROP TABLE IF EXISTS astreinte_bis_periodes")
        db.execute("DROP TABLE IF EXISTS astreinte_bis_membres")
        db.execute("DROP TABLE IF EXISTS astreinte_bis_groupes")
        db.execute("DROP TABLE IF EXISTS astreinte_bis_planning")
        # Cleanup : entrées polluées par d'anciens imports (dates dans la colonne technicien) sur la v1
        for _patt in ('% 00:00:00', '%-__-__ %', '%-__-__'):
            try:
                db.execute("DELETE FROM astreinte_planning WHERE technicien LIKE ?", (_patt,))
            except Exception: pass
        db.commit()
    except Exception as _e:
        pass

    # Seed initial du catalogue d'équipements de sécurité HT (16 items des captures)
    try:
        db = get_db()
        existing = db.execute("SELECT COUNT(*) FROM securite_equipements").fetchone()[0]
        if existing == 0:
            seed = [
                "Panneau 1er soin FR",
                "Panneau 1er soin DE",
                "Panneau 5 règles de sécurités FR et DE",
                "Panneau DE : Bekampfung von branden (VDE 0132)",
                "Panneau DE : Starkstromanlagen, svt. (VDE 0105)",
                "Gants isolant 20kV",
                "Boite à gants avec talc",
                "Tabouret isolant 24kV",
                "Vérificateur Absence de Tension",
                "Luminaire de secours portatif avec chargeur et sa fixation murale",
                "Garniture MALT. (70mm2) + Perche",
                "Tapis isolants (10/24KV-6,3 A)",
                "Présence de la perche de sauvetage - longueur 1.65m",
                "Extincteur 20kV portatif au CO2, 5kg, avec sa fixation murale",
                "Fusibles 20kv de réserve",
                "Boite de 1er secours avec sa fixation murale (DIN 13457C)",
            ]
            for idx, lib in enumerate(seed, start=1):
                db.execute("INSERT INTO securite_equipements (libelle, ordre, actif) VALUES (?,?,1)", (lib, idx))
            db.commit()
    except Exception:
        pass
    # occupations.numero_projet (pour les occupations de type Offre, format P00000)
    _add_col("occupations", "numero_projet", "TEXT DEFAULT ''")
    # occupations.nom_chantier (obligatoire si type=Offre, accompagne numero_projet)
    _add_col("occupations", "nom_chantier", "TEXT DEFAULT ''")
    # occupations.accompagnants_ids (JSON array des ids de techniciens accompagnants, informatif)
    _add_col("occupations", "accompagnants_ids", "TEXT DEFAULT ''")
    # occupation_types.autorise_accompagnants : 0/1 pour afficher le champ accompagnants
    _add_col("occupation_types", "autorise_accompagnants", "INTEGER DEFAULT 0")
    # v218.52 : Table de liaison occupation ↔ techniciens (multi-tech sur une occupation)
    db.execute("""
        CREATE TABLE IF NOT EXISTS occupation_techniciens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            occupation_id INTEGER NOT NULL REFERENCES occupations(id) ON DELETE CASCADE,
            technicien_id INTEGER NOT NULL REFERENCES utilisateurs(id),
            UNIQUE(occupation_id, technicien_id)
        )
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_occ_tech_occid ON occupation_techniciens(occupation_id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_occ_tech_techid ON occupation_techniciens(technicien_id)")
    # Migration v218.52 : peupler la table à partir de technicien_id + accompagnants_ids
    # On le fait UNE seule fois (idempotent : si la table contient déjà des liens, on ne refait rien)
    nb_links = one(db.execute("SELECT COUNT(*) AS n FROM occupation_techniciens")).get("n", 0)
    if nb_links == 0:
        all_occs = rows(db.execute("SELECT id, technicien_id, accompagnants_ids FROM occupations"))
        for o in all_occs:
            ids_set = set()
            tid = o.get("technicien_id")
            if tid: ids_set.add(int(tid))
            raw = o.get("accompagnants_ids") or ""
            if raw and isinstance(raw, str):
                # Parse robuste : tente JSON d'abord, sinon CSV ou int seul
                try:
                    parsed = json.loads(raw)
                    if isinstance(parsed, list):
                        for x in parsed:
                            try: ids_set.add(int(x))
                            except: pass
                except Exception:
                    # Tentative en chaîne brute (ex: "8" ou "8,5")
                    for tok in raw.replace("[", "").replace("]", "").split(","):
                        tok = tok.strip()
                        if tok:
                            try: ids_set.add(int(tok))
                            except: pass
            for uid in ids_set:
                try:
                    db.execute("INSERT OR IGNORE INTO occupation_techniciens (occupation_id, technicien_id) VALUES (?, ?)",
                               (o["id"], uid))
                except Exception as e:
                    logger.warning(f"[MIG_OCC] échec lien occ={o['id']} tech={uid} : {e}")
        db.commit()
        logger.info(f"[MIG_OCC] migration accompagnants → occupation_techniciens : {len(all_occs)} occupations traitées")
    # utilisateurs.matricule (pour la fiche de pointage SOCOM)
    _add_col("utilisateurs", "matricule", "TEXT DEFAULT ''")
    # utilisateurs.manager_id (le manager d'un technicien pour la fiche de pointage)
    _add_col("utilisateurs", "manager_id", "INTEGER")
    # utilisateurs.techniques : CSV des techniques gérées par un manager technique (ex: "CVC,Electricite")
    _add_col("utilisateurs", "techniques", "TEXT DEFAULT ''")
    # utilisateurs.poste_id : référence à la table postes (organigramme)
    _add_col("utilisateurs", "poste_id", "INTEGER")
    # utilisateurs.superieur_id : supérieur hiérarchique dans l'organigramme visuel
    # (distinct de manager_id qui sert à l'attribution GMAO / fiche de pointage)
    _add_col("utilisateurs", "superieur_id", "INTEGER")
    # postes.couleur : couleur hex (#RRGGBB) utilisée dans l'organigramme
    _add_col("postes", "couleur", "TEXT DEFAULT '#3B82F6'")
    # projets.deplacement_km : distance A/R pour les dépannages
    # (le nombre de déplacements est saisi au niveau du CR pour les dépannages)
    _add_col("projets", "deplacement_km", "REAL DEFAULT 0")
    _add_col("projets", "nb_deplacements", "INTEGER DEFAULT 0")
    _add_col("projets", "logo_filename", "TEXT DEFAULT ''")
    # comptes_rendus.nb_deplacements : nombre de déplacements effectués sur ce CR
    # (repris × km du projet pour la facturation - dépannage uniquement)
    _add_col("comptes_rendus", "nb_deplacements", "INTEGER DEFAULT 0")
    # Garantir la colonne 'numero' sur comptes_rendus (utilisée pour affichage planning CR000032)
    _add_col("comptes_rendus", "numero", "TEXT")
    # Migration one-shot : attribuer un numéro aux CRs qui n'en ont pas
    try:
        missing = rows(db.execute("SELECT id FROM comptes_rendus WHERE numero IS NULL OR numero='' ORDER BY id"))
        for cr_row in missing:
            try:
                new_num = next_numero(db, "CR", "comptes_rendus", "numero")
                db.execute("UPDATE comptes_rendus SET numero=? WHERE id=?", (new_num, cr_row["id"]))
            except Exception as ex:
                logger.warning(f"[migration CR numero] skip id={cr_row['id']} : {ex}")
        if missing:
            db.commit()
            logger.info(f"[migration CR numero] {len(missing)} CR(s) numérotés")
    except Exception as ex:
        logger.warning(f"[migration CR numero] erreur globale : {ex}")

    # Migration : retirer le type OFFRE si présent dans la CHECK constraint.
    # Supprime d'abord tous les bons de type OFFRE (et leurs dépendances en cascade).
    try:
        schema_row = one(db.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='interventions'"))
        if schema_row and "'OFFRE'" in (schema_row.get("sql") or ""):
            logger.info("[MIGRATION] retrait du type OFFRE")
            # 1. Compter puis supprimer les bons OFFRE existants
            nb_offre = one(db.execute("SELECT COUNT(*) AS n FROM interventions WHERE type='OFFRE'"))
            nb = (nb_offre and nb_offre.get("n")) or 0
            if nb:
                logger.info(f"[MIGRATION] suppression de {nb} bon(s) de type OFFRE")
                # Supprimer les CR liés + intervenants
                db.execute("""DELETE FROM cr_intervenants WHERE cr_id IN
                              (SELECT cr.id FROM comptes_rendus cr
                               JOIN interventions i ON cr.intervention_id=i.id
                               WHERE i.type='OFFRE')""")
                db.execute("""DELETE FROM comptes_rendus WHERE intervention_id IN
                              (SELECT id FROM interventions WHERE type='OFFRE')""")
                db.execute("DELETE FROM intervention_creneaux WHERE intervention_id IN (SELECT id FROM interventions WHERE type='OFFRE')")
                # Table intervention_techniciens si elle existe
                try:
                    db.execute("DELETE FROM intervention_techniciens WHERE intervention_id IN (SELECT id FROM interventions WHERE type='OFFRE')")
                except Exception:
                    pass
                db.execute("DELETE FROM interventions WHERE type='OFFRE'")
                db.commit()
            # 2. Recréer la table avec CHECK resserré
            db.execute("PRAGMA foreign_keys=OFF")
            db.executescript("""
                CREATE TABLE interventions_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, numero TEXT UNIQUE,
                    equipement_id INTEGER NOT NULL REFERENCES equipements(id),
                    technicien_id INTEGER REFERENCES utilisateurs(id),
                    equipe_id INTEGER REFERENCES equipes(id),
                    type TEXT NOT NULL DEFAULT 'MAINTENANCE'
                        CHECK(type IN ('MAINTENANCE','DEPANNAGE')),
                    statut TEXT NOT NULL DEFAULT 'PLANIFIEE'
                        CHECK(statut IN ('A_PLANIFIER','PLANIFIEE','EN_COURS','TERMINEE','ANNULEE')),
                    date_prevue TEXT, date_realisation TEXT,
                    description TEXT DEFAULT '', rapport TEXT DEFAULT '',
                    created_at TEXT DEFAULT (datetime('now')),
                    heure_prevue TEXT DEFAULT '08:00'
                );
                INSERT INTO interventions_new
                    (id, numero, equipement_id, technicien_id, equipe_id, type, statut,
                     date_prevue, date_realisation, description, rapport, created_at, heure_prevue)
                SELECT id, numero, equipement_id, technicien_id, equipe_id, type, statut,
                       date_prevue, date_realisation, description, rapport, created_at,
                       COALESCE(heure_prevue, '08:00')
                FROM interventions;
                DROP TABLE interventions;
                ALTER TABLE interventions_new RENAME TO interventions;
            """)
            db.execute("PRAGMA foreign_keys=ON")
            db.commit()
            logger.info("[MIGRATION] table interventions recreee sans OFFRE")
    except Exception as e:
        logger.error(f"[MIGRATION] echec retrait OFFRE: {e}")

    # Migration : élargir CHECK statut pour inclure A_PLANIFIER
    # (utilisé quand un bon est créé sans date/tech, à planifier ultérieurement)
    try:
        r = one(db.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='interventions'"))
        if r and r.get("sql") and "A_PLANIFIER" not in r["sql"]:
            logger.info("[MIGRATION] ajout A_PLANIFIER au CHECK statut interventions")
            db.execute("PRAGMA foreign_keys=OFF")
            db.executescript("""
                CREATE TABLE interventions_new2 (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, numero TEXT UNIQUE,
                    equipement_id INTEGER NOT NULL REFERENCES equipements(id),
                    technicien_id INTEGER REFERENCES utilisateurs(id),
                    equipe_id INTEGER REFERENCES equipes(id),
                    type TEXT NOT NULL DEFAULT 'MAINTENANCE'
                        CHECK(type IN ('MAINTENANCE','DEPANNAGE')),
                    statut TEXT NOT NULL DEFAULT 'PLANIFIEE'
                        CHECK(statut IN ('A_PLANIFIER','PLANIFIEE','EN_COURS','TERMINEE','ANNULEE')),
                    date_prevue TEXT, date_realisation TEXT,
                    description TEXT DEFAULT '', rapport TEXT DEFAULT '',
                    created_at TEXT DEFAULT (datetime('now')),
                    heure_prevue TEXT DEFAULT '08:00'
                );
                INSERT INTO interventions_new2
                    (id, numero, equipement_id, technicien_id, equipe_id, type, statut,
                     date_prevue, date_realisation, description, rapport, created_at, heure_prevue)
                SELECT id, numero, equipement_id, technicien_id, equipe_id, type, statut,
                       date_prevue, date_realisation, description, rapport, created_at,
                       COALESCE(heure_prevue, '08:00')
                FROM interventions;
                DROP TABLE interventions;
                ALTER TABLE interventions_new2 RENAME TO interventions;
            """)
            db.execute("PRAGMA foreign_keys=ON")
            db.commit()
            logger.info("[MIGRATION] table interventions recreee avec A_PLANIFIER")
    except Exception as e:
        logger.error(f"[MIGRATION] echec ajout A_PLANIFIER: {e}")

    # v218.23 : élargir CHECK type interventions pour inclure ASTREINTE
    try:
        r = one(db.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='interventions'"))
        if r and r.get("sql") and "ASTREINTE" not in r["sql"]:
            logger.info("[MIGRATION] ajout ASTREINTE au CHECK type interventions")
            db.execute("PRAGMA foreign_keys=OFF")
            db.executescript("""
                CREATE TABLE interventions_new3 (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, numero TEXT UNIQUE,
                    equipement_id INTEGER NOT NULL REFERENCES equipements(id),
                    technicien_id INTEGER REFERENCES utilisateurs(id),
                    equipe_id INTEGER REFERENCES equipes(id),
                    type TEXT NOT NULL DEFAULT 'MAINTENANCE'
                        CHECK(type IN ('MAINTENANCE','DEPANNAGE','ASTREINTE')),
                    statut TEXT NOT NULL DEFAULT 'PLANIFIEE'
                        CHECK(statut IN ('A_PLANIFIER','PLANIFIEE','EN_COURS','TERMINEE','ANNULEE')),
                    date_prevue TEXT, date_realisation TEXT,
                    description TEXT DEFAULT '', rapport TEXT DEFAULT '',
                    created_at TEXT DEFAULT (datetime('now')),
                    heure_prevue TEXT DEFAULT '08:00',
                    tableau_id INTEGER
                );
                INSERT INTO interventions_new3
                    (id, numero, equipement_id, technicien_id, equipe_id, type, statut,
                     date_prevue, date_realisation, description, rapport, created_at,
                     heure_prevue, tableau_id)
                SELECT id, numero, equipement_id, technicien_id, equipe_id, type, statut,
                       date_prevue, date_realisation, description, rapport, created_at,
                       COALESCE(heure_prevue, '08:00'),
                       tableau_id
                FROM interventions;
                DROP TABLE interventions;
                ALTER TABLE interventions_new3 RENAME TO interventions;
            """)
            db.execute("PRAGMA foreign_keys=ON")
            db.commit()
            logger.info("[MIGRATION] table interventions recreee avec ASTREINTE")
    except Exception as e:
        logger.error(f"[MIGRATION] echec ajout ASTREINTE: {e}")

    # v218.31 : élargir CHECK frequence causeries pour inclure ANNUEL
    try:
        r = one(db.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='causeries'"))
        if r and r.get("sql") and "ANNUEL" not in r["sql"]:
            logger.info("[MIGRATION] ajout ANNUEL au CHECK frequence causeries")
            db.execute("PRAGMA foreign_keys=OFF")
            db.executescript("""
                CREATE TABLE causeries_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    titre TEXT NOT NULL,
                    pdf_filename TEXT DEFAULT '',
                    frequence TEXT DEFAULT 'MENSUEL'
                        CHECK(frequence IN ('HEBDO','BI_MENSUEL','MENSUEL','TRIMESTRIEL','ANNUEL')),
                    date_publication TEXT,
                    date_cloture TEXT,
                    mail_destinataire TEXT DEFAULT '',
                    actif INTEGER DEFAULT 1,
                    cree_par INTEGER REFERENCES utilisateurs(id),
                    created_at TEXT DEFAULT (datetime('now'))
                );
                INSERT INTO causeries_new SELECT * FROM causeries;
                DROP TABLE causeries;
                ALTER TABLE causeries_new RENAME TO causeries;
            """)
            db.execute("PRAGMA foreign_keys=ON")
            db.commit()
            logger.info("[MIGRATION] table causeries recreee avec ANNUEL")
    except Exception as e:
        logger.error(f"[MIGRATION] echec ajout ANNUEL causeries: {e}")

    # Nettoyage table orpheline utilisateurs_new (si présente)
    try:
        r = one(db.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='utilisateurs_new'"))
        if r:
            db.execute("DROP TABLE utilisateurs_new")
            logger.info("[MIGRATION] table orpheline 'utilisateurs_new' supprimee")
    except Exception as e:
        logger.warning(f"[MIGRATION] drop utilisateurs_new: {e}")

    db.commit()

    # Seed types d'occupation par défaut (première installation uniquement)
    try:
        r = one(db.execute("SELECT COUNT(*) AS n FROM occupation_types"))
        if r and r["n"] == 0:
            defaults = [
                ("Congés",    "#10B981"),  # vert
                ("Formation", "#6366F1"),  # violet
                ("Maladie",   "#EF4444"),  # rouge
                ("RTT",       "#F59E0B"),  # orange
                ("Réunion",   "#3B82F6"),  # bleu
                ("Offre",     "#A855F7"),  # pourpre - chiffrage / préparation commerciale
                ("Autre",     "#64748B"),  # gris
            ]
            for nom, couleur in defaults:
                db.execute("INSERT OR IGNORE INTO occupation_types (nom, couleur) VALUES (?,?)", (nom, couleur))
            db.commit()
            logger.info("[SEED] types d'occupation par défaut crées")
        # S'assurer que "Offre" existe (installations antérieures)
        has_offre = one(db.execute("SELECT id FROM occupation_types WHERE nom='Offre'"))
        if not has_offre:
            db.execute("INSERT INTO occupation_types (nom, couleur) VALUES (?,?)", ("Offre","#A855F7"))
            db.commit()
            logger.info("[SEED] type 'Offre' ajoute (installation existante)")
        # S'assurer que "Congés" existe (utilise par les demandes de conges)
        has_conges = one(db.execute("SELECT id FROM occupation_types WHERE nom='Congés'"))
        if not has_conges:
            db.execute("INSERT INTO occupation_types (nom, couleur) VALUES (?,?)", ("Congés","#10B981"))
            db.commit()
            logger.info("[SEED] type 'Congés' ajoute (installation existante)")
    except Exception as e:
        logger.error(f"[SEED occupations] echec: {e}")

    # ══════════════════════════════════════════════════════════════════════
    # SOUS-TYPES DE MAINTENANCE (Entretien / Visite / etc.)
    # Configurable comme les types d'occupation. Affecté à chaque gamme.
    # Le sous-type d'un bon est déduit de ses gammes (cf. _bon_sous_type_label)
    # ══════════════════════════════════════════════════════════════════════
    db.execute("""
    CREATE TABLE IF NOT EXISTS maintenance_sous_types (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL UNIQUE,
        couleur TEXT DEFAULT '#1E3A8A',
        ordre INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    )
    """)
    # Colonne sous_type_id sur les gammes (NULL = pas de sous-type)
    _add_col("gammes", "sous_type_id", "INTEGER")
    # v211 : nombre de chaînes pour pièces de type batterie
    _add_col("pieces", "nbr_chaine", "INTEGER")
    # Seed initial : Entretien + Visite (uniquement si table vide)
    try:
        nb = one(db.execute("SELECT COUNT(*) AS n FROM maintenance_sous_types"))
        if nb and nb["n"] == 0:
            db.execute("INSERT INTO maintenance_sous_types (nom, couleur, ordre) VALUES (?, ?, ?)",
                       ("Entretien", "#10B981", 1))
            db.execute("INSERT INTO maintenance_sous_types (nom, couleur, ordre) VALUES (?, ?, ?)",
                       ("Visite", "#F59E0B", 2))
            db.commit()
            logger.info("[SEED] sous-types de maintenance initialisés (Entretien, Visite)")
    except Exception as e:
        logger.warning(f"[SEED maintenance_sous_types] échec : {e}")

    # ══════════════════════════════════════════════════════════════════════
    # MESURES TECHNIQUES PAR (technique × sous-type)
    # mesure_blocs    : un tableau de mesures (ex. "Mesure d'entrée")
    # mesure_lignes   : une ligne dans un tableau (ex. "Tension L1" en V)
    # intervention_mesures : valeurs saisies par l'utilisateur pour un bon
    # ══════════════════════════════════════════════════════════════════════
    db.execute("""
    CREATE TABLE IF NOT EXISTS mesure_blocs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        technique_id INTEGER NOT NULL REFERENCES techniques(id) ON DELETE CASCADE,
        sous_type_id INTEGER NOT NULL REFERENCES maintenance_sous_types(id) ON DELETE CASCADE,
        nom TEXT NOT NULL,
        ordre INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    )
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_mb_tech_st ON mesure_blocs(technique_id, sous_type_id)")
    # v206 : type de module (tableau / texte / checklist / graphique)
    _add_col("mesure_blocs", "type", "TEXT DEFAULT 'tableau'")
    # v209 : config globale du module (JSON) — utilisé pour les modules 'graphique'
    _add_col("mesure_blocs", "field_options", "TEXT DEFAULT ''")
    # v213 : largeur du module dans le rapport ('pleine' ou 'demi')
    _add_col("mesure_blocs", "largeur", "TEXT DEFAULT 'pleine'")
    # v214 : numéro de page sur lequel le module doit s'afficher dans le rapport
    _add_col("mesure_blocs", "page_num", "INTEGER DEFAULT 1")
    # v215 : icône PNG à afficher dans le titre du module (slug, ex: 'bolt', 'battery-100')
    _add_col("mesure_blocs", "icon", "TEXT DEFAULT ''")
    # v218.109 : flag indiquant un module BC (dépannage). Si is_bc=1, technique_id et sous_type_id
    # sont ignorés (le module est global pour la société). Permet de gérer des modules BC sans
    # dupliquer toute la table mesure_blocs et toute la chaîne mesure_lignes/intervention_mesures.
    _add_col("mesure_blocs", "is_bc", "INTEGER DEFAULT 0")
    db.execute("CREATE INDEX IF NOT EXISTS idx_mb_bc ON mesure_blocs(is_bc, societe_id)")
    # v207 : type de champ par ligne pour les modules 'tableau' (numeric/text_short/text_long/binary/select)
    _add_col("mesure_lignes", "field_type", "TEXT DEFAULT 'numeric'")
    # field_options : JSON pour stocker la config selon le type (labels binaire, options select, etc.)
    _add_col("mesure_lignes", "field_options", "TEXT DEFAULT ''")
    db.execute("""
    CREATE TABLE IF NOT EXISTS mesure_lignes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bloc_id INTEGER NOT NULL REFERENCES mesure_blocs(id) ON DELETE CASCADE,
        libelle TEXT NOT NULL,
        unite TEXT DEFAULT '',
        ordre INTEGER DEFAULT 0
    )
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_ml_bloc ON mesure_lignes(bloc_id)")
    db.execute("""
    CREATE TABLE IF NOT EXISTS intervention_mesures (
        intervention_id INTEGER NOT NULL REFERENCES interventions(id) ON DELETE CASCADE,
        ligne_id INTEGER NOT NULL REFERENCES mesure_lignes(id) ON DELETE CASCADE,
        valeur TEXT DEFAULT '',
        PRIMARY KEY (intervention_id, ligne_id)
    )
    """)

    # v217 : Bibliothèque d'images réutilisables
    db.execute("""
    CREATE TABLE IF NOT EXISTS image_library (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL,
        filename TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now')),
        created_by INTEGER REFERENCES utilisateurs(id) ON DELETE SET NULL
    )
    """)
    # v217 : image principale d'un équipement (FK vers image_library)
    _add_col("equipements", "image_id", "INTEGER REFERENCES image_library(id) ON DELETE SET NULL")

    # v218.33 : tracking de l'envoi du mail récap pour les causeries (idempotent)
    _add_col("causeries", "mail_envoye", "INTEGER DEFAULT 0")
    _add_col("causeries", "mail_envoye_at", "TEXT DEFAULT ''")
    # v218.37 : planification précise (rang + jour de la semaine + mois)
    _add_col("causeries", "planif_rang", "TEXT DEFAULT ''")     # 'PREMIER', 'DEUXIEME', 'TROISIEME', 'QUATRIEME', 'DERNIER'
    _add_col("causeries", "planif_jour", "TEXT DEFAULT ''")     # 'LUNDI'..'DIMANCHE'
    _add_col("causeries", "planif_mois", "INTEGER DEFAULT 0")   # 1..12 (0 = non défini)
    _add_col("causeries", "planif_annee", "INTEGER DEFAULT 0")  # année cible (0 = année courante au calcul)
    # v218.44 : temps de réponse en secondes (déduction de 1pt/sec)
    _add_col("causeries_reponses", "temps_secondes", "INTEGER DEFAULT 0")

    # v218 : Statuts d'équipement configurables
    db.execute("""
    CREATE TABLE IF NOT EXISTS equipement_statuts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT NOT NULL UNIQUE,
        label TEXT NOT NULL,
        couleur TEXT DEFAULT '#10B981',
        ordre INTEGER DEFAULT 0
    )
    """)
    # Seed initial : ne s'exécute qu'une seule fois (table vide)
    nb_st = one(db.execute("SELECT COUNT(*) AS n FROM equipement_statuts"))
    if not nb_st or not nb_st["n"]:
        db.execute("INSERT INTO equipement_statuts (code, label, couleur, ordre) VALUES (?, ?, ?, ?)",
                   ("EN_SERVICE", "En service", "#10B981", 1))
        db.execute("INSERT INTO equipement_statuts (code, label, couleur, ordre) VALUES (?, ?, ?, ?)",
                   ("HORS_SERVICE", "Hors service", "#EF4444", 2))

    # Seed UPS × Entretien : 4 tableaux par défaut (uniquement si pas déjà fait)
    try:
        # Trouver les IDs nécessaires
        tech_ups = one(db.execute("SELECT id FROM techniques WHERE LOWER(nom) IN ('ups','onduleur','onduleurs')"))
        st_entretien = one(db.execute("SELECT id FROM maintenance_sous_types WHERE LOWER(nom)='entretien'"))
        if tech_ups and st_entretien:
            tech_id = tech_ups["id"]
            st_id = st_entretien["id"]
            # Vérifier si on n'a pas déjà fait le seed (existence d'au moins un bloc UPS×Entretien)
            existing = one(db.execute(
                "SELECT COUNT(*) AS n FROM mesure_blocs WHERE technique_id=? AND sous_type_id=?",
                (tech_id, st_id)
            ))
            if not existing or existing["n"] == 0:
                blocs_seed = [
                    ("Mesure d'entrée", 1, [
                        ("Tension L1", "V"),
                        ("Tension L2", "V"),
                        ("Tension L3", "V"),
                        ("Fréquence", "Hz"),
                    ]),
                    ("Mesure by-pass", 2, [
                        ("Tension L1", "V"),
                        ("Tension L2", "V"),
                        ("Tension L3", "V"),
                        ("Fréquence", "Hz"),
                    ]),
                    ("Mesure de sortie", 3, [
                        ("Tension L1", "V"),
                        ("Tension L2", "V"),
                        ("Tension L3", "V"),
                        ("Fréquence", "Hz"),
                        ("Courant L1", "A"),
                        ("Courant L2", "A"),
                        ("Courant L3", "A"),
                    ]),
                    ("Mesure de la batterie", 4, [
                        ("Tension branche +", "Vdc"),
                        ("Tension branche -", "Vdc"),
                        ("Tension +/-", "Vdc"),
                        ("Courant de charge", "A"),
                        ("Courant de décharge", "A"),
                        ("Autonomie théorique", "min"),
                    ]),
                ]
                for bloc_nom, bloc_ordre, lignes in blocs_seed:
                    db.execute(
                        "INSERT INTO mesure_blocs (technique_id, sous_type_id, nom, ordre) VALUES (?, ?, ?, ?)",
                        (tech_id, st_id, bloc_nom, bloc_ordre)
                    )
                    bid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
                    for i, (libelle, unite) in enumerate(lignes):
                        db.execute(
                            "INSERT INTO mesure_lignes (bloc_id, libelle, unite, ordre) VALUES (?, ?, ?, ?)",
                            (bid, libelle, unite, i)
                        )
                db.commit()
                logger.info("[SEED] mesures UPS Entretien : 4 tableaux initialisés")
    except Exception as e:
        logger.warning(f"[SEED mesures UPS] échec : {e}")

    # ══════════════════════════════════════════════════════════════════════
    # SYSTÈME DE PERMISSIONS (RBAC)
    # roles : 4 rôles built-in (admin/manager/technicien/acl) + rôles personnalisés
    # permissions : grille rôle × module × action
    # ══════════════════════════════════════════════════════════════════════
    db.execute("""
    CREATE TABLE IF NOT EXISTS roles (
        code TEXT PRIMARY KEY,
        label TEXT NOT NULL,
        parent_role TEXT NOT NULL DEFAULT 'technicien',
        builtin INTEGER NOT NULL DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    )
    """)
    db.execute("""
    CREATE TABLE IF NOT EXISTS permissions (
        role_code TEXT NOT NULL,
        module TEXT NOT NULL,
        action TEXT NOT NULL,
        allowed INTEGER NOT NULL DEFAULT 0,
        PRIMARY KEY (role_code, module, action)
    )
    """)
    # Initialisation : rôles built-in + grille par défaut (idempotent)
    _init_roles_and_permissions(db)

    # ═══════════════════════════════════════════════════════════════════
    # v218.62 — MULTI-TENANT PHASE 1 : Fondations DB
    # Crée la table societes, la table de liaison utilisateur_societes,
    # et ajoute societe_id sur toutes les tables métier.
    # SOCOM devient la société id=1 et reçoit toutes les données existantes.
    # AUCUN CHANGEMENT FONCTIONNEL : le code actuel continue à fonctionner.
    # Idempotent : peut tourner plusieurs fois sans casser.
    # ═══════════════════════════════════════════════════════════════════
    _migrate_multitenant_phase1(db)
    _migrate_multitenant_phase2(db)
    _migrate_cover_pages(db)

    # v218.117 : Champs personnalisés sur les équipements
    # Catalogue : 1 ligne par champ défini par l'admin
    db.execute("""
    CREATE TABLE IF NOT EXISTS equipement_champs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        societe_id INTEGER NOT NULL DEFAULT 1,
        technique_id INTEGER,  -- NULL = champ global, sinon spécifique à la technique
        slug TEXT NOT NULL,    -- identifiant interne, ex: 'fournisseur'
        label TEXT NOT NULL,   -- libellé affiché, ex: 'Fournisseur'
        type TEXT NOT NULL,    -- text|number|date|boolean|select|textarea|file
        options_json TEXT DEFAULT '{}',  -- {choices:[...], placeholder, required, ...}
        ordre INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (technique_id) REFERENCES techniques(id) ON DELETE CASCADE,
        FOREIGN KEY (societe_id) REFERENCES societes(id) ON DELETE CASCADE,
        UNIQUE(societe_id, slug)
    )""")
    db.execute("CREATE INDEX IF NOT EXISTS idx_eq_champs_societe ON equipement_champs(societe_id, technique_id, ordre)")
    # Valeurs : 1 ligne par (équipement, champ)
    db.execute("""
    CREATE TABLE IF NOT EXISTS equipement_valeurs_custom (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        equipement_id INTEGER NOT NULL,
        champ_id INTEGER NOT NULL,
        valeur TEXT DEFAULT '',
        updated_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (equipement_id) REFERENCES equipements(id) ON DELETE CASCADE,
        FOREIGN KEY (champ_id) REFERENCES equipement_champs(id) ON DELETE CASCADE,
        UNIQUE(equipement_id, champ_id)
    )""")
    db.execute("CREATE INDEX IF NOT EXISTS idx_eq_val_eq ON equipement_valeurs_custom(equipement_id)")
    # Fichiers : pour les champs type=file (stockage binaire)
    db.execute("""
    CREATE TABLE IF NOT EXISTS equipement_fichiers_custom (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        valeur_id INTEGER NOT NULL,
        filename TEXT NOT NULL,
        mime_type TEXT DEFAULT 'application/octet-stream',
        data BLOB NOT NULL,
        size INTEGER DEFAULT 0,
        uploaded_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (valeur_id) REFERENCES equipement_valeurs_custom(id) ON DELETE CASCADE
    )""")
    db.execute("CREATE INDEX IF NOT EXISTS idx_eq_fich_val ON equipement_fichiers_custom(valeur_id)")

    # v218.124 : Jours fériés (société-level)
    db.execute("""
    CREATE TABLE IF NOT EXISTS jours_feries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        societe_id INTEGER NOT NULL DEFAULT 1,
        date TEXT NOT NULL,
        label TEXT NOT NULL,
        heure_debut TEXT DEFAULT '08:00',
        heure_fin TEXT DEFAULT '16:00',
        created_at TEXT DEFAULT (datetime('now')),
        UNIQUE(societe_id, date),
        FOREIGN KEY (societe_id) REFERENCES societes(id) ON DELETE CASCADE
    )""")
    db.execute("CREATE INDEX IF NOT EXISTS idx_jf_societe_date ON jours_feries(societe_id, date)")
    # Config par société : heures par défaut pour les nouveaux jours fériés
    db.execute("""
    CREATE TABLE IF NOT EXISTS jours_feries_config (
        societe_id INTEGER PRIMARY KEY,
        heure_debut_default TEXT DEFAULT '08:00',
        heure_fin_default TEXT DEFAULT '16:00',
        updated_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (societe_id) REFERENCES societes(id) ON DELETE CASCADE
    )""")
    # Init config par société (1 ligne par société existante)
    try:
        for s in rows(db.execute("SELECT id FROM societes")):
            if not one(db.execute("SELECT societe_id FROM jours_feries_config WHERE societe_id=?", (s["id"],))):
                db.execute("INSERT INTO jours_feries_config (societe_id) VALUES (?)", (s["id"],))
    except Exception as e:
        logger.warning(f"[jours_feries_config] init société : {e}")

    # v218.124 : Migration des anciennes occupations "Jour férié" → jours_feries puis suppression
    try:
        jf_type = one(db.execute("SELECT id FROM occupation_types WHERE nom=?", ("Jour férié",)))
        if jf_type:
            old_occs = rows(db.execute("""SELECT DISTINCT date, societe_id FROM occupations
                                          WHERE type_id=?""", (jf_type["id"],)))
            migrated = 0
            for occ in old_occs:
                try:
                    sid_occ = occ.get("societe_id") or 1
                    db.execute("""INSERT OR IGNORE INTO jours_feries
                                  (societe_id, date, label, heure_debut, heure_fin)
                                  VALUES (?,?,?,?,?)""",
                               (sid_occ, occ["date"], "Jour férié", "08:00", "16:00"))
                    migrated += 1
                except Exception: pass
            # Supprimer les anciennes occupations puis le type lui-même
            db.execute("DELETE FROM occupations WHERE type_id=?", (jf_type["id"],))
            db.execute("DELETE FROM occupation_types WHERE id=?", (jf_type["id"],))
            logger.info(f"[jours_feries] migration : {migrated} occupation(s) 'Jour férié' migrées et type supprimé")
    except Exception as e:
        logger.warning(f"[jours_feries] migration échec : {e}")

    db.commit()

    # Seed admin par défaut (première installation uniquement)
    try:
        r = one(db.execute("SELECT COUNT(*) AS n FROM utilisateurs"))
        if r and r["n"] == 0:
            db.execute("INSERT INTO utilisateurs (nom,email,password,role) VALUES (?,?,?,?)",
                       ("Administrateur","admin@gmao.fr",hash_password("admin"),"admin"))
            db.commit()
            logger.warning("[SEED] compte admin par defaut cree (admin@gmao.fr / admin) — A CHANGER !")
    except Exception as e:
        logger.error(f"[SEED] echec: {e}")
    # v218.160 : Suppression des données obsolètes (les tables sont conservées pour compat code,
    # mais vidées. Le UI ne permet plus d'en créer de nouvelles - voir v218.159).
    # Le user a explicitement validé la perte des liens entre interventions/CR et tableaux.
    try:
        # Vider cr_tableaux puis equipement_tableaux puis equipement_cellules (ordre FK)
        db.execute("DELETE FROM cr_tableaux")
        db.execute("DELETE FROM equipement_tableaux")
        # Nettoyer les références orphelines dans interventions
        db.execute("UPDATE interventions SET tableau_id=NULL WHERE tableau_id IS NOT NULL")
        # equipement_cellules peut être complètement DROP (pas de FK externe)
        db.execute("DROP TABLE IF EXISTS equipement_cellules")
        db.commit()
        logger.info("[MIG v218.160] Données purgées : cr_tableaux, equipement_tableaux, equipement_cellules")
    except Exception as e:
        logger.error(f"[MIG v218.160] purge échec: {e}")

    # v218.169 : pré-remplir types_pieces avec les 5 types historiques pour chaque société qui n'en a aucun
    try:
        defaults_types = [
            ("BATTERIES", 1, 0),
            ("VENTILATEURS", 0, 1),
            ("CONDENSATEURS_AC", 0, 2),
            ("CONDENSATEURS_DC", 0, 3),
            ("CARTE_ALIMENTATION", 0, 4),
        ]
        societes_rows = rows(db.execute("SELECT id FROM societes"))
        for s in societes_rows:
            cnt = one(db.execute("SELECT COUNT(*) AS n FROM types_pieces WHERE societe_id=?", (s["id"],)))
            if cnt and cnt.get("n", 0) == 0:
                for nom, is_bat, ordre in defaults_types:
                    db.execute("INSERT INTO types_pieces (societe_id, nom, is_batterie, ordre) VALUES (?,?,?,?)",
                               (s["id"], nom, is_bat, ordre))
                logger.info(f"[MIG v218.169] {len(defaults_types)} types_pieces par défaut créés pour societe_id={s['id']}")
        db.commit()
    except Exception as e:
        logger.error(f"[MIG v218.169] seed types_pieces échec: {e}")

    # v218.173 : Suppression du CHECK constraint hérité sur pieces.type_piece
    # (l'ancienne version limitait à 5 types fixes ; maintenant types_pieces est dynamique)
    try:
        schema_row = one(db.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='pieces'"
        ))
        if schema_row and "CHECK" in (schema_row.get("sql") or "") and "type_piece IN" in schema_row["sql"]:
            logger.info("[MIG v218.173] CHECK constraint détecté sur pieces.type_piece — migration en cours…")
            db.execute("PRAGMA foreign_keys=OFF")
            db.execute("BEGIN TRANSACTION")
            try:
                # Recréer la table sans le CHECK
                db.execute("""
                    CREATE TABLE pieces_new (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        equipement_id INTEGER NOT NULL REFERENCES equipements(id),
                        type_piece TEXT NOT NULL,
                        date_installation TEXT,
                        duree_vie_estimee INTEGER,
                        date_fin_de_vie TEXT,
                        statut TEXT NOT NULL DEFAULT 'OK' CHECK(statut IN ('OK','A_SURVEILLER','A_REMPLACER')),
                        commentaire TEXT,
                        created_at TEXT DEFAULT (datetime('now')),
                        quantite INTEGER DEFAULT 1,
                        numero_serie TEXT DEFAULT '',
                        reference TEXT DEFAULT '',
                        nbr_chaine INTEGER,
                        societe_id INTEGER NOT NULL DEFAULT 1
                    )
                """)
                # Copier les colonnes existantes (intersection avec la nouvelle table)
                existing_cols = [c["name"] for c in rows(db.execute("PRAGMA table_info(pieces)"))]
                new_cols = ["id","equipement_id","type_piece","date_installation","duree_vie_estimee",
                            "date_fin_de_vie","statut","commentaire","created_at","quantite",
                            "numero_serie","reference","nbr_chaine","societe_id"]
                common = [c for c in new_cols if c in existing_cols]
                cols_str = ",".join(common)
                db.execute(f"INSERT INTO pieces_new ({cols_str}) SELECT {cols_str} FROM pieces")
                db.execute("DROP TABLE pieces")
                db.execute("ALTER TABLE pieces_new RENAME TO pieces")
                db.execute("COMMIT")
                logger.info("[MIG v218.173] CHECK constraint supprimé avec succès")
            except Exception as e_inner:
                db.execute("ROLLBACK")
                logger.error(f"[MIG v218.173] échec migration: {e_inner}")
            db.execute("PRAGMA foreign_keys=ON")
    except Exception as e:
        logger.error(f"[MIG v218.173] détection échec: {e}")

with app.app_context(): init_db()


# ══ DASHBOARD ══
@app.route("/api/dashboard")
@require_auth
def get_dashboard():
    db = get_db()
    u  = request.user
    td = today()
    sid = current_societe_id()
    # v218.72 : filtre société au niveau le plus haut
    role_filter = " AND i.societe_id=?"
    role_params = [sid]
    if u["role"] == "technicien":
        # Technicien : uniquement ses bons (colonne legacy, équipe, ou liaison multi-tech)
        role_filter += (" AND (i.technicien_id=? "
                       "OR i.equipe_id IN (SELECT equipe_id FROM equipe_membres WHERE technicien_id=?) "
                       "OR EXISTS (SELECT 1 FROM intervention_techniciens it WHERE it.intervention_id=i.id AND it.utilisateur_id=?))")
        role_params += [u["id"], u["id"], u["id"]]
    elif u["role"] == "manager":
        # Manager : bons dont il gère le projet, OU équipe (techs dont manager_id=lui), OU techniques gérées
        # Récupérer ses techniques
        udb = one(db.execute("SELECT techniques FROM utilisateurs WHERE id=?", (u["id"],)))
        tech_csv = (udb and udb.get("techniques")) or ""
        tech_list = [t.strip() for t in tech_csv.split(",") if t.strip()]
        conditions = [
            "i.equipement_id IN (SELECT e.id FROM equipements e JOIN projets p ON e.projet_id=p.id WHERE p.manager_id=?)",
            "i.technicien_id=?",
            "i.technicien_id IN (SELECT id FROM utilisateurs WHERE manager_id=?)",
            "EXISTS (SELECT 1 FROM intervention_techniciens it WHERE it.intervention_id=i.id AND "
            "  (it.utilisateur_id=? OR it.utilisateur_id IN (SELECT id FROM utilisateurs WHERE manager_id=?)))"
        ]
        role_params += [u["id"], u["id"], u["id"], u["id"], u["id"]]
        if tech_list:
            placeholders = ",".join(["?"] * len(tech_list))
            conditions.append(f"i.equipement_id IN (SELECT e.id FROM equipements e WHERE e.type_technique IN ({placeholders}))")
            role_params += tech_list
        role_filter += " AND (" + " OR ".join(conditions) + ")"
    # Filtre manager optionnel pour admin (liste déroulante dashboard)
    manager_id = to_int(request.args.get("manager_id")) if u["role"] in ("admin","superadmin") else None

    # KPI compteurs
    kpi = {
        "a_planifier": one(db.execute(f"SELECT COUNT(*) AS n FROM interventions i WHERE i.statut='PLANIFIEE'{role_filter}", role_params))["n"],
        "en_cours":    one(db.execute(f"SELECT COUNT(*) AS n FROM interventions i WHERE i.statut='EN_COURS'{role_filter}", role_params))["n"],
        "terminees":   one(db.execute(f"SELECT COUNT(*) AS n FROM interventions i WHERE i.statut='TERMINEE' AND strftime('%Y-%m',i.date_realisation)=strftime('%Y-%m','now'){role_filter}", role_params))["n"],
    }

    # Interventions du jour
    manager_filter = ""
    manager_params = []
    if manager_id:
        manager_filter = " AND p.manager_id=?"
        manager_params = [manager_id]
    sql_today = f"""SELECT i.*,e.designation AS equip_nom,e.type_technique,
                    p.nom AS projet_nom,u2.nom AS technicien_nom
                    FROM interventions i JOIN equipements e ON i.equipement_id=e.id
                    JOIN projets p ON e.projet_id=p.id
                    LEFT JOIN utilisateurs u2 ON i.technicien_id=u2.id
                    WHERE i.statut NOT IN ('TERMINEE','ANNULEE')
                    AND date(i.date_prevue)=?{role_filter}{manager_filter}
                    ORDER BY i.date_prevue"""
    today_ivs = rows(db.execute(sql_today, [td] + role_params + manager_params))

    # Retards
    sql_retards = f"""SELECT i.*,e.designation AS equip_nom,e.type_technique,
                    p.nom AS projet_nom,u2.nom AS technicien_nom
                    FROM interventions i JOIN equipements e ON i.equipement_id=e.id
                    JOIN projets p ON e.projet_id=p.id
                    LEFT JOIN utilisateurs u2 ON i.technicien_id=u2.id
                    WHERE i.statut NOT IN ('TERMINEE','ANNULEE')
                    AND i.date_prevue IS NOT NULL AND date(i.date_prevue) < ?{role_filter}{manager_filter}
                    ORDER BY i.date_prevue"""
    retards = rows(db.execute(sql_retards, [td] + role_params + manager_params))

    # Alertes pièces
    alertes = rows(db.execute("""SELECT p.*,e.designation AS equip_nom,pr.nom AS projet_nom
        FROM pieces p JOIN equipements e ON p.equipement_id=e.id
        JOIN projets pr ON e.projet_id=pr.id
        WHERE p.statut IN ('A_SURVEILLER','A_REMPLACER') ORDER BY p.statut DESC LIMIT 10"""))

    # Équipements critiques
    critiques = rows(db.execute("""SELECT e.*,p.nom AS projet_nom FROM equipements e
        JOIN projets p ON e.projet_id=p.id
        WHERE e.id IN (SELECT DISTINCT equipement_id FROM pieces WHERE statut='A_REMPLACER') LIMIT 10"""))

    # Stats délais
    r_dep = one(db.execute("""SELECT COUNT(*) AS nb,
        AVG(CAST((JULIANDAY(date_realisation)-JULIANDAY(created_at)) AS REAL)) AS moy
        FROM interventions WHERE type='DEPANNAGE' AND statut='TERMINEE' AND date_realisation IS NOT NULL"""))
    r_mai = one(db.execute("""SELECT COUNT(*) AS nb,
        AVG(CAST((JULIANDAY(date_realisation)-JULIANDAY(created_at)) AS REAL)) AS moy
        FROM interventions WHERE type='MAINTENANCE' AND statut='TERMINEE' AND date_realisation IS NOT NULL"""))

    stats = {
        "moy_jours_depannage":  round(r_dep["moy"] or 0, 1) if r_dep else 0,
        "nb_depannages":        r_dep["nb"] if r_dep else 0,
        "moy_jours_maintenance": round(r_mai["moy"] or 0, 1) if r_mai else 0,
        "nb_maintenances":      r_mai["nb"] if r_mai else 0,
    }

    return jsonify({
        "kpi": kpi,
        "today": today_ivs,
        "retards": retards,
        "alertes": alertes,
        "critiques": critiques,
        "stats": stats
    })


# ══ MOT DE PASSE OUBLIÉ ══
@app.route("/api/reset-password", methods=["POST"])
@limiter.limit("5 per hour")
def request_reset():
    d  = request.json or {}
    email = (d.get("email","") or "").strip().lower()
    if not email: return jsonify({"error":"Email requis"}),400
    db = get_db()
    u = one(db.execute("SELECT * FROM utilisateurs WHERE lower(email)=? AND actif=1",(email,)))
    if not u:
        # Ne pas révéler si l'email existe ou non
        return jsonify({"ok":True})
    token = secrets.token_urlsafe(32)
    expires = (datetime.utcnow() + timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")
    # Les colonnes reset_token et reset_expires sont créées dans init_db (migrations)
    db.execute("UPDATE utilisateurs SET reset_token=?,reset_expires=? WHERE id=?",(token,expires,u["id"]))
    db.commit()
    base_url = request.headers.get("Origin","http://"+request.host)
    # Construire le lien : si la demande vient de /mobile, renvoyer vers /mobile
    referer = request.headers.get("Referer","") or ""
    path_suffix = "/mobile" if "/mobile" in referer else ""
    reset_link = f"{base_url}{path_suffix}?reset_token={token}"
    body = f"""Bonjour {u['nom']},

Vous avez demandé la réinitialisation de votre mot de passe SOCOM GMAO.

Cliquez sur ce lien pour définir un nouveau mot de passe (valable 1 heure) :
{reset_link}

Si vous n'avez pas fait cette demande, ignorez cet email.

Cordialement,
SOCOM GMAO"""
    send_mail(u["email"],"[GMAO] Réinitialisation de mot de passe",body)
    tok_short = (token[:12] + "…") if token else "(vide)"
    logger.info(f"[RESET] demande pour uid={u['id']} token={tok_short} lien_base={path_suffix or '/'}")
    return jsonify({"ok":True})

@app.route("/api/reset-password/confirm", methods=["POST"])
@limiter.limit("10 per hour")
def confirm_reset():
    d = request.json or {}
    token = d.get("token","")
    new_pw = d.get("password","")
    # Log pour diagnostic
    tok_short = (token[:12] + "…") if token else "(vide)"
    logger.info(f"[RESET confirm] token recu: {tok_short} (len={len(token)})")
    if not token or not new_pw: return jsonify({"error":"Token et mot de passe requis"}),400
    if len(new_pw) < 8: return jsonify({"error":"Mot de passe trop court (8 caractères minimum)"}),400
    db = get_db()
    u = one(db.execute("SELECT * FROM utilisateurs WHERE reset_token=?",(token,)))
    if not u:
        # Cas : le token n'existe plus en base — peut-être déjà utilisé ou demande écrasée
        active = db.execute("SELECT id, nom, LENGTH(reset_token) AS tk_len FROM utilisateurs WHERE reset_token IS NOT NULL AND reset_token != ''").fetchall()
        logger.warning(f"[RESET confirm] token introuvable. Tokens actifs en base: {len(active)} utilisateur(s)")
        return jsonify({"error":"Token invalide ou déjà utilisé. Refaites une demande de réinitialisation."}),400
    # Vérifier expiration
    try:
        expires = datetime.strptime(u["reset_expires"],"%Y-%m-%d %H:%M:%S")
        if datetime.utcnow() > expires:
            logger.info(f"[RESET confirm] token expire pour uid={u['id']}")
            return jsonify({"error":"Lien expiré (valable 1 heure). Refaites une demande de réinitialisation."}),400
    except Exception as e:
        logger.error(f"[RESET confirm] erreur parsing expiration: {e}")
        return jsonify({"error":"Token invalide (données corrompues)"}),400
    # Invalider toutes les sessions existantes (bump token_version)
    new_tv = (u.get("token_version") or 0) + 1
    db.execute("UPDATE utilisateurs SET password=?, reset_token=NULL, reset_expires=NULL, token_version=? WHERE id=?",
               (hash_password(new_pw), new_tv, u["id"]))
    db.commit()
    logger.info(f"[RESET] MDP change pour uid={u['id']} (sessions revoquees)")
    return jsonify({"ok":True})


# ══ MAILS ══
@app.route("/api/mails")
@require_role("admin","manager")
def get_mails():
    db = get_db()
    return jsonify(rows(db.execute(
        "SELECT * FROM mail_log ORDER BY created_at DESC LIMIT 200")))

@app.route("/api/mails/<int:mid>", methods=["DELETE"])
@require_role("admin")
def delete_mail(mid):
    db = get_db()
    db.execute("DELETE FROM mail_log WHERE id=?", (mid,))
    db.commit()
    return jsonify({"ok": True})


# ══ POSTES (organigramme) ══
@app.route("/api/postes")
@require_auth
def get_postes():
    """Liste des postes définis (pour tous les rôles car utilisés dans le form utilisateur)."""
    db = get_db()
    return jsonify(rows(db.execute(
        "SELECT * FROM postes ORDER BY ordre, nom")))

@app.route("/api/postes", methods=["POST"])
@require_role("admin")
def create_poste():
    d = request.json or {}
    nom = (d.get("nom") or "").strip()
    if not nom: return jsonify({"error":"Nom requis"}), 400
    couleur = (d.get("couleur") or "#3B82F6").strip()
    # Validation simple : doit ressembler à #RRGGBB
    if not (couleur.startswith("#") and len(couleur) == 7):
        couleur = "#3B82F6"
    db = get_db()
    try:
        db.execute("INSERT INTO postes (nom, ordre, couleur) VALUES (?, ?, ?)",
                   (nom, to_int(d.get("ordre")) or 0, couleur))
        db.commit()
        new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        log_action(request.user, "CREATE", "poste", new_id, nom)
        return jsonify({"id": new_id}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/postes/<int:pid>", methods=["PATCH"])
@require_role("admin")
def update_poste(pid):
    d = request.json or {}
    db = get_db()
    sets, params = [], []
    if "nom" in d:
        nom = (d.get("nom") or "").strip()
        if not nom: return jsonify({"error":"Nom requis"}), 400
        sets.append("nom=?"); params.append(nom)
    if "ordre" in d:
        sets.append("ordre=?"); params.append(to_int(d.get("ordre")) or 0)
    if "couleur" in d:
        c = (d.get("couleur") or "").strip()
        if c.startswith("#") and len(c) == 7:
            sets.append("couleur=?"); params.append(c)
    if not sets: return jsonify({"error":"Rien à modifier"}), 400
    params.append(pid)
    try:
        db.execute(f"UPDATE postes SET {','.join(sets)} WHERE id=?", params)
        db.commit()
        p_info = one(db.execute("SELECT nom FROM postes WHERE id=?", (pid,)))
        log_action(request.user, "UPDATE", "poste", pid, (p_info or {}).get("nom",""))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/postes/<int:pid>", methods=["DELETE"])
@require_role("admin")
def delete_poste(pid):
    db = get_db()
    # Vérifier si poste utilisé
    cnt = one(db.execute("SELECT COUNT(*) AS n FROM utilisateurs WHERE poste_id=?", (pid,)))
    if cnt and cnt["n"] > 0:
        return jsonify({"error": f"Ce poste est attribué à {cnt['n']} utilisateur(s). Retirez-le d'abord."}), 400
    p_info = one(db.execute("SELECT nom FROM postes WHERE id=?", (pid,)))
    label = (p_info or {}).get("nom", f"pid={pid}")
    db.execute("DELETE FROM postes WHERE id=?", (pid,))
    db.commit()
    log_action(request.user, "DELETE", "poste", pid, label)
    return jsonify({"ok": True})


# ══ LOGS (audit) ══
@app.route("/api/logs")
@require_role("admin")
def get_logs():
    """Liste paginée des logs avec filtres.
    
    Params : user_id, action, entity_type, depuis (YYYY-MM-DD), jusqua (YYYY-MM-DD),
             q (recherche libre sur user_nom/action/entity_label/details), limit (default 200, max 1000)
    """
    db = get_db()
    sql = "SELECT * FROM logs WHERE 1=1"
    params = []
    if request.args.get("user_id"):
        sql += " AND user_id=?"; params.append(to_int(request.args.get("user_id")))
    if request.args.get("action"):
        sql += " AND action=?"; params.append(request.args.get("action"))
    if request.args.get("entity_type"):
        sql += " AND entity_type=?"; params.append(request.args.get("entity_type"))
    if request.args.get("depuis"):
        sql += " AND date(created_at) >= date(?)"; params.append(request.args.get("depuis"))
    if request.args.get("jusqua"):
        sql += " AND date(created_at) <= date(?)"; params.append(request.args.get("jusqua"))
    if request.args.get("q"):
        qtxt = f"%{request.args.get('q')}%"
        sql += " AND (user_nom LIKE ? OR action LIKE ? OR entity_label LIKE ? OR details LIKE ?)"
        params += [qtxt, qtxt, qtxt, qtxt]
    try:
        limit = min(int(request.args.get("limit", 200)), 1000)
    except:
        limit = 200
    sql += f" ORDER BY created_at DESC, id DESC LIMIT {limit}"
    return jsonify(rows(db.execute(sql, params)))

@app.route("/api/logs/stats")
@require_role("admin")
def get_logs_stats():
    """Stats pour alimenter les filtres : actions distinctes, entity_types distincts, users distincts."""
    db = get_db()
    actions = [r["action"] for r in rows(db.execute("SELECT DISTINCT action FROM logs WHERE action != '' ORDER BY action"))]
    entity_types = [r["entity_type"] for r in rows(db.execute("SELECT DISTINCT entity_type FROM logs WHERE entity_type != '' ORDER BY entity_type"))]
    users = rows(db.execute("SELECT DISTINCT user_id, user_nom FROM logs WHERE user_id IS NOT NULL ORDER BY user_nom"))
    return jsonify({"actions": actions, "entity_types": entity_types, "users": users})



# ══ TECHNICIENS ASTREINTE ══
@app.route("/api/astreinte/techniciens")
@require_auth
def get_astreinte_techniciens():
    return jsonify(rows(get_db().execute("SELECT * FROM techniciens_astreinte ORDER BY nom")))

@app.route("/api/astreinte/techniciens", methods=["POST"])
@require_role("admin","manager")
def create_astreinte_technicien():
    d=request.json or {}
    if not d.get("nom"): return jsonify({"error":"nom requis"}),400
    db=get_db()
    try:
        db.execute("INSERT INTO techniciens_astreinte (nom,gsm,couleur) VALUES (?,?,?)",
                   (d["nom"],d.get("gsm",""),d.get("couleur","#3b82f6")))
        db.commit()
        return jsonify({"id":db.execute("SELECT last_insert_rowid()").fetchone()[0]}),201
    except Exception as e: return jsonify({"error":str(e)}),400

@app.route("/api/astreinte/techniciens/<int:tid>", methods=["PATCH"])
@require_role("admin","manager")
def update_astreinte_technicien(tid):
    d = request.json or {}
    db = get_db()
    n = safe_update(db, "techniciens_astreinte", tid, {k:v for k,v in d.items() if k in ("nom","gsm","couleur","actif")})
    if n == 0: return jsonify({"error":"Rien"}), 400
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/astreinte/techniciens/<int:tid>", methods=["DELETE"])
@require_role("admin","manager")
def delete_astreinte_technicien(tid):
    db=get_db(); db.execute("DELETE FROM techniciens_astreinte WHERE id=?",(tid,)); db.commit()
    return jsonify({"ok":True})

# ══ ASTREINTE ══
@app.route("/api/astreinte/specialites")
@require_auth
def get_astreinte_specialites():
    return jsonify(rows(get_db().execute("SELECT * FROM astreinte_specialites ORDER BY ordre,nom")))

@app.route("/api/astreinte/specialites", methods=["POST"])
@require_role("admin","manager")
def create_astreinte_specialite():
    d=request.json or {}
    if not d.get("nom"): return jsonify({"error":"nom requis"}),400
    db=get_db()
    try:
        db.execute("INSERT INTO astreinte_specialites (nom,description,ordre) VALUES (?,?,?)",
                   (d["nom"],d.get("description",""),d.get("ordre",0)))
        db.commit()
        return jsonify({"id":db.execute("SELECT last_insert_rowid()").fetchone()[0]}),201
    except Exception as e: return jsonify({"error":str(e)}),400

@app.route("/api/astreinte/specialites/<int:sid>", methods=["PATCH"])
@require_role("admin","manager")
def update_astreinte_specialite(sid):
    d=request.json or {}; db=get_db(); sets,params=[],[]
    for f in ["nom","description","ordre"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    if not sets: return jsonify({"error":"Rien"}),400
    params.append(sid); db.execute(f"UPDATE astreinte_specialites SET {chr(44).join(sets)} WHERE id=?",params); db.commit()
    return jsonify({"ok":True})

@app.route("/api/astreinte/specialites/<int:sid>", methods=["DELETE"])
@require_role("admin","manager")
def delete_astreinte_specialite(sid):
    db=get_db(); db.execute("DELETE FROM astreinte_specialites WHERE id=?",(sid,)); db.commit()
    return jsonify({"ok":True})

@app.route("/api/astreinte/planning")
@require_auth
def get_astreinte_planning():
    db=get_db()
    date_from = request.args.get("from","")
    date_to   = request.args.get("to","")
    sql = """SELECT p.*,s.nom AS specialite_nom,s.ordre
             FROM astreinte_planning p
             JOIN astreinte_specialites s ON p.specialite_id=s.id
             WHERE 1=1"""
    params=[]
    if date_from: sql+=" AND p.date>=?"; params.append(date_from)
    if date_to:   sql+=" AND p.date<=?"; params.append(date_to)
    return jsonify(rows(db.execute(sql+" ORDER BY p.date,s.ordre",params)))

@app.route("/api/astreinte/planning/today")
@require_auth
def get_astreinte_today():
    db=get_db()
    td=today()
    return jsonify(rows(db.execute("""SELECT p.*,s.nom AS specialite_nom,s.description,s.ordre
        FROM astreinte_planning p
        JOIN astreinte_specialites s ON p.specialite_id=s.id
        WHERE p.date=? ORDER BY s.ordre""",(td,))))

@app.route("/api/astreinte/planning", methods=["POST"])
@require_role("admin","manager")
def update_astreinte_planning():
    d=request.json or {}
    if not d.get("date") or not d.get("entries"): return jsonify({"error":"date et entries requis"}),400
    db=get_db()
    for entry in d["entries"]:
        sid=to_int(entry.get("specialite_id"))
        tech=entry.get("technicien","")
        if not sid: continue
        db.execute("""INSERT INTO astreinte_planning (date,specialite_id,technicien)
                      VALUES (?,?,?) ON CONFLICT(date,specialite_id) DO UPDATE SET technicien=?""",
                   (d["date"],sid,tech,tech))
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/astreinte/import", methods=["POST"])
@require_role("admin","manager")
def import_astreinte():
    try:
        import openpyxl
        f=request.files.get("file")
        if not f: return jsonify({"error":"Fichier requis"}),400
        wb=openpyxl.load_workbook(f,data_only=True)
        db=get_db()
        # Colonnes du planning (Col Excel D=4 → Back-Up Général, etc.)
        COLS = [
            (4,"Back-Up Général"),(6,"Back-Up LUXAIRPORT"),(8,"Back-Up SES"),
            (10,"Haute tension"),(12,"Basse tension"),
            (14,"Courant faible (EBRC-BGL-CFL)"),(16,"Courant faible (BIL-SES ASTRA)"),
            (18,"Courant faible (LUXAIRPORT)"),(20,"Courant faible"),
            (22,"Détection incendie BOSCH, NSC"),(24,"Détection incendie ESSER"),
            (26,"HVAC"),(28,"KNX / EIB"),(30,"LITENET/ZUMTOBEL"),
            (32,"DETECTION INCENDIE CEL"),(34,"DETECTION INCENDIE SOLELEC"),
            (36,"DETECTION INCENDIE SCHAUSS"),(38,"HVAC SOCLIMA"),
            (40,"DETECTION INCENDIE GE SOLUTION")
        ]
        # Créer spécialités si nécessaire
        for idx,(col,nom) in enumerate(COLS):
            try:
                db.execute("INSERT OR IGNORE INTO astreinte_specialites (nom,ordre) VALUES (?,?)",(nom,idx))
            except Exception: pass
        db.commit()
        # Récupérer les IDs
        specs={r["nom"]:r["id"] for r in rows(db.execute("SELECT id,nom FROM astreinte_specialites"))}
        count=0
        if "Planning Techniciens" in wb.sheetnames:
            ws=wb["Planning Techniciens"]
            for row in ws.iter_rows(min_row=4,values_only=True):
                date_val=row[0]
                if not date_val: continue
                try:
                    if hasattr(date_val,'strftime'): date_str=date_val.strftime("%Y-%m-%d")
                    else: date_str=str(date_val)[:10]
                except: continue
                for col_idx,nom in COLS:
                    tech=row[col_idx-1] if col_idx-1 < len(row) else None
                    # Filtrer : si la cellule contient un datetime/date, c'est une cellule mal formatée
                    # dans le fichier source (case vide formatée en date) → on ignore
                    if tech is None: continue
                    if hasattr(tech, 'strftime'):
                        # C'est un objet datetime ou date → on ignore (case vide formatée date)
                        continue
                    tech_str = str(tech).strip()
                    if not tech_str or tech_str == "nan": continue
                    # Garde-fou supplémentaire : si la chaîne ressemble à un timestamp ISO
                    import re as _re
                    if _re.match(r'^\d{4}-\d{2}-\d{2}([\sT]\d{2}:\d{2}(:\d{2})?)?$', tech_str):
                        continue
                    sid=specs.get(nom)
                    if sid:
                        db.execute("""INSERT INTO astreinte_planning (date,specialite_id,technicien)
                            VALUES (?,?,?) ON CONFLICT(date,specialite_id) DO UPDATE SET technicien=?""",
                            (date_str,sid,tech_str,tech_str))
                        count+=1
        db.commit()
        return jsonify({"ok":True,"imported":count})
    except ImportError: return jsonify({"error":"openpyxl non installe"}),500
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/api/astreinte/export")
@require_auth
def export_astreinte():
    try:
        import openpyxl
        from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
        db=get_db()
        specs=rows(db.execute("SELECT * FROM astreinte_specialites ORDER BY ordre"))
        planning=rows(db.execute("""SELECT p.*,s.nom AS specialite_nom FROM astreinte_planning p
            JOIN astreinte_specialites s ON p.specialite_id=s.id ORDER BY p.date,s.ordre"""))
        wb=openpyxl.Workbook()
        ws=wb.active; ws.title="Planning Astreinte"
        BF=PatternFill("solid",fgColor="244298"); WF=Font(bold=True,color="FFFFFF")
        # Headers
        headers=["Date"]+[s["nom"] for s in specs]
        ws.append(headers)
        for cell in ws[1]: cell.font=WF; cell.fill=BF; cell.alignment=Alignment(horizontal="center")
        # Data
        from itertools import groupby
        dates=sorted(set(p["date"] for p in planning))
        for date in dates:
            # Date forcée en string ISO
            row_data=[str(date)[:10] if date else ""]
            day_data={p["specialite_id"]:p["technicien"] for p in planning if p["date"]==date}
            for s in specs:
                v = day_data.get(s["id"],"")
                # Si la valeur ressemble à un timestamp (ex: '2026-04-25 00:00:00'), garder juste la date courte
                # mais surtout forcer string pour éviter conversion auto par openpyxl
                if v is None: v = ""
                v = str(v).strip()
                row_data.append(v)
            ws.append(row_data)
        # Forcer toutes les cellules en text (number_format @) pour empêcher conversion auto
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.number_format = '@'
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width=max(len(str(c.value or "")) for c in col)+4
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        as_attachment=True,download_name=f"astreinte_export_{today()}.xlsx")
    except Exception as e: return jsonify({"error":str(e)}),500


@app.route("/api/astreinte/import-direct", methods=["POST"])
@require_role("admin")
def import_astreinte_direct():
    """Import direct des données astreinte sans fichier"""
    try:
        d = request.json or {}
        specs_list = d.get("specs", [])
        data = d.get("data", {})
        if not specs_list or not data:
            return jsonify({"error": "specs et data requis"}), 400
        db = get_db()
        # Créer spécialités
        for i, nom in enumerate(specs_list):
            db.execute("INSERT OR IGNORE INTO astreinte_specialites (nom,ordre) VALUES (?,?)", (nom, i))
        db.commit()
        specs = {r["nom"]: r["id"] for r in rows(db.execute("SELECT id,nom FROM astreinte_specialites"))}
        count = 0
        for date_str, day_data in data.items():
            for nom, tech in day_data.items():
                sid = specs.get(nom)
                if sid and tech:
                    db.execute("""INSERT INTO astreinte_planning (date,specialite_id,technicien)
                        VALUES (?,?,?) ON CONFLICT(date,specialite_id) DO UPDATE SET technicien=?""",
                        (date_str, sid, tech, tech))
                    count += 1
        db.commit()
        return jsonify({"ok": True, "imported": count})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/dashboard/managers")
@require_auth
def get_dashboard_managers():
    """Retourne la liste des managers ayant des projets avec des retards"""
    db = get_db()
    td = today()
    mgrs = rows(db.execute("""
        SELECT DISTINCT u.id, u.nom
        FROM utilisateurs u
        JOIN projets p ON p.manager_id=u.id
        JOIN equipements e ON e.projet_id=p.id
        JOIN interventions i ON i.equipement_id=e.id
        WHERE i.statut NOT IN ('TERMINEE','ANNULEE')
        AND i.date_prevue IS NOT NULL AND date(i.date_prevue) < ?
        ORDER BY u.nom
    """, (td,)))
    return jsonify(mgrs)

# ══════════════════════════════════════════════════════════════════════
# WEB PUSH — Notifications PWA
# ══════════════════════════════════════════════════════════════════════
@app.route("/api/push/vapid_public_key")
def push_vapid_key():
    """Clé publique VAPID pour le frontend (endpoint non authentifié car clé publique)."""
    return jsonify({"key": VAPID_PUBLIC_KEY})


@app.route("/api/push/subscribe", methods=["POST"])
@require_auth
def push_subscribe():
    """Enregistre (ou met à jour) un abonnement push pour l'utilisateur connecté.
    Body : {endpoint, keys:{p256dh, auth}, user_agent?}"""
    u = request.user
    d = request.get_json() or {}
    endpoint = d.get("endpoint")
    keys = d.get("keys") or {}
    p256dh = keys.get("p256dh")
    auth = keys.get("auth")
    if not endpoint or not p256dh or not auth:
        return jsonify({"error":"endpoint + keys.p256dh + keys.auth requis"}), 400
    ua = (d.get("user_agent") or request.headers.get("User-Agent") or "")[:500]
    db = get_db()
    # UPSERT via ON CONFLICT (endpoint UNIQUE)
    db.execute("""
        INSERT INTO push_subscriptions (utilisateur_id, endpoint, p256dh, auth, user_agent)
        VALUES (?,?,?,?,?)
        ON CONFLICT(endpoint) DO UPDATE SET
            utilisateur_id=excluded.utilisateur_id,
            p256dh=excluded.p256dh,
            auth=excluded.auth,
            user_agent=excluded.user_agent
    """, (u["id"], endpoint, p256dh, auth, ua))
    db.commit()
    logger.info(f"[push] abonnement enregistré user={u['id']} ua={ua[:60]}")
    return jsonify({"ok": True})


@app.route("/api/push/unsubscribe", methods=["POST"])
@require_auth
def push_unsubscribe():
    """Désabonnement d'un endpoint précis."""
    u = request.user
    d = request.get_json() or {}
    endpoint = d.get("endpoint")
    if not endpoint:
        return jsonify({"error":"endpoint requis"}), 400
    db = get_db()
    db.execute("DELETE FROM push_subscriptions WHERE endpoint=? AND utilisateur_id=?", (endpoint, u["id"]))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/push/test", methods=["POST"])
@require_auth
def push_test():
    """Envoie une notification de test à l'utilisateur connecté."""
    u = request.user
    n = send_push_to_users([u["id"]], {
        "title": "Test SOCOM GMAO",
        "body": "Notification de test reçue ✓",
        "icon": "/icons/socom-icon-192.png",
        "url": "/"
    })
    return jsonify({"ok": True, "sent": n})


def send_push_to_users(user_ids, payload):
    """Envoie une notification push à tous les abonnements des utilisateurs ciblés.
    Retourne le nombre d'envois réussis. Les abonnements invalides (410/404) sont supprimés.

    payload : {title, body, icon?, url?, tag?}
    """
    if not HAS_PYWEBPUSH:
        logger.warning("[push] pywebpush non installé, push ignoré")
        return 0
    if not VAPID_PUBLIC_KEY or not VAPID_PRIVATE_KEY:
        logger.warning("[push] clés VAPID non configurées, push ignoré")
        return 0
    if not user_ids:
        return 0
    try:
        db = get_db()
        ph = ",".join(["?"] * len(user_ids))
        subs = rows(db.execute(
            f"SELECT id, endpoint, p256dh, auth FROM push_subscriptions WHERE utilisateur_id IN ({ph})",
            user_ids
        ))
    except Exception as ex:
        logger.exception(f"[push] Erreur chargement abonnements : {ex}")
        return 0

    sent = 0
    data_json = json.dumps(payload)
    for sub in subs:
        try:
            webpush(
                subscription_info={
                    "endpoint": sub["endpoint"],
                    "keys": {"p256dh": sub["p256dh"], "auth": sub["auth"]}
                },
                data=data_json,
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims={"sub": VAPID_CLAIM_EMAIL}
            )
            sent += 1
        except WebPushException as ex:
            # 410 Gone / 404 Not Found = abonnement périmé → supprimer
            status = getattr(ex.response, "status_code", None)
            if status in (404, 410):
                try:
                    db.execute("DELETE FROM push_subscriptions WHERE id=?", (sub["id"],))
                    db.commit()
                    logger.info(f"[push] abonnement supprimé (périmé) id={sub['id']}")
                except Exception: pass
            else:
                logger.warning(f"[push] Erreur envoi sub={sub['id']} : {ex}")
        except Exception as ex:
            logger.exception(f"[push] Erreur inattendue sub={sub['id']} : {ex}")
    logger.info(f"[push] {sent}/{len(subs)} notifications envoyées à {len(user_ids)} user(s)")
    return sent


def _notify_intervention(iv_id, event_type, details="", extra_user_ids=None):
    """Envoie une notification push à tous les techniciens concernés par une intervention.
    event_type : 'assigned' (nouveau), 'modified' (changement date/heure), 'removed' (retiré)
    details : message complémentaire (ex: nouvelle date)
    extra_user_ids : liste d'ids supplémentaires à notifier (ex: tech d'un créneau spécifique)
    """
    try:
        db = get_db()
        iv = one(db.execute(
            """SELECT i.id, i.numero, i.type, i.date_prevue, i.heure_prevue, i.technicien_id,
                      e.designation AS equip_nom, p.nom AS projet_nom, p.numero AS projet_numero
               FROM interventions i
               LEFT JOIN equipements e ON i.equipement_id=e.id
               LEFT JOIN projets p ON e.projet_id=p.id
               WHERE i.id=?""", (iv_id,)
        ))
        if not iv: return
        user_ids = set()
        if iv.get("technicien_id"): user_ids.add(iv["technicien_id"])
        try:
            rows_it = rows(db.execute(
                "SELECT utilisateur_id FROM intervention_techniciens WHERE intervention_id=?",
                (iv_id,)))
            for r in rows_it:
                if r.get("utilisateur_id"): user_ids.add(r["utilisateur_id"])
        except Exception: pass
        try:
            rows_cr = rows(db.execute(
                "SELECT DISTINCT technicien_id FROM intervention_creneaux WHERE intervention_id=? AND technicien_id IS NOT NULL",
                (iv_id,)))
            for r in rows_cr:
                if r.get("technicien_id"): user_ids.add(r["technicien_id"])
        except Exception: pass
        if extra_user_ids:
            for uid in extra_user_ids:
                if uid: user_ids.add(uid)
        if not user_ids: return

        # ═══ Format de la notification ═══
        # Title : Type + numéro (Bon maintenance BP000042)
        # Body ligne 1 : N° projet · Nom projet
        # Body ligne 2 : Date + horaires
        type_label = "Bon maintenance" if iv.get("type") == "MAINTENANCE" else "Bon dépannage"
        if iv.get("numero"): type_label += " " + iv["numero"]
        projet_line = ""
        if iv.get("projet_numero") and iv.get("projet_nom"):
            projet_line = f"{iv['projet_numero']} · {iv['projet_nom']}"
        elif iv.get("projet_nom"):
            projet_line = iv["projet_nom"]
        elif iv.get("projet_numero"):
            projet_line = iv["projet_numero"]
        date_line = ""
        if iv.get("date_prevue"):
            try:
                parts = iv["date_prevue"][:10].split("-")
                if len(parts) == 3:
                    date_line = f"{parts[2]}/{parts[1]}/{parts[0]}"
                    if iv.get("heure_prevue"):
                        date_line += f" à {iv['heure_prevue']}"
            except Exception: pass
        body_lines = []
        if projet_line: body_lines.append(projet_line)
        if date_line: body_lines.append(date_line)
        if details: body_lines.append(details)
        payload = {
            "title": type_label,
            "body": "\n".join(body_lines),
            "icon": "/icons/socom-icon-192.png",
            "url": f"/mobile#iv/{iv_id}",
            "tag": f"iv-{iv_id}"
        }
        send_push_to_users(list(user_ids), payload)
    except Exception as ex:
        logger.exception(f"[push notify_intervention] Erreur : {ex}")


@app.route("/api/dashboard/heures")
@require_auth
def get_dashboard_heures():
    """Retourne le total des heures groupées par catégorie avec filtrage par équipe."""
    u = request.user
    db = get_db()
    today_str = today()
    year = today_str[:4]
    debut = request.args.get("debut") or f"{year}-01-01"
    fin = request.args.get("fin") or f"{year}-12-31"
    logger.info(f"[dashboard/heures] user={u['id']} role={u['role']} debut={debut} fin={fin}")

    try:
        # Déterminer la liste des techniciens visibles selon le rôle
        if u["role"] in ("admin", "superadmin"):
            tech_ids = None  # None = pas de filtre
        elif u["role"] == "manager":
            team = rows(db.execute("SELECT id FROM utilisateurs WHERE manager_id=?", (u["id"],)))
            tech_ids = [u["id"]] + [t["id"] for t in team]
            logger.info(f"[dashboard/heures] manager team = {tech_ids}")
        else:  # technicien
            tech_ids = [u["id"]]

        # Si aucun tech visible (cas impossible sauf bug)
        if tech_ids is not None and not tech_ids:
            return jsonify({"debut":debut, "fin":fin, "bc":0, "bp":0, "occupations":[]})

        def _sum_bons(type_bon):
            """Somme des heures des intervenants CR pour un type de bon."""
            if tech_ids is None:
                r = one(db.execute("""
                    SELECT COALESCE(SUM(ci.total_heures),0) AS h
                    FROM cr_intervenants ci
                    JOIN comptes_rendus cr ON ci.cr_id=cr.id
                    JOIN interventions i ON cr.intervention_id=i.id
                    WHERE i.type=?
                      AND (
                          (cr.date_intervention BETWEEN ? AND ?)
                       OR (NULLIF(ci.date,'') BETWEEN ? AND ?)
                      )
                """, (type_bon, debut, fin, debut, fin)))
            else:
                ph = ",".join(["?"] * len(tech_ids))
                r = one(db.execute(f"""
                    SELECT COALESCE(SUM(ci.total_heures),0) AS h
                    FROM cr_intervenants ci
                    JOIN comptes_rendus cr ON ci.cr_id=cr.id
                    JOIN interventions i ON cr.intervention_id=i.id
                    WHERE i.type=?
                      AND (
                          (cr.date_intervention BETWEEN ? AND ?)
                       OR (NULLIF(ci.date,'') BETWEEN ? AND ?)
                      )
                      AND ci.utilisateur_id IN ({ph})
                """, [type_bon, debut, fin, debut, fin] + tech_ids))
            return float((r or {}).get("h") or 0)

        heures_bc = _sum_bons("DEPANNAGE")
        heures_bp = _sum_bons("MAINTENANCE")
        logger.info(f"[dashboard/heures] bc={heures_bc} bp={heures_bp}")

        # Occupations par type (fallback : utiliser total_heures stocké, sinon calculer depuis horaires)
        if tech_ids is None:
            # Vue globale : 1 fois par occupation (pas de jointure liaison sinon double comptage)
            occs = rows(db.execute("""
                SELECT ot.nom AS type_nom,
                       ot.couleur AS type_couleur,
                       COALESCE(SUM(o.total_heures),0) AS h
                FROM occupations o
                JOIN occupation_types ot ON o.type_id=ot.id
                WHERE o.date BETWEEN ? AND ?
                GROUP BY ot.id, ot.nom, ot.couleur
                HAVING h > 0
                ORDER BY h DESC
            """, (debut, fin)))
        else:
            # v218.57 : filtre par tech via occupation_techniciens (multi-tech).
            # Chaque tech voit ses heures comptabilisées (ex: occ 8h × 3 techs = 8h pour chaque tech).
            ph = ",".join(["?"] * len(tech_ids))
            occs = rows(db.execute(f"""
                SELECT ot.nom AS type_nom,
                       ot.couleur AS type_couleur,
                       COALESCE(SUM(o.total_heures),0) AS h
                FROM occupations o
                JOIN occupation_types ot ON o.type_id=ot.id
                JOIN occupation_techniciens link ON link.occupation_id = o.id
                WHERE o.date BETWEEN ? AND ?
                  AND link.technicien_id IN ({ph})
                GROUP BY ot.id, ot.nom, ot.couleur
                HAVING h > 0
                ORDER BY h DESC
            """, [debut, fin] + tech_ids))
        logger.info(f"[dashboard/heures] occupations count={len(occs)}")

        for o in occs:
            o["h"] = round(float(o.get("h") or 0), 2)

        return jsonify({
            "debut": debut,
            "fin": fin,
            "bc": round(heures_bc, 2),
            "bp": round(heures_bp, 2),
            "occupations": occs
        })
    except Exception as ex:
        logger.exception(f"[dashboard/heures] ERREUR : {ex}")
        return jsonify({"error": str(ex)}), 500


# ══ COMPTEURS ══
@app.route("/api/compteur_types")
@require_auth
def get_compteur_types():
    return jsonify(rows(get_db().execute("SELECT * FROM compteur_types ORDER BY nom")))

@app.route("/api/compteur_types", methods=["POST"])
@require_role("admin","manager")
def create_compteur_type():
    d=request.json or {}
    if not d.get("nom"): return jsonify({"error":"nom requis"}),400
    db=get_db()
    try:
        db.execute("INSERT INTO compteur_types (nom,description) VALUES (?,?)",(d["nom"],d.get("description","")))
        db.commit()
        return jsonify({"id":db.execute("SELECT last_insert_rowid()").fetchone()[0]}),201
    except Exception as e: return jsonify({"error":str(e)}),400

@app.route("/api/compteur_types/<int:tid>", methods=["PATCH"])
@require_role("admin","manager")
def update_compteur_type(tid):
    d=request.json or {}; db=get_db(); sets,params=[],[]
    for f in ["nom","description"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    if not sets: return jsonify({"error":"Rien"}),400
    params.append(tid); db.execute(f"UPDATE compteur_types SET {chr(44).join(sets)} WHERE id=?",params); db.commit()
    return jsonify({"ok":True})

@app.route("/api/compteur_types/<int:tid>", methods=["DELETE"])
@require_role("admin","manager")
def delete_compteur_type(tid):
    db=get_db(); db.execute("DELETE FROM compteur_types WHERE id=?",(tid,)); db.commit()
    return jsonify({"ok":True})

@app.route("/api/compteurs")
@require_auth
def get_compteurs():
    eq_id=to_int(request.args.get("equipement_id"))
    if eq_id:
        return jsonify(rows(get_db().execute("""
            SELECT c.*,ct.nom AS type_nom FROM compteurs c
            LEFT JOIN compteur_types ct ON c.type_id=ct.id
            WHERE c.equipement_id=? ORDER BY c.nom""",(eq_id,))))
    return jsonify(rows(get_db().execute("""
        SELECT c.*,ct.nom AS type_nom FROM compteurs c
        LEFT JOIN compteur_types ct ON c.type_id=ct.id ORDER BY c.nom""")))

@app.route("/api/compteurs", methods=["POST"])
@require_role("admin","manager","technicien")
def create_compteur():
    d=request.json or {}
    if not d.get("equipement_id") or not d.get("nom"):
        return jsonify({"error":"equipement_id et nom requis"}),400
    db=get_db()
    db.execute("INSERT INTO compteurs (equipement_id,type_id,nom,numero,unite,localisation) VALUES (?,?,?,?,?,?)",
               (d["equipement_id"],d.get("type_id"),d["nom"],d.get("numero",""),d.get("unite",""),d.get("localisation","")))
    db.commit()
    return jsonify({"id":db.execute("SELECT last_insert_rowid()").fetchone()[0]}),201

@app.route("/api/compteurs/<int:cid>", methods=["PATCH"])
@require_role("admin","manager","technicien")
def update_compteur(cid):
    d=request.json or {}; db=get_db(); sets,params=[],[]
    for f in ["type_id","nom","numero","unite","localisation"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    if not sets: return jsonify({"error":"Rien"}),400
    params.append(cid); db.execute(f"UPDATE compteurs SET {chr(44).join(sets)} WHERE id=?",params); db.commit()
    return jsonify({"ok":True})

@app.route("/api/compteurs/<int:cid>", methods=["DELETE"])
@require_role("admin","manager")
def delete_compteur(cid):
    db=get_db(); db.execute("DELETE FROM compteurs WHERE id=?",(cid,)); db.commit()
    return jsonify({"ok":True})


@app.route("/api/interventions/<int:iid>/compteurs")
@require_auth
def get_intervention_compteurs(iid):
    db=get_db()
    iv=one(db.execute("SELECT equipement_id FROM interventions WHERE id=?",(iid,)))
    if not iv: return jsonify([])
    return jsonify(rows(db.execute("SELECT c.*,ct.nom AS type_nom FROM compteurs c LEFT JOIN compteur_types ct ON c.type_id=ct.id WHERE c.equipement_id=? ORDER BY c.nom",(iv["equipement_id"],))))

@app.route("/api/releves_compteurs/<int:cr_id>")
@require_auth
def get_releves(cr_id):
    return jsonify(rows(get_db().execute("SELECT r.*,c.nom AS compteur_nom,c.numero,c.unite,c.localisation,ct.nom AS type_nom FROM releves_compteurs r JOIN compteurs c ON r.compteur_id=c.id LEFT JOIN compteur_types ct ON c.type_id=ct.id WHERE r.cr_id=? ORDER BY c.nom",(cr_id,))))

@app.route("/api/releves_compteurs",methods=["POST"])
@require_auth
def save_releves():
    d=request.json or {}
    cr_id=to_int(d.get("cr_id")); releves=d.get("releves",[])
    if not cr_id: return jsonify({"error":"cr_id requis"}),400
    db=get_db()
    for r in releves:
        cid=to_int(r.get("compteur_id"))
        if not cid: continue
        existing=one(db.execute("SELECT id FROM releves_compteurs WHERE cr_id=? AND compteur_id=?",(cr_id,cid)))
        if existing:
            db.execute("UPDATE releves_compteurs SET valeur=?,date_releve=?,notes=? WHERE id=?",(r.get("valeur"),r.get("date_releve",""),r.get("notes",""),existing["id"]))
        else:
            db.execute("INSERT INTO releves_compteurs (cr_id,compteur_id,valeur,date_releve,notes) VALUES (?,?,?,?,?)",(cr_id,cid,r.get("valeur"),r.get("date_releve",""),r.get("notes","")))
    db.commit()
    return jsonify({"ok":True})


# ══ PERSONNALISATION ══
@app.route("/api/personnalisation", methods=["GET"])
@require_auth
def get_personnalisation():
    db = get_db()
    rows_list = rows(db.execute("SELECT cle, valeur FROM parametres_app"))
    return jsonify({r["cle"]: r["valeur"] for r in rows_list})

@app.route("/api/personnalisation", methods=["POST"])
@require_role("admin")
def save_personnalisation():
    d = request.json or {}
    db = get_db()
    for cle, valeur in d.items():
        existing = one(db.execute("SELECT id FROM parametres_app WHERE cle=?", (cle,)))
        if existing:
            db.execute("UPDATE parametres_app SET valeur=? WHERE cle=?", (valeur, cle))
        else:
            db.execute("INSERT INTO parametres_app (cle, valeur) VALUES (?,?)", (cle, valeur))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/personnalisation/logo", methods=["POST"])
@require_role("admin")
def upload_logo():
    d = request.json or {}
    logo_data = d.get("logo", "")
    db = get_db()
    existing = one(db.execute("SELECT id FROM parametres_app WHERE cle='logo'"))
    if existing:
        db.execute("UPDATE parametres_app SET valeur=? WHERE cle='logo'", (logo_data,))
    else:
        db.execute("INSERT INTO parametres_app (cle, valeur) VALUES ('logo',?)", (logo_data,))
    db.commit()
    return jsonify({"ok": True})


# ══ DJU ══
@app.route("/api/dju")
@require_auth
def get_dju():
    """Calcule les DJU chaud et froid depuis Open-Meteo pour un mois donné"""
    import urllib.request, json as json_lib, math
    lat = request.args.get("lat", type=float)
    lon = request.args.get("lon", type=float)
    ville = request.args.get("ville", "")
    annee = request.args.get("annee", type=int)
    mois = request.args.get("mois", type=int)
    if not lat or not lon or not annee or not mois:
        # Géocoder la ville si lat/lon manquants
        if ville:
            try:
                geo_url = f"https://geocoding-api.open-meteo.com/v1/search?name={urllib.parse.quote(ville)}&count=1&language=fr"
                with urllib.request.urlopen(geo_url, timeout=5) as resp:
                    geo = json_lib.loads(resp.read())
                if geo.get("results"):
                    lat = geo["results"][0]["latitude"]
                    lon = geo["results"][0]["longitude"]
            except Exception as e:
                return jsonify({"error": f"Géocodage échoué: {str(e)}"}), 400
        if not lat or not lon:
            return jsonify({"error": "lat/lon ou ville requis"}), 400
    # Construire les dates du mois
    import calendar
    nb_jours = calendar.monthrange(annee, mois)[1]
    date_debut = f"{annee}-{mois:02d}-01"
    date_fin = f"{annee}-{mois:02d}-{nb_jours:02d}"
    try:
        import urllib.parse
        url = (f"https://archive-api.open-meteo.com/v1/archive?"
               f"latitude={lat}&longitude={lon}"
               f"&start_date={date_debut}&end_date={date_fin}"
               f"&daily=temperature_2m_max,temperature_2m_min"
               f"&timezone=Europe%2FParis")
        with urllib.request.urlopen(url, timeout=10) as resp:
            data = json_lib.loads(resp.read())
        t_max = data["daily"]["temperature_2m_max"]
        t_min = data["daily"]["temperature_2m_min"]
        BASE_CHAUD = 18.0
        BASE_FROID = 18.0
        dju_chaud = 0.0
        dju_froid = 0.0
        for tmax, tmin in zip(t_max, t_min):
            if tmax is None or tmin is None:
                continue
            t_moy = (tmax + tmin) / 2
            if t_moy > BASE_CHAUD:
                dju_chaud += t_moy - BASE_CHAUD
            elif t_moy < BASE_FROID:
                dju_froid += BASE_FROID - t_moy
        return jsonify({
            "dju_chaud": round(dju_chaud, 2),
            "dju_froid": round(dju_froid, 2),
            "lat": lat,
            "lon": lon,
            "annee": annee,
            "mois": mois,
            "nb_jours": nb_jours
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/geocode")
@require_auth  
def geocode():
    """Géocode une ville via Open-Meteo"""
    import urllib.request, urllib.parse, json as json_lib
    ville = request.args.get("ville", "")
    cp = request.args.get("cp", "")
    q = cp + " " + ville if cp else ville
    if not q.strip():
        return jsonify({"error": "ville requis"}), 400
    try:
        url = f"https://geocoding-api.open-meteo.com/v1/search?name={urllib.parse.quote(q.strip())}&count=1&language=fr"
        with urllib.request.urlopen(url, timeout=5) as resp:
            geo = json_lib.loads(resp.read())
        if geo.get("results"):
            r = geo["results"][0]
            return jsonify({"lat": r["latitude"], "lon": r["longitude"], "nom": r.get("name",""), "pays": r.get("country","")})
        return jsonify({"error": "Ville non trouvée"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══ GRAPHIQUES CONFIG ══
@app.route("/api/graphiques/<int:projet_id>")
@require_auth
def get_graphiques(projet_id):
    import json as _json
    rows_g = rows(get_db().execute("SELECT * FROM graphiques_config WHERE projet_id=? ORDER BY ordre,id",(projet_id,)))
    for g in rows_g:
        try: g["series"] = _json.loads(g["series"])
        except: g["series"] = []
    return jsonify(rows_g)

@app.route("/api/graphiques", methods=["POST"])
@require_role("admin","manager")
def create_graphique():
    import json as _json
    d = request.json or {}
    if not d.get("projet_id"): return jsonify({"error":"projet_id requis"}),400
    db = get_db()
    db.execute("INSERT INTO graphiques_config (projet_id,titre,ordre,series,comparaison) VALUES (?,?,?,?,?)",
               (d["projet_id"],d.get("titre","Graphique"),d.get("ordre",0),_json.dumps(d.get("series",[])),d.get("comparaison",0)))
    db.commit()
    return jsonify({"id":db.execute("SELECT last_insert_rowid()").fetchone()[0]}),201

@app.route("/api/graphiques/<int:gid>", methods=["PATCH"])
@require_role("admin","manager")
def update_graphique(gid):
    import json as _json
    d = request.json or {}; db = get_db()
    sets,params = [],[]
    if "titre" in d: sets.append("titre=?"); params.append(d["titre"])
    if "ordre" in d: sets.append("ordre=?"); params.append(d["ordre"])
    if "series" in d: sets.append("series=?"); params.append(_json.dumps(d["series"]))
    if "comparaison" in d: sets.append("comparaison=?"); params.append(d["comparaison"])
    if not sets: return jsonify({"error":"Rien"}),400
    params.append(gid)
    db.execute(f"UPDATE graphiques_config SET {chr(44).join(sets)} WHERE id=?",params)
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/graphiques/<int:gid>", methods=["DELETE"])
@require_role("admin","manager")
def delete_graphique(gid):
    db = get_db()
    db.execute("DELETE FROM graphiques_config WHERE id=?",(gid,))
    db.commit()
    return jsonify({"ok":True})


# ══ ANALYSES RAPPORT ══
@app.route("/api/analyses/<int:projet_id>/<int:annee>")
@require_auth
def get_analyses(projet_id, annee):
    rows_a = rows(get_db().execute(
        "SELECT * FROM analyses_rapport WHERE projet_id=? AND annee=?",
        (projet_id, annee)))
    return jsonify({r["type_analyse"]: r["texte"] for r in rows_a})

@app.route("/api/analyses/<int:projet_id>/<int:annee>", methods=["POST"])
@require_auth
def save_analyse(projet_id, annee):
    d = request.json or {}
    type_analyse = d.get("type_analyse","")
    texte = d.get("texte","")
    if not type_analyse: return jsonify({"error":"type_analyse requis"}),400
    db = get_db()
    db.execute("""INSERT INTO analyses_rapport (projet_id,annee,type_analyse,texte,updated_at)
        VALUES (?,?,?,?,datetime('now'))
        ON CONFLICT(projet_id,annee,type_analyse)
        DO UPDATE SET texte=excluded.texte, updated_at=excluded.updated_at""",
        (projet_id, annee, type_analyse, texte))
    db.commit()
    return jsonify({"ok": True})


# ══ PARAMETRES RAPPORT ══
@app.route("/api/parametres_rapport/<int:projet_id>/<int:annee>")
@require_auth
def get_parametres_rapport(projet_id, annee):
    rows_p = rows(get_db().execute(
        "SELECT * FROM parametres_rapport WHERE projet_id=? AND annee=? ORDER BY chapitre,ordre,id",
        (projet_id, annee)))
    result = {}
    for r in rows_p:
        ch = r["chapitre"]
        if ch not in result: result[ch] = []
        result[ch].append({"id":r["id"],"parametre":r["parametre"],"valeur":r["valeur"],"ordre":r["ordre"]})
    return jsonify(result)

@app.route("/api/parametres_rapport", methods=["POST"])
@require_auth
def create_parametre_rapport():
    d = request.json or {}
    if not d.get("projet_id") or not d.get("chapitre"):
        return jsonify({"error":"projet_id et chapitre requis"}),400
    db = get_db()
    db.execute("INSERT INTO parametres_rapport (projet_id,annee,chapitre,parametre,valeur,ordre) VALUES (?,?,?,?,?,?)",
               (d["projet_id"],d.get("annee",0),d["chapitre"],d.get("parametre",""),d.get("valeur",""),d.get("ordre",0)))
    db.commit()
    return jsonify({"id":db.execute("SELECT last_insert_rowid()").fetchone()[0]}),201

@app.route("/api/parametres_rapport/<int:rid>", methods=["PATCH"])
@require_auth
def update_parametre_rapport(rid):
    d = request.json or {}; db = get_db(); sets,params=[],[]
    for f in ["parametre","valeur","ordre"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    if not sets: return jsonify({"error":"Rien"}),400
    params.append(rid)
    db.execute(f"UPDATE parametres_rapport SET {chr(44).join(sets)} WHERE id=?",params)
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/parametres_rapport/<int:rid>", methods=["DELETE"])
@require_auth
def delete_parametre_rapport(rid):
    db = get_db()
    db.execute("DELETE FROM parametres_rapport WHERE id=?",(rid,))
    db.commit()
    return jsonify({"ok":True})


# ══ RAPPORT ENERGETIQUE PDF COMPLET ══
@app.route("/api/rapport-energie/<int:projet_id>/pdf-complet", methods=["POST"])
@require_auth
def generate_rapport_complet(projet_id):
    try:
        import matplotlib; matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        from matplotlib.patches import FancyBboxPatch, Circle, Wedge
        import numpy as np
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, PageBreak, HRFlowable, KeepTogether
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.utils import ImageReader
        import base64, io, calendar, re as _re
        from datetime import datetime
        from itertools import groupby

        db = get_db()
        proj = one(db.execute("SELECT * FROM projets WHERE id=?", (projet_id,)))
        if not proj: return jsonify({"error":"Projet non trouve"}),404
        data = request.json or {}
        annee = int(data.get("annee", datetime.now().year))
        PAGE_W, PAGE_H = landscape(A4)
        BLUE = colors.HexColor("#213e9a")  # SOCOM Blue
        ACCENT = colors.HexColor("#213e9a")  # SOCOM Blue
        LIGHT = colors.HexColor("#e8ecf8")  # SOCOM Blue light
        BORDER = colors.HexColor("#c5cef0")  # SOCOM Blue border
        buf = io.BytesIO()
        buf_garde = io.BytesIO()
        buf_data = io.BytesIO()
        proj_nom = proj.get("nom","") or ""
        proj_loc = (str(proj.get("ville",""))+" "+str(proj.get("code_postal",""))).strip() if proj.get("ville") else ""

        # ═══════════════════════════════════════
        # PAGE DE GARDE (canvas dans buf_garde)
        # ═══════════════════════════════════════
        cv = rl_canvas.Canvas(buf_garde, pagesize=landscape(A4))
        SOCOM_BLUE = "#213e9a"
        SOCOM_DARK = "#0d1728"

        # ══ FOND BLANC ══
        cv.setFillColor(colors.white)
        cv.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0)

        # ══ IMAGE DROITE ══
        import PIL.Image as PILImage
        import numpy as np2
        try:
            _img_path = "/var/www/gmao/rapport_cover.jpg"
            import os
            if not os.path.exists(_img_path):
                raise FileNotFoundError
            _pimg = PILImage.open(_img_path).convert("RGB")
        except Exception:
            _pimg = None

        if _pimg:
            _iw, _ih = _pimg.size
            _iz_x = PAGE_W * 0.40; _iz_w = PAGE_W * 0.60; _iz_h = PAGE_H
            _rz = _iz_w / _iz_h; _ri = _iw / _ih
            if _ri > _rz:
                _nw = int(_ih * _rz); _off = (_iw - _nw) // 2
                _ic = _pimg.crop((_off, 0, _off + _nw, _ih))
            else:
                _nh = int(_iw / _rz); _off = (_ih - _nh) // 2
                _ic = _pimg.crop((0, _off, _iw, _off + _nh))
            _ic = _ic.resize((int(_iz_w * 2), int(_iz_h * 2)), PILImage.LANCZOS)
            _ib = io.BytesIO(); _ic.save(_ib, format="JPEG", quality=90); _ib.seek(0)
            cv.drawImage(ImageReader(_ib), _iz_x, 0, width=_iz_w, height=_iz_h, preserveAspectRatio=False)
            # Fondu blanc
            _fw = 130
            _fp = np2.zeros((int(_iz_h * 3), _fw, 4), dtype=np2.uint8)
            for _x in range(_fw):
                _a = int(255 * (1 - _x/_fw) ** 1.6)
                _fp[:, _x, :] = [255, 255, 255, _a]
            _fi = PILImage.fromarray(_fp, "RGBA")
            _fb = io.BytesIO(); _fi.save(_fb, format="PNG"); _fb.seek(0)
            cv.drawImage(ImageReader(_fb), _iz_x - 5, 0, width=_fw/3, height=_iz_h, mask="auto")

        # ══ ZONE BLANCHE GAUCHE ══
        cv.setFillColor(colors.white)
        cv.rect(0, 0, PAGE_W * 0.415, PAGE_H, fill=1, stroke=0)
        cv.setFillColor(colors.HexColor(SOCOM_BLUE))
        cv.rect(PAGE_W * 0.415, 0, 3, PAGE_H, fill=1, stroke=0)
        cv.rect(0, PAGE_H - 5, PAGE_W * 0.415, 5, fill=1, stroke=0)

        # ══ LOGO SOCOM TEXTE ══
        try:
            html_c2 = open("/var/www/gmao/index.html").read()
            import re as _re2
            _matches = list(_re2.finditer(r'src="data:image/([^;]+);base64,([^"]{100,})"', html_c2))
            import base64 as _b64
            # Image 1 = logo texte 650x148
            if len(_matches) >= 2:
                _logo_bytes = _b64.b64decode(_matches[1].group(2))
                _limg = ImageReader(io.BytesIO(_logo_bytes))
            else:
                _limg = ImageReader(io.BytesIO(_b64.b64decode(_matches[0].group(2))))
            _lw = 200; _lh = _lw * 148/650
            cv.drawImage(_limg, 36, PAGE_H - _lh - 28, width=_lw, height=_lh)
        except Exception: pass

        # Ligne sous logo
        cv.setStrokeColor(colors.HexColor("#e2e8f0"))
        cv.setLineWidth(0.8)
        cv.line(36, PAGE_H - 80, PAGE_W * 0.40, PAGE_H - 80)

        # ══ TITRE ══
        cv.setFillColor(colors.HexColor(SOCOM_DARK))
        cv.setFont("Helvetica-Bold", 36)
        cv.drawString(36, PAGE_H * 0.535, "SUIVI")
        cv.setFont("Helvetica-Bold", 36)
        cv.drawString(36, PAGE_H * 0.455, "ENERGETIQUE")

        # Ligne accent
        cv.setStrokeColor(colors.HexColor(SOCOM_BLUE))
        cv.setLineWidth(3)
        cv.line(36, PAGE_H * 0.432, 220, PAGE_H * 0.432)

        # Sous-titre
        cv.setFillColor(colors.HexColor("#64748b"))
        cv.setFont("Helvetica", 10)
        cv.drawString(36, PAGE_H * 0.395, "Analyse & Comparaison annuelle")

        # Séparateur
        cv.setStrokeColor(colors.HexColor("#e2e8f0"))
        cv.setLineWidth(0.8)
        cv.line(36, PAGE_H * 0.368, PAGE_W * 0.38, PAGE_H * 0.368)

        # ══ INFOS PROJET ══
        cv.setFillColor(colors.HexColor(SOCOM_DARK))
        cv.setFont("Helvetica-Bold", 18)
        cv.drawString(36, PAGE_H * 0.330, proj_nom)
        cv.setFillColor(colors.HexColor("#64748b"))
        cv.setFont("Helvetica", 10)
        if proj_loc: cv.drawString(36, PAGE_H * 0.290, proj_loc)
        cv.drawString(36, PAGE_H * 0.255, "Annee "+str(annee-1)+" vs "+str(annee))

        # ══ PIED DE PAGE ══
        cv.setFillColor(colors.HexColor(SOCOM_BLUE))
        cv.rect(0, 0, PAGE_W * 0.415, 22, fill=1, stroke=0)
        cv.setFillColor(colors.HexColor("#1a3180"))
        cv.rect(0, 0, 4, 22, fill=1, stroke=0)
        cv.setFillColor(colors.white)
        cv.setFont("Helvetica-Bold", 7.5)
        cv.drawString(12, 7, "SOCOM - Maintenance Multi-Techniques")
        cv.setFillColor(colors.HexColor("#64748b"))
        cv.setFont("Helvetica", 7.5)
        cv.drawString(PAGE_W-155, 7, "Genere le "+datetime.now().strftime("%d/%m/%Y"))

        cv.showPage()
        # Finaliser la page de garde
        cv.save()

        # ═══════════════════════════════════════
        # PAGES DE DONNÉES
        # ═══════════════════════════════════════
        styles2=getSampleStyleSheet()
        title_s=ParagraphStyle("T",parent=styles2["Heading1"],fontSize=13,textColor=BLUE,spaceBefore=0,spaceAfter=8)
        section_s=ParagraphStyle("S",parent=styles2["Heading2"],fontSize=10,textColor=ACCENT,spaceBefore=10,spaceAfter=4)
        normal_s=ParagraphStyle("N",parent=styles2["Normal"],fontSize=8,spaceAfter=4)
        cell_s=ParagraphStyle("C",parent=styles2["Normal"],fontSize=7,wordWrap="CJK")

        def mk_cell(txt):
            return Paragraph(str(txt) if txt else "—", cell_s)

        def mk_table(data_r, total_w=None):
            # Largeurs adaptatives : calculer max chars par colonne
            if not data_r or len(data_r)<1: return Spacer(1,1)
            n_cols=len(data_r[0])
            avail=total_w or (PAGE_W-2.4*cm)
            # Largeur min par col selon contenu
            col_chars=[0]*n_cols
            for row in data_r:
                for j,cell in enumerate(row):
                    txt=str(cell) if cell else ""
                    col_chars[j]=max(col_chars[j],len(txt))
            total_chars=sum(col_chars) or 1
            cws=[max(1.5*cm, min(6*cm, avail*col_chars[j]/total_chars)) for j in range(n_cols)]
            # Normaliser pour que la somme = avail
            s_cws=sum(cws)
            cws=[w*avail/s_cws for w in cws]
            # Wrapper les cellules en Paragraph
            wrapped=[]
            for i3,row in enumerate(data_r):
                wrapped.append([mk_cell(cell) if i3>0 else Paragraph(str(cell) if cell else "",
                    ParagraphStyle("H",parent=styles2["Normal"],fontSize=7,textColor=colors.white,fontName="Helvetica-Bold")) for cell in row])
            t=Table(wrapped,colWidths=cws,repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),BLUE),
                ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                ("GRID",(0,0),(-1,-1),0.5,BORDER),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,LIGHT]),
                ("ALIGN",(0,0),(-1,-1),"CENTER"),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                ("PADDING",(0,0),(-1,-1),4),
            ]))
            return t

        # Charger données
        equips=rows(db.execute("SELECT e.* FROM equipements e WHERE e.projet_id=? AND e.type_technique='Compteurs'",(projet_id,)))
        story_p=[]
        # v218.6 : diagnostic
        logger.info(f"[RAPPORT_ENERGIE] projet_id={projet_id} annee={annee} equips_compteurs={len(equips)}")

        for eq in equips:
            eq["compteurs"]=rows(db.execute("SELECT c.*,ct.nom AS type_nom FROM compteurs c LEFT JOIN compteur_types ct ON c.type_id=ct.id WHERE c.equipement_id=? ORDER BY c.nom",(eq["id"],)))
            ivs=rows(db.execute("SELECT * FROM interventions WHERE equipement_id=? ORDER BY date_prevue",(eq["id"],)))
            cols_d=[]
            for iv in ivs:
                dp=str(iv.get("date_prevue","") or "")
                if not dp or str(dp)[:4] not in [str(annee),str(annee-1)]: continue
                # v218.5 : Récupérer les CR de cette intervention (s'il y en a)
                crs_iv = rows(db.execute("SELECT * FROM comptes_rendus WHERE intervention_id=?",(iv["id"],)))
                if crs_iv:
                    # Au moins un CR : créer une colonne par CR avec ses relevés
                    for cr in crs_iv:
                        cr["releves"]=rows(db.execute("SELECT r.*,c.nom AS cnom FROM releves_compteurs r JOIN compteurs c ON r.compteur_id=c.id WHERE r.cr_id=?",(cr["id"],)))
                        cr["numero_bt"]=str(iv.get("numero","") or ""); cr["date_planif"]=dp
                        cols_d.append(cr)
                else:
                    # Pas de CR : créer une colonne vide pour montrer l'intervention dans le tableau
                    cols_d.append({"numero_bt": str(iv.get("numero","") or ""), "date_planif": dp, "releves": [], "id": None})
            cols_d.sort(key=lambda x:str(x.get("date_planif","")))
            eq["cols"]=cols_d[:12]
            # v218.6 : diagnostic
            logger.info(f"[RAPPORT_ENERGIE]   eq id={eq['id']} '{eq.get('designation','')}' compteurs={len(eq['compteurs'])} cols={len(eq['cols'])} ivs_total={len(ivs)}")
            # v218.5 : on garde l'équipement si au moins des compteurs sont présents (même sans relevés)
            if not eq["compteurs"]:
                logger.info(f"[RAPPORT_ENERGIE]   -> SKIP (pas de compteurs)")
                continue
            if not eq["cols"]:
                logger.info(f"[RAPPORT_ENERGIE]   -> SKIP (pas de cols dans annee {annee} ou {annee-1})")
                continue

            # v218.7 : try/except autour des append pour capturer toute exception silencieuse
            try:
                story_p.append(Paragraph("Equipement : "+str(eq.get("designation","")),title_s))
                logger.info(f"[RAPPORT_ENERGIE]     append(designation) OK -> len={len(story_p)}")
                if eq.get("localisation"): story_p.append(Paragraph("Localisation : "+str(eq["localisation"]),normal_s))
                story_p.append(HRFlowable(width="100%",thickness=1,color=ACCENT,spaceAfter=6))
                cpts=eq["compteurs"]; cols=eq["cols"]

                def bt_lbl(c4): return str(c4.get("numero_bt",""))+" "+str(c4.get("date_planif",""))[:10]

                # ── Relevés ──
                story_p.append(Paragraph("Releves",section_s))
                hdrs=["Type","Compteur","Localisation","Unite"]+[bt_lbl(c4) for c4 in cols]
                rrows=[hdrs]
                for cpt in cpts:
                    row=[cpt.get("type_nom","—"),cpt.get("nom","—"),cpt.get("localisation","—"),cpt.get("unite","—")]
                    for c4 in cols:
                        r4=next((r for r in c4.get("releves",[]) if r["compteur_id"]==cpt["id"]),None)
                        row.append(str(round(r4["valeur"],1)) if r4 and r4.get("valeur") is not None else "—")
                    rrows.append(row)
                story_p.append(mk_table(rrows)); story_p.append(Spacer(1,8))
                logger.info(f"[RAPPORT_ENERGIE]     append(releves) OK -> len={len(story_p)}")
            except Exception as eq_err:
                import traceback
                logger.error(f"[RAPPORT_ENERGIE]   EXCEPTION dans append releves : {eq_err}\n{traceback.format_exc()}")
                continue

            # ── Consommations Brutes ──
            if len(cols)>1:
                story_p.append(Paragraph("Consommations Brutes",section_s))
                hdrs2=["Type","Compteur","Localisation","Unite"]+[bt_lbl(cols[i4]) for i4 in range(1,len(cols))]
                cbrows=[hdrs2]
                for cpt in cpts:
                    row=[cpt.get("type_nom","—"),cpt.get("nom","—"),cpt.get("localisation","—"),cpt.get("unite","—")]
                    for i4 in range(1,len(cols)):
                        rp4=next((r for r in cols[i4-1].get("releves",[]) if r["compteur_id"]==cpt["id"]),None)
                        rc4=next((r for r in cols[i4].get("releves",[]) if r["compteur_id"]==cpt["id"]),None)
                        if rp4 and rc4 and rp4.get("valeur") is not None and rc4.get("valeur") is not None:
                            row.append(str(round(rc4["valeur"]-rp4["valeur"],2)))
                        else: row.append("—")
                    cbrows.append(row)
                story_p.append(mk_table(cbrows)); story_p.append(Spacer(1,8))

            # ── Consommations Corrigées ──
            if len(cols)>1:
                story_p.append(Paragraph("Consommations Corrigees",section_s))
                hdrs3=["Type","Compteur","Localisation","Unite"]+[bt_lbl(cols[i4-1]) for i4 in range(1,len(cols))]
                ccrows=[hdrs3]
                for cpt in cpts:
                    row=[cpt.get("type_nom","—"),cpt.get("nom","—"),cpt.get("localisation","—"),cpt.get("unite","—")]
                    for i4 in range(1,len(cols)):
                        rp4=next((r for r in cols[i4-1].get("releves",[]) if r["compteur_id"]==cpt["id"]),None)
                        rc4=next((r for r in cols[i4].get("releves",[]) if r["compteur_id"]==cpt["id"]),None)
                        dp4s=str(cols[i4-1].get("date_planif","") or ""); dc4s=str(cols[i4].get("date_planif","") or "")
                        if rp4 and rc4 and rp4.get("valeur") is not None and rc4.get("valeur") is not None and dp4s and dc4s:
                            try:
                                dp4=datetime.strptime(dp4s[:10],"%Y-%m-%d"); dc4_=datetime.strptime(dc4s[:10],"%Y-%m-%d")
                                nj=abs((dc4_-dp4).days); njm=calendar.monthrange(dp4.year,dp4.month)[1]
                                row.append(str(round((rc4["valeur"]-rp4["valeur"])/nj*njm,2)) if nj>0 else "—")
                            except: row.append("—")
                        else: row.append("—")
                    ccrows.append(row)
                story_p.append(mk_table(ccrows))
            story_p.append(PageBreak())

        # ── Graphiques matplotlib (recalculés côté serveur) ──
        graphiques=rows(db.execute("SELECT * FROM graphiques_config WHERE projet_id=? ORDER BY ordre,id",(projet_id,)))
        import json as _json2
        if graphiques:
            # v218.9 : pas de PageBreak ici — celui de la fin de "Conso corrigées" (ligne 3311) suffit
            story_p.append(Paragraph("Graphiques",title_s))
            story_p.append(HRFlowable(width="100%",thickness=1,color=ACCENT,spaceAfter=8))
            # Charger données compteurs pour graphiques
            equips_g=rows(db.execute("SELECT e.* FROM equipements e WHERE e.projet_id=? AND e.type_technique='Compteurs'",(projet_id,)))
            cpt_data={}
            for eq in equips_g:
                cpts_g=rows(db.execute("SELECT c.*,ct.nom AS type_nom FROM compteurs c LEFT JOIN compteur_types ct ON c.type_id=ct.id WHERE c.equipement_id=?",(eq["id"],)))
                ivs_g=rows(db.execute("SELECT * FROM interventions WHERE equipement_id=? ORDER BY date_prevue",(eq["id"],)))
                cols_g=[]
                for iv in ivs_g:
                    dp=str(iv.get("date_prevue","") or "")
                    if not dp or str(dp)[:4] not in [str(annee),str(annee-1)]: continue
                    for cr in rows(db.execute("SELECT * FROM comptes_rendus WHERE intervention_id=?",(iv["id"],))):
                        cr["releves"]=rows(db.execute("SELECT r.*,c.nom AS cnom FROM releves_compteurs r JOIN compteurs c ON r.compteur_id=c.id WHERE r.cr_id=?",(cr["id"],)))
                        cr["numero_bt"]=str(iv.get("numero","") or ""); cr["date_planif"]=dp
                        cols_g.append(cr)
                cols_g.sort(key=lambda x:str(x.get("date_planif","")))
                cols_g=cols_g[:12]
                for cpt in cpts_g:
                    cpt_data[cpt["id"]]={"nom":cpt["nom"],"unite":cpt.get("unite",""),"pts":[]}
                    for i4 in range(1,len(cols_g)):
                        rp4=next((r for r in cols_g[i4-1].get("releves",[]) if r["compteur_id"]==cpt["id"]),None)
                        rc4=next((r for r in cols_g[i4].get("releves",[]) if r["compteur_id"]==cpt["id"]),None)
                        dp4s=str(cols_g[i4-1].get("date_planif","") or ""); dc4s=str(cols_g[i4].get("date_planif","") or "")
                        if rp4 and rc4 and rp4.get("valeur") is not None and rc4.get("valeur") is not None and dp4s and dc4s:
                            try:
                                dp4=datetime.strptime(dp4s[:10],"%Y-%m-%d"); dc4_=datetime.strptime(dc4s[:10],"%Y-%m-%d")
                                nj=abs((dc4_-dp4).days); njm=calendar.monthrange(dp4.year,dp4.month)[1]
                                val=round((rc4["valeur"]-rp4["valeur"])/nj*njm,2) if nj>0 else None
                                lbl=dp4s[:7]
                                cpt_data[cpt["id"]]["pts"].append({"label":lbl,"value":val})
                            except Exception: pass

            COLORS_G=["#3b82f6","#ef4444","#22c55e","#f59e0b","#8b5cf6","#14b8a6","#f97316"]
            GRAPH_W = (PAGE_W - 2.4*cm) * 0.75  # Largeur -25%
            GRAPH_H = PAGE_H * 0.62 * 0.75   # Hauteur -25%

            for g5 in graphiques:
                try:
                    series5=_json2.loads(g5.get("series","[]"))
                    if not series5: continue
                    titre5 = str(g5.get("titre","Graphique"))
                    types5=[s.get("type","histogramme") for s in series5]
                    has_dual=len(set(types5))>1 and len(series5)>1

                    fig5,ax5=plt.subplots(figsize=(11,5),facecolor="white")
                    ax5.set_facecolor("#f8fafc")
                    ax5_r=ax5.twinx() if has_dual else None
                    has_data=False

                    for si5,serie5 in enumerate(series5):
                        cid5=serie5.get("compteur_id")
                        col5=COLORS_G[si5%len(COLORS_G)]
                        ax_use=ax5_r if (has_dual and si5>0) else ax5
                        pts5=[]
                        # Chercher par id int ou string
                        cid5_key=None
                        for k in cpt_data.keys():
                            if str(k)==str(cid5): cid5_key=k; break
                        if cid5_key is not None:
                            pts5=cpt_data[cid5_key]["pts"]
                        if not pts5: continue
                        lbls5=[p["label"] for p in pts5 if p["value"] is not None]
                        vals5=[p["value"] for p in pts5 if p["value"] is not None]
                        if not lbls5: continue
                        lbl5=str(serie5.get("label","") or serie5.get("nom_compteur","") or "Serie ")+str(si5+1)
                        t5=serie5.get("type","histogramme")
                        x5=range(len(lbls5))
                        if t5=="courbe":
                            ax_use.plot(x5,vals5,"o-",color=col5,lw=2.5,ms=6,label=lbl5)
                        elif t5=="histogramme_empile":
                            ax_use.bar(x5,vals5,color=col5,alpha=0.78,label=lbl5)
                        else:
                            offset=si5*0.35 if not has_dual else 0
                            ax_use.bar([x+offset for x in x5],vals5,width=0.35,color=col5,alpha=0.78,label=lbl5)
                        ax5.set_xticks(range(len(lbls5))); ax5.set_xticklabels(lbls5,rotation=30,ha="right",fontsize=8)
                        has_data=True

                    if has_data:
                        ax5.set_title(titre5,fontsize=12,fontweight="bold",color="#0d1b2a",pad=12)
                        ax5.grid(axis="y",alpha=0.3,linestyle="--",color="#e2e8f0")
                        ax5.spines["top"].set_visible(False); ax5.spines["right"].set_visible(False)
                        ax5.spines["left"].set_color("#e2e8f0"); ax5.spines["bottom"].set_color("#e2e8f0")
                        ax5.tick_params(colors="#64748b",labelsize=8)
                        handles5,labels5_=ax5.get_legend_handles_labels()
                        if ax5_r:
                            h2,l2=ax5_r.get_legend_handles_labels()
                            handles5+=h2; labels5_+=l2
                            ax5_r.spines["top"].set_visible(False)
                            ax5_r.tick_params(colors="#64748b",labelsize=8)
                        if handles5: ax5.legend(handles5,labels5_,fontsize=9,loc="upper right",framealpha=0.9,edgecolor="#e2e8f0")
                        plt.tight_layout(pad=1.5)
                        gi=io.BytesIO()
                        plt.savefig(gi,format="png",dpi=150,bbox_inches="tight",facecolor="white",edgecolor="none")
                        gi.seek(0)
                        # Un graphique par page, centré, titre et graphique ensemble
                        story_p.append(KeepTogether([
                            Spacer(1, (PAGE_H - 2.6*cm - GRAPH_H - 40) / 2),  # Centrage vertical
                            Paragraph(titre5, ParagraphStyle("GT",parent=styles2["Heading2"],fontSize=14,textColor=BLUE,spaceAfter=12,alignment=1)),
                            RLImage(gi, width=GRAPH_W, height=GRAPH_H),
                            Spacer(1,10),
                        ]))
                        story_p.append(PageBreak())
                    plt.close(fig5)
                except Exception as eg:
                    import traceback; pass

                # ── Analyses ──
        analyses=rows(db.execute("SELECT * FROM analyses_rapport WHERE projet_id=? AND annee=?",(projet_id,annee)))
        if analyses and any(a.get("texte") for a in analyses):
            story_p.append(Paragraph("Analyses",title_s))
            story_p.append(HRFlowable(width="100%",thickness=1,color=ACCENT,spaceAfter=6))
            for a2 in analyses:
                if a2.get("texte"):
                    story_p.append(Paragraph(str(a2.get("type_analyse","")),section_s))
                    story_p.append(Paragraph(str(a2["texte"]),normal_s)); story_p.append(Spacer(1,4))
            story_p.append(PageBreak())

        # ── Paramètres ──
        params=rows(db.execute("SELECT * FROM parametres_rapport WHERE projet_id=? AND annee=? ORDER BY chapitre,ordre",(projet_id,annee)))
        if params:
            story_p.append(Paragraph("Parametres",title_s))
            story_p.append(HRFlowable(width="100%",thickness=1,color=ACCENT,spaceAfter=6))
            for ch3,items3 in groupby(params,key=lambda x:x["chapitre"]):
                story_p.append(Paragraph(str(ch3),section_s))
                pr=[["Parametre","Valeur"]]
                for it in items3:
                    if it.get("parametre") or it.get("valeur"):
                        pr.append([str(it.get("parametre","")),str(it.get("valeur",""))])
                if len(pr)>1: story_p.append(mk_table(pr))
                story_p.append(Spacer(1,6))

        def on_pg(c6,doc):
            c6.saveState()
            c6.setFillColor(BLUE); c6.rect(0,PAGE_H-26,PAGE_W,26,fill=1,stroke=0)
            c6.setFillColor(colors.white); c6.setFont("Helvetica-Bold",9)
            c6.drawString(16,PAGE_H-17,"SUIVI ENERGETIQUE - "+proj_nom)
            c6.setFont("Helvetica",8); c6.drawRightString(PAGE_W-16,PAGE_H-17,"Annee "+str(annee))
            c6.setFillColor(colors.HexColor("#f8fafc")); c6.rect(0,0,PAGE_W,18,fill=1,stroke=0)
            c6.setFillColor(ACCENT); c6.rect(0,0,4,18,fill=1,stroke=0)
            c6.setFillColor(colors.HexColor("#64748b")); c6.setFont("Helvetica",7)
            c6.drawString(12,5,"SOCOM"); c6.drawRightString(PAGE_W-12,5,"Page "+str(doc.page))
            c6.restoreState()

        # v218.7 : ReportLab vide la liste story_p pendant doc.build !
        # On mémorise un flag AVANT le build pour savoir si on a des pages de données.
        has_data_pages = bool(story_p)
        if has_data_pages:
            doc=SimpleDocTemplate(buf_data,pagesize=landscape(A4),leftMargin=1.2*cm,rightMargin=1.2*cm,topMargin=1.8*cm,bottomMargin=1.2*cm)
            doc.build(story_p,onFirstPage=on_pg,onLaterPages=on_pg)
        # v218.6 : diagnostic
        logger.info(f"[RAPPORT_ENERGIE] has_data_pages={has_data_pages} -> pages de donnees {'GENEREES' if has_data_pages else 'VIDES'}")

        # ═══════════════════════════════════════
        # FUSION : page de garde + pages de données
        # ═══════════════════════════════════════
        from pypdf import PdfWriter, PdfReader
        writer = PdfWriter()
        # Page de garde
        buf_garde.seek(0)
        try:
            for page in PdfReader(buf_garde).pages:
                writer.add_page(page)
        except Exception: pass
        # Pages de données (si existantes)
        if has_data_pages:
            buf_data.seek(0)
            try:
                for page in PdfReader(buf_data).pages:
                    writer.add_page(page)
            except Exception as e_pdf:
                logger.error(f"[RAPPORT_ENERGIE] erreur lecture buf_data : {e_pdf}")
        writer.write(buf)
        buf.seek(0)
        fname="rapport_energie_"+proj_nom.replace(" ","_")+"_"+str(annee)+".pdf"
        return send_file(buf,mimetype="application/pdf",as_attachment=True,download_name=fname)

    except Exception as e:
        import traceback
        return jsonify({"error":str(e),"trace":traceback.format_exc()}),500


@app.route("/api/login",methods=["POST"])
@limiter.limit("10 per minute")
def login():
    d = request.json or {}
    email = (d.get("email","") or "").strip().lower()
    pw    = d.get("password","") or ""
    requested_societe_id = d.get("societe_id")  # v218.64 : optionnel pour multi-société
    db = get_db()
    u = one(db.execute("SELECT * FROM utilisateurs WHERE lower(email)=? AND actif=1", (email,)))
    if not u or not verify_password(pw, u["password"]):
        # Délai constant pour freiner timing attacks
        time.sleep(0.3)
        logger.warning(f"[LOGIN] echec pour {email} depuis {request.remote_addr}")
        return jsonify({"error":"Email ou mot de passe incorrect"}), 401

    # Rehash automatique vers bcrypt si MDP legacy SHA-256
    if needs_rehash(u["password"]):
        try:
            new_hash = hash_password(pw)
            db.execute("UPDATE utilisateurs SET password=? WHERE id=?", (new_hash, u["id"]))
            db.commit()
            u["password"] = new_hash
            logger.info(f"[REHASH] MDP migre vers bcrypt pour uid={u['id']}")
        except Exception as e:
            logger.warning(f"[REHASH] echec uid={u['id']}: {e}")

    # v218.64 : récupérer les sociétés actives de l'utilisateur
    # v218.72 : inclure menus_actifs pour piloter les permissions de menus côté frontend
    societes_dispo = rows(db.execute("""
        SELECT us.societe_id, us.role, s.nom, s.code, s.logo_path, s.couleur_primaire, s.menus_actifs
        FROM utilisateur_societes us
        JOIN societes s ON s.id = us.societe_id
        WHERE us.utilisateur_id = ? AND us.actif = 1 AND s.actif = 1
        ORDER BY s.nom
    """, (u["id"],)))

    if not societes_dispo:
        # Filet de sécurité : aucune société associée → comportement legacy (SOCOM)
        logger.warning(f"[LOGIN] uid={u['id']} sans société associée, fallback SOCOM")
        target_sid = 1
        societe_role = u.get("role") or "technicien"
    elif len(societes_dispo) == 1:
        # Une seule société : connexion directe
        target_sid = societes_dispo[0]["societe_id"]
        societe_role = societes_dispo[0]["role"]
    elif requested_societe_id:
        # L'utilisateur a précisé la société à utiliser
        match = next((s for s in societes_dispo if s["societe_id"] == requested_societe_id), None)
        if not match:
            return jsonify({"error":"Société non autorisée pour cet utilisateur"}), 403
        target_sid = match["societe_id"]
        societe_role = match["role"]
    else:
        # Plusieurs sociétés et pas de choix : retourner la liste pour que le client demande
        return jsonify({
            "needs_societe_choice": True,
            "user": {"id": u["id"], "nom": u["nom"], "email": u["email"]},
            "societes": [
                {
                    "id": s["societe_id"],
                    "nom": s["nom"],
                    "code": s["code"],
                    "logo_path": s.get("logo_path") or "",
                    "couleur_primaire": s.get("couleur_primaire") or "#0F1E3D",
                    "role": s["role"]
                }
                for s in societes_dispo
            ]
        })

    token = make_token(u, societe_id=target_sid)

    # v218.66 : si l'utilisateur a le rôle global superadmin (utilisateurs.role),
    # ce rôle est PRIORITAIRE sur celui dans utilisateur_societes.
    # Le superadmin est transversal aux sociétés.
    effective_role = "superadmin" if (u.get("role") == "superadmin") else societe_role

    logger.info(f"[LOGIN] succes uid={u['id']} role={effective_role} societe={target_sid} depuis {request.remote_addr}")
    log_action(u, "LOGIN", entity_type="utilisateur", entity_id=u["id"], entity_label=u["nom"],
               details=f"depuis {request.remote_addr} societe={target_sid}")

    # v218.64 : enrichir la réponse avec la société active et la liste des sociétés disponibles
    societe_active = next((s for s in societes_dispo if s["societe_id"] == target_sid), None)
    # v218.72 : parser menus_actifs (JSON string → liste). Superadmin = tous les menus.
    try:
        menus_actifs = json.loads((societe_active or {}).get("menus_actifs") or "[]")
        if not isinstance(menus_actifs, list): menus_actifs = []
    except Exception:
        menus_actifs = []
    if effective_role == "superadmin":
        menus_actifs = list(_MENUS_CONFIGURABLES)  # superadmin voit tout, partout
    return jsonify({
        "token": token,
        "user": {
            "id": u["id"],
            "nom": u["nom"],
            "email": u["email"],
            "role": effective_role,  # superadmin global OU rôle dans la société active
            "societe_id": target_sid,
        },
        "societe": {
            "id": target_sid,
            "nom": (societe_active or {}).get("nom", "SOCOM"),
            "code": (societe_active or {}).get("code", "socom"),
            "logo_path": (societe_active or {}).get("logo_path", "") or "",
            "couleur_primaire": (societe_active or {}).get("couleur_primaire", "#0F1E3D") or "#0F1E3D",
            "menus_actifs": menus_actifs,
        },
        "societes_disponibles": [
            {
                "id": s["societe_id"],
                "nom": s["nom"],
                "code": s["code"],
                "role": s["role"]
            }
            for s in societes_dispo
        ]
    })


@app.route("/api/auth/switch-societe", methods=["POST"])
@require_auth
def switch_societe():
    """Bascule l'utilisateur connecté vers une autre de ses sociétés.
    Retourne un nouveau token pointant sur la société choisie."""
    u = request.user
    d = request.json or {}
    target_sid = to_int(d.get("societe_id"))
    if not target_sid:
        return jsonify({"error": "societe_id requis"}), 400
    db = get_db()
    access = one(db.execute("""
        SELECT us.role, s.nom, s.code, s.logo_path, s.couleur_primaire, s.menus_actifs
        FROM utilisateur_societes us
        JOIN societes s ON s.id = us.societe_id
        WHERE us.utilisateur_id = ? AND us.societe_id = ? AND us.actif = 1 AND s.actif = 1
    """, (u["id"], target_sid)))
    if not access:
        return jsonify({"error": "Société non autorisée"}), 403
    new_token = make_token(u, societe_id=target_sid)
    # v218.66 : superadmin global prioritaire
    db_user = one(db.execute("SELECT role FROM utilisateurs WHERE id=?", (u["id"],)))
    effective_role = "superadmin" if (db_user and db_user.get("role") == "superadmin") else access["role"]
    # v218.72 : parser menus_actifs (JSON string → liste). Superadmin = tous les menus.
    try:
        menus_actifs = json.loads(access.get("menus_actifs") or "[]")
        if not isinstance(menus_actifs, list): menus_actifs = []
    except Exception:
        menus_actifs = []
    if effective_role == "superadmin":
        menus_actifs = list(_MENUS_CONFIGURABLES)
    logger.info(f"[SWITCH-SOCIETE] uid={u['id']} → societe={target_sid}")
    return jsonify({
        "token": new_token,
        "user": {
            "id": u["id"], "nom": u["nom"], "email": u["email"],
            "role": effective_role, "societe_id": target_sid,
        },
        "societe": {
            "id": target_sid,
            "nom": access["nom"],
            "code": access["code"],
            "logo_path": access.get("logo_path") or "",
            "couleur_primaire": access.get("couleur_primaire") or "#0F1E3D",
            "menus_actifs": menus_actifs,
        }
    })


def current_societe_id():
    """Retourne le societe_id de la requête en cours.
    Utilisé par les routes pour filtrer les données par société."""
    u = getattr(request, "user", None)
    if u and "societe_id" in u:
        return u["societe_id"]
    return 1  # fallback SOCOM


def is_superadmin(u):
    """Un user est superadmin si son rôle (utilisateurs.role) vaut 'superadmin'.
    Le rôle dans utilisateur_societes est secondaire pour cette vérification :
    on regarde le rôle global."""
    if not u: return False
    return u.get("role") == "superadmin"


def require_superadmin(fn):
    """Décorateur : seul un user avec utilisateurs.role='superadmin' peut appeler."""
    @wraps(fn)
    def wrapper(*a, **k):
        u = _authenticate()
        if not u:
            return jsonify({"error": "Non authentifie"}), 401
        if not is_superadmin(u):
            return jsonify({"error": "Acces reserve au superadmin"}), 403
        request.user = u
        return fn(*a, **k)
    return wrapper


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES SUPERADMIN — Gestion des sociétés (v218.65)
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/api/societes/public", methods=["GET"])
def get_societes_public():
    """Endpoint PUBLIC (sans auth) : liste des sociétés actives pour la page
    d'accueil avec tuiles. Ne retourne aucune info sensible — uniquement les
    éléments d'identité visuelle nécessaires à l'affichage des tuiles."""
    db = get_db()
    res = rows(db.execute("""
        SELECT id, nom, code, logo_path, couleur_primaire
        FROM societes
        WHERE actif = 1
        ORDER BY nom
    """))
    return jsonify(res)


@app.route("/api/societes", methods=["GET"])
@require_superadmin
def get_societes():
    """Liste toutes les sociétés (superadmin uniquement)."""
    db = get_db()
    res = rows(db.execute("""
        SELECT s.*,
               (SELECT COUNT(*) FROM utilisateur_societes us WHERE us.societe_id=s.id AND us.actif=1) AS nb_users,
               (SELECT COUNT(*) FROM interventions WHERE societe_id=s.id) AS nb_interventions,
               (SELECT COUNT(*) FROM equipements WHERE societe_id=s.id) AS nb_equipements,
               (SELECT COUNT(*) FROM projets WHERE societe_id=s.id) AS nb_projets
        FROM societes s
        ORDER BY s.nom
    """))
    return jsonify(res)


@app.route("/api/societes/<int:sid>", methods=["GET"])
@require_superadmin
def get_societe(sid):
    db = get_db()
    s = one(db.execute("SELECT * FROM societes WHERE id=?", (sid,)))
    if not s: return jsonify({"error": "Société introuvable"}), 404
    # Liste des utilisateurs de cette société
    users = rows(db.execute("""
        SELECT u.id, u.nom, u.email, u.actif AS user_actif,
               us.role, us.manager_id, us.techniques, us.actif AS lien_actif
        FROM utilisateur_societes us
        JOIN utilisateurs u ON u.id = us.utilisateur_id
        WHERE us.societe_id = ?
        ORDER BY u.nom
    """, (sid,)))
    s["utilisateurs"] = users
    return jsonify(s)


@app.route("/api/societes", methods=["POST"])
@require_superadmin
def create_societe():
    d = request.json or {}
    nom = (d.get("nom") or "").strip()
    if not nom: return jsonify({"error": "nom requis"}), 400
    code = (d.get("code") or "").strip().lower()
    if not code:
        # Génération automatique du code à partir du nom
        import re as _re
        code = _re.sub(r"[^a-z0-9]+", "-", nom.lower()).strip("-")
    db = get_db()
    # Vérifier unicité du code
    if one(db.execute("SELECT id FROM societes WHERE code=?", (code,))):
        return jsonify({"error": f"Code '{code}' déjà utilisé"}), 400
    # v218.72 : menus actifs essentiels par défaut à la création
    menus_default = json.dumps(_MENUS_DEFAULT_NEW_SOCIETE)
    db.execute("""
        INSERT INTO societes (nom, code, couleur_primaire, adresse, telephone, email, site_web, config_json, actif, menus_actifs)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
    """, (
        nom, code,
        d.get("couleur_primaire") or "#0F1E3D",
        d.get("adresse") or "",
        d.get("telephone") or "",
        d.get("email") or "",
        d.get("site_web") or "",
        d.get("config_json") or "{}",
        menus_default,
    ))
    new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    db.commit()
    log_action(request.user, "CREATE", "societe", new_id, nom)
    logger.info(f"[SUPERADMIN] Société créée : id={new_id} nom='{nom}' code='{code}' menus_default={_MENUS_DEFAULT_NEW_SOCIETE}")
    return jsonify({"id": new_id, "nom": nom, "code": code}), 201


@app.route("/api/societes/<int:sid>", methods=["PATCH"])
@require_superadmin
def update_societe(sid):
    d = request.json or {}
    db = get_db()
    s = one(db.execute("SELECT * FROM societes WHERE id=?", (sid,)))
    if not s: return jsonify({"error": "Société introuvable"}), 404
    sets, params = [], []
    for f in ["nom", "code", "couleur_primaire", "adresse", "telephone", "email", "site_web", "config_json", "actif"]:
        if f in d:
            sets.append(f"{f}=?")
            params.append(d[f])
    if not sets: return jsonify({"error": "Rien à modifier"}), 400
    params.append(sid)
    db.execute(f"UPDATE societes SET {','.join(sets)} WHERE id=?", params)
    db.commit()
    log_action(request.user, "UPDATE", "societe", sid, s["nom"])
    return jsonify({"ok": True})


@app.route("/api/societes/<int:sid>", methods=["DELETE"])
@require_superadmin
def delete_societe(sid):
    """Supprime une société. Bloqué si elle contient des données."""
    if sid == 1:
        return jsonify({"error": "La société principale (id=1) ne peut pas être supprimée"}), 400
    db = get_db()
    s = one(db.execute("SELECT * FROM societes WHERE id=?", (sid,)))
    if not s: return jsonify({"error": "Société introuvable"}), 404
    # Vérifier qu'aucune donnée métier n'est attachée
    counts = {}
    for table in ["interventions", "equipements", "projets", "clients", "occupations", "causeries"]:
        try:
            r = one(db.execute(f"SELECT COUNT(*) AS n FROM {table} WHERE societe_id=?", (sid,)))
            if r and r["n"] > 0:
                counts[table] = r["n"]
        except Exception:
            pass
    if counts:
        return jsonify({
            "error": "La société contient des données et ne peut pas être supprimée",
            "details": counts,
            "hint": "Désactivez la société (actif=0) ou supprimez d'abord les données liées."
        }), 400
    # OK : supprimer la société et ses liaisons utilisateurs
    db.execute("DELETE FROM utilisateur_societes WHERE societe_id=?", (sid,))
    db.execute("DELETE FROM societes WHERE id=?", (sid,))
    db.commit()
    log_action(request.user, "DELETE", "societe", sid, s["nom"])
    logger.info(f"[SUPERADMIN] Société supprimée : id={sid} nom='{s['nom']}'")
    return jsonify({"ok": True})


# ─── Permissions de menus par société (v218.72) ─────────────────────────────

@app.route("/api/societes/<int:sid>/menus", methods=["GET"])
@require_superadmin
def get_societe_menus(sid):
    """Retourne la liste des codes de menus actifs pour la société.
    Format : {"all_menus":[...], "menus_actifs":[...]} pour piloter l'UI."""
    db = get_db()
    s = one(db.execute("SELECT id, nom, menus_actifs FROM societes WHERE id=?", (sid,)))
    if not s:
        return jsonify({"error": "Société introuvable"}), 404
    try:
        actifs = json.loads(s.get("menus_actifs") or "[]")
        if not isinstance(actifs, list):
            actifs = []
    except Exception:
        actifs = []
    return jsonify({
        "id": sid,
        "nom": s["nom"],
        "all_menus": _MENUS_CONFIGURABLES,
        "menus_actifs": actifs,
    })


@app.route("/api/societes/<int:sid>/menus", methods=["PUT"])
@require_superadmin
def set_societe_menus(sid):
    """Met à jour la liste des menus actifs pour la société.
    Body : {"menus_actifs": ["interventions","equipements",...]}"""
    d = request.json or {}
    menus = d.get("menus_actifs")
    if not isinstance(menus, list):
        return jsonify({"error": "menus_actifs doit être une liste"}), 400
    # Filtrer pour ne garder que les codes valides
    menus_valides = [m for m in menus if m in _MENUS_CONFIGURABLES]
    db = get_db()
    s = one(db.execute("SELECT id, nom FROM societes WHERE id=?", (sid,)))
    if not s:
        return jsonify({"error": "Société introuvable"}), 404
    db.execute("UPDATE societes SET menus_actifs=? WHERE id=?",
               (json.dumps(menus_valides), sid))
    db.commit()
    log_action(request.user, "UPDATE", "societe-menus", sid, s["nom"], f"menus={','.join(menus_valides)}")
    return jsonify({"ok": True, "menus_actifs": menus_valides})


# ─── Page de garde personnalisée par société (v218.81) ──────────────────────

@app.route("/api/cover-page", methods=["GET"])
@require_auth
def get_cover_page():
    """Retourne la configuration de la page de garde, en-tête et pied de la société active.
    v218.85 : retourne aussi header_blocks et footer_blocks.
    Tous les utilisateurs peuvent consulter (utile pour preview), seuls
    admin et superadmin peuvent modifier (cf PUT)."""
    sid = current_societe_id()
    db = get_db()
    row = one(db.execute("""SELECT blocks_json, header_blocks_json, footer_blocks_json
                            FROM cover_pages WHERE societe_id=?""", (sid,)))
    if not row:
        # Pas encore configuré pour cette société : retourner les défauts
        return jsonify({
            "societe_id": sid,
            "blocks": list(_COVER_DEFAULT_BLOCKS),
            "header_blocks": list(_HEADER_DEFAULT_BLOCKS),
            "footer_blocks": list(_FOOTER_DEFAULT_BLOCKS),
            "available_fields": _COVER_DYNAMIC_FIELDS,
            "available_fields_hf": _HF_DYNAMIC_FIELDS,
            "is_default": True,
        })
    def _parse(s, default):
        try:
            v = json.loads(s or "[]")
            return v if isinstance(v, list) else list(default)
        except Exception:
            return list(default)
    return jsonify({
        "societe_id": sid,
        "blocks": _parse(row["blocks_json"], []),
        "header_blocks": _parse(row.get("header_blocks_json"), _HEADER_DEFAULT_BLOCKS),
        "footer_blocks": _parse(row.get("footer_blocks_json"), _FOOTER_DEFAULT_BLOCKS),
        "available_fields": _COVER_DYNAMIC_FIELDS,
        "available_fields_hf": _HF_DYNAMIC_FIELDS,
        "is_default": False,
    })


@app.route("/api/cover-page", methods=["PUT"])
@require_role("admin")
def set_cover_page():
    """Met à jour la configuration de la page de garde, en-tête et pied de la société active.
    Body: { blocks?, header_blocks?, footer_blocks? }
    Réservé aux admins et superadmins."""
    d = request.json or {}
    valid_cover_types = {"title", "subtitle", "image", "info_box", "text", "spacer", "separator",
                         "gamme_maintenance", "observations"}
    valid_hf_types = {"row", "text", "image", "spacer", "separator"}
    valid_cols = {"left", "center", "right"}

    def _clean_cover(blocks):
        out = []
        for b in blocks or []:
            if not isinstance(b, dict): continue
            if b.get("type") not in valid_cover_types: continue
            out.append(b)
        return out

    def _clean_hf(blocks):
        out = []
        for b in blocks or []:
            if not isinstance(b, dict): continue
            t = b.get("type")
            if t not in valid_hf_types: continue
            if t == "row":
                items = b.get("items") or []
                cleaned_items = []
                for it in items:
                    if not isinstance(it, dict): continue
                    if it.get("col") not in valid_cols: continue
                    if it.get("type") not in {"text", "image", "spacer"}: continue
                    cleaned_items.append(it)
                out.append({**b, "items": cleaned_items})
            else:
                out.append(b)
        return out

    sid = current_societe_id()
    db = get_db()
    # Upsert : on charge l'existant pour ne pas écraser ce qui n'est pas envoyé
    existing = one(db.execute("""SELECT blocks_json, header_blocks_json, footer_blocks_json
                                  FROM cover_pages WHERE societe_id=?""", (sid,)))
    if existing:
        cur_cover = json.loads(existing.get("blocks_json") or "[]")
        cur_header = json.loads(existing.get("header_blocks_json") or "[]")
        cur_footer = json.loads(existing.get("footer_blocks_json") or "[]")
    else:
        cur_cover = list(_COVER_DEFAULT_BLOCKS)
        cur_header = list(_HEADER_DEFAULT_BLOCKS)
        cur_footer = list(_FOOTER_DEFAULT_BLOCKS)

    if "blocks" in d:        cur_cover = _clean_cover(d.get("blocks"))
    if "header_blocks" in d: cur_header = _clean_hf(d.get("header_blocks"))
    if "footer_blocks" in d: cur_footer = _clean_hf(d.get("footer_blocks"))

    cover_json = json.dumps(cur_cover)
    header_json = json.dumps(cur_header)
    footer_json = json.dumps(cur_footer)

    if existing:
        db.execute("""UPDATE cover_pages SET
                        blocks_json=?, header_blocks_json=?, footer_blocks_json=?,
                        updated_at=datetime('now')
                      WHERE societe_id=?""",
                   (cover_json, header_json, footer_json, sid))
    else:
        db.execute("""INSERT INTO cover_pages
                        (societe_id, blocks_json, header_blocks_json, footer_blocks_json)
                        VALUES (?, ?, ?, ?)""",
                   (sid, cover_json, header_json, footer_json))
    db.commit()
    log_action(request.user, "UPDATE", "cover-page", sid, f"sid={sid}",
               f"cover={len(cur_cover)} header={len(cur_header)} footer={len(cur_footer)}")
    return jsonify({"ok": True, "blocks": cur_cover, "header_blocks": cur_header, "footer_blocks": cur_footer})


# ─── Page de garde des bons de DÉPANNAGE (BC) — v218.109 ─────────────────────
# Même structure que cover_pages mais stockée dans cover_pages_bc. La logique
# de validation est identique, on factorise.

def _cover_bc_validators():
    """Renvoie les sets de validation partagés avec cover-page."""
    valid_cover_types = {"title", "subtitle", "image", "info_box", "text", "spacer", "separator",
                         "gamme_maintenance", "observations"}
    valid_hf_types = {"row", "text", "image", "spacer", "separator"}
    valid_cols = {"left", "center", "right"}
    return valid_cover_types, valid_hf_types, valid_cols


@app.route("/api/cover-page-bc", methods=["GET"])
@require_auth
def get_cover_page_bc():
    """Page de garde des bons DÉPANNAGE (1 layout par société).
    Symétrique à get_cover_page mais sur la table cover_pages_bc."""
    sid = current_societe_id()
    db = get_db()
    row = one(db.execute("""SELECT blocks_json, header_blocks_json, footer_blocks_json
                            FROM cover_pages_bc WHERE societe_id=?""", (sid,)))
    if not row:
        return jsonify({
            "societe_id": sid,
            "blocks": list(_COVER_BC_DEFAULT_BLOCKS),
            "header_blocks": list(_HEADER_DEFAULT_BLOCKS),
            "footer_blocks": list(_FOOTER_DEFAULT_BLOCKS),
            "available_fields": _COVER_DYNAMIC_FIELDS,
            "available_fields_hf": _HF_DYNAMIC_FIELDS,
            "is_default": True,
        })
    def _parse(s, default):
        try:
            v = json.loads(s or "[]")
            return v if isinstance(v, list) else list(default)
        except Exception:
            return list(default)
    return jsonify({
        "societe_id": sid,
        "blocks": _parse(row["blocks_json"], []),
        "header_blocks": _parse(row.get("header_blocks_json"), _HEADER_DEFAULT_BLOCKS),
        "footer_blocks": _parse(row.get("footer_blocks_json"), _FOOTER_DEFAULT_BLOCKS),
        "available_fields": _COVER_DYNAMIC_FIELDS,
        "available_fields_hf": _HF_DYNAMIC_FIELDS,
        "is_default": False,
    })


@app.route("/api/cover-page-bc", methods=["PUT"])
@require_role("admin")
def set_cover_page_bc():
    """Met à jour la page de garde BC. Body: { blocks?, header_blocks?, footer_blocks? }."""
    d = request.json or {}
    valid_cover_types, valid_hf_types, valid_cols = _cover_bc_validators()

    def _clean_cover(blocks):
        out = []
        for b in blocks or []:
            if not isinstance(b, dict): continue
            if b.get("type") not in valid_cover_types: continue
            out.append(b)
        return out

    def _clean_hf(blocks):
        out = []
        for b in blocks or []:
            if not isinstance(b, dict): continue
            t = b.get("type")
            if t not in valid_hf_types: continue
            if t == "row":
                items = b.get("items") or []
                cleaned_items = []
                for it in items:
                    if not isinstance(it, dict): continue
                    if it.get("col") not in valid_cols: continue
                    if it.get("type") not in {"text", "image", "spacer"}: continue
                    cleaned_items.append(it)
                out.append({**b, "items": cleaned_items})
            else:
                out.append(b)
        return out

    sid = current_societe_id()
    db = get_db()
    existing = one(db.execute("""SELECT blocks_json, header_blocks_json, footer_blocks_json
                                  FROM cover_pages_bc WHERE societe_id=?""", (sid,)))
    if existing:
        cur_cover = json.loads(existing.get("blocks_json") or "[]")
        cur_header = json.loads(existing.get("header_blocks_json") or "[]")
        cur_footer = json.loads(existing.get("footer_blocks_json") or "[]")
    else:
        cur_cover = list(_COVER_BC_DEFAULT_BLOCKS)
        cur_header = list(_HEADER_DEFAULT_BLOCKS)
        cur_footer = list(_FOOTER_DEFAULT_BLOCKS)

    if "blocks" in d:        cur_cover = _clean_cover(d.get("blocks"))
    if "header_blocks" in d: cur_header = _clean_hf(d.get("header_blocks"))
    if "footer_blocks" in d: cur_footer = _clean_hf(d.get("footer_blocks"))

    if existing:
        db.execute("""UPDATE cover_pages_bc SET
                        blocks_json=?, header_blocks_json=?, footer_blocks_json=?,
                        updated_at=datetime('now')
                      WHERE societe_id=?""",
                   (json.dumps(cur_cover), json.dumps(cur_header), json.dumps(cur_footer), sid))
    else:
        db.execute("""INSERT INTO cover_pages_bc
                        (societe_id, blocks_json, header_blocks_json, footer_blocks_json)
                        VALUES (?, ?, ?, ?)""",
                   (sid, json.dumps(cur_cover), json.dumps(cur_header), json.dumps(cur_footer)))
    db.commit()
    log_action(request.user, "UPDATE", "cover-page-bc", sid, f"sid={sid}",
               f"cover={len(cur_cover)} header={len(cur_header)} footer={len(cur_footer)}")
    return jsonify({"ok": True, "blocks": cur_cover, "header_blocks": cur_header, "footer_blocks": cur_footer})


# ─── Upload d'image pour la page de garde ───────────────────────────────────

@app.route("/api/cover-page/image", methods=["POST"])
@require_role("admin")
def upload_cover_image():
    """Téléverse une image (logo, illustration) pour la page de garde de la
    société active. Retourne le chemin relatif à utiliser dans un bloc 'image'."""
    if "file" not in request.files:
        return jsonify({"error": "Aucun fichier reçu"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Nom de fichier vide"}), 400
    sid = current_societe_id()
    # Extension : on garde celle d'origine si valide
    ext = (f.filename.rsplit(".", 1)[-1] or "").lower()
    if ext not in {"png", "jpg", "jpeg", "gif", "webp"}:
        return jsonify({"error": "Format non supporté (png/jpg/jpeg/gif/webp uniquement)"}), 400
    # Dossier dédié
    cover_dir = os.path.join(BASE_DIR, "cover_images")
    os.makedirs(cover_dir, exist_ok=True)
    # Nom unique : societe_<sid>_<timestamp>.<ext>
    fname = f"societe_{sid}_{int(time.time())}.{ext}"
    fpath = os.path.join(cover_dir, fname)
    f.save(fpath)
    rel_path = f"cover_images/{fname}"
    logger.info(f"[cover-page] Image téléversée : sid={sid} path={rel_path}")
    return jsonify({"ok": True, "path": rel_path})


@app.route("/cover_images/<path:filename>")
def serve_cover_image(filename):
    """Sert les images de la page de garde."""
    cover_dir = os.path.join(BASE_DIR, "cover_images")
    return send_file(os.path.join(cover_dir, filename))


@app.route("/api/module-images", methods=["POST"])
@require_role("admin")
def upload_module_image():
    """v218.92 — Téléverse une image utilisée dans un module 'image_checklist'.
    Retourne le chemin relatif à stocker dans field_options.image_path."""
    if "file" not in request.files:
        return jsonify({"error": "Aucun fichier reçu"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Nom de fichier vide"}), 400
    sid = current_societe_id()
    ext = (f.filename.rsplit(".", 1)[-1] or "").lower()
    if ext not in {"png", "jpg", "jpeg", "gif", "webp"}:
        return jsonify({"error": "Format non supporté (png/jpg/jpeg/gif/webp uniquement)"}), 400
    mod_dir = os.path.join(BASE_DIR, "module_images")
    os.makedirs(mod_dir, exist_ok=True)
    fname = f"sid{sid}_{int(time.time())}_{os.urandom(3).hex()}.{ext}"
    fpath = os.path.join(mod_dir, fname)
    f.save(fpath)
    rel_path = f"module_images/{fname}"
    logger.info(f"[module-images] Image téléversée : sid={sid} path={rel_path}")
    return jsonify({"ok": True, "path": rel_path})


@app.route("/module_images/<path:filename>")
def serve_module_image(filename):
    """Sert les images de modules image_checklist."""
    mod_dir = os.path.join(BASE_DIR, "module_images")
    return send_file(os.path.join(mod_dir, filename))


# ─── Gestion des liaisons User ↔ Société ────────────────────────────────────

@app.route("/api/societes/<int:sid>/utilisateurs", methods=["POST"])
@require_superadmin
def add_user_to_societe(sid):
    """Lie un utilisateur existant à une société avec un rôle donné.
    Si le lien existe déjà, met à jour le rôle."""
    d = request.json or {}
    uid = to_int(d.get("utilisateur_id"))
    role = d.get("role") or "technicien"
    if not uid: return jsonify({"error": "utilisateur_id requis"}), 400
    db = get_db()
    if not one(db.execute("SELECT id FROM societes WHERE id=?", (sid,))):
        return jsonify({"error": "Société introuvable"}), 404
    if not one(db.execute("SELECT id FROM utilisateurs WHERE id=?", (uid,))):
        return jsonify({"error": "Utilisateur introuvable"}), 404
    # Upsert
    existing = one(db.execute(
        "SELECT id FROM utilisateur_societes WHERE utilisateur_id=? AND societe_id=?",
        (uid, sid)
    ))
    if existing:
        db.execute("""
            UPDATE utilisateur_societes
            SET role=?, manager_id=?, techniques=?, actif=1
            WHERE id=?
        """, (role, to_int(d.get("manager_id")), d.get("techniques") or "", existing["id"]))
    else:
        db.execute("""
            INSERT INTO utilisateur_societes (utilisateur_id, societe_id, role, manager_id, techniques, actif)
            VALUES (?, ?, ?, ?, ?, 1)
        """, (uid, sid, role, to_int(d.get("manager_id")), d.get("techniques") or ""))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/societes/<int:sid>/utilisateurs/<int:uid>", methods=["DELETE"])
@require_superadmin
def remove_user_from_societe(sid, uid):
    """Retire un utilisateur d'une société (désactive le lien).
    Le compte utilisateur lui-même n'est pas supprimé."""
    db = get_db()
    db.execute(
        "UPDATE utilisateur_societes SET actif=0 WHERE utilisateur_id=? AND societe_id=?",
        (uid, sid)
    )
    db.commit()
    return jsonify({"ok": True})


# ─── Stats consolidées superadmin ───────────────────────────────────────────

@app.route("/api/superadmin/stats", methods=["GET"])
@require_superadmin
def superadmin_stats():
    """Stats consolidées par société pour le tableau de bord superadmin."""
    db = get_db()
    societes = rows(db.execute("SELECT id, nom, code, actif FROM societes ORDER BY nom"))
    result = []
    for s in societes:
        sid = s["id"]
        stats = {
            "id": sid,
            "nom": s["nom"],
            "code": s["code"],
            "actif": s["actif"],
        }
        # Comptages
        for table, label in [
            ("interventions", "nb_interventions"),
            ("equipements", "nb_equipements"),
            ("projets", "nb_projets"),
            ("clients", "nb_clients"),
            ("comptes_rendus", "nb_crs"),
        ]:
            try:
                r = one(db.execute(f"SELECT COUNT(*) AS n FROM {table} WHERE societe_id=?", (sid,)))
                stats[label] = (r and r["n"]) or 0
            except Exception:
                stats[label] = 0
        # Nb utilisateurs
        r = one(db.execute(
            "SELECT COUNT(*) AS n FROM utilisateur_societes WHERE societe_id=? AND actif=1", (sid,)
        ))
        stats["nb_users"] = (r and r["n"]) or 0
        # Heures totales (somme cr_intervenants pour les CRs de cette société)
        try:
            r = one(db.execute("""
                SELECT COALESCE(SUM(ci.total_heures), 0) AS total
                FROM cr_intervenants ci
                JOIN comptes_rendus cr ON cr.id = ci.cr_id
                WHERE cr.societe_id = ?
            """, (sid,)))
            stats["total_heures"] = round((r and r["total"]) or 0, 2)
        except Exception:
            stats["total_heures"] = 0
        result.append(stats)
    return jsonify(result)


@app.route("/api/superadmin/utilisateurs", methods=["GET"])
@require_superadmin
def superadmin_list_users():
    """Liste tous les utilisateurs avec leurs sociétés associées."""
    db = get_db()
    users = rows(db.execute("""
        SELECT id, nom, email, role, actif
        FROM utilisateurs
        ORDER BY nom
    """))
    for u in users:
        societes = rows(db.execute("""
            SELECT us.societe_id, s.nom AS societe_nom, s.code AS societe_code,
                   us.role AS role_societe, us.actif AS lien_actif
            FROM utilisateur_societes us
            JOIN societes s ON s.id = us.societe_id
            WHERE us.utilisateur_id = ?
            ORDER BY s.nom
        """, (u["id"],)))
        u["societes"] = societes
    return jsonify(users)


@app.route("/api/societes/<int:sid>/logo", methods=["POST"])
@require_superadmin
def upload_societe_logo(sid):
    """Upload du logo d'une société (PNG/JPG). Stocké dans logos/societe_<sid>.<ext>."""
    db = get_db()
    s = one(db.execute("SELECT id, nom FROM societes WHERE id=?", (sid,)))
    if not s: return jsonify({"error": "Société introuvable"}), 404
    if "logo" not in request.files:
        return jsonify({"error": "Fichier 'logo' requis"}), 400
    f = request.files["logo"]
    if not f.filename:
        return jsonify({"error": "Fichier vide"}), 400
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else "png"
    if ext not in ("png", "jpg", "jpeg", "webp", "svg"):
        return jsonify({"error": "Format non supporté (png/jpg/webp/svg attendus)"}), 400
    # Dossier logos/
    logos_dir = BASE_DIR / "logos"
    logos_dir.mkdir(exist_ok=True)
    filename = f"societe_{sid}.{ext}"
    filepath = logos_dir / filename
    f.save(str(filepath))
    relative_path = f"logos/{filename}"
    db.execute("UPDATE societes SET logo_path=? WHERE id=?", (relative_path, sid))
    db.commit()
    logger.info(f"[SUPERADMIN] Logo uploadé pour société {sid} : {filename}")
    return jsonify({"ok": True, "logo_path": relative_path})


@app.route("/logos/<filename>", methods=["GET"])
def serve_logo(filename):
    """Sert les logos des sociétés (publique pour affichage UI)."""
    from flask import send_from_directory
    logos_dir = BASE_DIR / "logos"
    return send_from_directory(str(logos_dir), filename)

# ══ UTILISATEURS ══
def _techniques_to_list(csv_val):
    """Normalise la valeur techniques (CSV) en liste d'abord en évitant doublons et vides."""
    if not csv_val:
        return []
    parts = [x.strip() for x in str(csv_val).split(",")]
    seen = set(); out = []
    for p in parts:
        if p and p not in seen:
            seen.add(p); out.append(p)
    return out

def _techniques_from_body(val):
    """Accepte liste, CSV, ou None. Retourne un CSV propre."""
    if val is None:
        return ""
    if isinstance(val, list):
        items = val
    else:
        items = [x.strip() for x in str(val).split(",")]
    seen = set(); out = []
    for it in items:
        it = str(it).strip()
        if it and it not in seen:
            seen.add(it); out.append(it)
    return ",".join(out)

@app.route("/api/utilisateurs")
@require_auth
def get_utilisateurs():
    sid = current_societe_id()
    # v218.72 : filtrer les utilisateurs liés à la société active
    # Le rôle, manager_id et techniques affichés sont ceux de la société active (utilisateur_societes)
    result = rows(get_db().execute("""
        SELECT u.id, u.nom, u.email, u.actif, u.matricule,
               us.role AS role,
               us.manager_id AS manager_id,
               us.techniques AS techniques,
               u.poste_id, u.superieur_id,
               m.nom AS manager_nom,
               p.nom AS poste_nom,
               p.couleur AS poste_couleur,
               s.nom AS superieur_nom
        FROM utilisateurs u
        JOIN utilisateur_societes us ON us.utilisateur_id = u.id
        LEFT JOIN utilisateurs m ON us.manager_id = m.id
        LEFT JOIN postes p ON u.poste_id = p.id
        LEFT JOIN utilisateurs s ON u.superieur_id = s.id
        WHERE us.societe_id = ? AND us.actif = 1
        ORDER BY u.nom
    """, (sid,)))
    # Convertir le CSV en liste pour le frontend
    for r in result:
        r["techniques_list"] = _techniques_to_list(r.get("techniques"))
    return jsonify(result)

@app.route("/api/type_techniques")
@require_auth
def get_type_techniques():
    """Liste des techniques disponibles, issue de la table techniques (source de vérité).
    Retourne juste les noms (strings)."""
    db = get_db()
    result = rows(db.execute("SELECT nom FROM techniques ORDER BY nom"))
    return jsonify([r["nom"] for r in result])

@app.route("/api/utilisateurs",methods=["POST"])
@require_role("admin","manager")
def create_utilisateur():
    d=request.json or {}
    if not all([d.get("nom"),d.get("email"),d.get("password")]): return jsonify({"error":"nom,email,password requis"}),400
    if len(d["password"]) < 8: return jsonify({"error":"Mot de passe trop court (8 caracteres minimum)"}),400
    role_demande = d.get("role","technicien")
    # v218.65 : seul un superadmin peut créer un autre superadmin
    if role_demande == "superadmin" and not is_superadmin(request.user):
        return jsonify({"error":"Seul un superadmin peut créer un autre superadmin"}), 403
    sid = current_societe_id()
    db=get_db()
    matricule = (d.get("matricule") or "").strip()
    manager_id = to_int(d.get("manager_id")) or None
    techniques_csv = _techniques_from_body(d.get("techniques"))
    poste_id = to_int(d.get("poste_id")) or None
    superieur_id = to_int(d.get("superieur_id")) or None
    try:
        db.execute("INSERT INTO utilisateurs (nom,email,password,role,matricule,manager_id,techniques,poste_id,superieur_id) VALUES (?,?,?,?,?,?,?,?,?)",
                   (d["nom"],d["email"].strip().lower(),hash_password(d["password"]),role_demande,matricule,manager_id,techniques_csv,poste_id,superieur_id))
        new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        # v218.72 : créer la liaison utilisateur_societes pour la société active
        # Si le rôle global est superadmin, on stocke quand même un rôle de société (admin par défaut)
        societe_role = role_demande if role_demande != "superadmin" else "admin"
        db.execute("""
            INSERT INTO utilisateur_societes (utilisateur_id, societe_id, role, manager_id, techniques, actif)
            VALUES (?, ?, ?, ?, ?, 1)
        """, (new_id, sid, societe_role, manager_id, techniques_csv))
        db.commit()
        log_action(request.user, "CREATE", "utilisateur", new_id, d["nom"], f"role={role_demande} societe={sid}")
        return jsonify({"id":new_id}),201
    except Exception as e: return jsonify({"error":str(e)}),400

@app.route("/api/utilisateurs/<int:uid>",methods=["PATCH"])
@require_role("admin","manager")
def update_utilisateur(uid):
    d = request.json or {}
    db = get_db()
    # v218.65 : protection rôle superadmin
    if "role" in d:
        nouveau_role = d.get("role")
        ancien = one(db.execute("SELECT role FROM utilisateurs WHERE id=?", (uid,)))
        ancien_role = (ancien or {}).get("role")
        # Si on tente d'attribuer le rôle superadmin → exige que l'appelant soit superadmin
        if nouveau_role == "superadmin" and not is_superadmin(request.user):
            return jsonify({"error":"Seul un superadmin peut promouvoir au rôle superadmin"}), 403
        # Si on tente de RETIRER le rôle superadmin à quelqu'un → exige aussi superadmin
        if ancien_role == "superadmin" and nouveau_role != "superadmin" and not is_superadmin(request.user):
            return jsonify({"error":"Seul un superadmin peut retirer le rôle superadmin"}), 403
    fields = {}
    for f in ["nom","email","role","actif","matricule"]:
        if f in d: fields[f] = d[f]
    # manager_id : None si vide, sinon int
    if "manager_id" in d:
        mid = d.get("manager_id")
        fields["manager_id"] = to_int(mid) if mid else None
    # techniques : liste ou CSV → CSV propre
    if "techniques" in d:
        fields["techniques"] = _techniques_from_body(d.get("techniques"))
    # poste_id et superieur_id : None si vide, sinon int
    if "poste_id" in d:
        pid = d.get("poste_id")
        fields["poste_id"] = to_int(pid) if pid else None
    if "superieur_id" in d:
        sid = d.get("superieur_id")
        fields["superieur_id"] = to_int(sid) if sid else None
    if fields.get("email"): fields["email"] = fields["email"].strip().lower()
    if d.get("password"):
        if len(d["password"]) < 8:
            return jsonify({"error":"Mot de passe trop court (8 caracteres minimum)"}),400
        fields["password"] = hash_password(d["password"])
        # Révoquer toutes les sessions existantes
        current = one(db.execute("SELECT token_version FROM utilisateurs WHERE id=?", (uid,)))
        fields["token_version"] = (current["token_version"] or 0) + 1 if current else 1
    n = safe_update(db, "utilisateurs", uid, fields)
    if n == 0: return jsonify({"error":"Rien"}), 400
    db.commit()
    # Log : on récupère le nom pour tracer
    u_info = one(db.execute("SELECT nom FROM utilisateurs WHERE id=?", (uid,)))
    label = (u_info or {}).get("nom", f"uid={uid}")
    changes = ", ".join([f"{k}={v}" for k,v in fields.items() if k != "password"])
    if "password" in fields: changes += (", " if changes else "") + "password=***"
    log_action(request.user, "UPDATE", "utilisateur", uid, label, changes)
    return jsonify({"ok":True})

@app.route("/api/utilisateurs/<int:uid>",methods=["DELETE"])
@require_role("admin")
def delete_utilisateur(uid):
    if uid==request.user["id"]: return jsonify({"error":"Impossible de supprimer son propre compte"}),400
    db=get_db()
    # Récupérer le nom avant suppression pour le log
    u_info = one(db.execute("SELECT nom FROM utilisateurs WHERE id=?", (uid,)))
    label = (u_info or {}).get("nom", f"uid={uid}")
    db.execute("DELETE FROM utilisateurs WHERE id=?",(uid,)); db.commit()
    log_action(request.user, "DELETE", "utilisateur", uid, label)
    return jsonify({"ok":True})

@app.route("/api/utilisateurs/me",methods=["PATCH"])
@require_auth
def update_me():
    d = request.json or {}
    db = get_db()
    uid = request.user["id"]
    fields = {}
    if "nom" in d: fields["nom"] = d["nom"]
    if d.get("password"):
        if len(d["password"]) < 8:
            return jsonify({"error":"Mot de passe trop court (8 caracteres minimum)"}),400
        fields["password"] = hash_password(d["password"])
        current = one(db.execute("SELECT token_version FROM utilisateurs WHERE id=?", (uid,)))
        fields["token_version"] = (current["token_version"] or 0) + 1 if current else 1
    n = safe_update(db, "utilisateurs", uid, fields)
    if n == 0: return jsonify({"error":"Rien"}), 400
    db.commit()
    # Si MDP changé, le client devra se reconnecter (token invalide immédiatement)
    return jsonify({"ok":True, "password_changed": "password" in fields})

# ══════════════════════════════════════════════════════════════════════
# RBAC — Rôles et permissions
# ══════════════════════════════════════════════════════════════════════
@app.route("/api/roles")
@require_auth
def get_roles():
    """Liste tous les rôles (built-in + personnalisés)."""
    db = get_db()
    rs = rows(db.execute("SELECT code, label, parent_role, builtin, created_at FROM roles ORDER BY builtin DESC, code"))
    # Comptabiliser les utilisateurs par rôle
    counts = {}
    for r in rows(db.execute("SELECT role, COUNT(*) AS n FROM utilisateurs GROUP BY role")):
        counts[r["role"]] = r["n"]
    for r in rs:
        r["nb_utilisateurs"] = counts.get(r["code"], 0)
    return jsonify(rs)

@app.route("/api/roles", methods=["POST"])
@require_role("admin")
def create_role():
    """Crée un nouveau rôle personnalisé.
    Body : { code: 'sous_traitant', label: 'Sous-traitant', parent_role: 'technicien' }
    Les permissions sont initialisées à partir du parent_role."""
    d = request.json or {}
    code = (d.get("code") or "").strip().lower()
    label = (d.get("label") or "").strip()
    parent = (d.get("parent_role") or "technicien").strip().lower()
    if not code: return jsonify({"error": "Code requis"}), 400
    if not label: return jsonify({"error": "Libellé requis"}), 400
    if parent not in ("admin", "manager", "technicien", "acl"):
        return jsonify({"error": "Rôle parent invalide"}), 400
    # Validation : code = lettres/chiffres/underscore uniquement
    import re
    if not re.match(r'^[a-z0-9_]+$', code):
        return jsonify({"error": "Code invalide (lettres minuscules, chiffres et _ uniquement)"}), 400
    if len(code) > 50: return jsonify({"error": "Code trop long (max 50)"}), 400
    db = get_db()
    if one(db.execute("SELECT code FROM roles WHERE code=?", (code,))):
        return jsonify({"error": "Ce code de rôle existe déjà"}), 400
    db.execute("INSERT INTO roles (code, label, parent_role, builtin) VALUES (?, ?, ?, 0)",
               (code, label, parent))
    # Cloner les permissions du rôle parent
    db.execute("""INSERT INTO permissions (role_code, module, action, allowed)
                  SELECT ?, module, action, allowed FROM permissions WHERE role_code=?""",
               (code, parent))
    db.commit()
    log_action(request.user, "CREATE", "role", None, f"Rôle '{label}' ({code}) cloné depuis {parent}")
    return jsonify({"ok": True, "code": code}), 201

@app.route("/api/roles/<code>", methods=["PATCH"])
@require_role("admin")
def update_role(code):
    """Modifie le libellé d'un rôle (le code et builtin sont immuables)."""
    d = request.json or {}
    db = get_db()
    r = one(db.execute("SELECT code, builtin FROM roles WHERE code=?", (code,)))
    if not r: return jsonify({"error": "Rôle introuvable"}), 404
    fields = {}
    if "label" in d:
        new_label = (d["label"] or "").strip()
        if not new_label: return jsonify({"error": "Libellé requis"}), 400
        fields["label"] = new_label
    if "parent_role" in d and not r["builtin"]:
        np = (d["parent_role"] or "").strip().lower()
        if np not in ("admin", "manager", "technicien", "acl"):
            return jsonify({"error": "Rôle parent invalide"}), 400
        fields["parent_role"] = np
    if not fields: return jsonify({"error": "Rien à modifier"}), 400
    placeholders = ", ".join(f"{k}=?" for k in fields.keys())
    db.execute(f"UPDATE roles SET {placeholders} WHERE code=?", (*fields.values(), code))
    db.commit()
    log_action(request.user, "UPDATE", "role", None, f"Rôle {code}: {fields}")
    return jsonify({"ok": True})

@app.route("/api/roles/<code>", methods=["DELETE"])
@require_role("admin")
def delete_role(code):
    """Supprime un rôle personnalisé. Les rôles built-in ne peuvent pas être supprimés.
    Si des utilisateurs ont ce rôle, on refuse (l'admin doit d'abord les réassigner)."""
    db = get_db()
    r = one(db.execute("SELECT code, builtin, label FROM roles WHERE code=?", (code,)))
    if not r: return jsonify({"error": "Rôle introuvable"}), 404
    if r["builtin"]:
        return jsonify({"error": "Impossible de supprimer un rôle système"}), 400
    nb = one(db.execute("SELECT COUNT(*) AS n FROM utilisateurs WHERE role=?", (code,)))
    if nb and nb["n"] > 0:
        return jsonify({
            "error": f"{nb['n']} utilisateur(s) ont ce rôle. Réassignez-les d'abord."
        }), 400
    db.execute("DELETE FROM permissions WHERE role_code=?", (code,))
    db.execute("DELETE FROM roles WHERE code=?", (code,))
    db.commit()
    log_action(request.user, "DELETE", "role", None, f"Rôle '{r['label']}' ({code}) supprimé")
    return jsonify({"ok": True})

@app.route("/api/permissions/<role_code>")
@require_auth
def get_role_permissions(role_code):
    """Retourne la grille de permissions pour un rôle.
    Format : { module: { action: bool } }
    Lecture autorisée à tous (utile pour l'UI), modification réservée à l'admin."""
    db = get_db()
    if not one(db.execute("SELECT code FROM roles WHERE code=?", (role_code,))):
        return jsonify({"error": "Rôle introuvable"}), 404
    return jsonify({
        "role_code": role_code,
        "modules": [
            {"code": m[0], "label": m[1], "actions": m[2]}
            for m in RBAC_MODULES
        ],
        "permissions": _perms_for_role(role_code),
    })

@app.route("/api/permissions/<role_code>", methods=["PUT"])
@require_role("admin")
def update_role_permissions(role_code):
    """Met à jour la grille de permissions pour un rôle.
    Body : { permissions: { module: { action: bool } } }
    Le rôle 'admin' ne peut PAS être modifié (sécurité)."""
    if role_code == "admin":
        return jsonify({"error": "Les permissions admin ne peuvent pas être modifiées"}), 400
    d = request.json or {}
    perms = d.get("permissions") or {}
    db = get_db()
    if not one(db.execute("SELECT code FROM roles WHERE code=?", (role_code,))):
        return jsonify({"error": "Rôle introuvable"}), 404
    # Valider la structure et appliquer
    valid_modules = {m[0]: m[2] for m in RBAC_MODULES}
    n_updated = 0
    for module, actions_dict in perms.items():
        if module not in valid_modules: continue
        valid_actions = valid_modules[module]
        for action, allowed in (actions_dict or {}).items():
            if action not in valid_actions: continue
            allowed_int = 1 if allowed else 0
            db.execute("""INSERT INTO permissions (role_code, module, action, allowed)
                          VALUES (?, ?, ?, ?)
                          ON CONFLICT(role_code, module, action)
                          DO UPDATE SET allowed=excluded.allowed""",
                       (role_code, module, action, allowed_int))
            n_updated += 1
    db.commit()
    log_action(request.user, "UPDATE", "permissions", None,
               f"Rôle '{role_code}' : {n_updated} permission(s) mises à jour")
    return jsonify({"ok": True, "updated": n_updated})

@app.route("/api/me/permissions")
@require_auth
def get_my_permissions():
    """Retourne les permissions de l'utilisateur courant (utile pour l'UI client).
    Format : { role: code, parent_role: code, permissions: { module: { action: bool } } }"""
    u = request.user
    role_code = _user_role_code(u)
    return jsonify({
        "role": role_code,
        "parent_role": _resolve_parent_role(role_code),
        "modules": [
            {"code": m[0], "label": m[1], "actions": m[2]}
            for m in RBAC_MODULES
        ],
        "permissions": _perms_for_role(role_code),
    })

# ══ CLIENTS ══
@app.route("/api/clients")
@require_auth
def get_clients():
    sid = current_societe_id()
    return jsonify(rows(get_db().execute("SELECT * FROM clients WHERE societe_id=? ORDER BY societe", (sid,))))

@app.route("/api/clients",methods=["POST"])
@require_role("admin","manager")
def create_client():
    d=request.json or {}
    if not d.get("societe"): return jsonify({"error":"societe requis"}),400
    sid = current_societe_id()
    db=get_db()
    db.execute("INSERT INTO clients (societe,nom,prenom,email,telephone,notes,societe_id) VALUES (?,?,?,?,?,?,?)",
               (d["societe"],d.get("nom",""),d.get("prenom",""),d.get("email",""),d.get("telephone",""),d.get("notes",""),sid))
    db.commit(); return jsonify({"id":db.execute("SELECT last_insert_rowid()").fetchone()[0]}),201

@app.route("/api/clients/<int:cid>",methods=["PATCH"])
@require_role("admin","manager")
def update_client(cid):
    d=request.json or {}; db=get_db(); sets,params=[],[]
    sid = current_societe_id()
    for f in ["societe","nom","prenom","email","telephone","notes"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    if not sets: return jsonify({"error":"Rien"}),400
    params.extend([cid, sid]); db.execute(f"UPDATE clients SET {chr(44).join(sets)} WHERE id=? AND societe_id=?",params); db.commit()
    return jsonify({"ok":True})

@app.route("/api/clients/<int:cid>",methods=["DELETE"])
@require_role("admin")
def delete_client(cid):
    db=get_db(); sid = current_societe_id()
    db.execute("DELETE FROM clients WHERE id=? AND societe_id=?",(cid, sid)); db.commit()
    return jsonify({"ok":True})

# ══ PROJETS ══
@app.route("/api/projets/all")
@require_auth
def get_projets_all():
    """Retourne TOUS les projets de la société active."""
    db = get_db()
    sid = current_societe_id()
    return jsonify(rows(db.execute("""SELECT p.*,c.societe AS client_nom,m.nom AS manager_nom
        FROM projets p LEFT JOIN clients c ON p.client_id=c.id
        LEFT JOIN utilisateurs m ON p.manager_id=m.id
        WHERE p.societe_id=? ORDER BY p.nom""", (sid,))))

@app.route("/api/projets")
@require_auth
def get_projets():
    db=get_db(); u=request.user
    sid = current_societe_id()
    if u["role"]=="technicien":
        return jsonify(rows(db.execute("""SELECT p.*,c.societe AS client_nom,m.nom AS manager_nom
            FROM projets p LEFT JOIN clients c ON p.client_id=c.id
            LEFT JOIN utilisateurs m ON p.manager_id=m.id
            WHERE p.societe_id=? AND p.id IN (
                SELECT DISTINCT e.projet_id FROM equipements e
                JOIN interventions i ON i.equipement_id=e.id WHERE i.technicien_id=?
                UNION
                SELECT DISTINCT e.projet_id FROM equipements e
                JOIN interventions i ON i.equipement_id=e.id
                JOIN equipe_membres em ON em.equipe_id=i.equipe_id WHERE em.technicien_id=?
            ) ORDER BY p.nom""", (sid, u["id"], u["id"]))))
    if u["role"]=="manager":
        udb = one(db.execute("SELECT techniques FROM utilisateurs WHERE id=?", (u["id"],)))
        tech_csv = (udb and udb.get("techniques")) or ""
        tech_list = [t.strip() for t in tech_csv.split(",") if t.strip()]
        if tech_list:
            placeholders = ",".join(["?"] * len(tech_list))
            sql = f"""SELECT p.*,c.societe AS client_nom,m.nom AS manager_nom
                FROM projets p LEFT JOIN clients c ON p.client_id=c.id
                LEFT JOIN utilisateurs m ON p.manager_id=m.id
                WHERE p.societe_id=? AND (p.manager_id=?
                   OR p.id IN (SELECT DISTINCT e.projet_id FROM equipements e WHERE e.type_technique IN ({placeholders})))
                ORDER BY p.nom"""
            return jsonify(rows(db.execute(sql, [sid, u["id"]] + tech_list)))
        else:
            return jsonify(rows(db.execute("""SELECT p.*,c.societe AS client_nom,m.nom AS manager_nom
                FROM projets p LEFT JOIN clients c ON p.client_id=c.id
                LEFT JOIN utilisateurs m ON p.manager_id=m.id
                WHERE p.societe_id=? AND p.manager_id=? ORDER BY p.nom""", (sid, u["id"]))))
    # admin / superadmin
    return jsonify(rows(db.execute("""SELECT p.*,c.societe AS client_nom,m.nom AS manager_nom
        FROM projets p LEFT JOIN clients c ON p.client_id=c.id
        LEFT JOIN utilisateurs m ON p.manager_id=m.id
        WHERE p.societe_id=? ORDER BY p.nom""", (sid,))))

@app.route("/api/projets",methods=["POST"])
@require_role("admin","manager")
def create_projet():
    d=request.json or {}
    if not d.get("nom"): return jsonify({"error":"nom requis"}),400
    sid = current_societe_id()
    db=get_db(); num=d.get("numero_projet") or next_numero(db,"P","projets","numero_projet")
    db.execute("""INSERT INTO projets (numero_projet,nom,client_id,manager_id,description,
               date_debut,date_fin,statut,ville,code_postal,deplacement_km,nb_deplacements,societe_id)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
               (num,d["nom"],to_int(d.get("client_id")),to_int(d.get("manager_id")),
                d.get("description",""),d.get("date_debut") or None,d.get("date_fin") or None,
                d.get("statut","EN_COURS"),d.get("ville",""),d.get("code_postal",""),
                float(d.get("deplacement_km") or 0), int(d.get("nb_deplacements") or 0), sid))
    db.commit()
    new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    log_action(request.user, "CREATE", "projet", new_id, f"{num} - {d['nom']}")
    return jsonify({"id":new_id}),201


@app.route("/api/projets/<int:pid>")
@require_auth
def get_projet(pid):
    db = get_db()
    p = one(db.execute(
        "SELECT p.*,c.societe AS client_nom,m.nom AS manager_nom"
        " FROM projets p LEFT JOIN clients c ON p.client_id=c.id"
        " LEFT JOIN utilisateurs m ON p.manager_id=m.id WHERE p.id=?",(pid,)))
    if not p: return jsonify({"error":"Non trouve"}),404
    p["equipements"] = rows(db.execute(
        "SELECT e.* FROM equipements e WHERE e.projet_id=? ORDER BY e.designation",(pid,)))
    for e in p["equipements"]:
        e["criticite"] = criticite(e["id"],db)
    return jsonify(p)

@app.route("/api/projets/<int:pid>",methods=["PATCH"])
@require_role("admin","manager")
def update_projet(pid):
    d=request.json or {}; db=get_db(); sets,params=[],[]
    logger.info(f"[PATCH projet] pid={pid} body={d}")
    for f in ["nom","client_id","manager_id","description","date_debut","date_fin","statut","numero_projet","ville","code_postal","deplacement_km"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    if not sets: return jsonify({"error":"Rien"}),400
    params.append(pid); db.execute(f"UPDATE projets SET {chr(44).join(sets)} WHERE id=?",params); db.commit()
    logger.info(f"[PATCH projet] pid={pid} champs={sets}")
    p_info = one(db.execute("SELECT nom,numero_projet FROM projets WHERE id=?", (pid,)))
    label = f"{(p_info or {}).get('numero_projet','')} - {(p_info or {}).get('nom','')}" if p_info else f"pid={pid}"
    changes = ", ".join([f"{k}={v}" for k,v in d.items()])
    log_action(request.user, "UPDATE", "projet", pid, label, changes[:500])
    return jsonify({"ok":True})

@app.route("/api/projets/<int:pid>",methods=["DELETE"])
@require_role("admin")
def delete_projet(pid):
    db=get_db()
    p_info = one(db.execute("SELECT nom,numero_projet,logo_filename FROM projets WHERE id=?", (pid,)))
    label = f"{(p_info or {}).get('numero_projet','')} - {(p_info or {}).get('nom','')}" if p_info else f"pid={pid}"
    # Supprimer le logo associé si existe
    if p_info and p_info.get("logo_filename"):
        try:
            logo_path = BASE_DIR / "uploads" / "projet_logos" / p_info["logo_filename"]
            if logo_path.exists(): logo_path.unlink()
        except Exception as e:
            logger.warning(f"[delete_projet] suppression logo échec : {e}")
    db.execute("DELETE FROM projets WHERE id=?",(pid,)); db.commit()
    log_action(request.user, "DELETE", "projet", pid, label)
    return jsonify({"ok":True})

# ══════════════════════════════════════════════════════════════════════
# LOGO DE PROJET — upload, serve, suppression
# Stockage : /var/www/gmao/uploads/projet_logos/<projet_id>_<timestamp>.<ext>
# ══════════════════════════════════════════════════════════════════════
PROJET_LOGOS_DIR = BASE_DIR / "uploads" / "projet_logos"
ALLOWED_LOGO_EXTS = {"png", "jpg", "jpeg"}
MAX_LOGO_SIZE = 5 * 1024 * 1024  # 5 Mo

def _ensure_logos_dir():
    PROJET_LOGOS_DIR.mkdir(parents=True, exist_ok=True)

def _logo_ext(filename):
    """Retourne l'extension en minuscule (sans le point) ou None."""
    if not filename or "." not in filename: return None
    return filename.rsplit(".", 1)[1].lower()

@app.route("/api/projets/<int:pid>/logo", methods=["POST"])
@require_auth
def upload_projet_logo(pid):
    """Upload le logo d'un projet (PNG/JPG, max 5 Mo).
    Multipart form-data avec champ 'logo'.
    Réservé aux admins et managers (du projet).
    """
    u = request.user; db = get_db()
    proj = one(db.execute("SELECT id, manager_id, logo_filename, numero_projet, nom FROM projets WHERE id=?", (pid,)))
    if not proj: return jsonify({"error":"Projet introuvable"}), 404
    # Permissions
    if u["role"] == "technicien":
        return jsonify({"error":"Non autorisé"}), 403
    if u["role"] == "manager" and proj.get("manager_id") != u["id"]:
        return jsonify({"error":"Vous n'êtes pas le manager de ce projet"}), 403
    # Récupérer le fichier
    if "logo" not in request.files:
        return jsonify({"error":"Aucun fichier reçu (champ 'logo' attendu)"}), 400
    f = request.files["logo"]
    if not f or not f.filename:
        return jsonify({"error":"Fichier vide"}), 400
    # Validation extension
    ext = _logo_ext(f.filename)
    if ext not in ALLOWED_LOGO_EXTS:
        return jsonify({"error":f"Extension non autorisée. Formats acceptés : {', '.join(sorted(ALLOWED_LOGO_EXTS))}"}), 400
    # Validation taille (lecture en mémoire pour vérifier la taille avant écriture)
    f.seek(0, 2); size = f.tell(); f.seek(0)
    if size > MAX_LOGO_SIZE:
        return jsonify({"error":f"Fichier trop volumineux (max {MAX_LOGO_SIZE // (1024*1024)} Mo)"}), 400
    if size == 0:
        return jsonify({"error":"Fichier vide"}), 400
    _ensure_logos_dir()
    # Nom de fichier : <pid>_<timestamp>.<ext>
    ts = int(time.time())
    new_filename = f"{pid}_{ts}.{ext}"
    dest_path = PROJET_LOGOS_DIR / new_filename
    try:
        f.save(str(dest_path))
    except Exception as e:
        logger.error(f"[upload_logo] échec écriture {dest_path}: {e}")
        return jsonify({"error":"Erreur lors de l'enregistrement"}), 500
    # Supprimer l'ancien logo si existait
    old = proj.get("logo_filename")
    if old and old != new_filename:
        try:
            old_path = PROJET_LOGOS_DIR / old
            if old_path.exists(): old_path.unlink()
        except Exception as e:
            logger.warning(f"[upload_logo] suppression ancien logo échec : {e}")
    # Mise à jour BDD
    db.execute("UPDATE projets SET logo_filename=? WHERE id=?", (new_filename, pid))
    db.commit()
    label = f"{proj.get('numero_projet','')} - {proj.get('nom','')}"
    log_action(u, "UPDATE", "projet_logo", pid, label, f"Logo : {new_filename}")
    return jsonify({"ok": True, "logo_filename": new_filename, "logo_url": f"/api/projets/{pid}/logo"})

@app.route("/api/projets/<int:pid>/logo", methods=["GET"])
def serve_projet_logo(pid):
    """Sert le logo d'un projet (public, mis en cache).
    Pas d'auth requise pour pouvoir l'embarquer dans des PDFs / pages publiques.
    """
    db = get_db()
    proj = one(db.execute("SELECT logo_filename FROM projets WHERE id=?", (pid,)))
    if not proj or not proj.get("logo_filename"):
        return "", 404
    fname = proj["logo_filename"]
    # Sécurité : pas de path traversal
    if "/" in fname or "\\" in fname or ".." in fname:
        return "", 404
    fpath = PROJET_LOGOS_DIR / fname
    if not fpath.exists() or not fpath.is_file():
        return "", 404
    resp = send_file(str(fpath))
    resp.headers["Cache-Control"] = "public, max-age=86400"
    return resp

@app.route("/api/projets/<int:pid>/logo", methods=["DELETE"])
@require_auth
def delete_projet_logo(pid):
    """Supprime le logo d'un projet."""
    u = request.user; db = get_db()
    proj = one(db.execute("SELECT id, manager_id, logo_filename, numero_projet, nom FROM projets WHERE id=?", (pid,)))
    if not proj: return jsonify({"error":"Projet introuvable"}), 404
    if u["role"] == "technicien":
        return jsonify({"error":"Non autorisé"}), 403
    if u["role"] == "manager" and proj.get("manager_id") != u["id"]:
        return jsonify({"error":"Vous n'êtes pas le manager de ce projet"}), 403
    fname = proj.get("logo_filename")
    if fname:
        try:
            fpath = PROJET_LOGOS_DIR / fname
            if fpath.exists(): fpath.unlink()
        except Exception as e:
            logger.warning(f"[delete_logo] suppression fichier échec : {e}")
    db.execute("UPDATE projets SET logo_filename='' WHERE id=?", (pid,))
    db.commit()
    label = f"{proj.get('numero_projet','')} - {proj.get('nom','')}"
    log_action(u, "DELETE", "projet_logo", pid, label)
    return jsonify({"ok": True})


@app.route("/api/projets/<int:pid>/planning-pdf", methods=["POST"])
@require_auth
def generate_projet_planning_pdf(pid):
    """Génère un PDF de planning client pour un projet.
    Body : { annee, lignes: [{equipement, gamme, periodicite, date, heure, occ_idx, occ_total}, ...] }
    Retourne : application/pdf"""
    d = request.json or {}
    db = get_db()
    proj = one(db.execute("""SELECT p.id, p.nom, p.numero_projet,
                                    c.societe AS client_nom
                             FROM projets p LEFT JOIN clients c ON p.client_id=c.id
                             WHERE p.id=?""", (pid,)))
    if not proj:
        return jsonify({"error": "Projet introuvable"}), 404
    try:
        from rapport_pdf import generate_planning_client_pdf
        data = {
            "projet": {
                "nom": proj.get("nom") or "",
                "numero_projet": proj.get("numero_projet") or "",
                "client_nom": proj.get("client_nom") or "",
                "adresse": "",  # pas de colonne adresse en BDD
            },
            "annee": d.get("annee") or date.today().year,
            "lignes": d.get("lignes") or [],
            "date_emission": date.today().strftime("%Y-%m-%d"),
        }
        pdf_bytes = generate_planning_client_pdf(data)
        # v218.139 : nommage Planning_<Nom du projet>_<N° projet>_<Année>.pdf
        # Caractères interdits dans les noms de fichiers (/ \ : * ? " < > |) → remplacés par '-'
        import re as _re_fname
        def _sanitize(s):
            return _re_fname.sub(r'[\\/:*?"<>|]', '-', (s or '').strip()).strip('-_ ') or 'X'
        nom_safe = _sanitize(proj.get('nom') or 'Projet')
        num_safe = _sanitize(proj.get('numero_projet') or '')
        annee_v = data['annee']
        if num_safe:
            fname = f"Planning_{nom_safe}_{num_safe}_{annee_v}.pdf"
        else:
            fname = f"Planning_{nom_safe}_{annee_v}.pdf"
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f'inline; filename="{fname}"'}
        )
    except Exception as e:
        logger.error(f"[planning-pdf] échec : {e}", exc_info=True)
        return jsonify({"error": f"Erreur génération PDF : {e}"}), 500

# ══ EQUIPEMENTS ══
@app.route("/api/equipements")
@require_auth
def get_equipements():
    db=get_db(); u=request.user
    sid = current_societe_id()
    sql="""SELECT e.*,p.nom AS projet_nom,p.numero_projet,g.nom AS gamme_nom,t.nom AS technique_nom
           FROM equipements e JOIN projets p ON e.projet_id=p.id
           LEFT JOIN gammes g ON e.gamme_id=g.id
           LEFT JOIN techniques t ON e.technique_id=t.id WHERE e.societe_id=?"""
    params=[sid]
    if request.args.get("projet_id"): sql+=" AND e.projet_id=?"; params.append(request.args["projet_id"])
    # Filtrage par rôle : manager = projets qu'il gère + techniques qu'il gère
    if u["role"]=="manager":
        udb = one(db.execute("SELECT techniques FROM utilisateurs WHERE id=?", (u["id"],)))
        tech_csv = (udb and udb.get("techniques")) or ""
        tech_list = [t.strip() for t in tech_csv.split(",") if t.strip()]
        if tech_list:
            placeholders = ",".join(["?"] * len(tech_list))
            sql += f" AND (p.manager_id=? OR e.type_technique IN ({placeholders}))"
            params = params + [u["id"]] + tech_list
        else:
            sql += " AND p.manager_id=?"
            params.append(u["id"])
    equips=rows(db.execute(sql+" ORDER BY p.nom,e.designation",params))
    for e in equips:
        e["criticite"]=criticite(e["id"],db)
        e["nb_maintenance"]=db.execute("SELECT COUNT(*) FROM interventions WHERE equipement_id=? AND type='MAINTENANCE'",(e["id"],)).fetchone()[0]
        e["nb_depannage"]=db.execute("SELECT COUNT(*) FROM interventions WHERE equipement_id=? AND type='DEPANNAGE'",(e["id"],)).fetchone()[0]
        e["gammes"]=rows(db.execute("""SELECT g.*,
            eg.planning_id AS eg_planning_id,
            eg.semaine_planif AS eg_semaine_planif,
            eg.jour_semaine_planif AS eg_jour_semaine_planif,
            eg.intervention_samedi AS eg_intervention_samedi,
            eg.intervention_dimanche AS eg_intervention_dimanche,
            eg.planif_mode AS eg_planif_mode,
            eg.nth_semaine_mois AS eg_nth_semaine_mois,
            eg.force_planif AS eg_force_planif
            FROM gammes g JOIN equipement_gammes eg ON g.id=eg.gamme_id
            WHERE eg.equipement_id=? ORDER BY g.nom""",(e["id"],)))
    return jsonify(equips)

@app.route("/api/equipements/<int:eid>")
@require_auth
def get_equipement(eid):
    db=get_db()
    e=one(db.execute("""SELECT e.*,p.nom AS projet_nom,p.numero_projet,g.nom AS gamme_nom,t.nom AS technique_nom,
                              il.filename AS image_filename, il.nom AS image_nom
        FROM equipements e JOIN projets p ON e.projet_id=p.id
        LEFT JOIN gammes g ON e.gamme_id=g.id LEFT JOIN techniques t ON e.technique_id=t.id
        LEFT JOIN image_library il ON e.image_id = il.id
        WHERE e.id=?""",(eid,)))
    if not e: return jsonify({"error":"Non trouve"}),404
    pieces_raw=rows(db.execute("SELECT * FROM pieces WHERE equipement_id=? ORDER BY type_piece",(eid,)))
    for p in pieces_raw:
        ns=statut_piece(p.get("date_fin_de_vie"))
        if ns!=p.get("statut") and p.get("date_fin_de_vie"):
            db.execute("UPDATE pieces SET statut=? WHERE id=?",(ns,p["id"])); p["statut"]=ns
    db.commit()
    e["pieces"]=pieces_raw; e["criticite"]=criticite(eid,db)
    # Tableaux électriques (sous-équipements pour Basse tension)
    try:
        e["tableaux"]=rows(db.execute("SELECT id,nom,localisation,ordre FROM equipement_tableaux WHERE equipement_id=? ORDER BY ordre,id",(eid,)))
    except Exception:
        e["tableaux"]=[]
    # Cellules (sous-équipements pour Haute tension)
    try:
        e["cellules"]=rows(db.execute("SELECT id,designation,marque,type,ordre FROM equipement_cellules WHERE equipement_id=? ORDER BY ordre,id",(eid,)))
    except Exception:
        e["cellules"]=[]
    # v218.157 : Sous-équipements universels (toutes techniques)
    try:
        e["sous_equipements"]=rows(db.execute("SELECT id,designation,marque,type,nombre,ordre FROM equipement_sous_equipements WHERE equipement_id=? ORDER BY ordre,id",(eid,)))
    except Exception:
        e["sous_equipements"]=[]
    # Gammes : récupérer aussi la planif par gamme (colonnes ajoutées sur equipement_gammes)
    e["gammes"]=rows(db.execute("""SELECT g.*,
        eg.planning_id AS eg_planning_id,
        eg.semaine_planif AS eg_semaine_planif,
        eg.jour_semaine_planif AS eg_jour_semaine_planif,
        eg.intervention_samedi AS eg_intervention_samedi,
        eg.intervention_dimanche AS eg_intervention_dimanche,
        eg.planif_mode AS eg_planif_mode,
        eg.nth_semaine_mois AS eg_nth_semaine_mois,
        eg.force_planif AS eg_force_planif
        FROM gammes g JOIN equipement_gammes eg ON g.id=eg.gamme_id
        WHERE eg.equipement_id=? ORDER BY g.nom""",(eid,)))
    return jsonify(e)

@app.route("/api/equipements",methods=["POST"])
@require_role("admin","manager")
def create_equipement():
    d=request.json or {}
    if not all([d.get("designation"),d.get("projet_id"),d.get("type_technique")]): return jsonify({"error":"designation,projet_id,type_technique requis"}),400
    sid = current_societe_id()
    db=get_db()
    db.execute("INSERT INTO equipements (projet_id,designation,type_technique,localisation,marque,modele,puissance,numero_serie,in_out,date_mise_en_service,statut,notes,societe_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
               (d["projet_id"],d["designation"],d["type_technique"],d.get("localisation",""),d.get("marque",""),d.get("modele",""),d.get("puissance",""),d.get("numero_serie",""),d.get("in_out",""),d.get("date_mise_en_service") or None,d.get("statut","EN_SERVICE"),d.get("notes",""),sid))
    new_eid=db.execute("SELECT last_insert_rowid()").fetchone()[0]
    # Champs Transformateur (Haute tension) : update si fournis
    trafo_sets, trafo_params = [], []
    for f in ["trafo_marque","trafo_annee","trafo_numero_serie","trafo_puissance_kva",
              "trafo_refroidissement","trafo_poids_kg","trafo_tension_entree_v",
              "trafo_courant_a","trafo_norme","trafo_couplage",
              "trafo_tension_service_v","trafo_reglage_tension_kv"]:
        if f in d:
            trafo_sets.append(f"{f}=?"); trafo_params.append(d[f] or "")
    if trafo_sets:
        trafo_params.append(new_eid)
        db.execute(f"UPDATE equipements SET {','.join(trafo_sets)} WHERE id=?", trafo_params)
    # Format gammes : peut être [id, id, ...] (ancien) ou [{gamme_id, semaine_planif, ...}, ...] (nouveau)
    for g in d.get("gammes",[]):
        try:
            if isinstance(g, dict):
                gid = int(g.get("gamme_id") or g.get("id"))
                db.execute("""INSERT OR IGNORE INTO equipement_gammes
                    (equipement_id, gamme_id, planning_id, semaine_planif, jour_semaine_planif,
                     intervention_samedi, intervention_dimanche, planif_mode, nth_semaine_mois, force_planif)
                    VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (new_eid, gid,
                     to_int(g.get("planning_id")),
                     to_int(g.get("semaine_planif")),
                     to_int(g.get("jour_semaine_planif")),
                     1 if g.get("intervention_samedi") else 0,
                     1 if g.get("intervention_dimanche") else 0,
                     g.get("planif_mode") or 'SEMAINE',
                     to_int(g.get("nth_semaine_mois")),
                     1 if g.get("force_planif") else 0))
            else:
                db.execute("INSERT OR IGNORE INTO equipement_gammes (equipement_id, gamme_id) VALUES (?,?)",
                           (new_eid, int(g)))
        except Exception: pass
    db.commit()
    # Générer le premier bon si au moins une gamme a une planif configurée
    has_planif = False
    try:
        has_planif = bool(one(db.execute("""SELECT 1 FROM equipement_gammes
            WHERE equipement_id=? AND planning_id IS NOT NULL
            AND semaine_planif IS NOT NULL AND jour_semaine_planif IS NOT NULL
            LIMIT 1""", (new_eid,))))
    except Exception: pass
    if has_planif:
        try:
            from datetime import date as _date
            generate_next_bon(new_eid, db, _date.today().isoformat())
        except Exception as e_gen:
            pass
    return jsonify({"id":new_eid}),201

@app.route("/api/equipements/<int:eid>",methods=["PATCH"])
@require_role("admin","manager")
def update_equipement(eid):
    d=request.json or {}; db=get_db(); sets,params=[],[]
    for f in ["designation","type_technique","projet_id","localisation","marque","modele","puissance","numero_serie","in_out","date_mise_en_service","statut","technique_id","notes","planning_id","semaine_planif","jour_semaine_planif","intervention_samedi","intervention_dimanche","image_id",
              "trafo_marque","trafo_annee","trafo_numero_serie","trafo_puissance_kva","trafo_refroidissement","trafo_poids_kg","trafo_tension_entree_v","trafo_courant_a","trafo_norme","trafo_couplage","trafo_tension_service_v","trafo_reglage_tension_kv"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    if sets: params.append(eid); db.execute(f"UPDATE equipements SET {chr(44).join(sets)} WHERE id=?",params)
    if "gammes" in d:
        db.execute("DELETE FROM equipement_gammes WHERE equipement_id=?",(eid,))
        for g in d["gammes"]:
            try:
                if isinstance(g, dict):
                    gid = int(g.get("gamme_id") or g.get("id"))
                    db.execute("""INSERT OR IGNORE INTO equipement_gammes
                        (equipement_id, gamme_id, planning_id, semaine_planif, jour_semaine_planif,
                         intervention_samedi, intervention_dimanche, planif_mode, nth_semaine_mois, force_planif)
                        VALUES (?,?,?,?,?,?,?,?,?,?)""",
                        (eid, gid,
                         to_int(g.get("planning_id")),
                         to_int(g.get("semaine_planif")),
                         to_int(g.get("jour_semaine_planif")),
                         1 if g.get("intervention_samedi") else 0,
                         1 if g.get("intervention_dimanche") else 0,
                         g.get("planif_mode") or 'SEMAINE',
                         to_int(g.get("nth_semaine_mois")),
                         1 if g.get("force_planif") else 0))
                else:
                    db.execute("INSERT OR IGNORE INTO equipement_gammes (equipement_id, gamme_id) VALUES (?,?)",
                               (eid, int(g)))
            except Exception: pass
    db.commit(); return jsonify({"ok":True})

@app.route("/api/equipements/<int:eid>",methods=["DELETE"])
@require_role("admin","manager")
def delete_equipement(eid):
    db=get_db(); db.execute("DELETE FROM pieces WHERE equipement_id=?",(eid,))
    for i in rows(db.execute("SELECT id FROM interventions WHERE equipement_id=?",(eid,))):
        db.execute("DELETE FROM comptes_rendus WHERE intervention_id=?",(i["id"],))
    db.execute("DELETE FROM interventions WHERE equipement_id=?",(eid,))
    db.execute("DELETE FROM equipements WHERE id=?",(eid,)); db.commit()
    return jsonify({"ok":True})

# ══ TABLEAUX ÉLECTRIQUES (sous-équipements BT) ══
@app.route("/api/equipements/<int:eid>/tableaux", methods=["GET"])
@require_auth
def list_tableaux(eid):
    db = get_db()
    return jsonify(rows(db.execute(
        "SELECT id,nom,localisation,ordre FROM equipement_tableaux WHERE equipement_id=? ORDER BY ordre,id",
        (eid,)
    )))

@app.route("/api/equipements/<int:eid>/tableaux", methods=["POST"])
@require_role("admin","manager")
def create_tableau(eid):
    d = request.json or {}
    nom = (d.get("nom","") or "").strip()
    if not nom:
        return jsonify({"error": "Nom requis"}), 400
    db = get_db()
    # Vérifier que l'équipement existe et est de type Basse tension
    eq = one(db.execute("SELECT type_technique FROM equipements WHERE id=?", (eid,)))
    if not eq:
        return jsonify({"error": "Équipement introuvable"}), 404
    db.execute(
        "INSERT INTO equipement_tableaux (equipement_id,nom,localisation,ordre) VALUES (?,?,?,?)",
        (eid, nom, d.get("localisation","") or "", int(d.get("ordre",0) or 0))
    )
    db.commit()
    tid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    return jsonify({"id": tid}), 201

@app.route("/api/tableaux/<int:tid>", methods=["PATCH"])
@require_role("admin","manager")
def update_tableau(tid):
    d = request.json or {}
    db = get_db()
    sets, params = [], []
    for k in ("nom","localisation","ordre"):
        if k in d:
            sets.append(f"{k}=?"); params.append(d[k])
    if not sets:
        return jsonify({"ok": True})
    params.append(tid)
    db.execute(f"UPDATE equipement_tableaux SET {','.join(sets)} WHERE id=?", params)
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/tableaux/<int:tid>", methods=["DELETE"])
@require_role("admin","manager")
def delete_tableau(tid):
    db = get_db()
    # Détacher les interventions qui référencent ce tableau
    db.execute("UPDATE interventions SET tableau_id=NULL WHERE tableau_id=?", (tid,))
    db.execute("DELETE FROM equipement_tableaux WHERE id=?", (tid,))
    db.commit()
    return jsonify({"ok": True})


# v218.160 : endpoints CRUD /cellules SUPPRIMÉS (table equipement_cellules DROP).
# Utiliser /api/equipements/<id>/sous-equipements à la place.


# ══ SOUS-ÉQUIPEMENTS (v218.157) ══
# Sous-équipements universels (toutes techniques) : désignation, marque, type, nombre
@app.route("/api/equipements/<int:eid>/sous-equipements", methods=["GET"])
@require_auth
def list_sous_equipements(eid):
    db = get_db()
    return jsonify(rows(db.execute(
        "SELECT id, designation, marque, type, nombre, ordre FROM equipement_sous_equipements WHERE equipement_id=? ORDER BY ordre, id",
        (eid,)
    )))

@app.route("/api/equipements/<int:eid>/sous-equipements", methods=["PUT"])
@require_role("admin","manager")
def set_sous_equipements(eid):
    """Remplace toute la liste des sous-équipements d'un équipement.
    Body : { items: [{designation, marque, type, nombre}, ...] }"""
    d = request.json or {}
    items = d.get("items") or []
    if not isinstance(items, list):
        return jsonify({"error": "items doit être une liste"}), 400
    db = get_db()
    if not one(db.execute("SELECT 1 FROM equipements WHERE id=?", (eid,))):
        return jsonify({"error": "Équipement introuvable"}), 404
    # Stratégie simple : supprimer tout puis réinsérer
    db.execute("DELETE FROM equipement_sous_equipements WHERE equipement_id=?", (eid,))
    inserted = 0
    for idx, item in enumerate(items):
        designation = (item.get("designation") or "").strip()
        if not designation:
            continue  # ignorer lignes vides
        marque = (item.get("marque") or "").strip()
        type_ = (item.get("type") or "").strip()
        try:
            nombre = int(item.get("nombre") or 1)
            if nombre < 1: nombre = 1
        except Exception:
            nombre = 1
        db.execute(
            "INSERT INTO equipement_sous_equipements (equipement_id, designation, marque, type, nombre, ordre) VALUES (?,?,?,?,?,?)",
            (eid, designation, marque, type_, nombre, idx)
        )
        inserted += 1
    db.commit()
    return jsonify({"ok": True, "count": inserted})


# ══ SÉCURITÉ HT — Catalogue admin ══
@app.route("/api/securite_equipements", methods=["GET"])
@require_auth
def list_securite_equipements():
    """Liste tous les équipements de sécurité (catalogue admin).
    Le param `actif=1` filtre uniquement les actifs (par défaut tous)."""
    db = get_db()
    only_actif = request.args.get("actif") == "1"
    sql = "SELECT id, libelle, photo_mime, ordre, actif, (CASE WHEN photo_data!='' THEN 1 ELSE 0 END) AS has_photo FROM securite_equipements"
    if only_actif:
        sql += " WHERE actif=1"
    sql += " ORDER BY ordre, id"
    return jsonify(rows(db.execute(sql)))

@app.route("/api/securite_equipements/<int:eid>/photo", methods=["GET"])
@require_auth
def get_securite_photo(eid):
    """Retourne l'image en bytes."""
    db = get_db()
    r = one(db.execute("SELECT photo_data, photo_mime FROM securite_equipements WHERE id=?", (eid,)))
    if not r or not r.get("photo_data"):
        return ("", 404)
    import base64
    try:
        data = base64.b64decode(r["photo_data"])
        return (data, 200, {"Content-Type": r.get("photo_mime") or "image/jpeg",
                            "Cache-Control": "private, max-age=300"})
    except Exception:
        return ("", 500)

@app.route("/api/securite_equipements", methods=["POST"])
@require_role("admin","manager")
def create_securite_equipement():
    d = request.json or {}
    libelle = (d.get("libelle","") or "").strip()
    if not libelle:
        return jsonify({"error":"libelle requis"}), 400
    db = get_db()
    photo_data = d.get("photo_data") or ""
    photo_mime = d.get("photo_mime") or ""
    # Nettoyer le préfixe data:image/...;base64,
    if photo_data and photo_data.startswith("data:"):
        comma = photo_data.find(",")
        if comma > 0:
            head = photo_data[:comma]
            photo_data = photo_data[comma+1:]
            if not photo_mime and "image/" in head:
                photo_mime = head.split(";")[0].split(":")[1]
    # Calculer ordre auto = max+1
    nx = (one(db.execute("SELECT COALESCE(MAX(ordre),0)+1 AS nx FROM securite_equipements")) or {}).get("nx",1)
    db.execute("INSERT INTO securite_equipements (libelle, photo_data, photo_mime, ordre, actif) VALUES (?,?,?,?,1)",
               (libelle, photo_data, photo_mime, int(d.get("ordre", nx) or nx)))
    db.commit()
    sid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    return jsonify({"id": sid}), 201

@app.route("/api/securite_equipements/<int:eid>", methods=["PATCH"])
@require_role("admin","manager")
def update_securite_equipement(eid):
    d = request.json or {}
    db = get_db()
    sets, params = [], []
    if "libelle" in d:
        lib = (d.get("libelle","") or "").strip()
        if not lib:
            return jsonify({"error":"libelle ne peut pas etre vide"}), 400
        sets.append("libelle=?"); params.append(lib)
    if "ordre" in d:
        sets.append("ordre=?"); params.append(int(d.get("ordre") or 0))
    if "actif" in d:
        sets.append("actif=?"); params.append(1 if d.get("actif") else 0)
    if "photo_data" in d:
        photo_data = d.get("photo_data") or ""
        photo_mime = d.get("photo_mime") or ""
        if photo_data and photo_data.startswith("data:"):
            comma = photo_data.find(",")
            if comma > 0:
                head = photo_data[:comma]
                photo_data = photo_data[comma+1:]
                if not photo_mime and "image/" in head:
                    photo_mime = head.split(";")[0].split(":")[1]
        sets.append("photo_data=?"); params.append(photo_data)
        sets.append("photo_mime=?"); params.append(photo_mime)
    if not sets:
        return jsonify({"ok": True})
    params.append(eid)
    db.execute(f"UPDATE securite_equipements SET {','.join(sets)} WHERE id=?", params)
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/securite_equipements/<int:eid>", methods=["DELETE"])
@require_role("admin","manager")
def delete_securite_equipement(eid):
    db = get_db()
    db.execute("DELETE FROM securite_equipements WHERE id=?", (eid,))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/securite_equipements/reorder", methods=["POST"])
@require_role("admin","manager")
def reorder_securite_equipements():
    """Body: {ids: [3, 1, 2, ...]} — réordonne selon l'ordre passé."""
    d = request.json or {}
    ids = d.get("ids") or []
    if not isinstance(ids, list):
        return jsonify({"error":"ids must be a list"}), 400
    db = get_db()
    for idx, sid in enumerate(ids, start=1):
        try:
            db.execute("UPDATE securite_equipements SET ordre=? WHERE id=?", (idx, int(sid)))
        except Exception: pass
    db.commit()
    return jsonify({"ok": True})


# ══ SÉCURITÉ HT — Statut sur un bon ══
@app.route("/api/interventions/<int:iid>/securite", methods=["GET"])
@require_auth
def get_intervention_securite(iid):
    """Liste tous les équipements de sécurité actifs + statut sur ce bon."""
    db = get_db()
    if not one(db.execute("SELECT 1 FROM interventions WHERE id=?", (iid,))):
        return jsonify({"error":"Intervention introuvable"}), 404
    items = rows(db.execute("""
        SELECT se.id, se.libelle, se.ordre,
               (CASE WHEN se.photo_data!='' THEN 1 ELSE 0 END) AS has_photo,
               isec.present AS present,
               isec.conforme AS conforme
        FROM securite_equipements se
        LEFT JOIN intervention_securite isec ON isec.securite_equipement_id=se.id AND isec.intervention_id=?
        WHERE se.actif=1
        ORDER BY se.ordre, se.id
    """, (iid,)))
    return jsonify(items)

@app.route("/api/interventions/<int:iid>/securite/<int:eq_id>", methods=["POST"])
@require_auth
def set_intervention_securite(iid, eq_id):
    """Body: {present: 0/1, conforme: 0/1}. Upsert."""
    d = request.json or {}
    db = get_db()
    if not one(db.execute("SELECT 1 FROM interventions WHERE id=?", (iid,))):
        return jsonify({"error":"Intervention introuvable"}), 404
    if not one(db.execute("SELECT 1 FROM securite_equipements WHERE id=?", (eq_id,))):
        return jsonify({"error":"Équipement de sécurité introuvable"}), 404
    # Accepte 0/1/null pour chaque axe (null = indéterminé)
    def _norm(v):
        if v is None: return None
        return 1 if v else 0
    present = _norm(d.get("present"))
    conforme = _norm(d.get("conforme"))
    # Upsert
    existing = one(db.execute(
        "SELECT id FROM intervention_securite WHERE intervention_id=? AND securite_equipement_id=?",
        (iid, eq_id)
    ))
    if existing:
        db.execute("UPDATE intervention_securite SET present=?, conforme=?, updated_at=datetime('now') WHERE id=?",
                   (present, conforme, existing["id"]))
    else:
        db.execute("INSERT INTO intervention_securite (intervention_id, securite_equipement_id, present, conforme) VALUES (?,?,?,?)",
                   (iid, eq_id, present, conforme))
    db.commit()
    return jsonify({"ok": True})


# ══ BILAN ANNUEL ══
def _compute_bilan_annuel(projet_id, annee):
    """Calcule toutes les données du bilan annuel pour un projet sur une année donnée."""
    db = get_db()
    annee = int(annee)
    date_debut = f"{annee}-01-01"
    date_fin = f"{annee}-12-31"
    # Projet
    projet = one(db.execute("""
        SELECT p.id, p.nom, p.numero_projet, p.logo_filename, c.societe AS client_nom
        FROM projets p LEFT JOIN clients c ON p.client_id=c.id
        WHERE p.id=?
    """, (projet_id,)))
    if not projet:
        return None
    # Filtre interventions : par date_realisation si présente sinon date_prevue, dans l'année
    iv_filter = """
        FROM interventions i JOIN equipements e ON i.equipement_id=e.id
        WHERE e.projet_id=?
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) >= ?
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) <= ?
    """
    # Totaux BC / BP
    totaux = {
        "total": (one(db.execute("SELECT COUNT(*) AS n " + iv_filter, (projet_id, date_debut, date_fin))) or {}).get("n", 0),
        "depannage": (one(db.execute("SELECT COUNT(*) AS n " + iv_filter + " AND i.type='DEPANNAGE'", (projet_id, date_debut, date_fin))) or {}).get("n", 0),
        "maintenance": (one(db.execute("SELECT COUNT(*) AS n " + iv_filter + " AND i.type='MAINTENANCE'", (projet_id, date_debut, date_fin))) or {}).get("n", 0),
        "terminees": (one(db.execute("SELECT COUNT(*) AS n " + iv_filter + " AND i.statut='TERMINEE'", (projet_id, date_debut, date_fin))) or {}).get("n", 0),
        "en_cours": (one(db.execute("SELECT COUNT(*) AS n " + iv_filter + " AND i.statut='EN_COURS'", (projet_id, date_debut, date_fin))) or {}).get("n", 0),
    }
    # Répartition par technique
    par_technique = rows(db.execute("""
        SELECT e.type_technique AS technique, COUNT(*) AS nb
        FROM interventions i JOIN equipements e ON i.equipement_id=e.id
        WHERE e.projet_id=?
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) >= ?
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) <= ?
        GROUP BY e.type_technique ORDER BY nb DESC
    """, (projet_id, date_debut, date_fin)))
    # Évolution mois par mois (12 mois)
    par_mois = []
    for m in range(1, 13):
        d1 = f"{annee}-{m:02d}-01"
        d2 = f"{annee}-{m:02d}-31"
        bc = (one(db.execute("SELECT COUNT(*) AS n " + iv_filter + " AND i.type='DEPANNAGE'", (projet_id, d1, d2))) or {}).get("n", 0)
        bp = (one(db.execute("SELECT COUNT(*) AS n " + iv_filter + " AND i.type='MAINTENANCE'", (projet_id, d1, d2))) or {}).get("n", 0)
        par_mois.append({"mois": m, "bc": bc, "bp": bp})
    # Heures passées (somme des total_heures de tous les CR de l'année)
    heures = (one(db.execute("""
        SELECT COALESCE(SUM(cr.total_heures),0) AS total
        FROM comptes_rendus cr
        JOIN interventions i ON cr.intervention_id=i.id
        JOIN equipements e ON i.equipement_id=e.id
        WHERE e.projet_id=?
          AND cr.date_intervention >= ? AND cr.date_intervention <= ?
    """, (projet_id, date_debut, date_fin))) or {}).get("total", 0)
    # Équipements concernés + leurs CR
    equipements = rows(db.execute("""
        SELECT DISTINCT e.id, e.designation, e.type_technique, e.localisation
        FROM equipements e JOIN interventions i ON i.equipement_id=e.id
        WHERE e.projet_id=?
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) >= ?
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) <= ?
        ORDER BY e.designation
    """, (projet_id, date_debut, date_fin)))
    for eq in equipements:
        eq["interventions"] = rows(db.execute("""
            SELECT i.id, i.numero, i.type, i.statut, i.date_prevue, i.date_realisation,
                   (SELECT COUNT(*) FROM comptes_rendus WHERE intervention_id=i.id) AS nb_cr
            FROM interventions i
            WHERE i.equipement_id=?
              AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) >= ?
              AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) <= ?
            ORDER BY COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,''))
        """, (eq["id"], date_debut, date_fin)))
    # Liste complète des interventions
    interventions = rows(db.execute("""
        SELECT i.id, i.numero, i.type, i.statut, i.date_prevue, i.date_realisation,
               e.designation AS equipement, e.type_technique,
               u.nom AS technicien_nom,
               (SELECT COUNT(*) FROM comptes_rendus WHERE intervention_id=i.id) AS nb_cr,
               (SELECT COALESCE(SUM(total_heures),0) FROM comptes_rendus WHERE intervention_id=i.id) AS heures
        FROM interventions i JOIN equipements e ON i.equipement_id=e.id
        LEFT JOIN utilisateurs u ON i.technicien_id=u.id
        WHERE e.projet_id=?
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) >= ?
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) <= ?
        ORDER BY COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,''))
    """, (projet_id, date_debut, date_fin)))
    # Complétion gammes maintenance préventive (sur tous les BP de l'année)
    completion_bp = []
    for iv in [x for x in interventions if x["type"]=="MAINTENANCE"]:
        # Récupérer les opérations de la gamme de l'équipement de cette intervention
        eq_id = (one(db.execute("SELECT equipement_id FROM interventions WHERE id=?", (iv["id"],))) or {}).get("equipement_id")
        if not eq_id: continue
        gids = [g["gamme_id"] for g in rows(db.execute("SELECT gamme_id FROM equipement_gammes WHERE equipement_id=?", (eq_id,)))]
        eq_row = one(db.execute("SELECT gamme_id FROM equipements WHERE id=?", (eq_id,)))
        if eq_row and eq_row.get("gamme_id") and eq_row["gamme_id"] not in gids:
            gids.append(eq_row["gamme_id"])
        if not gids: continue
        ph = ",".join(["?"]*len(gids))
        nb_ops = (one(db.execute(f"SELECT COUNT(*) AS n FROM gamme_operations WHERE gamme_id IN ({ph})", gids)) or {}).get("n", 0)
        if not nb_ops: continue
        nb_done = (one(db.execute(
            f"SELECT COUNT(*) AS n FROM intervention_operations io WHERE io.intervention_id=? AND io.gamme_operation_id IN (SELECT id FROM gamme_operations WHERE gamme_id IN ({ph}))",
            [iv["id"]] + gids
        )) or {}).get("n", 0)
        completion_bp.append({
            "intervention_id": iv["id"],
            "numero": iv["numero"],
            "equipement": iv["equipement"],
            "date": iv.get("date_realisation") or iv.get("date_prevue") or "",
            "nb_ops": nb_ops,
            "nb_done": nb_done,
            "pct": round(100 * nb_done / nb_ops) if nb_ops else 0,
        })
    return {
        "projet": projet,
        "annee": annee,
        "totaux": totaux,
        "par_technique": par_technique,
        "par_mois": par_mois,
        "heures": float(heures or 0),
        "equipements": equipements,
        "interventions": interventions,
        "completion_bp": completion_bp,
    }

@app.route("/api/bilan_annuel", methods=["GET"])
@require_auth
def get_bilan_annuel():
    projet_id = request.args.get("projet_id", type=int)
    annee = request.args.get("annee", type=int)
    if not projet_id or not annee:
        return jsonify({"error": "projet_id et annee requis"}), 400
    data = _compute_bilan_annuel(projet_id, annee)
    if not data:
        return jsonify({"error": "Projet introuvable"}), 404
    return jsonify(data)

@app.route("/api/bilan_annuel/pdf", methods=["GET"])
@require_auth
def get_bilan_annuel_pdf():
    projet_id = request.args.get("projet_id", type=int)
    annee = request.args.get("annee", type=int)
    if not projet_id or not annee:
        return jsonify({"error": "projet_id et annee requis"}), 400
    data = _compute_bilan_annuel(projet_id, annee)
    if not data:
        return jsonify({"error": "Projet introuvable"}), 404
    try:
        from rapport_pdf import generate_bilan_annuel_pdf
        pdf_bytes = generate_bilan_annuel_pdf(data)
    except Exception as e:
        return jsonify({"error": f"Génération PDF échouée: {e}"}), 500
    fname = f"bilan_{(data['projet'].get('numero_projet') or 'projet')}_{annee}.pdf"
    return (pdf_bytes, 200, {
        "Content-Type": "application/pdf",
        "Content-Disposition": f'attachment; filename="{fname}"',
    })


# ══════════════════════════════════════════════════════════════════════
# v218.16 : MON ÉQUIPE — fiches techniciens
# ══════════════════════════════════════════════════════════════════════
@app.route("/api/equipe/techniciens", methods=["GET"])
@require_auth
def list_equipe_techniciens():
    """Retourne la liste des techniciens avec quelques stats résumées pour la grille de tuiles.
    - admin/manager : tous les techniciens
    - technicien : uniquement lui-même
    v218.73 : filtré par société active via utilisateur_societes.
    """
    db = get_db()
    annee = int(request.args.get("annee") or datetime.now().year)
    user = request.user
    role = (user.get("role") or "").lower()
    sid = current_societe_id()
    # v218.73 : filtre via utilisateur_societes (rôle de la société active)
    where = """WHERE us.societe_id=? AND us.actif=1 AND us.role='technicien'
               AND COALESCE(u.actif,1)=1"""
    params = [sid]
    if role == "technicien":
        where += " AND u.id=?"
        params.append(user["id"])
    techs = rows(db.execute(f"""
        SELECT u.id, u.nom, u.email, us.role AS role
        FROM utilisateurs u
        JOIN utilisateur_societes us ON us.utilisateur_id = u.id
        {where}
        ORDER BY u.nom
    """, params))
    logger.info(f"[EQUIPE] user={user.get('id')} role={role} sid={sid} → {len(techs)} techniciens trouvés")
    # Stats par technicien : heures BC + BP + nb interventions
    date_debut = f"{annee}-01-01"
    date_fin = f"{annee}-12-31"
    for t in techs:
        # Heures via cr_intervenants (sum total_heures de ses CR de l'année)
        h = one(db.execute("""
            SELECT
              COALESCE(SUM(CASE WHEN i.type='DEPANNAGE' THEN ci.total_heures ELSE 0 END), 0) AS h_bc,
              COALESCE(SUM(CASE WHEN i.type='MAINTENANCE' THEN ci.total_heures ELSE 0 END), 0) AS h_bp
            FROM cr_intervenants ci
            JOIN comptes_rendus cr ON ci.cr_id=cr.id
            JOIN interventions i ON cr.intervention_id=i.id
            WHERE ci.utilisateur_id=? AND ci.date>=? AND ci.date<=?
        """, (t["id"], date_debut, date_fin))) or {}
        t["h_bc"] = round(float(h.get("h_bc") or 0), 1)
        t["h_bp"] = round(float(h.get("h_bp") or 0), 1)
        # Nombre d'interventions où le technicien apparaît (CR ou tech principal)
        nb = one(db.execute("""
            SELECT COUNT(DISTINCT i.id) AS n
            FROM interventions i
            LEFT JOIN comptes_rendus cr ON cr.intervention_id=i.id
            LEFT JOIN cr_intervenants ci ON ci.cr_id=cr.id
            WHERE (ci.utilisateur_id=? OR i.technicien_id=?)
              AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) >= ?
              AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) <= ?
        """, (t["id"], t["id"], date_debut, date_fin))) or {}
        t["nb_interventions"] = nb.get("n", 0)
    return jsonify(techs)


@app.route("/api/equipe/techniciens/<int:tech_id>/stats", methods=["GET"])
@require_auth
def get_technicien_stats(tech_id):
    """Stats détaillées d'un technicien sur une année donnée."""
    db = get_db()
    annee = int(request.args.get("annee") or datetime.now().year)
    user = request.user
    role = (user.get("role") or "").lower()
    # Permissions : technicien ne peut voir que lui-même
    if role == "technicien" and user["id"] != tech_id:
        return jsonify({"error": "Accès refusé"}), 403
    # Vérif tech existe
    tech = one(db.execute("SELECT id, nom, email, role FROM utilisateurs WHERE id=?", (tech_id,)))
    if not tech:
        return jsonify({"error": "Technicien introuvable"}), 404
    date_debut = f"{annee}-01-01"
    date_fin = f"{annee}-12-31"

    # KPI heures (BC, BP)
    heures = one(db.execute("""
        SELECT
          COALESCE(SUM(CASE WHEN i.type='DEPANNAGE' THEN ci.total_heures ELSE 0 END), 0) AS h_bc,
          COALESCE(SUM(CASE WHEN i.type='MAINTENANCE' THEN ci.total_heures ELSE 0 END), 0) AS h_bp,
          COALESCE(SUM(ci.total_heures), 0) AS h_total
        FROM cr_intervenants ci
        JOIN comptes_rendus cr ON ci.cr_id=cr.id
        JOIN interventions i ON cr.intervention_id=i.id
        WHERE ci.utilisateur_id=? AND ci.date>=? AND ci.date<=?
    """, (tech_id, date_debut, date_fin))) or {}

    # Congés (jours validés sur l'année)
    h_conges = 0
    try:
        cg = one(db.execute("""
            SELECT COALESCE(SUM(nb_jours), 0) AS n
            FROM demandes_conges
            WHERE utilisateur_id=? AND statut='APPROUVEE'
              AND date_debut <= ? AND date_fin >= ?
        """, (tech_id, date_fin, date_debut))) or {}
        h_conges = float(cg.get("n") or 0)
    except Exception:
        # Si la table demandes_conges n'existe pas dans cette installation, on ignore
        pass

    # Astreintes effectuées (semaines)
    nb_astreintes = 0
    try:
        ast = one(db.execute("""
            SELECT COUNT(*) AS n FROM astreintes
            WHERE utilisateur_id=? AND date_debut>=? AND date_debut<=?
        """, (tech_id, date_debut, date_fin))) or {}
        nb_astreintes = ast.get("n", 0)
    except Exception:
        pass

    # v218.17 : Toutes les occupations groupées par type, avec total_heures
    occupations_par_type = []
    try:
        # v218.54 : utilise la table de liaison occupation_techniciens (multi-tech)
        occupations_par_type = rows(db.execute("""
            SELECT
              ot.nom AS type_nom,
              ot.couleur AS type_couleur,
              COALESCE(SUM(o.total_heures), 0) AS total_heures,
              COUNT(o.id) AS nb_occurrences
            FROM occupations o
            JOIN occupation_types ot ON o.type_id=ot.id
            JOIN occupation_techniciens link ON link.occupation_id=o.id
            WHERE link.technicien_id=? AND o.date>=? AND o.date<=?
            GROUP BY ot.id, ot.nom, ot.couleur
            ORDER BY total_heures DESC
        """, (tech_id, date_debut, date_fin)))
        for o in occupations_par_type:
            o["total_heures"] = round(float(o.get("total_heures") or 0), 1)
    except Exception:
        pass

    # v218.17 : Heures supplémentaires (>40h/semaine)
    # On agrège toutes les heures travaillées (cr_intervenants + occupations hors congés/maladie)
    # par semaine ISO, puis on calcule le surplus au-delà de 40h/semaine.
    heures_supp = 0.0
    detail_supp = []  # liste {semaine, total_heures, supp}
    try:
        # Récupérer toutes les lignes de temps : interventions (cr_intervenants) + occupations
        # CR intervenants (heures travaillées sur bons)
        lignes_iv = rows(db.execute("""
            SELECT ci.date, ci.total_heures
            FROM cr_intervenants ci
            WHERE ci.utilisateur_id=? AND ci.date>=? AND ci.date<=?
        """, (tech_id, date_debut, date_fin)))
        # Occupations qui comptent comme heures travaillées (Formation, Réunion, Offre, Autre)
        # On exclut Congés, Maladie, RTT (= absences, pas du travail)
        lignes_occ = []
        try:
            lignes_occ = rows(db.execute("""
                SELECT o.date, o.total_heures
                FROM occupations o
                JOIN occupation_types ot ON o.type_id=ot.id
                JOIN occupation_techniciens link ON link.occupation_id=o.id
                WHERE link.technicien_id=? AND o.date>=? AND o.date<=?
                  AND ot.nom NOT IN ('Congés', 'Maladie', 'RTT')
            """, (tech_id, date_debut, date_fin)))
        except Exception:
            pass
        # Agréger par semaine ISO
        from collections import defaultdict
        heures_par_semaine = defaultdict(float)
        for l in (lignes_iv + lignes_occ):
            d_str = l.get("date") or ""
            if not d_str: continue
            try:
                d_obj = datetime.strptime(d_str[:10], "%Y-%m-%d")
                iso_year, iso_week, _ = d_obj.isocalendar()
                key = f"{iso_year}-S{iso_week:02d}"
                heures_par_semaine[key] += float(l.get("total_heures") or 0)
            except Exception:
                pass
        SEUIL_HEBDO = 40.0
        for sem, total in sorted(heures_par_semaine.items()):
            if total > SEUIL_HEBDO:
                supp = total - SEUIL_HEBDO
                heures_supp += supp
                detail_supp.append({
                    "semaine": sem,
                    "total_heures": round(total, 1),
                    "supp": round(supp, 1)
                })
    except Exception as e:
        logger.warning(f"[TECHNICIEN_STATS] erreur calcul heures supp tech={tech_id} : {e}")

    # Bons par statut
    statuts = {}
    for st in ["TERMINEE", "EN_COURS", "ANNULEE", "PLANIFIEE", "A_PLANIFIER"]:
        r = one(db.execute("""
            SELECT COUNT(DISTINCT i.id) AS n
            FROM interventions i
            LEFT JOIN comptes_rendus cr ON cr.intervention_id=i.id
            LEFT JOIN cr_intervenants ci ON ci.cr_id=cr.id
            WHERE (ci.utilisateur_id=? OR i.technicien_id=?)
              AND i.statut=?
              AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) >= ?
              AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) <= ?
        """, (tech_id, tech_id, st, date_debut, date_fin))) or {}
        statuts[st] = r.get("n", 0)

    # Répartition par technique (équipement)
    par_technique = rows(db.execute("""
        SELECT
          COALESCE(NULLIF(e.type_technique,''), '—') AS technique,
          SUM(CASE WHEN i.type='DEPANNAGE' THEN 1 ELSE 0 END) AS nb_bc,
          SUM(CASE WHEN i.type='MAINTENANCE' THEN 1 ELSE 0 END) AS nb_bp,
          COUNT(DISTINCT i.id) AS nb_total
        FROM interventions i
        LEFT JOIN equipements e ON i.equipement_id=e.id
        LEFT JOIN comptes_rendus cr ON cr.intervention_id=i.id
        LEFT JOIN cr_intervenants ci ON ci.cr_id=cr.id
        WHERE (ci.utilisateur_id=? OR i.technicien_id=?)
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) >= ?
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) <= ?
        GROUP BY technique ORDER BY nb_total DESC
    """, (tech_id, tech_id, date_debut, date_fin)))

    # Évolution mensuelle (heures BC, BP par mois)
    par_mois = []
    for m in range(1, 13):
        d_d = f"{annee}-{m:02d}-01"
        # dernier jour du mois
        if m == 12:
            d_f = f"{annee}-12-31"
        else:
            d_f = f"{annee}-{m+1:02d}-01"
        r = one(db.execute("""
            SELECT
              COALESCE(SUM(CASE WHEN i.type='DEPANNAGE' THEN ci.total_heures ELSE 0 END), 0) AS h_bc,
              COALESCE(SUM(CASE WHEN i.type='MAINTENANCE' THEN ci.total_heures ELSE 0 END), 0) AS h_bp
            FROM cr_intervenants ci
            JOIN comptes_rendus cr ON ci.cr_id=cr.id
            JOIN interventions i ON cr.intervention_id=i.id
            WHERE ci.utilisateur_id=? AND ci.date>=? AND ci.date<?
        """, (tech_id, d_d, d_f))) or {}
        par_mois.append({
            "mois": m,
            "h_bc": round(float(r.get("h_bc") or 0), 1),
            "h_bp": round(float(r.get("h_bp") or 0), 1),
        })

    # Complétion gammes (sur ses bons préventifs)
    completion = rows(db.execute("""
        SELECT i.id AS iv_id, i.numero, e.designation AS equipement,
          COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) AS date,
          (SELECT COUNT(*) FROM intervention_operations io WHERE io.intervention_id=i.id) AS nb_done,
          (SELECT COUNT(*) FROM equipement_gammes eg
             JOIN gamme_operations go ON go.gamme_id=eg.gamme_id
             WHERE eg.equipement_id=i.equipement_id) AS nb_ops_total
        FROM interventions i
        LEFT JOIN equipements e ON i.equipement_id=e.id
        LEFT JOIN comptes_rendus cr ON cr.intervention_id=i.id
        LEFT JOIN cr_intervenants ci ON ci.cr_id=cr.id
        WHERE (ci.utilisateur_id=? OR i.technicien_id=?)
          AND i.type='MAINTENANCE'
          AND i.statut IN ('TERMINEE','EN_COURS')
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) >= ?
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) <= ?
        GROUP BY i.id ORDER BY date DESC
    """, (tech_id, tech_id, date_debut, date_fin)))
    for c in completion:
        nb_total = int(c.get("nb_ops_total") or 0)
        nb_done = int(c.get("nb_done") or 0)
        c["pct"] = round(100 * nb_done / nb_total) if nb_total > 0 else 0

    # Top équipements travaillés (top 10)
    top_equips = rows(db.execute("""
        SELECT e.id, e.designation, e.type_technique,
          COUNT(DISTINCT i.id) AS nb_interventions
        FROM interventions i
        JOIN equipements e ON i.equipement_id=e.id
        LEFT JOIN comptes_rendus cr ON cr.intervention_id=i.id
        LEFT JOIN cr_intervenants ci ON ci.cr_id=cr.id
        WHERE (ci.utilisateur_id=? OR i.technicien_id=?)
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) >= ?
          AND COALESCE(NULLIF(i.date_realisation,''), NULLIF(i.date_prevue,'')) <= ?
        GROUP BY e.id ORDER BY nb_interventions DESC LIMIT 10
    """, (tech_id, tech_id, date_debut, date_fin)))

    # Délais moyens (date_prevue → date_realisation pour les bons terminés)
    delais = one(db.execute("""
        SELECT
          AVG(julianday(i.date_realisation) - julianday(i.date_prevue)) AS delai_moyen_jours,
          MIN(julianday(i.date_realisation) - julianday(i.date_prevue)) AS delai_min,
          MAX(julianday(i.date_realisation) - julianday(i.date_prevue)) AS delai_max,
          COUNT(*) AS nb
        FROM interventions i
        LEFT JOIN comptes_rendus cr ON cr.intervention_id=i.id
        LEFT JOIN cr_intervenants ci ON ci.cr_id=cr.id
        WHERE (ci.utilisateur_id=? OR i.technicien_id=?)
          AND i.statut='TERMINEE'
          AND i.date_prevue IS NOT NULL AND i.date_prevue!=''
          AND i.date_realisation IS NOT NULL AND i.date_realisation!=''
          AND i.date_prevue >= ? AND i.date_prevue <= ?
    """, (tech_id, tech_id, date_debut, date_fin))) or {}

    # v218.40/44 : Score QSE — sommes des points (% - temps_secondes)
    # Règle v218.44 : pourcentage - temps_secondes (clampé à 0). Ex: 80% - 10s = 70pts.
    qse_data = rows(db.execute("""
        SELECT c.id, c.titre, c.planif_mois, c.date_publication,
               r.score, r.total_questions, r.completed_at, r.temps_secondes
        FROM causeries c
        LEFT JOIN causeries_reponses r ON r.causerie_id=c.id AND r.utilisateur_id=?
        WHERE (c.planif_annee=? OR (c.planif_annee=0 AND substr(c.date_publication,1,4)=?))
          AND c.actif=1
        ORDER BY c.planif_mois, c.id
    """, (tech_id, annee, str(annee))))
    qse_total = 0
    qse_details = []
    qse_nb_repondues = 0
    qse_nb_total = len(qse_data)
    for q in qse_data:
        completed = bool(q.get("completed_at"))
        if completed:
            score_q = int(q.get("score") or 0)
            total_q = int(q.get("total_questions") or 0)
            temps = int(q.get("temps_secondes") or 0)
            pct = round(100 * score_q / total_q) if total_q else 0
            # Points QSE = % - temps (jamais négatif)
            points = max(0, pct - temps)
            qse_total += points
            qse_nb_repondues += 1
        else:
            pct = None
            points = None
            temps = 0
        qse_details.append({
            "causerie_id": q["id"],
            "titre": q.get("titre"),
            "mois": q.get("planif_mois") or 0,
            "date_publication": q.get("date_publication"),
            "complet": bool(q.get("completed_at")),
            "score": q.get("score"),
            "total_questions": q.get("total_questions"),
            "pourcentage": pct,
            "temps_secondes": temps,
            "points_qse": points,
            "completed_at": q.get("completed_at"),
        })

    return jsonify({
        "technicien": tech,
        "annee": annee,
        "heures": {
            "bc": round(float(heures.get("h_bc") or 0), 1),
            "bp": round(float(heures.get("h_bp") or 0), 1),
            "total": round(float(heures.get("h_total") or 0), 1),
        },
        "conges_jours": h_conges,
        "nb_astreintes": nb_astreintes,
        # v218.17 : occupations détaillées + heures supplémentaires
        "occupations": occupations_par_type,
        "heures_supp": {
            "total": round(heures_supp, 1),
            "detail": detail_supp,
        },
        "statuts": statuts,
        "par_technique": par_technique,
        "par_mois": par_mois,
        "completion": completion,
        "top_equipements": top_equips,
        "delais": {
            "moyen_jours": round(float(delais.get("delai_moyen_jours") or 0), 1),
            "min_jours": round(float(delais.get("delai_min") or 0), 1),
            "max_jours": round(float(delais.get("delai_max") or 0), 1),
            "nb": delais.get("nb", 0),
        },
        # v218.40 : Score QSE
        "qse": {
            "score_total": qse_total,
            "nb_repondues": qse_nb_repondues,
            "nb_total": qse_nb_total,
            "details": qse_details,
        },
    })


# ══════════════════════════════════════════════════════════════════════
# v218.18 : ASTREINTE — Déclenchement et gestion
# ══════════════════════════════════════════════════════════════════════
@app.route("/api/astreinte/specialite_du_jour", methods=["GET"])
@require_auth
def get_astreinte_specialite_du_jour():
    """Pour chaque spécialité d'astreinte, retourne le technicien de garde
    pour la date donnée (default = aujourd'hui).
    """
    db = get_db()
    date = (request.args.get("date") or "").strip() or datetime.now().strftime("%Y-%m-%d")
    specialites = rows(db.execute("""
        SELECT s.id, s.nom, s.description, s.ordre,
          (SELECT ap.technicien FROM astreinte_planning ap
            WHERE ap.specialite_id=s.id AND ap.date=?
            LIMIT 1) AS technicien_de_garde
        FROM astreinte_specialites s
        ORDER BY s.ordre, s.nom
    """, (date,)))
    return jsonify(specialites)

@app.route("/api/astreinte/declencher", methods=["POST"])
@require_role("admin", "manager", "acl")
def declencher_astreinte():
    """Déclenche une astreinte : crée un bon de type ASTREINTE + un astreinte_declenchements lié.
    v218.20 : Le technicien GMAO est résolu AUTOMATIQUEMENT à partir de la spécialité + date :
      - On lit astreinte_planning pour la date donnée et la spécialité
      - On cherche l'utilisateur GMAO (role=technicien) dont le nom correspond
      - Si trouvé → assigné comme technicien_id de l'intervention
      - Sinon → l'intervention est créée sans tech_id (visible par tous, à assigner manuellement)
    Body :
      - description (requis)
      - demandeur_nom, demandeur_tel, demandeur_email
      - criticite (BLOQUANT, CRITIQUE, MINEUR — default CRITIQUE)
      - heure_appel (HH:MM)
      - specialite_id (requis — pour identifier le tech d'astreinte du jour)
    """
    d = request.json or {}
    if not (d.get("description") or "").strip():
        return jsonify({"error": "Description du problème requise"}), 400
    specialite_id = to_int(d.get("specialite_id"))
    if not specialite_id:
        return jsonify({"error": "Spécialité d'astreinte requise"}), 400
    db = get_db()
    logger.info(f"[ASTREINTE_DECLENCH] début specialite_id={specialite_id} desc='{(d.get('description') or '')[:40]}'")
    try:
        # v218.20 : résoudre le technicien d'astreinte du jour
        date_prevue = (d.get("date_prevue") or datetime.now().strftime("%Y-%m-%d"))
        planning = one(db.execute(
            "SELECT technicien FROM astreinte_planning WHERE specialite_id=? AND date=? LIMIT 1",
            (specialite_id, date_prevue)
        ))
        tech_astreinte_nom = (planning or {}).get("technicien", "") or ""
        # Chercher l'utilisateur GMAO correspondant (par nom)
        # v218.25 : on cherche dans technicien OU manager (un manager peut être de garde)
        tech_id = None
        if tech_astreinte_nom:
            # 1) Match exact
            user_match = one(db.execute(
                "SELECT id FROM utilisateurs WHERE role IN ('technicien','manager') AND nom=? AND actif=1 LIMIT 1",
                (tech_astreinte_nom,)
            ))
            if user_match:
                tech_id = user_match["id"]
            else:
                # 2) Tentative en LIKE simple (insensible à la casse, espaces)
                user_match = one(db.execute(
                    "SELECT id FROM utilisateurs WHERE role IN ('technicien','manager') AND LOWER(nom) LIKE ? AND actif=1 LIMIT 1",
                    (f"%{tech_astreinte_nom.lower().strip()}%",)
                ))
                if user_match:
                    tech_id = user_match["id"]
                else:
                    # v218.24 : 3) Match par mots dans n'importe quel ordre
                    # On extrait les "mots significatifs" du nom du planning et on cherche un utilisateur dont
                    # le nom contient TOUS ces mots (peu importe l'ordre).
                    # v218.25 : on préfère les noms les plus COURTS (moins de mots = match plus précis)
                    # Ex : planning "SALMON Nicolas" → préférer "Nicolas SALMON" (id=4) à "Nicolas SALMON TECHNICIEN" (id=182)
                    mots = [m.strip().lower() for m in tech_astreinte_nom.replace(",", " ").split() if len(m.strip()) >= 2]
                    if mots:
                        conditions = " AND ".join(["LOWER(nom) LIKE ?"] * len(mots))
                        params_match = [f"%{m}%" for m in mots]
                        # ORDER BY length(nom) ASC : on prend le nom le plus court qui matche tous les mots
                        user_match = one(db.execute(
                            f"SELECT id, nom FROM utilisateurs WHERE role IN ('technicien','manager') AND actif=1 AND ({conditions}) ORDER BY length(nom) ASC LIMIT 1",
                            params_match
                        ))
                        if user_match:
                            tech_id = user_match["id"]
                            logger.info(f"[ASTREINTE] match par mots : '{tech_astreinte_nom}' → utilisateur '{user_match['nom']}' (id={tech_id})")
        # Si toujours pas de tech_id : on log un warning mais on crée quand même le bon
        if not tech_id:
            logger.warning(f"[ASTREINTE] tech non trouvé pour spécialité={specialite_id} date={date_prevue} nom='{tech_astreinte_nom}' — bon créé sans assignation")
        logger.info(f"[ASTREINTE_DECLENCH] tech résolu : nom='{tech_astreinte_nom}' tech_id={tech_id}")

        # 1) Créer l'intervention de type ASTREINTE
        # v218.19 : statut PLANIFIEE — le technicien devra "Démarrer" pour confirmer la prise en compte
        num = next_numero(db, "BA", "interventions", "numero")
        heure_prevue = (d.get("heure_appel") or datetime.now().strftime("%H:%M")).strip() or "08:00"
        # equipement_id : si non fourni, on cherche un équipement bidon ou on insère NULL
        # Mais la colonne equipement_id n'est pas NULL dans la table — on doit trouver une solution.
        # Solution : on utilise un équipement "Astreinte (à compléter)" placeholder
        eq_id = to_int(d.get("equipement_id"))
        if not eq_id:
            # Chercher ou créer un équipement placeholder "_ASTREINTE_PLACEHOLDER_"
            placeholder = one(db.execute(
                "SELECT id FROM equipements WHERE designation='_ASTREINTE_PLACEHOLDER_' LIMIT 1"
            ))
            if placeholder:
                eq_id = placeholder["id"]
            else:
                # Créer le placeholder dans le 1er projet existant (ou créer un projet placeholder)
                proj = one(db.execute("SELECT id FROM projets LIMIT 1"))
                if not proj:
                    return jsonify({"error": "Aucun projet en base — impossible de créer le placeholder"}), 500
                cur_eq = db.execute("""
                    INSERT INTO equipements (projet_id, designation, type_technique, statut, notes)
                    VALUES (?, '_ASTREINTE_PLACEHOLDER_', 'ASTREINTE', 'EN_SERVICE',
                            'Placeholder système pour bons astreinte sans équipement assigné')
                """, (proj["id"],))
                eq_id = cur_eq.lastrowid
        logger.info(f"[ASTREINTE_DECLENCH] eq_id={eq_id} num={num}")
        cur = db.execute("""
            INSERT INTO interventions
              (numero, equipement_id, technicien_id, type, statut,
               date_prevue, heure_prevue, description)
            VALUES (?, ?, ?, 'ASTREINTE', 'PLANIFIEE', ?, ?, ?)
        """, (num, eq_id, tech_id, date_prevue, heure_prevue, d["description"]))
        iid = cur.lastrowid
        logger.info(f"[ASTREINTE_DECLENCH] intervention créée iid={iid}")
        # Enregistrer le tech comme intervenant (si trouvé)
        if tech_id:
            try:
                db.execute("INSERT OR IGNORE INTO intervention_techniciens (intervention_id, utilisateur_id) VALUES (?, ?)",
                           (iid, tech_id))
            except Exception: pass
        # 2) Créer l'astreinte_declenchement lié (sans heures arrivee/fin/km — tech remplira)
        db.execute("""
            INSERT INTO astreinte_declenchements
              (intervention_id, demandeur_nom, demandeur_tel, demandeur_email,
               criticite, heure_appel, specialite_id, technicien_astreinte, declenche_par)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            iid,
            (d.get("demandeur_nom") or "").strip(),
            (d.get("demandeur_tel") or "").strip(),
            (d.get("demandeur_email") or "").strip(),
            (d.get("criticite") or "CRITIQUE").strip(),
            (d.get("heure_appel") or "").strip(),
            specialite_id,
            tech_astreinte_nom,
            request.user.get("id"),
        ))
        db.commit()
        log_action(request.user, "CREATE", "astreinte_declenchement", iid,
                   f"BA {num} - {d.get('demandeur_nom','')[:40]}")
        return jsonify({
            "ok": True,
            "intervention_id": iid,
            "numero": num,
            "tech_assigne": tech_astreinte_nom,
            "tech_id": tech_id,  # null si pas de mapping trouvé
        })
    except Exception as e:
        db.rollback()
        import traceback
        logger.error(f"[ASTREINTE] échec déclenchement : {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

@app.route("/api/astreinte/declenchement/<int:iv_id>", methods=["GET"])
@require_auth
def get_astreinte_declenchement(iv_id):
    """Retourne les infos de déclenchement liées à une intervention ASTREINTE."""
    db = get_db()
    d = one(db.execute("""
        SELECT ad.*,
          s.nom AS specialite_nom,
          u.nom AS declenche_par_nom
        FROM astreinte_declenchements ad
        LEFT JOIN astreinte_specialites s ON ad.specialite_id=s.id
        LEFT JOIN utilisateurs u ON ad.declenche_par=u.id
        WHERE ad.intervention_id=?
    """, (iv_id,)))
    if not d:
        return jsonify({"error": "Pas de déclenchement pour cette intervention"}), 404
    return jsonify(d)

@app.route("/api/astreinte/declenchement/<int:iv_id>", methods=["PATCH"])
@require_role("admin", "manager", "acl")
def update_astreinte_declenchement(iv_id):
    """Met à jour les infos de déclenchement (horaires, kilométrage, etc.)."""
    d = request.json or {}
    db = get_db()
    existing = one(db.execute("SELECT * FROM astreinte_declenchements WHERE intervention_id=?", (iv_id,)))
    if not existing:
        return jsonify({"error": "Introuvable"}), 404
    fields = []
    params = []
    for col in ("demandeur_nom", "demandeur_tel", "demandeur_email", "criticite",
                "heure_appel", "heure_arrivee", "heure_fin",
                "specialite_id", "technicien_astreinte"):
        if col in d:
            fields.append(f"{col}=?")
            params.append(d[col])
    if "kilometrage" in d:
        fields.append("kilometrage=?")
        params.append(float(d["kilometrage"] or 0))
    if not fields:
        return jsonify({"ok": True})
    params.append(iv_id)
    db.execute(f"UPDATE astreinte_declenchements SET {','.join(fields)} WHERE intervention_id=?", params)
    db.commit()
    log_action(request.user, "UPDATE", "astreinte_declenchement", iv_id, "")
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════
# v218.30 : QSE/RSE — Causeries avec QCM
# ══════════════════════════════════════════════════════════════════════
CAUSERIES_PDF_DIR = BASE_DIR / "uploads" / "causeries"
CAUSERIES_PDF_DIR.mkdir(parents=True, exist_ok=True)

# v218.37 : helper calcul date de publication selon planif (rang + jour + mois)
_RANG_MAP = {"PREMIER": 1, "DEUXIEME": 2, "TROISIEME": 3, "QUATRIEME": 4, "DERNIER": -1}
_JOUR_MAP = {"LUNDI": 0, "MARDI": 1, "MERCREDI": 2, "JEUDI": 3, "VENDREDI": 4, "SAMEDI": 5, "DIMANCHE": 6}

def _calcul_date_planif(rang, jour, mois, annee):
    """Calcule la date YYYY-MM-DD selon : rang ('PREMIER'..'DERNIER'),
    jour ('LUNDI'..'DIMANCHE'), mois (1..12), annee (4 chiffres).
    Ex: PREMIER, LUNDI, 1, 2026 → '2026-01-05' (1er lundi de janvier 2026)
    Retourne None si paramètres invalides."""
    if not rang or not jour or not mois:
        return None
    rang_n = _RANG_MAP.get(rang.upper())
    jour_n = _JOUR_MAP.get(jour.upper())
    if rang_n is None or jour_n is None:
        return None
    try:
        mois = int(mois); annee = int(annee or datetime.now().year)
        if mois < 1 or mois > 12: return None
    except Exception:
        return None
    if rang_n > 0:
        # n-ième jour du mois (1er, 2ème, 3ème, 4ème)
        # On parcourt les jours du mois et on compte les jours-cible
        d = datetime(annee, mois, 1)
        count = 0
        while d.month == mois:
            if d.weekday() == jour_n:
                count += 1
                if count == rang_n:
                    return d.strftime("%Y-%m-%d")
            d += timedelta(days=1)
        return None  # pas trouvé (ex: 5e lundi inexistant en février)
    else:
        # DERNIER : on part de la fin du mois
        # Trouver le dernier jour du mois
        if mois == 12:
            next_first = datetime(annee + 1, 1, 1)
        else:
            next_first = datetime(annee, mois + 1, 1)
        d = next_first - timedelta(days=1)
        # Reculer jusqu'au jour-cible
        while d.weekday() != jour_n:
            d -= timedelta(days=1)
        return d.strftime("%Y-%m-%d")


def _causerie_a_repondre(user_id):
    """Retourne la causerie active non répondue par cet utilisateur (la plus récente),
    ou None si tout est OK."""
    db = get_db()
    return one(db.execute("""
        SELECT c.id, c.titre, c.pdf_filename, c.date_publication
        FROM causeries c
        WHERE c.actif=1
          AND (c.date_publication IS NULL OR c.date_publication <= date('now'))
          AND NOT EXISTS (
            SELECT 1 FROM causeries_reponses r
            WHERE r.causerie_id=c.id AND r.utilisateur_id=?
          )
        ORDER BY c.date_publication DESC, c.id DESC
        LIMIT 1
    """, (user_id,)))

@app.route("/api/causeries", methods=["GET"])
@require_auth
def list_causeries():
    """Admin : liste toutes les causeries (avec filtre annee facultatif).
    Tech/manager : retourne la causerie en cours non répondue (1 seule).
    Query params :
      - annee (facultatif, ex: 2026) : filtre par planif_annee OU par année de date_publication si planif vide
    """
    db = get_db()
    user = request.user
    sid = current_societe_id()
    role = (user.get("role") or "").lower()
    if role == "admin" or role == "superadmin":
        annee_str = request.args.get("annee")
        params = [sid]
        where = "WHERE c.societe_id=?"
        if annee_str:
            try:
                annee = int(annee_str)
                where += " AND (c.planif_annee=? OR (c.planif_annee=0 AND substr(c.date_publication,1,4)=?))"
                params += [annee, str(annee)]
            except Exception:
                pass
        cs = rows(db.execute(f"""
            SELECT c.*,
              (SELECT COUNT(*) FROM causeries_questions q WHERE q.causerie_id=c.id) AS nb_questions,
              (SELECT COUNT(*) FROM causeries_reponses r WHERE r.causerie_id=c.id) AS nb_reponses
            FROM causeries c
            {where}
            ORDER BY c.planif_annee DESC, c.planif_mois ASC, c.created_at DESC, c.id DESC
        """, params))
        return jsonify(cs)
    else:
        # Pour les non-admin : la causerie à répondre (s'il y en a une)
        c = _causerie_a_repondre(user["id"])
        return jsonify([c] if c else [])

@app.route("/api/causeries/a_repondre", methods=["GET"])
@require_auth
def get_causerie_a_repondre():
    """Endpoint utilisé par tous les utilisateurs au démarrage de l'app pour savoir
    s'ils ont une causerie à répondre. Retourne null sinon."""
    c = _causerie_a_repondre(request.user["id"])
    return jsonify(c)

# v218.40 : profil utilisateur (utilisé par le mobile pour la fiche Profil)
@app.route("/api/mon_profil", methods=["GET"])
@require_auth
def get_mon_profil():
    """Retourne les statistiques de l'utilisateur connecté pour l'année donnée.
    Plus léger que /api/equipe/techniciens/<id>/stats — pour mobile.
    """
    db = get_db()
    user = request.user
    tech_id = user["id"]
    annee = int(request.args.get("annee") or datetime.now().year)
    date_debut = f"{annee}-01-01"
    date_fin = f"{annee}-12-31"
    # Infos user
    u = one(db.execute("SELECT id, nom, email, role, matricule FROM utilisateurs WHERE id=?", (tech_id,)))
    if not u:
        return jsonify({"error": "Utilisateur introuvable"}), 404
    # Heures BC/BP totales (BC=DEPANNAGE, BP=MAINTENANCE selon convention existante)
    heures = one(db.execute("""
        SELECT
          COALESCE(SUM(CASE WHEN i.type='DEPANNAGE' THEN ci.total_heures ELSE 0 END), 0) AS h_bc,
          COALESCE(SUM(CASE WHEN i.type IN ('MAINTENANCE','ASTREINTE') THEN ci.total_heures ELSE 0 END), 0) AS h_bp,
          COALESCE(SUM(ci.total_heures), 0) AS h_total
        FROM cr_intervenants ci
        JOIN comptes_rendus cr ON ci.cr_id=cr.id
        JOIN interventions i ON cr.intervention_id=i.id
        WHERE ci.utilisateur_id=? AND ci.date>=? AND ci.date<=?
    """, (tech_id, date_debut, date_fin))) or {}
    # Statuts interventions
    statuts = one(db.execute("""
        SELECT
          COALESCE(SUM(CASE WHEN i.statut='TERMINEE' THEN 1 ELSE 0 END), 0) AS terminees,
          COALESCE(SUM(CASE WHEN i.statut='EN_COURS' THEN 1 ELSE 0 END), 0) AS en_cours,
          COALESCE(SUM(CASE WHEN i.statut='PLANIFIEE' THEN 1 ELSE 0 END), 0) AS planifiees,
          COUNT(DISTINCT i.id) AS total
        FROM interventions i
        WHERE (i.technicien_id=?
               OR EXISTS (SELECT 1 FROM cr_intervenants ci JOIN comptes_rendus cr ON ci.cr_id=cr.id
                          WHERE cr.intervention_id=i.id AND ci.utilisateur_id=?))
          AND i.date_prevue >= ? AND i.date_prevue <= ?
    """, (tech_id, tech_id, date_debut, date_fin))) or {}
    # Congés (jours validés sur l'année) — colonne = nb_jours
    h_conges = 0
    try:
        cg = one(db.execute("""
            SELECT COALESCE(SUM(nb_jours), 0) AS n
            FROM demandes_conges
            WHERE utilisateur_id=? AND statut='APPROUVEE'
              AND date_debut <= ? AND date_fin >= ?
        """, (tech_id, date_fin, date_debut))) or {}
        h_conges = float(cg.get("n") or 0)
    except Exception:
        pass
    # v218.43 : Heures supplémentaires (>40h/semaine) — même algo que get_technicien_stats
    heures_supp = 0.0
    try:
        lignes_iv = rows(db.execute("""
            SELECT ci.date, ci.total_heures
            FROM cr_intervenants ci
            WHERE ci.utilisateur_id=? AND ci.date>=? AND ci.date<=?
        """, (tech_id, date_debut, date_fin)))
        lignes_occ = []
        try:
            lignes_occ = rows(db.execute("""
                SELECT o.date, o.total_heures
                FROM occupations o
                JOIN occupation_types ot ON o.type_id=ot.id
                JOIN occupation_techniciens link ON link.occupation_id=o.id
                WHERE link.technicien_id=? AND o.date>=? AND o.date<=?
                  AND ot.nom NOT IN ('Congés', 'Maladie', 'RTT')
            """, (tech_id, date_debut, date_fin)))
        except Exception:
            pass
        from collections import defaultdict
        heures_par_semaine = defaultdict(float)
        for l in (lignes_iv + lignes_occ):
            d_str = l.get("date") or ""
            if not d_str: continue
            try:
                d_obj = datetime.strptime(d_str[:10], "%Y-%m-%d")
                iso_year, iso_week, _ = d_obj.isocalendar()
                key = f"{iso_year}-S{iso_week:02d}"
                heures_par_semaine[key] += float(l.get("total_heures") or 0)
            except Exception:
                pass
        SEUIL_HEBDO = 40.0
        for sem, total in sorted(heures_par_semaine.items()):
            if total > SEUIL_HEBDO:
                heures_supp += (total - SEUIL_HEBDO)
    except Exception as e:
        logger.warning(f"[MON_PROFIL] erreur calcul heures supp uid={tech_id}: {e}")
    # Score QSE
    qse_data = rows(db.execute("""
        SELECT c.id, c.titre, c.planif_mois, c.date_publication,
               r.score, r.total_questions, r.completed_at, r.temps_secondes
        FROM causeries c
        LEFT JOIN causeries_reponses r ON r.causerie_id=c.id AND r.utilisateur_id=?
        WHERE (c.planif_annee=? OR (c.planif_annee=0 AND substr(c.date_publication,1,4)=?))
          AND c.actif=1
        ORDER BY c.planif_mois, c.id
    """, (tech_id, annee, str(annee))))
    qse_total = 0
    qse_nb_repondues = 0
    qse_details = []
    for q in qse_data:
        completed = bool(q.get("completed_at"))
        if completed:
            score_q = int(q.get("score") or 0)
            total_q = int(q.get("total_questions") or 0)
            temps = int(q.get("temps_secondes") or 0)
            pct = round(100 * score_q / total_q) if total_q else 0
            points = max(0, pct - temps)
            qse_total += points
            qse_nb_repondues += 1
        else:
            pct = None
            points = None
            temps = 0
        qse_details.append({
            "causerie_id": q["id"],
            "titre": q.get("titre"),
            "mois": q.get("planif_mois") or 0,
            "complet": completed,
            "score": q.get("score"),
            "total_questions": q.get("total_questions"),
            "pourcentage": pct,
            "temps_secondes": temps,
            "points_qse": points,
        })
    return jsonify({
        "user": u,
        "annee": annee,
        "heures": {
            "bc": round(float(heures.get("h_bc") or 0), 1),
            "bp": round(float(heures.get("h_bp") or 0), 1),
            "total": round(float(heures.get("h_total") or 0), 1),
        },
        "statuts": statuts,
        "conges_jours": h_conges,
        "heures_supp": round(heures_supp, 1),
        "qse": {
            "score_total": qse_total,
            "nb_repondues": qse_nb_repondues,
            "nb_total": len(qse_data),
            "details": qse_details,
        },
    })

@app.route("/api/causeries/<int:cid>", methods=["GET"])
@require_auth
def get_causerie(cid):
    """Détail d'une causerie : infos + questions (sans la bonne réponse pour les non-admins
    qui n'ont pas encore répondu)."""
    db = get_db()
    c = one(db.execute("SELECT * FROM causeries WHERE id=?", (cid,)))
    if not c:
        return jsonify({"error": "Introuvable"}), 404
    role = (request.user.get("role") or "").lower()
    qs = rows(db.execute("SELECT * FROM causeries_questions WHERE causerie_id=? ORDER BY ordre, id", (cid,)))
    if role not in ("admin", "superadmin"):
        # Masquer correct_idx pour les non-admins
        for q in qs:
            q.pop("correct_idx", None)
    c["questions"] = qs
    return jsonify(c)

@app.route("/api/causeries", methods=["POST"])
@require_role("admin")
def create_causerie():
    """Crée une causerie avec ses questions QCM (sans PDF, upload séparé).
    Body :
      - titre (requis)
      - frequence ('HEBDO','BI_MENSUEL','MENSUEL','TRIMESTRIEL')
      - date_publication (YYYY-MM-DD)
      - date_cloture (YYYY-MM-DD facultatif)
      - mail_destinataire (email pour le récap)
      - questions: [{texte, opt_a, opt_b, opt_c, opt_d, correct_idx}]
    """
    d = request.json or {}
    titre = (d.get("titre") or "").strip()
    if not titre:
        return jsonify({"error": "Titre requis"}), 400
    questions = d.get("questions") or []
    if not questions or len(questions) < 1:
        return jsonify({"error": "Au moins 1 question requise"}), 400
    db = get_db()
    # v218.37 : calcul auto de date_publication depuis planif si fournie
    planif_rang = (d.get("planif_rang") or "").strip().upper()
    planif_jour = (d.get("planif_jour") or "").strip().upper()
    planif_mois = int(d.get("planif_mois") or 0)
    planif_annee = int(d.get("planif_annee") or datetime.now().year)
    date_pub_calculee = None
    if planif_rang and planif_jour and planif_mois:
        date_pub_calculee = _calcul_date_planif(planif_rang, planif_jour, planif_mois, planif_annee)
        if not date_pub_calculee:
            return jsonify({"error": "Planification invalide (ex: 5ème lundi inexistant)"}), 400
    date_publication = date_pub_calculee or (d.get("date_publication") or "").strip() or None
    sid = current_societe_id()
    try:
        cur = db.execute("""
            INSERT INTO causeries (titre, frequence, date_publication, date_cloture,
                                   mail_destinataire, actif, cree_par,
                                   planif_rang, planif_jour, planif_mois, planif_annee, societe_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            titre,
            (d.get("frequence") or "MENSUEL").upper(),
            date_publication,
            (d.get("date_cloture") or "").strip() or None,
            (d.get("mail_destinataire") or "").strip(),
            1 if d.get("actif", True) else 0,
            request.user["id"],
            planif_rang, planif_jour, planif_mois, planif_annee, sid,
        ))
        cid = cur.lastrowid
        for idx, q in enumerate(questions):
            db.execute("""
                INSERT INTO causeries_questions
                  (causerie_id, ordre, texte, opt_a, opt_b, opt_c, opt_d, correct_idx, societe_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                cid, idx,
                (q.get("texte") or "").strip(),
                (q.get("opt_a") or "").strip(),
                (q.get("opt_b") or "").strip(),
                (q.get("opt_c") or "").strip(),
                (q.get("opt_d") or "").strip(),
                int(q.get("correct_idx") or 0),
                sid,
            ))
        db.commit()
        log_action(request.user, "CREATE", "causerie", cid, titre[:60])
        return jsonify({"ok": True, "id": cid})
    except Exception as e:
        db.rollback()
        logger.error(f"[CAUSERIE] échec création : {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/causeries/<int:cid>", methods=["PATCH"])
@require_role("admin")
def update_causerie(cid):
    """Met à jour les méta-données + remplace les questions (transaction)."""
    d = request.json or {}
    db = get_db()
    # v218.37 : recalcul de date_publication si planif fournie
    if "planif_rang" in d or "planif_jour" in d or "planif_mois" in d:
        rang = (d.get("planif_rang") or "").strip().upper()
        jour = (d.get("planif_jour") or "").strip().upper()
        mois = int(d.get("planif_mois") or 0)
        annee = int(d.get("planif_annee") or datetime.now().year)
        if rang and jour and mois:
            new_date = _calcul_date_planif(rang, jour, mois, annee)
            if not new_date:
                return jsonify({"error": "Planification invalide (ex: 5ème lundi inexistant)"}), 400
            d["date_publication"] = new_date
    sets, params = [], []
    for col in ("titre", "frequence", "date_publication", "date_cloture", "mail_destinataire",
                "planif_rang", "planif_jour"):
        if col in d:
            v = d[col]
            if isinstance(v, str): v = v.strip() or None
            sets.append(f"{col}=?")
            params.append(v)
    for col in ("planif_mois", "planif_annee"):
        if col in d:
            sets.append(f"{col}=?")
            params.append(int(d[col] or 0))
    if "actif" in d:
        sets.append("actif=?")
        params.append(1 if d["actif"] else 0)
    if sets:
        params.append(cid)
        db.execute(f"UPDATE causeries SET {','.join(sets)} WHERE id=?", params)
    # Si questions fournies, on les remplace toutes
    if "questions" in d:
        # v218.36 : supprimer dans le bon ordre pour respecter les FK
        # 1) Détails de réponses (FK → causeries_questions)
        db.execute("""DELETE FROM causeries_reponses_detail
                      WHERE reponse_id IN (SELECT id FROM causeries_reponses WHERE causerie_id=?)""", (cid,))
        # 2) Réponses (FK → causeries)
        db.execute("DELETE FROM causeries_reponses WHERE causerie_id=?", (cid,))
        # 3) Questions
        db.execute("DELETE FROM causeries_questions WHERE causerie_id=?", (cid,))
        # 4) Reset du flag mail_envoye (nouveau cycle car questions modifiées)
        db.execute("UPDATE causeries SET mail_envoye=0, mail_envoye_at='' WHERE id=?", (cid,))
        # 5) Réinsertion des nouvelles questions
        for idx, q in enumerate(d["questions"] or []):
            db.execute("""
                INSERT INTO causeries_questions
                  (causerie_id, ordre, texte, opt_a, opt_b, opt_c, opt_d, correct_idx)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                cid, idx,
                (q.get("texte") or "").strip(),
                (q.get("opt_a") or "").strip(),
                (q.get("opt_b") or "").strip(),
                (q.get("opt_c") or "").strip(),
                (q.get("opt_d") or "").strip(),
                int(q.get("correct_idx") or 0),
            ))
    db.commit()
    log_action(request.user, "UPDATE", "causerie", cid, "")
    return jsonify({"ok": True})

@app.route("/api/causeries/<int:cid>", methods=["DELETE"])
@require_role("admin")
def delete_causerie(cid):
    db = get_db()
    # Supprimer le PDF associé si existe
    c = one(db.execute("SELECT pdf_filename FROM causeries WHERE id=?", (cid,)))
    if c and c.get("pdf_filename"):
        try:
            (CAUSERIES_PDF_DIR / c["pdf_filename"]).unlink(missing_ok=True)
        except Exception: pass
    db.execute("DELETE FROM causeries WHERE id=?", (cid,))
    db.commit()
    log_action(request.user, "DELETE", "causerie", cid, "")
    return jsonify({"ok": True})

@app.route("/api/causeries/<int:cid>/pdf", methods=["POST"])
@require_role("admin")
def upload_causerie_pdf(cid):
    """Upload du PDF de la causerie (multipart/form-data, champ 'file')."""
    db = get_db()
    c = one(db.execute("SELECT id, pdf_filename FROM causeries WHERE id=?", (cid,)))
    if not c:
        return jsonify({"error": "Causerie introuvable"}), 404
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "Fichier requis"}), 400
    if not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Seuls les PDF sont acceptés"}), 400
    # Supprimer ancien fichier si existe
    if c.get("pdf_filename"):
        try:
            (CAUSERIES_PDF_DIR / c["pdf_filename"]).unlink(missing_ok=True)
        except Exception: pass
    # Nouveau nom : causerie_<id>_<timestamp>.pdf
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    new_fname = f"causerie_{cid}_{ts}.pdf"
    f.save(str(CAUSERIES_PDF_DIR / new_fname))
    db.execute("UPDATE causeries SET pdf_filename=? WHERE id=?", (new_fname, cid))
    db.commit()
    return jsonify({"ok": True, "filename": new_fname})

@app.route("/api/causeries/<int:cid>/pdf", methods=["GET"])
@require_auth
def get_causerie_pdf(cid):
    """Sert le PDF associé à une causerie."""
    db = get_db()
    c = one(db.execute("SELECT pdf_filename FROM causeries WHERE id=?", (cid,)))
    if not c or not c.get("pdf_filename"):
        return jsonify({"error": "PDF introuvable"}), 404
    fpath = CAUSERIES_PDF_DIR / c["pdf_filename"]
    if not fpath.exists():
        return jsonify({"error": "Fichier introuvable sur le serveur"}), 404
    return send_file(str(fpath), mimetype="application/pdf")

@app.route("/api/causeries/<int:cid>/repondre", methods=["POST"])
@require_auth
def repondre_causerie(cid):
    """Enregistre les réponses d'un utilisateur à une causerie.
    Body : {reponses: [{question_id, choix_index}], temps_secondes: int}
    v218.44 : temps_secondes envoyé par le client (chrono démarré au clic "Répondre au QCM")
    """
    d = request.json or {}
    rep_input = d.get("reponses") or []
    temps_secondes = int(d.get("temps_secondes") or 0)
    db = get_db()
    user_id = request.user["id"]
    # Vérifier qu'il n'a pas déjà répondu
    existing = one(db.execute(
        "SELECT id FROM causeries_reponses WHERE causerie_id=? AND utilisateur_id=?",
        (cid, user_id)
    ))
    if existing:
        return jsonify({"error": "Vous avez déjà répondu à cette causerie"}), 400
    # Charger les questions pour valider
    questions = rows(db.execute("SELECT * FROM causeries_questions WHERE causerie_id=?", (cid,)))
    qmap = {q["id"]: q for q in questions}
    score = 0
    details = []
    for r in rep_input:
        qid = int(r.get("question_id") or 0)
        choix = int(r.get("choix_index") or 0)
        q = qmap.get(qid)
        if not q: continue
        est_correct = 1 if choix == q.get("correct_idx") else 0
        score += est_correct
        details.append((qid, choix, est_correct))
    try:
        cur = db.execute("""
            INSERT INTO causeries_reponses (causerie_id, utilisateur_id, score, total_questions, temps_secondes)
            VALUES (?, ?, ?, ?, ?)
        """, (cid, user_id, score, len(questions), temps_secondes))
        rep_id = cur.lastrowid
        for qid, choix, est_correct in details:
            db.execute("""
                INSERT INTO causeries_reponses_detail (reponse_id, question_id, choix_index, est_correct)
                VALUES (?, ?, ?, ?)
            """, (rep_id, qid, choix, est_correct))
        db.commit()
        # Calcul des points QSE pour info au client
        pct = round(100 * score / len(questions)) if len(questions) else 0
        points_qse = max(0, pct - temps_secondes)  # négatif possible → clamp à 0
        log_action(request.user, "ANSWER", "causerie", cid, f"score={score}/{len(questions)} pct={pct}% temps={temps_secondes}s pts={points_qse}")
        # v218.39 : envoi mail récap individuel à l'utilisateur (avec ses réponses, couleurs rouge/vert)
        try:
            _envoyer_mail_recap_individuel(cid, user_id, score, questions, details, temps_secondes)
        except Exception as e:
            logger.warning(f"[CAUSERIE] erreur envoi mail individuel cid={cid} uid={user_id}: {e}")
        # v218.33 : envoi mail récap quand TOUT LE MONDE a répondu
        try:
            _check_and_send_causerie_recap(cid)
        except Exception as e:
            logger.warning(f"[CAUSERIE] erreur check recap cid={cid}: {e}")
        return jsonify({"ok": True, "score": score, "total": len(questions),
                        "temps_secondes": temps_secondes, "pourcentage": pct, "points_qse": points_qse})
    except Exception as e:
        db.rollback()
        logger.error(f"[CAUSERIE] échec réponse cid={cid} uid={user_id} : {e}")
        return jsonify({"error": str(e)}), 500


def _envoyer_mail_recap_individuel(cid, user_id, score, questions, details, temps_secondes=0):
    """v218.39 : envoie un mail HTML à l'utilisateur avec récap de ses réponses.
    - Réponses correctes en VERT
    - Réponses incorrectes en ROUGE + bonne réponse mise en évidence
    - Score total + temps + points QSE
    v218.44 : intègre le temps de réponse et les points QSE.
    """
    db = get_db()
    user = one(db.execute("SELECT id, nom, email FROM utilisateurs WHERE id=?", (user_id,)))
    if not user or not user.get("email"):
        logger.info(f"[CAUSERIE_MAIL] uid={user_id} pas d'email — pas de récap envoyé")
        return
    causerie = one(db.execute("SELECT id, titre FROM causeries WHERE id=?", (cid,)))
    if not causerie:
        return
    total_q = len(questions)
    pct = round(100 * score / total_q) if total_q else 0
    points_qse = max(0, pct - int(temps_secondes or 0))
    # Mapping question_id → (choix_user, est_correct)
    user_resp_map = {qid: (choix, est_correct) for (qid, choix, est_correct) in details}
    # Construction du HTML
    score_color = "#10B981" if pct >= 80 else ("#F59E0B" if pct >= 50 else "#DC2626")
    points_color = "#10B981" if points_qse >= 70 else ("#F59E0B" if points_qse >= 40 else "#DC2626")
    html_parts = []
    html_parts.append("""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Récapitulatif causerie</title></head>
<body style="font-family:Arial,Helvetica,sans-serif;background:#F1F5F9;margin:0;padding:20px;color:#0F172A">
<div style="max-width:700px;margin:0 auto;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 4px 12px rgba(0,0,0,0.08)">""")
    # Header
    html_parts.append(f"""
<div style="background:#1E3A8A;color:#fff;padding:24px 28px">
  <div style="font-size:14px;opacity:0.85;margin-bottom:4px">🛡 Causerie QSE / RSE — Récapitulatif individuel</div>
  <div style="font-size:22px;font-weight:700">{html_escape(causerie['titre'])}</div>
  <div style="font-size:13px;margin-top:6px;opacity:0.9">Bonjour <b>{html_escape(user['nom'] or '')}</b>, voici le récapitulatif de vos réponses au questionnaire.</div>
</div>""")
    # Score + Temps + Points QSE (3 cartes côte à côte)
    html_parts.append(f"""
<div style="padding:20px 28px;background:#F8FAFC;border-bottom:1px solid #E2E8F0">
  <table style="width:100%;border-collapse:collapse"><tr>
    <td style="text-align:center;padding:8px;border-right:1px solid #E2E8F0">
      <div style="font-size:11px;color:#64748B;margin-bottom:4px">Score</div>
      <div style="font-size:30px;font-weight:700;color:{score_color}">{score} / {total_q}</div>
      <div style="font-size:14px;color:{score_color};font-weight:600">{pct}%</div>
    </td>
    <td style="text-align:center;padding:8px;border-right:1px solid #E2E8F0">
      <div style="font-size:11px;color:#64748B;margin-bottom:4px">⏱ Temps</div>
      <div style="font-size:30px;font-weight:700;color:#1E3A8A">{int(temps_secondes or 0)}</div>
      <div style="font-size:14px;color:#1E3A8A;font-weight:600">secondes</div>
    </td>
    <td style="text-align:center;padding:8px">
      <div style="font-size:11px;color:#64748B;margin-bottom:4px">🛡 Points QSE</div>
      <div style="font-size:30px;font-weight:700;color:{points_color}">{points_qse}</div>
      <div style="font-size:11px;color:#64748B;margin-top:2px">{pct}% − {int(temps_secondes or 0)}s</div>
    </td>
  </tr></table>
</div>""")
    # Liste des questions
    html_parts.append('<div style="padding:24px 28px">')
    html_parts.append('<div style="font-size:14px;font-weight:700;color:#1E3A8A;margin-bottom:14px">📝 Détail des réponses</div>')
    for qi, q in enumerate(questions):
        qid = q["id"]
        user_choix, est_correct = user_resp_map.get(qid, (None, 0))
        correct_idx = q.get("correct_idx", 0)
        # Couleur de bordure
        border_color = "#10B981" if est_correct else "#DC2626"
        bg_header = "#DCFCE7" if est_correct else "#FEE2E2"
        text_header = "#065F46" if est_correct else "#7F1D1D"
        statut_label = "✓ Correct" if est_correct else "✗ Incorrect"
        html_parts.append(f"""
<div style="margin-bottom:18px;border:2px solid {border_color};border-radius:8px;overflow:hidden">
  <div style="background:{bg_header};color:{text_header};padding:10px 14px;font-weight:700;font-size:13px;display:flex;justify-content:space-between;align-items:center">
    <span>Question {qi+1}</span>
    <span>{statut_label}</span>
  </div>
  <div style="padding:14px;background:#fff">
    <div style="font-weight:600;margin-bottom:12px;color:#0F172A;font-size:14px">{html_escape(q['texte'])}</div>""")
        # Lister les options avec mise en forme
        for idx, letter in enumerate(['a', 'b', 'c', 'd']):
            opt_text = q.get(f"opt_{letter}") or ""
            if not opt_text.strip(): continue
            is_user_choice = (user_choix == idx)
            is_correct_answer = (idx == correct_idx)
            # Détermination du style
            if is_user_choice and is_correct_answer:
                # Bonne réponse cochée → VERT
                style = "background:#DCFCE7;border:2px solid #10B981;color:#065F46;font-weight:600"
                icon = "✅ Votre réponse (correcte)"
            elif is_user_choice and not is_correct_answer:
                # Mauvaise réponse cochée → ROUGE
                style = "background:#FEE2E2;border:2px solid #DC2626;color:#7F1D1D;font-weight:600;text-decoration:line-through"
                icon = "❌ Votre réponse (incorrecte)"
            elif is_correct_answer:
                # Bonne réponse non cochée → VERT pâle
                style = "background:#F0FDF4;border:2px solid #10B981;color:#065F46;font-weight:600"
                icon = "👉 La bonne réponse"
            else:
                # Autre option → neutre
                style = "background:#F8FAFC;border:1px solid #E2E8F0;color:#64748B"
                icon = ""
            html_parts.append(f"""
    <div style="margin-bottom:6px;padding:10px 12px;border-radius:6px;{style}">
      <div style="font-size:13px"><b>{letter.upper()}.</b> {html_escape(opt_text)}</div>
      {f'<div style="font-size:11px;margin-top:4px;opacity:0.85">{icon}</div>' if icon else ''}
    </div>""")
        html_parts.append("  </div>\n</div>")  # fin question
    html_parts.append("</div>")  # fin padding
    # Footer
    html_parts.append("""
<div style="background:#F8FAFC;padding:16px 28px;border-top:1px solid #E2E8F0;font-size:11px;color:#94A3B8;text-align:center">
  Mail généré automatiquement par SOCOM GMAO · Merci de votre participation 🙏
</div>
</div>
</body></html>""")
    html_body = "".join(html_parts)
    # Plain text fallback (simple, pour clients qui ne lisent pas le HTML)
    txt_lines = [
        f"Récapitulatif de la causerie : {causerie['titre']}",
        f"Bonjour {user['nom'] or ''},",
        f"",
        f"Votre score : {score}/{total_q} ({pct}%)",
        f"",
        f"Pour le détail visuel des réponses, consultez la version HTML de ce mail.",
    ]
    subj = f"[Causerie QSE] Votre score : {score}/{total_q} — {causerie['titre']}"
    send_mail(user["email"], subj, html_body, html=True)
    logger.info(f"[CAUSERIE_MAIL] récap individuel envoyé à {user['email']} pour cid={cid} (score={score}/{total_q})")


# Helper pour échapper le HTML dans les chaînes utilisateur
def html_escape(s):
    if s is None: return ""
    s = str(s)
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _check_and_send_causerie_recap(cid):
    """Vérifie si tous les utilisateurs concernés ont répondu, et envoie le mail récap si oui.
    Idempotent : ne renvoie pas le mail si déjà envoyé (champ mail_envoye=1)."""
    db = get_db()
    c = one(db.execute("SELECT * FROM causeries WHERE id=?", (cid,)))
    if not c:
        return
    # Si déjà envoyé, on ne refait pas
    if c.get("mail_envoye"):
        return
    if not c.get("mail_destinataire"):
        return
    # Compter les utilisateurs qui doivent répondre (techniciens + managers actifs)
    # v218.44 : récupère aussi temps_secondes pour calcul des points QSE
    users = rows(db.execute("""
        SELECT u.id, u.nom, u.email, u.role,
          r.score, r.total_questions, r.completed_at, r.temps_secondes
        FROM utilisateurs u
        LEFT JOIN causeries_reponses r ON r.causerie_id=? AND r.utilisateur_id=u.id
        WHERE u.actif=1 AND u.role IN ('technicien','manager')
        ORDER BY u.nom
    """, (cid,)))
    nb_total = len(users)
    nb_repondus = sum(1 for u in users if u.get("completed_at"))
    if nb_total == 0 or nb_repondus < nb_total:
        # Pas encore complet, on attend
        return
    # Tout le monde a répondu → envoi mail récap
    logger.info(f"[CAUSERIE] cid={cid} complet ({nb_repondus}/{nb_total}) → envoi mail récap à {c.get('mail_destinataire')}")
    nb_q = c.get("nb_questions") or one(db.execute(
        "SELECT COUNT(*) AS n FROM causeries_questions WHERE causerie_id=?", (cid,)
    )).get("n", 0)
    # Construction du corps mail (texte) avec tableau
    lines = []
    lines.append(f"Récapitulatif de la causerie QSE/RSE")
    lines.append(f"=" * 50)
    lines.append(f"")
    lines.append(f"Titre : {c.get('titre','')}")
    lines.append(f"Date publication : {c.get('date_publication','')}")
    lines.append(f"Fréquence : {c.get('frequence','')}")
    lines.append(f"Nombre de questions : {nb_q}")
    lines.append(f"Réponses : {nb_repondus} / {nb_total} (COMPLET)")
    lines.append(f"")
    lines.append(f"=== RÉSULTATS ===")
    lines.append(f"")
    # Tableau formaté avec colonnes Temps + Points QSE
    name_w = max((len(u.get("nom") or u.get("email","")) for u in users), default=10)
    name_w = max(name_w, 12)
    role_w = 12
    score_w = 10
    pct_w = 6
    temps_w = 7
    pts_w = 8
    date_w = 19
    lines.append(f"{'Nom':<{name_w}} | {'Rôle':<{role_w}} | {'Score':>{score_w}} | {'%':>{pct_w}} | {'Temps':>{temps_w}} | {'Pts QSE':>{pts_w}} | {'Date complétion':<{date_w}}")
    lines.append("-" * (name_w + role_w + score_w + pct_w + temps_w + pts_w + date_w + 18))
    for u in users:
        nom = u.get("nom") or u.get("email","")
        role = u.get("role","")
        score = u.get("score", 0) or 0
        total = u.get("total_questions") or nb_q or 1
        temps = int(u.get("temps_secondes") or 0)
        pct = round(100 * score / total) if total else 0
        points_qse = max(0, pct - temps)
        date_str = (u.get("completed_at") or "")[:19].replace("T", " ")
        lines.append(f"{nom[:name_w]:<{name_w}} | {role:<{role_w}} | {score:>4} / {total:<3} | {pct:>3}% | {temps:>4} s | {points_qse:>5} pts | {date_str:<{date_w}}")
    lines.append("")
    lines.append("=" * 50)
    lines.append("Note : Les points QSE = pourcentage − temps (en secondes), avec minimum 0.")
    lines.append("Mail généré automatiquement par SOCOM GMAO.")
    body_txt = "\n".join(lines)
    # Pièce jointe : le PDF de la causerie si dispo
    attachments = None
    if c.get("pdf_filename"):
        try:
            fpath = CAUSERIES_PDF_DIR / c["pdf_filename"]
            if fpath.exists():
                with open(str(fpath), "rb") as fp:
                    attachments = [(c["pdf_filename"], fp.read())]
        except Exception as e:
            logger.warning(f"[CAUSERIE] échec lecture PDF pour mail: {e}")
    subj = f"[Causerie QSE] Récap : {c.get('titre','')}"
    send_mail(c["mail_destinataire"], subj, body_txt, attachments=attachments)
    # Marquer comme envoyé
    db.execute("UPDATE causeries SET mail_envoye=1, mail_envoye_at=datetime('now') WHERE id=?", (cid,))
    db.commit()
    logger.info(f"[CAUSERIE] mail récap envoyé pour cid={cid}")


# v218.45 : Classement QSE — somme des points sur l'année pour tous les utilisateurs
@app.route("/api/causeries/classement", methods=["GET"])
@require_auth
def get_classement_qse():
    """Classement de tous les utilisateurs (techniciens + managers + admins actifs)
    par total de points QSE sur l'année donnée.
    v218.76 : filtré par société active (causeries + users de la société).
    Query params : annee (default = année courante)
    Retourne : { annee, classement: [{rank, user_id, nom, score_total, nb_repondues, nb_total}] }
    """
    db = get_db()
    annee = int(request.args.get("annee") or datetime.now().year)
    sid = current_societe_id()
    # Récupère toutes les causeries actives de l'année POUR LA SOCIÉTÉ ACTIVE
    causeries = rows(db.execute("""
        SELECT id FROM causeries
        WHERE actif=1 AND societe_id=?
          AND (planif_annee=? OR (planif_annee=0 AND substr(date_publication,1,4)=?))
    """, (sid, annee, str(annee))))
    causerie_ids = [c["id"] for c in causeries]
    nb_total = len(causerie_ids)
    # Tous les utilisateurs concernés DE LA SOCIÉTÉ ACTIVE (techniciens + managers, via utilisateur_societes)
    users = rows(db.execute("""
        SELECT u.id, u.nom, u.email, us.role AS role
        FROM utilisateurs u
        JOIN utilisateur_societes us ON us.utilisateur_id = u.id
        WHERE u.actif=1 AND us.actif=1 AND us.societe_id=?
          AND us.role IN ('technicien','manager')
        ORDER BY u.nom
    """, (sid,)))
    if not causerie_ids:
        # Pas de causerie cette année → tous à 0
        classement = [{
            "user_id": u["id"], "nom": u["nom"] or u["email"], "role": u["role"],
            "score_total": 0, "nb_repondues": 0, "nb_total": 0
        } for u in users]
        for i, u in enumerate(classement): u["rank"] = i + 1
        return jsonify({"annee": annee, "classement": classement, "nb_causeries": 0})
    # Charger toutes les réponses
    placeholders = ",".join(["?"] * len(causerie_ids))
    reponses = rows(db.execute(f"""
        SELECT utilisateur_id, causerie_id, score, total_questions, temps_secondes
        FROM causeries_reponses
        WHERE causerie_id IN ({placeholders})
    """, causerie_ids))
    # Indexer par utilisateur
    rep_by_user = {}
    for r in reponses:
        uid = r["utilisateur_id"]
        if uid not in rep_by_user:
            rep_by_user[uid] = []
        rep_by_user[uid].append(r)
    # Calcul des points par utilisateur
    classement = []
    for u in users:
        uid = u["id"]
        user_reps = rep_by_user.get(uid, [])
        score_total = 0
        for r in user_reps:
            score_q = int(r.get("score") or 0)
            total_q = int(r.get("total_questions") or 0)
            temps = int(r.get("temps_secondes") or 0)
            pct = round(100 * score_q / total_q) if total_q else 0
            score_total += max(0, pct - temps)
        classement.append({
            "user_id": uid,
            "nom": u["nom"] or u["email"],
            "role": u["role"],
            "score_total": score_total,
            "nb_repondues": len(user_reps),
            "nb_total": nb_total,
        })
    # Tri par score décroissant, puis par nb de réponses, puis par nom
    classement.sort(key=lambda x: (-x["score_total"], -x["nb_repondues"], x["nom"].lower()))
    # Attribuer le rang (gestion des ex-aequo : même score = même rang)
    last_score = None
    last_rank = 0
    for i, item in enumerate(classement):
        if item["score_total"] != last_score:
            last_rank = i + 1
            last_score = item["score_total"]
        item["rank"] = last_rank
    return jsonify({
        "annee": annee,
        "classement": classement,
        "nb_causeries": nb_total,
    })


@app.route("/api/causeries/<int:cid>/resultats", methods=["GET"])
@require_role("admin")
def get_causerie_resultats(cid):
    """Tableau des résultats : tous les techniciens + leur statut/score/temps/points."""
    db = get_db()
    c = one(db.execute("SELECT * FROM causeries WHERE id=?", (cid,)))
    if not c:
        return jsonify({"error": "Introuvable"}), 404
    nb_q = one(db.execute("SELECT COUNT(*) AS n FROM causeries_questions WHERE causerie_id=?", (cid,))).get("n", 0)
    # v218.44 : récupère aussi temps_secondes
    users_raw = rows(db.execute("""
        SELECT u.id, u.nom, u.email, u.role,
          r.score, r.total_questions, r.completed_at, r.temps_secondes
        FROM utilisateurs u
        LEFT JOIN causeries_reponses r ON r.causerie_id=? AND r.utilisateur_id=u.id
        WHERE u.actif=1 AND u.role IN ('technicien','manager','admin')
        ORDER BY r.completed_at DESC, u.nom
    """, (cid,)))
    # Calcul des points QSE par user
    users = []
    for u in users_raw:
        u_out = dict(u)
        if u.get("completed_at") and u.get("total_questions"):
            score_q = int(u.get("score") or 0)
            total_q = int(u.get("total_questions") or 0)
            temps = int(u.get("temps_secondes") or 0)
            pct = round(100 * score_q / total_q) if total_q else 0
            u_out["pourcentage"] = pct
            u_out["points_qse"] = max(0, pct - temps)
        else:
            u_out["pourcentage"] = None
            u_out["points_qse"] = None
        users.append(u_out)
    nb_total = len(users)
    nb_repondus = sum(1 for u in users if u.get("completed_at"))
    return jsonify({
        "causerie": c,
        "nb_questions": nb_q,
        "nb_total_users": nb_total,
        "nb_repondus": nb_repondus,
        "complet": (nb_repondus == nb_total),
        "users": users,
    })

@app.route("/api/causeries/<int:cid>/reset", methods=["POST"])
@require_role("admin")
def reset_causerie_reponses(cid):
    """v218.33 : Réinitialise les réponses (utile pour les causeries récurrentes — nouveau cycle).
    Tous les techniciens devront répondre à nouveau, et le mail récap sera renvoyé à la fin."""
    db = get_db()
    c = one(db.execute("SELECT id FROM causeries WHERE id=?", (cid,)))
    if not c:
        return jsonify({"error": "Introuvable"}), 404
    # v218.36 : supprimer les détails AVANT les réponses (FK ON DELETE CASCADE le ferait,
    # mais on est explicites pour éviter les surprises)
    db.execute("""DELETE FROM causeries_reponses_detail
                  WHERE reponse_id IN (SELECT id FROM causeries_reponses WHERE causerie_id=?)""", (cid,))
    db.execute("DELETE FROM causeries_reponses WHERE causerie_id=?", (cid,))
    db.execute("UPDATE causeries SET mail_envoye=0, mail_envoye_at='' WHERE id=?", (cid,))
    db.commit()
    log_action(request.user, "RESET", "causerie", cid, "")
    return jsonify({"ok": True})

@app.route("/api/causeries/<int:cid>/renvoyer_mail", methods=["POST"])
@require_role("admin")
def renvoyer_causerie_mail(cid):
    """v218.33 : Force le renvoi du mail récap (admin), même si déjà envoyé.
    Utile si le mail s'est perdu ou pour test."""
    db = get_db()
    c = one(db.execute("SELECT id FROM causeries WHERE id=?", (cid,)))
    if not c:
        return jsonify({"error": "Introuvable"}), 404
    # Reset du flag pour forcer l'envoi
    db.execute("UPDATE causeries SET mail_envoye=0 WHERE id=?", (cid,))
    db.commit()
    try:
        _check_and_send_causerie_recap(cid)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══ CR ↔ TABLEAUX (sous-équipements cochés sur un CR) ══
@app.route("/api/comptes_rendus/cr/<int:cr_id>/tableaux", methods=["GET"])
@require_auth
def list_cr_tableaux(cr_id):
    """Retourne les sous-équipements cochés sur un CR."""
    db = get_db()
    return jsonify(rows(db.execute("""
        SELECT ct.tableau_id, et.nom, et.localisation
        FROM cr_tableaux ct JOIN equipement_tableaux et ON ct.tableau_id=et.id
        WHERE ct.cr_id=? ORDER BY et.ordre, et.id
    """, (cr_id,))))

@app.route("/api/comptes_rendus/cr/<int:cr_id>/tableaux", methods=["POST"])
@require_auth
def set_cr_tableaux(cr_id):
    """Remplace les sous-équipements cochés du CR. Body: {tableau_ids: [..]}."""
    d = request.json or {}
    ids = d.get("tableau_ids") or []
    if not isinstance(ids, list):
        return jsonify({"error": "tableau_ids doit être une liste"}), 400
    db = get_db()
    if not one(db.execute("SELECT 1 FROM comptes_rendus WHERE id=?", (cr_id,))):
        return jsonify({"error": "CR introuvable"}), 404
    db.execute("DELETE FROM cr_tableaux WHERE cr_id=?", (cr_id,))
    for tid in ids:
        try:
            db.execute("INSERT OR IGNORE INTO cr_tableaux (cr_id, tableau_id) VALUES (?,?)",
                       (cr_id, int(tid)))
        except Exception:
            continue
    db.commit()
    return jsonify({"ok": True, "count": len(ids)})

@app.route("/api/interventions/<int:iid>/tableaux_status", methods=["GET"])
@require_auth
def get_tableaux_status(iid):
    """Pour chaque sous-équipement de l'iv, indique s'il a déjà été coché par un CR (et lequel)."""
    db = get_db()
    iv = one(db.execute("SELECT equipement_id FROM interventions WHERE id=?", (iid,)))
    if not iv:
        return jsonify({"error":"Intervention introuvable"}), 404
    eq_id = iv["equipement_id"]
    tableaux = rows(db.execute("""
        SELECT id, nom, localisation, ordre
        FROM equipement_tableaux WHERE equipement_id=? ORDER BY ordre, id
    """, (eq_id,)))
    # Pour chaque tableau, chercher le 1er CR (de cette iv) qui l'a coché
    for t in tableaux:
        row = one(db.execute("""
            SELECT cr.id AS cr_id, cr.numero, cr.date_intervention
            FROM cr_tableaux ct
            JOIN comptes_rendus cr ON ct.cr_id=cr.id
            WHERE ct.tableau_id=? AND cr.intervention_id=?
            ORDER BY cr.date_intervention, cr.id LIMIT 1
        """, (t["id"], iid)))
        if row:
            t["done_cr_id"] = row["cr_id"]
            t["done_cr_numero"] = row["numero"]
            t["done_date"] = row["date_intervention"]
        else:
            t["done_cr_id"] = None
    return jsonify(tableaux)


# ══ OPÉRATIONS DE GAMME COCHÉES SUR UNE INTERVENTION ══
@app.route("/api/interventions/<int:iid>/operations", methods=["GET"])
@require_auth
def list_intervention_operations(iid):
    """Toutes les opérations des gammes liées à l'équipement de l'iv + statut (cochée par qui/quand)."""
    db = get_db()
    iv = one(db.execute("SELECT equipement_id FROM interventions WHERE id=?", (iid,)))
    if not iv:
        return jsonify({"error":"Intervention introuvable"}), 404
    eq_id = iv["equipement_id"]
    # Toutes les gammes liées à l'équipement
    gids_rows = rows(db.execute("SELECT gamme_id FROM equipement_gammes WHERE equipement_id=?", (eq_id,)))
    gids = [g["gamme_id"] for g in gids_rows]
    eq_row = one(db.execute("SELECT gamme_id FROM equipements WHERE id=?", (eq_id,)))
    if eq_row and eq_row.get("gamme_id") and eq_row["gamme_id"] not in gids:
        gids.append(eq_row["gamme_id"])
    if not gids:
        return jsonify([])
    placeholders = ",".join(["?"] * len(gids))
    operations = rows(db.execute(f"""
        SELECT go.id, go.description, go.ordre, go.gamme_id, g.nom AS gamme_nom
        FROM gamme_operations go
        JOIN gammes g ON go.gamme_id = g.id
        WHERE go.gamme_id IN ({placeholders})
        ORDER BY g.nom, go.ordre, go.id
    """, gids))
    # Pour chaque op : status sur cette intervention
    for op in operations:
        done = one(db.execute("""
            SELECT io.id, io.date_realisation, io.cr_id, io.technicien_id,
                   u.nom AS technicien_nom, cr.numero AS cr_numero
            FROM intervention_operations io
            LEFT JOIN utilisateurs u ON io.technicien_id = u.id
            LEFT JOIN comptes_rendus cr ON io.cr_id = cr.id
            WHERE io.intervention_id=? AND io.gamme_operation_id=?
        """, (iid, op["id"])))
        if done:
            op["done"] = True
            op["done_date"] = done.get("date_realisation","") or ""
            op["done_technicien_nom"] = done.get("technicien_nom","") or ""
            op["done_cr_numero"] = done.get("cr_numero","") or ""
        else:
            op["done"] = False
    return jsonify(operations)

@app.route("/api/interventions/<int:iid>/operations/<int:op_id>", methods=["POST"])
@require_auth
def check_intervention_operation(iid, op_id):
    """Coche une opération sur une intervention."""
    db = get_db()
    if not one(db.execute("SELECT 1 FROM interventions WHERE id=?", (iid,))):
        return jsonify({"error":"Intervention introuvable"}), 404
    if not one(db.execute("SELECT 1 FROM gamme_operations WHERE id=?", (op_id,))):
        return jsonify({"error":"Opération introuvable"}), 404
    user_id = request.user["id"]
    try:
        db.execute("""INSERT INTO intervention_operations
            (intervention_id, gamme_operation_id, technicien_id, date_realisation)
            VALUES (?,?,?, date('now'))""", (iid, op_id, user_id))
        db.commit()
    except Exception:
        # Déjà coché → on ne fait rien (idempotent)
        pass
    return jsonify({"ok": True})

@app.route("/api/interventions/<int:iid>/operations/<int:op_id>", methods=["DELETE"])
@require_auth
def uncheck_intervention_operation(iid, op_id):
    """Décoche une opération."""
    db = get_db()
    db.execute("DELETE FROM intervention_operations WHERE intervention_id=? AND gamme_operation_id=?",
               (iid, op_id))
    db.commit()
    return jsonify({"ok": True})


# ══ TYPES DE PIÈCES CRITIQUES (v218.169) ══
@app.route("/api/types-pieces", methods=["GET"])
@require_auth
def list_types_pieces():
    """Liste les types de pièces critiques de la société (actifs uniquement par défaut)."""
    db = get_db()
    sid = current_societe_id()
    incl_inactifs = request.args.get("incl_inactifs") == "1"
    if incl_inactifs:
        types = rows(db.execute(
            "SELECT id, nom, is_batterie, ordre, actif FROM types_pieces WHERE societe_id=? ORDER BY ordre, nom",
            (sid,)
        ))
    else:
        types = rows(db.execute(
            "SELECT id, nom, is_batterie, ordre, actif FROM types_pieces WHERE societe_id=? AND actif=1 ORDER BY ordre, nom",
            (sid,)
        ))
    return jsonify(types)

@app.route("/api/types-pieces", methods=["POST"])
@require_role("admin", "manager")
def create_type_piece():
    d = request.json or {}
    nom = (d.get("nom") or "").strip()
    if not nom:
        return jsonify({"error": "Nom requis"}), 400
    db = get_db()
    sid = current_societe_id()
    # Vérifier unicité
    existing = one(db.execute("SELECT id FROM types_pieces WHERE societe_id=? AND LOWER(nom)=LOWER(?)", (sid, nom)))
    if existing:
        return jsonify({"error": "Ce nom existe déjà"}), 400
    is_bat = 1 if d.get("is_batterie") else 0
    # Ordre = max+1
    res = one(db.execute("SELECT COALESCE(MAX(ordre),-1)+1 AS o FROM types_pieces WHERE societe_id=?", (sid,)))
    ordre = res.get("o") if res else 0
    cur = db.execute(
        "INSERT INTO types_pieces (societe_id, nom, is_batterie, ordre) VALUES (?,?,?,?)",
        (sid, nom, is_bat, ordre)
    )
    db.commit()
    return jsonify({"id": cur.lastrowid, "nom": nom, "is_batterie": is_bat, "ordre": ordre, "actif": 1}), 201

@app.route("/api/types-pieces/<int:tid>", methods=["PATCH"])
@require_role("admin", "manager")
def update_type_piece(tid):
    d = request.json or {}
    db = get_db()
    sid = current_societe_id()
    # Vérifier que le type appartient bien à la société
    t = one(db.execute("SELECT * FROM types_pieces WHERE id=? AND societe_id=?", (tid, sid)))
    if not t:
        return jsonify({"error": "Type introuvable"}), 404
    sets, params = [], []
    if "nom" in d:
        nom = (d["nom"] or "").strip()
        if not nom:
            return jsonify({"error": "Nom requis"}), 400
        # Unicité (sauf lui-même)
        ex = one(db.execute("SELECT id FROM types_pieces WHERE societe_id=? AND LOWER(nom)=LOWER(?) AND id!=?",
                            (sid, nom, tid)))
        if ex: return jsonify({"error": "Ce nom existe déjà"}), 400
        sets.append("nom=?"); params.append(nom)
    if "is_batterie" in d:
        sets.append("is_batterie=?"); params.append(1 if d["is_batterie"] else 0)
    if "ordre" in d:
        sets.append("ordre=?"); params.append(int(d["ordre"] or 0))
    if "actif" in d:
        sets.append("actif=?"); params.append(1 if d["actif"] else 0)
    if not sets:
        return jsonify({"ok": True})
    params.append(tid)
    db.execute(f"UPDATE types_pieces SET {','.join(sets)} WHERE id=?", params)
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/types-pieces/<int:tid>", methods=["DELETE"])
@require_role("admin", "manager")
def delete_type_piece(tid):
    """Suppression : refus si des pièces utilisent ce type.
    Sinon DELETE physique."""
    db = get_db()
    sid = current_societe_id()
    t = one(db.execute("SELECT nom FROM types_pieces WHERE id=? AND societe_id=?", (tid, sid)))
    if not t:
        return jsonify({"error": "Type introuvable"}), 404
    # Vérifier l'usage
    res = one(db.execute("SELECT COUNT(*) AS n FROM pieces WHERE type_piece=?", (t["nom"],)))
    n_used = (res and res.get("n", 0)) or 0
    if n_used > 0:
        return jsonify({"error": f"Ce type est utilisé par {n_used} pièce(s). Désactivez-le plutôt que de le supprimer."}), 400
    db.execute("DELETE FROM types_pieces WHERE id=?", (tid,))
    db.commit()
    return jsonify({"ok": True})


# ══ PIECES ══
@app.route("/api/pieces")
@require_auth
def get_pieces():
    db=get_db()
    sql="""SELECT p.*,e.designation AS equip_nom,pr.nom AS projet_nom
           FROM pieces p JOIN equipements e ON p.equipement_id=e.id
           JOIN projets pr ON e.projet_id=pr.id WHERE 1=1"""
    params=[]
    if request.args.get("equipement_id"): sql+=" AND p.equipement_id=?"; params.append(request.args["equipement_id"])
    if request.args.get("alertes"): sql+=" AND p.statut IN ('A_SURVEILLER','A_REMPLACER')"
    pieces=rows(db.execute(sql+" ORDER BY p.date_fin_de_vie",params))
    for p in pieces:
        ns=statut_piece(p.get("date_fin_de_vie"))
        if ns!=p.get("statut") and p.get("date_fin_de_vie"):
            db.execute("UPDATE pieces SET statut=? WHERE id=?",(ns,p["id"])); p["statut"]=ns
    db.commit(); return jsonify(pieces)

@app.route("/api/pieces",methods=["POST"])
@require_auth
def create_piece():
    d=request.json or {}
    if not all([d.get("equipement_id"),d.get("type_piece")]): return jsonify({"error":"equipement_id et type_piece requis"}),400
    db=get_db(); dv=d.get("date_fin_de_vie") or None
    if not dv and d.get("date_installation") and d.get("duree_vie_estimee"):
        try:
            di=datetime.strptime(d["date_installation"],"%Y-%m-%d")
            # v217.10 : ajouter N années (pas N×365 jours, pour gérer les bissextiles correctement)
            dv = _add_years(di, int(d["duree_vie_estimee"])).strftime("%Y-%m-%d")
        except Exception: pass
    st=statut_piece(dv)
    db.execute("INSERT INTO pieces (equipement_id,type_piece,date_installation,duree_vie_estimee,date_fin_de_vie,statut,quantite,numero_serie,reference,commentaire,nbr_chaine) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
               (d["equipement_id"],d["type_piece"],d.get("date_installation") or None,d.get("duree_vie_estimee") or None,dv,st,d.get("quantite",1),d.get("numero_serie",""),d.get("reference",""),d.get("commentaire",""),d.get("nbr_chaine") or None))
    db.commit(); return jsonify({"id":db.execute("SELECT last_insert_rowid()").fetchone()[0]}),201

@app.route("/api/pieces/<int:pid>",methods=["PATCH"])
@require_auth
def update_piece(pid):
    d=request.json or {}; db=get_db(); sets,params=[],[]
    for f in ["type_piece","date_installation","duree_vie_estimee","date_fin_de_vie","statut","quantite","numero_serie","reference","commentaire","nbr_chaine"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    if not sets: return jsonify({"error":"Rien"}),400
    params.append(pid)
    db.execute(f"UPDATE pieces SET {chr(44).join(sets)} WHERE id=?",params)
    # Recalculer date_fin_de_vie et statut si date_installation ou duree_vie changent
    if "date_installation" in d or "duree_vie_estimee" in d or "date_fin_de_vie" in d:
        piece = one(db.execute("SELECT date_installation,duree_vie_estimee,date_fin_de_vie FROM pieces WHERE id=?",(pid,)))
        if piece:
            fdv = piece["date_fin_de_vie"]
            # Si date_installation + duree_vie fournis, recalculer fdv
            if piece["date_installation"] and piece["duree_vie_estimee"] and not d.get("date_fin_de_vie"):
                try:
                    from datetime import datetime, timedelta
                    di = datetime.strptime(piece["date_installation"],"%Y-%m-%d")
                    # v217.10 : N années au lieu de N×365 jours
                    fdv = _add_years(di, int(piece["duree_vie_estimee"])).strftime("%Y-%m-%d")
                    db.execute("UPDATE pieces SET date_fin_de_vie=? WHERE id=?",(fdv,pid))
                except Exception: pass
            # Recalculer statut depuis fdv
            new_statut = statut_piece(fdv)
            # Ne recalculer que si statut pas explicitement fourni
            if "statut" not in d:
                db.execute("UPDATE pieces SET statut=? WHERE id=?",(new_statut,pid))
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/pieces/<int:pid>",methods=["DELETE"])
@require_auth
def delete_piece(pid):
    db=get_db(); db.execute("DELETE FROM pieces WHERE id=?",(pid,)); db.commit()
    return jsonify({"ok":True})

# v218.27 : GET pièce seule (utile après remplacement pour refresh)
@app.route("/api/pieces/<int:pid>", methods=["GET"])
@require_auth
def get_piece_single(pid):
    db = get_db()
    p = one(db.execute("""
        SELECT p.*, e.designation AS equip_nom, pr.id AS projet_id, pr.nom AS projet_nom
        FROM pieces p
        JOIN equipements e ON p.equipement_id=e.id
        JOIN projets pr ON e.projet_id=pr.id
        WHERE p.id=?
    """, (pid,)))
    if not p:
        return jsonify({"error": "Pièce introuvable"}), 404
    return jsonify(p)

# v218.27 : Remplacement de pièce + historique
@app.route("/api/pieces/<int:pid>/historique", methods=["GET"])
@require_auth
def get_piece_historique(pid):
    """Retourne l'historique des remplacements de la pièce, triés du plus récent au plus ancien."""
    db = get_db()
    hist = rows(db.execute("""
        SELECT h.*, u.nom AS remplace_par_nom
        FROM pieces_historique h
        LEFT JOIN utilisateurs u ON h.remplace_par = u.id
        WHERE h.piece_id = ?
        ORDER BY h.date_remplacement DESC, h.id DESC
    """, (pid,)))
    return jsonify(hist)

@app.route("/api/pieces/<int:pid>/remplacer", methods=["POST"])
@require_auth
def remplacer_piece(pid):
    """Remplace une pièce :
    1. Copie l'état actuel dans pieces_historique
    2. Met à jour la pièce avec les nouvelles infos
    Body :
      - reference, numero_serie, date_installation, duree_vie_estimee, date_fin_de_vie,
        statut, quantite, commentaire (nouvelles valeurs)
      - motif_categorie ('PANNE','FIN_DE_VIE','PREVENTIF','AUTRE') REQUIS
      - motif_detail (texte libre)
    Le type_piece n'est pas modifiable (Q1 = option 2).
    """
    d = request.json or {}
    motif_cat = (d.get("motif_categorie") or "").strip().upper()
    if motif_cat not in ("PANNE", "FIN_DE_VIE", "PREVENTIF", "AUTRE"):
        return jsonify({"error": "Motif de remplacement requis (PANNE, FIN_DE_VIE, PREVENTIF, AUTRE)"}), 400
    db = get_db()
    # Récupérer l'état actuel
    p_curr = one(db.execute("SELECT * FROM pieces WHERE id=?", (pid,)))
    if not p_curr:
        return jsonify({"error": "Pièce introuvable"}), 404
    try:
        date_remplacement = (d.get("date_remplacement") or datetime.now().strftime("%Y-%m-%d"))[:10]
        # 1) Copier l'état actuel dans historique
        db.execute("""
            INSERT INTO pieces_historique
              (piece_id, type_piece, date_installation, duree_vie_estimee, date_fin_de_vie,
               statut, quantite, numero_serie, reference, commentaire,
               date_remplacement, motif_categorie, motif_detail, remplace_par)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            pid,
            p_curr.get("type_piece") or "",
            p_curr.get("date_installation") or "",
            p_curr.get("duree_vie_estimee"),
            p_curr.get("date_fin_de_vie") or "",
            p_curr.get("statut") or "OK",
            p_curr.get("quantite") or 1,
            p_curr.get("numero_serie") or "",
            p_curr.get("reference") or "",
            p_curr.get("commentaire") or "",
            date_remplacement,
            motif_cat,
            (d.get("motif_detail") or "").strip(),
            request.user.get("id"),
        ))
        # 2) Mettre à jour la pièce avec les nouvelles infos (type_piece intouchable)
        sets, params = [], []
        for col in ("date_installation", "duree_vie_estimee", "date_fin_de_vie",
                    "statut", "quantite", "numero_serie", "reference", "commentaire"):
            if col in d:
                sets.append(f"{col}=?")
                params.append(d[col])
        if sets:
            params.append(pid)
            db.execute(f"UPDATE pieces SET {','.join(sets)} WHERE id=?", params)
        db.commit()
        log_action(request.user, "REPLACE", "piece", pid,
                   f"motif={motif_cat}, ref_old='{p_curr.get('reference','')[:30]}'")
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        logger.error(f"[PIECE_REMPLACER] échec pid={pid} : {e}")
        return jsonify({"error": str(e)}), 500

# ══ EQUIPES ══
@app.route("/api/equipes")
@require_auth
def get_equipes():
    db=get_db()
    equipes=rows(db.execute("SELECT e.*,u.nom AS manager_nom FROM equipes e LEFT JOIN utilisateurs u ON e.manager_id=u.id ORDER BY e.nom"))
    for eq in equipes:
        eq["membres"]=rows(db.execute("""SELECT u.id,u.nom,u.email FROM utilisateurs u
            JOIN equipe_membres em ON em.technicien_id=u.id WHERE em.equipe_id=?""",(eq["id"],)))
    return jsonify(equipes)

@app.route("/api/equipes",methods=["POST"])
@require_role("admin","manager")
def create_equipe():
    d=request.json or {}
    if not d.get("nom"): return jsonify({"error":"nom requis"}),400
    db=get_db(); db.execute("INSERT INTO equipes (nom,manager_id) VALUES (?,?)",(d["nom"],to_int(d.get("manager_id"))))
    db.commit(); eid=db.execute("SELECT last_insert_rowid()").fetchone()[0]
    for mid in d.get("membres",[]):
        try: db.execute("INSERT OR IGNORE INTO equipe_membres VALUES (?,?)",(eid,int(mid)))
        except Exception: pass
    db.commit(); return jsonify({"id":eid}),201

@app.route("/api/equipes/<int:eid>",methods=["PATCH"])
@require_role("admin","manager")
def update_equipe(eid):
    d=request.json or {}; db=get_db()
    if "nom" in d or "manager_id" in d:
        sets,params=[],[]
        if "nom" in d: sets.append("nom=?"); params.append(d["nom"])
        if "manager_id" in d: sets.append("manager_id=?"); params.append(to_int(d["manager_id"]))
        params.append(eid); db.execute(f"UPDATE equipes SET {chr(44).join(sets)} WHERE id=?",params)
    if "membres" in d:
        db.execute("DELETE FROM equipe_membres WHERE equipe_id=?",(eid,))
        for mid in d["membres"]:
            try: db.execute("INSERT OR IGNORE INTO equipe_membres VALUES (?,?)",(eid,int(mid)))
            except Exception: pass
    db.commit(); return jsonify({"ok":True})

@app.route("/api/equipes/<int:eid>",methods=["DELETE"])
@require_role("admin","manager")
def delete_equipe(eid):
    db=get_db(); db.execute("DELETE FROM equipe_membres WHERE equipe_id=?",(eid,))
    db.execute("DELETE FROM equipes WHERE id=?",(eid,)); db.commit()
    return jsonify({"ok":True})

# ══ INTERVENTIONS ══
@app.route("/api/interventions")
@require_auth
def get_interventions():
    db=get_db(); u=request.user
    sid = current_societe_id()
    sql="""SELECT i.*,e.designation AS equip_nom,e.type_technique,
                  p.nom AS projet_nom,p.numero_projet,c.societe AS client_nom,
                  u.nom AS technicien_nom,eq.nom AS equipe_nom,
                  et.nom AS tableau_nom, et.localisation AS tableau_localisation,
                  g.nom AS gamme_nom, g.temps AS gamme_temps,
                  (SELECT GROUP_CONCAT(it.utilisateur_id)
                   FROM intervention_techniciens it WHERE it.intervention_id=i.id) AS _tech_ids_csv
           FROM interventions i JOIN equipements e ON i.equipement_id=e.id
           JOIN projets p ON e.projet_id=p.id LEFT JOIN clients c ON p.client_id=c.id
           LEFT JOIN utilisateurs u ON i.technicien_id=u.id
           LEFT JOIN equipes eq ON i.equipe_id=eq.id
           LEFT JOIN equipement_tableaux et ON i.tableau_id=et.id
           LEFT JOIN gammes g ON i.gamme_id=g.id
           WHERE i.societe_id=?"""
    params=[sid]
    if u["role"]=="technicien":
        # v218.22 : les bons ASTREINTE en statut PLANIFIEE sont visibles par TOUS les techniciens
        # (premier qui démarre se l'attribue). Une fois EN_COURS/TERMINEE, seul le tech assigné le voit.
        sql+=(" AND ((i.type='ASTREINTE' AND i.statut='PLANIFIEE') "
              "OR i.technicien_id=? "
              "OR i.equipe_id IN (SELECT equipe_id FROM equipe_membres WHERE technicien_id=?) "
              "OR EXISTS (SELECT 1 FROM intervention_techniciens it WHERE it.intervention_id=i.id AND it.utilisateur_id=?) "
              "OR EXISTS (SELECT 1 FROM cr_intervenants ci JOIN comptes_rendus cr ON ci.cr_id=cr.id "
              "           WHERE cr.intervention_id=i.id AND ci.utilisateur_id=?))")
        params+=[u["id"],u["id"],u["id"],u["id"]]
    elif u["role"]=="manager":
        # Manager : bons dont le projet lui appartient, OU dont un tech de son équipe est assigné, OU lui-même,
        # OU bons liés à une technique qu'il gère (manager technique)
        # v218.25 : ajout — les bons ASTREINTE en PLANIFIEE sont visibles par tous les managers
        # (puisqu'un manager peut être de garde et recevoir un BA)
        # Récupérer ses techniques
        udb = one(get_db().execute("SELECT techniques FROM utilisateurs WHERE id=?", (u["id"],)))
        tech_csv = (udb and udb.get("techniques")) or ""
        tech_list = [t.strip() for t in tech_csv.split(",") if t.strip()]
        conditions = [
            "(i.type='ASTREINTE' AND i.statut='PLANIFIEE')",  # v218.25 : BA en attente visibles par tous les managers
            "p.manager_id=?",
            "i.technicien_id=?",
            "i.technicien_id IN (SELECT id FROM utilisateurs WHERE manager_id=?)",
            "EXISTS (SELECT 1 FROM intervention_techniciens it WHERE it.intervention_id=i.id AND "
            "  (it.utilisateur_id=? OR it.utilisateur_id IN (SELECT id FROM utilisateurs WHERE manager_id=?)))",
            "EXISTS (SELECT 1 FROM cr_intervenants ci JOIN comptes_rendus cr ON ci.cr_id=cr.id "
            "  WHERE cr.intervention_id=i.id AND (ci.utilisateur_id=? OR ci.utilisateur_id IN (SELECT id FROM utilisateurs WHERE manager_id=?)))"
        ]
        params_mgr = [u["id"], u["id"], u["id"], u["id"], u["id"], u["id"], u["id"]]
        if tech_list:
            placeholders = ",".join(["?"] * len(tech_list))
            conditions.append(f"e.type_technique IN ({placeholders})")
            params_mgr += tech_list
        sql += " AND (" + " OR ".join(conditions) + ")"
        params += params_mgr
        logger.info(f"[interventions manager u={u['id']}] techniques={tech_list} subordinates={rows(get_db().execute('SELECT id,nom FROM utilisateurs WHERE manager_id=?',(u['id'],)))}")
    # admin : pas de filtre, voit tout
    for arg,col in [("type","i.type"),("statut","i.statut"),("equipement_id","i.equipement_id"),("projet_id","e.projet_id")]:
        if request.args.get(arg): sql+=f" AND {col}=?"; params.append(request.args[arg])
    result = rows(db.execute(sql+" ORDER BY CASE WHEN i.date_prevue IS NULL THEN 1 ELSE 0 END, i.date_prevue ASC, i.created_at DESC",params))
    # Convertir la CSV en liste d'int pour chaque intervention
    for r in result:
        csv_ids = r.pop("_tech_ids_csv", None)
        if csv_ids:
            r["technicien_ids"] = [int(x) for x in csv_ids.split(",") if x]
        else:
            r["technicien_ids"] = [r["technicien_id"]] if r.get("technicien_id") else []
        # Pour les bons de dépannage ET maintenance avec CR : ajouter la liste des CRs
        # (utilisé pour le planning : dépannage = toujours CRs ; maintenance = CRs s'il y en a)
        if r.get("type") in ("DEPANNAGE", "MAINTENANCE"):
            try:
                crs_light = rows(db.execute(
                    "SELECT id, numero, date_intervention FROM comptes_rendus WHERE intervention_id=? ORDER BY date_intervention, id",
                    (r["id"],)))
                # Pour chaque CR, récupérer l'intervalle horaire global (min heure_debut → max heure_fin)
                # + la liste des intervenants (pour affichage planning multi-tech)
                for cr in crs_light:
                    try:
                        horaires = one(db.execute(
                            """SELECT MIN(NULLIF(heure_debut,'')) AS h_min, MAX(NULLIF(heure_fin,'')) AS h_max
                               FROM cr_intervenants WHERE cr_id=?""", (cr["id"],)))
                        cr["heure_debut"] = (horaires or {}).get("h_min") or ""
                        cr["heure_fin"] = (horaires or {}).get("h_max") or ""
                    except Exception:
                        cr["heure_debut"] = ""
                        cr["heure_fin"] = ""
                    # Liste des techniciens présents sur ce CR (ids uniquement, pour filtrage planning)
                    try:
                        tech_rows = rows(db.execute(
                            "SELECT DISTINCT utilisateur_id FROM cr_intervenants WHERE cr_id=? AND utilisateur_id IS NOT NULL",
                            (cr["id"],)))
                        cr["technicien_ids"] = [t["utilisateur_id"] for t in tech_rows if t.get("utilisateur_id")]
                    except Exception:
                        cr["technicien_ids"] = []
                    # Détail des intervenants : date + horaires individuels (pour éclater les tuiles par date si différentes)
                    try:
                        cr["intervenants_detail"] = rows(db.execute(
                            """SELECT utilisateur_id, nom,
                                      NULLIF(date,'') AS date,
                                      NULLIF(heure_debut,'') AS heure_debut,
                                      NULLIF(heure_fin,'') AS heure_fin
                               FROM cr_intervenants WHERE cr_id=?""",
                            (cr["id"],)))
                    except Exception:
                        cr["intervenants_detail"] = []
                r["comptes_rendus_light"] = crs_light
            except Exception:
                r["comptes_rendus_light"] = []
    return jsonify(result)

@app.route("/api/interventions/<int:iid>")
@require_auth
def get_intervention(iid):
    db=get_db()
    i=one(db.execute("""SELECT i.*,e.designation AS equip_nom,e.type_technique,
           e.trafo_marque, e.trafo_annee, e.trafo_numero_serie, e.trafo_puissance_kva,
           e.trafo_refroidissement, e.trafo_poids_kg, e.trafo_tension_entree_v,
           e.trafo_courant_a, e.trafo_norme, e.trafo_couplage,
           e.trafo_tension_service_v, e.trafo_reglage_tension_kv,
           p.nom AS projet_nom,c.societe AS client_nom,u.nom AS technicien_nom
           FROM interventions i JOIN equipements e ON i.equipement_id=e.id
           JOIN projets p ON e.projet_id=p.id LEFT JOIN clients c ON p.client_id=c.id
           LEFT JOIN utilisateurs u ON i.technicien_id=u.id WHERE i.id=?""",(iid,)))
    if not i: return jsonify({"error":"Non trouve"}),404
    crs=rows(db.execute("SELECT * FROM comptes_rendus WHERE intervention_id=? ORDER BY created_at",(iid,)))
    for cr in crs:
        cr["intervenants"]=rows(db.execute("""SELECT ci.*,COALESCE(u.nom,ci.nom,'') AS nom
            FROM cr_intervenants ci LEFT JOIN utilisateurs u ON ci.utilisateur_id=u.id
            WHERE ci.cr_id=?""",(cr["id"],)))
        try:
            cr["materiels"]=rows(db.execute("SELECT * FROM cr_materiels WHERE cr_id=? ORDER BY ordre,id",(cr["id"],)))
        except: cr["materiels"]=[]
    i["comptes_rendus"]=crs
    # Créneaux (pour interventions multi-jours)
    i["creneaux"]=rows(db.execute("""SELECT c.*, u.nom AS technicien_nom
        FROM intervention_creneaux c LEFT JOIN utilisateurs u ON c.technicien_id=u.id
        WHERE c.intervention_id=? ORDER BY c.date, c.heure_debut""",(iid,)))
    # Techniciens affectés (multi)
    techs = rows(db.execute("""SELECT it.utilisateur_id AS id, u.nom
        FROM intervention_techniciens it JOIN utilisateurs u ON it.utilisateur_id=u.id
        WHERE it.intervention_id=? ORDER BY it.id""",(iid,)))
    # Fallback : si la table liaison est vide mais qu'il y a un technicien_id historique, on l'expose
    if not techs and i.get("technicien_id"):
        u = one(db.execute("SELECT id,nom FROM utilisateurs WHERE id=?",(i["technicien_id"],)))
        if u: techs = [u]
    i["techniciens"] = techs
    i["technicien_ids"] = [t["id"] for t in techs]
    # Tableau électrique (sous-équipement BT) si renseigné
    try:
        if i.get("tableau_id"):
            t = one(db.execute("SELECT id,nom,localisation FROM equipement_tableaux WHERE id=?", (i["tableau_id"],)))
            if t:
                i["tableau_nom"] = t["nom"]
                i["tableau_localisation"] = t["localisation"] or ""
    except Exception:
        pass
    return jsonify(i)

# v218.126 : Helper pour trouver la prochaine date disponible
# selon les règles : pas de weekend, pas de férié, capacité du planning non atteinte.
# v218.137 : ajout de skip_weekend (défaut True, désactivable si l'utilisateur a choisi samedi/dimanche).
def _parse_duree_to_min(s):
    """Convertit '01h30' / '1:30' / '90' en minutes (int). Retourne 60 par défaut."""
    if not s: return 60
    s = str(s).strip().lower().replace(' ', '')
    try:
        # Format 'XhY' (ex. '1h30')
        if 'h' in s:
            parts = s.split('h')
            h = int(parts[0] or 0)
            m = int(parts[1] or 0) if len(parts) > 1 and parts[1] else 0
            return h * 60 + m
        # Format 'HH:MM'
        if ':' in s:
            parts = s.split(':')
            return int(parts[0]) * 60 + int(parts[1])
        # Nombre seul (minutes)
        return int(s)
    except Exception:
        return 60

def _hhmm_to_min(s):
    """'08:30' -> 510. Retourne 480 (08:00) par défaut."""
    if not s: return 480
    try:
        parts = str(s).split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except Exception:
        return 480


def _find_next_available_date(db, start_date_str, type_iv, sid, planning_id=None,
                              ignore_intervention_id=None, max_lookahead_days=365,
                              skip_weekend=True, force_planif=False,
                              heure_prevue=None, duree_min=None):
    """Trouve la première date >= start_date_str qui :
       - n'est ni samedi, ni dimanche (si skip_weekend=True)
       - n'est pas un jour férié (table jours_feries pour la société)
       - n'a pas plus de nb_simultane bons qui CHEVAUCHENT le créneau [heure_prevue, +duree_min]
         (v218.162 : la capacité ne compte plus les bons du jour mais les bons qui se chevauchent)
    Ne s'applique QU'aux interventions de type MAINTENANCE.
    Si force_planif=True : court-circuit complet (date retournée telle quelle).
    Retourne (date_str, was_shifted: bool).
    """
    if type_iv != "MAINTENANCE":
        return (start_date_str, False)
    if force_planif:
        return (start_date_str, False)
    try:
        cur = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    except Exception:
        return (start_date_str, False)
    original = cur
    end_window = cur + timedelta(days=max_lookahead_days)
    feries_rows = rows(db.execute(
        "SELECT date FROM jours_feries WHERE societe_id=? AND date BETWEEN ? AND ?",
        (sid, cur.strftime("%Y-%m-%d"), end_window.strftime("%Y-%m-%d"))
    ))
    feries_set = set(f["date"] for f in feries_rows)
    nb_simultane = 1
    if planning_id:
        plan = one(db.execute("SELECT nb_simultane FROM plannings WHERE id=? AND societe_id=?",
                              (planning_id, sid)))
        if plan and plan.get("nb_simultane"):
            try: nb_simultane = max(1, int(plan["nb_simultane"]))
            except: nb_simultane = 1
    # v218.162 : calcul de l'intervalle du nouveau bon en minutes (depuis minuit)
    new_start_min = _hhmm_to_min(heure_prevue) if heure_prevue else 480  # 08:00 défaut
    new_dur = duree_min if (duree_min and duree_min > 0) else 60  # 1h défaut
    new_end_min = new_start_min + new_dur
    iters = 0
    while iters < max_lookahead_days:
        iters += 1
        date_str = cur.strftime("%Y-%m-%d")
        if skip_weekend and cur.weekday() >= 5:
            cur += timedelta(days=1); continue
        if date_str in feries_set:
            cur += timedelta(days=1); continue
        if planning_id and nb_simultane >= 1:
            # v218.162 : charger tous les bons MAINTENANCE de ce jour sur le planning,
            # avec leur heure_prevue + durée de leur gamme. Compter les chevauchements.
            sql = """SELECT i.id, i.heure_prevue, g.temps AS gtemps
                     FROM interventions i
                     JOIN equipements e ON i.equipement_id = e.id
                     LEFT JOIN gammes g ON i.gamme_id = g.id
                     WHERE i.type='MAINTENANCE'
                       AND i.date_prevue = ?
                       AND i.societe_id = ?
                       AND e.planning_id = ?"""
            params = [date_str, sid, planning_id]
            if ignore_intervention_id:
                sql += " AND i.id != ?"
                params.append(ignore_intervention_id)
            day_bons = rows(db.execute(sql, params))
            # v218.163 : calculer le PIC de simultanéité dans l'intervalle [new_start, new_end].
            events = []
            events.append((new_start_min, +1))
            events.append((new_end_min, -1))
            considered = []  # debug
            for b in day_bons:
                bs_min = _hhmm_to_min(b.get("heure_prevue") or "08:00")
                bd = _parse_duree_to_min(b.get("gtemps") or "1h00")
                be_min = bs_min + bd
                if be_min <= new_start_min or bs_min >= new_end_min:
                    continue  # hors fenêtre
                events.append((bs_min, +1))
                events.append((be_min, -1))
                considered.append((b.get("id"), b.get("heure_prevue"), b.get("gtemps"), bs_min, be_min))
            events.sort(key=lambda e: (e[0], e[1]))
            peak = 0
            current = 0
            for t, delta in events:
                current += delta
                if current > peak: peak = current
            if peak > nb_simultane:
                logger.info(f"[next-avail-date] {date_str} REFUSÉ : peak={peak} > nb_simultane={nb_simultane}, "
                            f"nouveau bon=[{new_start_min}-{new_end_min}min ({heure_prevue}+{duree_min}min)], "
                            f"bons existants chevauchant: {considered}")
                cur += timedelta(days=1); continue
            else:
                logger.info(f"[next-avail-date] {date_str} OK : peak={peak} <= nb_simultane={nb_simultane}, "
                            f"nouveau=[{new_start_min}-{new_end_min}], bons={considered}")
        return (date_str, (cur != original))
    return (None, False)


@app.route("/api/interventions/find-next-slot", methods=["POST"])
@require_auth
def find_next_slot():
    """Calcule la prochaine date dispo pour un bon de maintenance, selon planning + fériés + weekends.
    Body : { date_souhaitee, equipement_id (pour récupérer le planning), type?, gamme_id? }
    Retourne : { date, was_shifted, original_date }"""
    d = request.json or {}
    date_souhaitee = (d.get("date_souhaitee") or "").strip()
    if not date_souhaitee:
        return jsonify({"error":"date_souhaitee requise"}), 400
    type_iv = (d.get("type") or "MAINTENANCE").upper()
    eq_id = to_int(d.get("equipement_id"))
    gamme_id = to_int(d.get("gamme_id"))
    # v218.162 : heure et durée pour calcul des chevauchements
    heure_prevue = d.get("heure_prevue") or "08:00"
    db = get_db()
    sid = current_societe_id()
    planning_id = None
    if eq_id:
        eq = one(db.execute("SELECT planning_id FROM equipements WHERE id=?", (eq_id,)))
        if eq and eq.get("planning_id"):
            planning_id = eq["planning_id"]
    # Récupérer la durée de la gamme (en minutes)
    duree_min = 60
    if gamme_id:
        g = one(db.execute("SELECT temps FROM gammes WHERE id=?", (gamme_id,)))
        if g and g.get("temps"):
            duree_min = _parse_duree_to_min(g["temps"])
    force_planif = False
    if eq_id and gamme_id:
        eg = one(db.execute("""SELECT force_planif FROM equipement_gammes
                               WHERE equipement_id=? AND gamme_id=?""", (eq_id, gamme_id)))
        if eg and eg.get("force_planif"):
            force_planif = True
    skip_weekend = True
    try:
        dt0 = datetime.strptime(date_souhaitee, "%Y-%m-%d").date()
        if dt0.weekday() >= 5:
            skip_weekend = False
    except Exception: pass
    new_date, was_shifted = _find_next_available_date(
        db, date_souhaitee, type_iv, sid, planning_id,
        skip_weekend=skip_weekend, force_planif=force_planif,
        heure_prevue=heure_prevue, duree_min=duree_min
    )
    if not new_date:
        return jsonify({"error":"Aucune date disponible dans les 365 prochains jours"}), 404
    return jsonify({"date": new_date, "was_shifted": was_shifted, "original_date": date_souhaitee})


# v218.161 : Version batch pour pré-calculer plusieurs dates en un seul appel
# Utilisé par "Planifier projet" pour afficher les décalages avant création.
@app.route("/api/interventions/find-next-slots-batch", methods=["POST"])
@require_auth
def find_next_slots_batch():
    """Body : { items: [{key, date_souhaitee, equipement_id, gamme_id?}, ...] }
    Retourne : { results: [{key, date, was_shifted, original_date}, ...] }
    'key' est juste un identifiant arbitraire pour le mapping côté frontend."""
    d = request.json or {}
    items = d.get("items") or []
    if not isinstance(items, list):
        return jsonify({"error": "items doit être une liste"}), 400
    db = get_db()
    sid = current_societe_id()
    results = []
    for it in items:
        key = it.get("key")
        date_souhaitee = (it.get("date_souhaitee") or "").strip()
        if not date_souhaitee:
            results.append({"key": key, "date": None, "was_shifted": False, "original_date": None})
            continue
        eq_id = to_int(it.get("equipement_id"))
        gamme_id = to_int(it.get("gamme_id"))
        # v218.162 : heure et durée pour chevauchements
        heure_prevue = it.get("heure_prevue") or "08:00"
        planning_id = None
        if eq_id:
            eq = one(db.execute("SELECT planning_id FROM equipements WHERE id=?", (eq_id,)))
            if eq and eq.get("planning_id"):
                planning_id = eq["planning_id"]
        duree_min = 60
        if gamme_id:
            g = one(db.execute("SELECT temps FROM gammes WHERE id=?", (gamme_id,)))
            if g and g.get("temps"):
                duree_min = _parse_duree_to_min(g["temps"])
        force_planif = False
        if eq_id and gamme_id:
            eg = one(db.execute("""SELECT force_planif FROM equipement_gammes
                                   WHERE equipement_id=? AND gamme_id=?""", (eq_id, gamme_id)))
            if eg and eg.get("force_planif"):
                force_planif = True
        skip_weekend = True
        try:
            dt0 = datetime.strptime(date_souhaitee, "%Y-%m-%d").date()
            if dt0.weekday() >= 5:
                skip_weekend = False
        except Exception: pass
        new_date, was_shifted = _find_next_available_date(
            db, date_souhaitee, "MAINTENANCE", sid, planning_id,
            skip_weekend=skip_weekend, force_planif=force_planif,
            heure_prevue=heure_prevue, duree_min=duree_min
        )
        results.append({
            "key": key,
            "date": new_date,
            "was_shifted": was_shifted,
            "original_date": date_souhaitee
        })
    return jsonify({"results": results})


@app.route("/api/interventions",methods=["POST"])
@require_auth
def create_intervention():
    d=request.json or {}
    if not d.get("equipement_id"): return jsonify({"error":"equipement_id requis"}),400
    db=get_db()
    type_iv = d.get("type","MAINTENANCE")
    # v218.18 : ajout du type ASTREINTE (déclenchement urgence)
    if type_iv not in ("MAINTENANCE","DEPANNAGE","ASTREINTE"):
        return jsonify({"error":"Type invalide"}),400
    try:
        # Préfixe numéro : BP (maintenance), BC (dépannage), BA (astreinte)
        prefix_map = {"MAINTENANCE": "BP", "DEPANNAGE": "BC", "ASTREINTE": "BA"}
        prefix = prefix_map.get(type_iv, "BT")
        num = next_numero(db, prefix, "interventions", "numero")
        # Normaliser la liste des techniciens
        # Accepte soit technicien_ids:[1,2,3] soit technicien_id:1 (legacy)
        tech_ids = d.get("technicien_ids") or []
        if not tech_ids and d.get("technicien_id"):
            tech_ids = [d.get("technicien_id")]
        tech_ids = [to_int(t) for t in tech_ids if t]
        # Dédoublonner en préservant l'ordre
        seen = set(); tech_ids_clean = []
        for t in tech_ids:
            if t and t not in seen: seen.add(t); tech_ids_clean.append(t)
        tech_ids = tech_ids_clean
        # Le "technicien principal" (colonne legacy) = premier de la liste
        main_tech = tech_ids[0] if tech_ids else None
        # Statut par défaut selon le type : MAINTENANCE = A_PLANIFIER, DEPANNAGE = PLANIFIEE
        # v218.18 : ASTREINTE = EN_COURS direct (urgence en cours, pas planifié)
        if type_iv == "MAINTENANCE":
            default_statut = "A_PLANIFIER"
        elif type_iv == "ASTREINTE":
            default_statut = "EN_COURS"
        else:
            default_statut = "PLANIFIEE"
        # v217.12 : inclure heure_prevue (était omise → toujours stockée à '08:00' par défaut)
        heure_prevue_v = (d.get("heure_prevue") or "").strip() or "08:00"
        # v218.116 : gamme_id optionnel — la gamme spécifique à laquelle ce bon est lié.
        # Permet au planning de calculer la bonne durée (sans heuristique).
        gamme_id_v = to_int(d.get("gamme_id")) if d.get("gamme_id") else None
        # v218.72 : injection automatique de societe_id
        sid = current_societe_id()
        # v218.126 : Décalage auto si jour férié / weekend / capacité atteinte.
        # Désactivable via flag force_date=true.
        date_prevue_final = d.get("date_prevue") or None
        was_shifted = False
        original_date = date_prevue_final
        force_date = bool(d.get("force_date"))
        if type_iv == "MAINTENANCE" and date_prevue_final and not force_date:
            # Récupérer le planning de l'équipement
            eq_row = one(db.execute("SELECT planning_id FROM equipements WHERE id=?", (d["equipement_id"],)))
            eq_planning_id = (eq_row and eq_row.get("planning_id")) or None
            # v218.137 : récupérer force_planif au niveau de la gamme (si fournie)
            eg_force_planif = False
            if gamme_id_v:
                eg = one(db.execute("""SELECT force_planif FROM equipement_gammes
                                       WHERE equipement_id=? AND gamme_id=?""", (d["equipement_id"], gamme_id_v)))
                if eg and eg.get("force_planif"):
                    eg_force_planif = True
            # v218.137 : si la date prévue tombe sur samedi/dimanche, ne pas décaler pour le weekend
            eg_skip_weekend = True
            try:
                dt0 = datetime.strptime(date_prevue_final, "%Y-%m-%d").date()
                if dt0.weekday() >= 5:
                    eg_skip_weekend = False
            except Exception: pass
            # v218.162 : récupérer la durée de la gamme pour calcul de chevauchement
            duree_gamme_min = 60
            if gamme_id_v:
                g = one(db.execute("SELECT temps FROM gammes WHERE id=?", (gamme_id_v,)))
                if g and g.get("temps"):
                    duree_gamme_min = _parse_duree_to_min(g["temps"])
            new_date, shifted = _find_next_available_date(
                db, date_prevue_final, type_iv, sid, eq_planning_id,
                skip_weekend=eg_skip_weekend, force_planif=eg_force_planif,
                heure_prevue=heure_prevue_v, duree_min=duree_gamme_min
            )
            if new_date and shifted:
                date_prevue_final = new_date
                was_shifted = True
        db.execute("""INSERT INTO interventions
            (numero,equipement_id,tableau_id,technicien_id,equipe_id,type,statut,date_prevue,heure_prevue,description,gamme_id,societe_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (num,d["equipement_id"],to_int(d.get("tableau_id")),main_tech,to_int(d.get("equipe_id")),
             type_iv,d.get("statut",default_statut),date_prevue_final,heure_prevue_v,d.get("description",""),gamme_id_v,sid))
        db.commit(); iid=db.execute("SELECT last_insert_rowid()").fetchone()[0]
        # Enregistrer les techniciens affectés (y compris le principal pour cohérence)
        for tid in tech_ids:
            try:
                db.execute("INSERT OR IGNORE INTO intervention_techniciens (intervention_id,utilisateur_id) VALUES (?,?)",(iid,tid))
            except Exception as e:
                logger.warning(f"[ivtech] insert fail iv={iid} tech={tid}: {e}")
        if tech_ids: db.commit()
        # Créneaux optionnels à la création (intervention multi-jours)
        creneaux = d.get("creneaux") or []
        for c in creneaux:
            if not c.get("date"): continue
            db.execute("""INSERT INTO intervention_creneaux
                (intervention_id,date,heure_debut,heure_fin,technicien_id,notes)
                VALUES (?,?,?,?,?,?)""",
                (iid, c["date"], c.get("heure_debut",""), c.get("heure_fin",""),
                 to_int(c.get("technicien_id")), c.get("notes","")))
        if creneaux: db.commit()
        # Actions post-insert non-critiques (log + notif) : ne doivent PAS empêcher la création
        try:
            log_action(request.user, "CREATE", "intervention", iid, num, f"type={type_iv}")
        except Exception as e:
            logger.warning(f"[create_intervention] log_action échec iid={iid}: {e}")
        try:
            notify(iid,"Nouvelle intervention creee")
        except Exception as e:
            logger.warning(f"[create_intervention] notify échec iid={iid}: {e}")
        # Notification push aux techniciens assignés
        try:
            _notify_intervention(iid, "assigned")
        except Exception as e:
            logger.warning(f"[create_intervention] push échec iid={iid}: {e}")
        return jsonify({"id":iid,"numero":num,
                        "was_shifted": was_shifted,
                        "original_date": original_date,
                        "new_date": date_prevue_final}),201
    except Exception as e:
        import traceback
        logger.error(f"[create_intervention] ERREUR: {e}\n{traceback.format_exc()}")
        return jsonify({"error": f"Erreur serveur : {str(e)}"}),500

@app.route("/api/interventions/<int:iid>",methods=["PATCH"])
@require_auth
def update_intervention(iid):
    d=request.json or {}; db=get_db(); sets,params=[],[]
    # Valider le type si fourni (v218.18 : ASTREINTE accepté)
    if "type" in d and d["type"] not in ("MAINTENANCE","DEPANNAGE","ASTREINTE"):
        return jsonify({"error":"Type invalide"}),400
    # Nettoyer les champs OFFRE obsolètes si envoyés par erreur
    d.pop("numero_offre", None)
    d.pop("numero_projet_offre", None)
    # Si technicien_ids fourni, extraire et synchroniser ensuite
    tech_ids = None
    if "technicien_ids" in d:
        raw = d.pop("technicien_ids") or []
        tech_ids = [to_int(t) for t in raw if t]
        # Dédoublonnage préservant l'ordre
        seen = set(); clean = []
        for t in tech_ids:
            if t and t not in seen: seen.add(t); clean.append(t)
        tech_ids = clean
        # Le premier est le "technicien principal" (compat colonne legacy)
        d["technicien_id"] = tech_ids[0] if tech_ids else None
    # v218.19 : equipement_id désormais modifiable (utile pour ASTREINTE — le tech complète l'équipement)
    # v218.22 : Si un technicien démarre un BA en statut PLANIFIEE → EN_COURS,
    # le bon lui est automatiquement attribué (technicien_id = lui-même)
    # v218.25 : autorise aussi les managers (un manager peut être tech de garde)
    if d.get("statut") == "EN_COURS" and request.user.get("role") in ("technicien", "manager"):
        iv_curr = one(db.execute("SELECT type, statut, technicien_id FROM interventions WHERE id=?", (iid,)))
        if iv_curr and iv_curr.get("type") == "ASTREINTE" and iv_curr.get("statut") == "PLANIFIEE":
            # Réassigner le tech à l'utilisateur courant
            d["technicien_id"] = request.user["id"]
            # Aussi insérer dans intervention_techniciens pour cohérence
            try:
                db.execute("INSERT OR IGNORE INTO intervention_techniciens (intervention_id,utilisateur_id) VALUES (?,?)",
                           (iid, request.user["id"]))
            except Exception: pass
            logger.info(f"[ASTREINTE] BA iid={iid} pris en charge par {request.user.get('role')} id={request.user['id']} ({request.user.get('nom','')})")

    for f in ["technicien_id","equipement_id","equipe_id","type","statut","date_prevue","heure_prevue","date_realisation","description","rapport","tableau_id","gamme_id"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    # Vérification : pour terminer un bon, il faut au minimum un compte rendu
    if "statut" in d and d["statut"]=="TERMINEE":
        nb_cr = (one(db.execute("SELECT COUNT(*) AS n FROM comptes_rendus WHERE intervention_id=?", (iid,))) or {}).get("n", 0)
        if nb_cr == 0:
            iv_row = one(db.execute("SELECT type FROM interventions WHERE id=?", (iid,)))
            type_label = "maintenance" if (iv_row and iv_row.get("type")=="MAINTENANCE") else "dépannage"
            return jsonify({"error": f"Impossible de terminer ce bon de {type_label} sans aucun compte rendu. Crée au moins un CR avant de passer le statut à Terminée."}), 400
        sets.append("date_realisation=?"); params.append(now())
    if not sets and tech_ids is None: return jsonify({"error":"Rien"}),400
    if sets:
        params.append(iid); db.execute(f"UPDATE interventions SET {chr(44).join(sets)} WHERE id=?",params); db.commit()
    # Synchroniser la table de liaison si tech_ids a été fourni
    if tech_ids is not None:
        db.execute("DELETE FROM intervention_techniciens WHERE intervention_id=?",(iid,))
        for tid in tech_ids:
            try:
                db.execute("INSERT OR IGNORE INTO intervention_techniciens (intervention_id,utilisateur_id) VALUES (?,?)",(iid,tid))
            except Exception as e:
                logger.warning(f"[ivtech] update fail iv={iid} tech={tid}: {e}")
        db.commit()
    notify(iid,f"Intervention mise a jour : {d.get('statut','')}")
    # Notification push si modification notable (date/heure/tech/statut)
    push_keys = {"date_prevue", "heure_prevue", "technicien_id", "statut", "technicien_ids"}
    if any(k in d for k in push_keys):
        try:
            details = ""
            if "date_prevue" in d or "heure_prevue" in d:
                details = "(date/heure modifiée)"
            elif "technicien_id" in d or "technicien_ids" in d:
                details = "(affectation modifiée)"
            elif "statut" in d:
                details = f"(statut : {d['statut']})"
            _notify_intervention(iid, "modified", details)
        except Exception as e:
            logger.warning(f"[update_intervention] push échec iid={iid}: {e}")
    # Log
    iv_info = one(db.execute("SELECT numero FROM interventions WHERE id=?", (iid,)))
    label = (iv_info or {}).get("numero", f"iv={iid}")
    changes = ", ".join([f"{k}={v}" for k,v in d.items() if k not in ("password",)])
    log_action(request.user, "UPDATE", "intervention", iid, label, changes[:500])
    # Générer le bon suivant si intervention terminée et planning configuré
    if d.get("statut") == "TERMINEE":
        iv = one(db.execute("SELECT * FROM interventions WHERE id=?", (iid,)))
        if iv and iv.get("equipement_id"):
            eq = one(db.execute("SELECT * FROM equipements WHERE id=?", (iv["equipement_id"],)))
            if eq and eq.get("planning_id") and eq.get("semaine_planif") is not None and eq.get("jour_semaine_planif") is not None:
                try:
                    generate_next_bon(eq["id"], db, iv.get("date_prevue") or iv.get("date_realisation"))
                except Exception as e_gen:
                    pass
    return jsonify({"ok":True})

@app.route("/api/interventions/<int:iid>",methods=["DELETE"])
@require_role("admin","manager")
def delete_intervention(iid):
    db=get_db()
    iv_info = one(db.execute("SELECT numero FROM interventions WHERE id=?", (iid,)))
    label = (iv_info or {}).get("numero", f"iv={iid}")
    db.execute("DELETE FROM comptes_rendus WHERE intervention_id=?",(iid,))
    db.execute("DELETE FROM intervention_creneaux WHERE intervention_id=?",(iid,))
    db.execute("DELETE FROM intervention_techniciens WHERE intervention_id=?",(iid,))
    db.execute("DELETE FROM interventions WHERE id=?",(iid,)); db.commit()
    log_action(request.user, "DELETE", "intervention", iid, label)
    return jsonify({"ok":True})

# ══ CRÉNEAUX D'INTERVENTION (planning multi-jours) ══
@app.route("/api/interventions/<int:iid>/creneaux")
@require_auth
def get_creneaux(iid):
    db=get_db()
    return jsonify(rows(db.execute("""SELECT c.*, u.nom AS technicien_nom
        FROM intervention_creneaux c LEFT JOIN utilisateurs u ON c.technicien_id=u.id
        WHERE c.intervention_id=? ORDER BY c.date, c.heure_debut""",(iid,))))

@app.route("/api/interventions/<int:iid>/creneaux",methods=["POST"])
@require_auth
def create_creneau(iid):
    d=request.json or {}
    if not d.get("date"): return jsonify({"error":"date requise"}),400
    db=get_db()
    if not one(db.execute("SELECT id FROM interventions WHERE id=?",(iid,))):
        return jsonify({"error":"intervention introuvable"}),404
    db.execute("""INSERT INTO intervention_creneaux
        (intervention_id,date,heure_debut,heure_fin,technicien_id,notes)
        VALUES (?,?,?,?,?,?)""",
        (iid,d["date"],d.get("heure_debut",""),d.get("heure_fin",""),
         to_int(d.get("technicien_id")),d.get("notes","")))
    db.commit()
    cid=db.execute("SELECT last_insert_rowid()").fetchone()[0]
    # Notification push aux techs concernés (incluant le nouveau tech si fourni)
    try:
        tech_id = to_int(d.get("technicien_id"))
        details = f"(nouveau créneau : {d['date']} {d.get('heure_debut','')})"
        extra = [tech_id] if tech_id else None
        _notify_intervention(iid, "modified", details, extra_user_ids=extra)
    except Exception as e:
        logger.warning(f"[create_creneau] push échec iid={iid}: {e}")
    return jsonify({"id":cid}),201

@app.route("/api/creneaux/<int:cid>",methods=["PATCH"])
@require_auth
def update_creneau(cid):
    d=request.json or {}; db=get_db(); sets,params=[],[]
    # Récupérer l'intervention parente + l'ancien technicien pour notification
    prev = one(db.execute("SELECT intervention_id, technicien_id, date, heure_debut FROM intervention_creneaux WHERE id=?", (cid,)))
    for f in ["date","heure_debut","heure_fin","technicien_id","notes"]:
        if f in d:
            sets.append(f"{f}=?"); params.append(d[f] if f!="technicien_id" else to_int(d[f]))
    if not sets: return jsonify({"error":"Rien"}),400
    params.append(cid)
    db.execute(f"UPDATE intervention_creneaux SET {chr(44).join(sets)} WHERE id=?",params); db.commit()
    # Notification push
    try:
        if prev and prev.get("intervention_id"):
            iid = prev["intervention_id"]
            # Inclure l'ancien ET le nouveau tech si changement
            extra = []
            if prev.get("technicien_id"): extra.append(prev["technicien_id"])
            if "technicien_id" in d:
                new_t = to_int(d.get("technicien_id"))
                if new_t and new_t not in extra: extra.append(new_t)
            details = "(créneau modifié"
            if "date" in d or "heure_debut" in d:
                new_date = d.get("date", prev.get("date",""))
                new_h = d.get("heure_debut", prev.get("heure_debut",""))
                details += f" : {new_date} {new_h}"
            details += ")"
            _notify_intervention(iid, "modified", details, extra_user_ids=extra if extra else None)
    except Exception as e:
        logger.warning(f"[update_creneau] push échec cid={cid}: {e}")
    return jsonify({"ok":True})

@app.route("/api/creneaux/<int:cid>",methods=["DELETE"])
@require_auth
def delete_creneau(cid):
    db=get_db()
    # Avant suppression : garder les infos pour notif
    prev = one(db.execute("SELECT intervention_id, technicien_id FROM intervention_creneaux WHERE id=?", (cid,)))
    db.execute("DELETE FROM intervention_creneaux WHERE id=?",(cid,)); db.commit()
    try:
        if prev and prev.get("intervention_id"):
            extra = [prev["technicien_id"]] if prev.get("technicien_id") else None
            _notify_intervention(prev["intervention_id"], "modified", "(créneau supprimé)", extra_user_ids=extra)
    except Exception as e:
        logger.warning(f"[delete_creneau] push échec cid={cid}: {e}")
    return jsonify({"ok":True})

# ══ OCCUPATIONS (congés, formation, etc.) ══

@app.route("/api/occupation_types")
@require_auth
def get_occupation_types():
    db = get_db()
    return jsonify(rows(db.execute("SELECT * FROM occupation_types WHERE actif=1 ORDER BY nom")))

@app.route("/api/occupation_types", methods=["POST"])
@require_role("admin")
def create_occupation_type():
    d = request.json or {}
    nom = (d.get("nom") or "").strip()
    if not nom: return jsonify({"error":"Nom requis"}), 400
    couleur = (d.get("couleur") or "#64748b").strip()
    aut_acc = 1 if d.get("autorise_accompagnants") else 0
    db = get_db()
    try:
        db.execute("INSERT INTO occupation_types (nom, couleur, autorise_accompagnants) VALUES (?,?,?)",
                   (nom, couleur, aut_acc))
        db.commit()
        return jsonify({"id": db.execute("SELECT last_insert_rowid()").fetchone()[0]}), 201
    except sqlite3.IntegrityError:
        return jsonify({"error":"Ce type existe déjà"}), 400

@app.route("/api/occupation_types/<int:tid>", methods=["PATCH"])
@require_role("admin")
def update_occupation_type(tid):
    d = request.json or {}; db = get_db(); sets, params = [], []
    for f in ["nom", "couleur", "actif", "autorise_accompagnants"]:
        if f in d:
            sets.append(f"{f}=?")
            val = d[f]
            if f == "autorise_accompagnants":
                val = 1 if val else 0
            params.append(val)
    if not sets: return jsonify({"error":"Rien"}), 400
    params.append(tid)
    db.execute(f"UPDATE occupation_types SET {chr(44).join(sets)} WHERE id=?", params)
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/occupation_types/<int:tid>", methods=["DELETE"])
@require_role("admin")
def delete_occupation_type(tid):
    db = get_db()
    # Soft-delete : on garde les occupations existantes, on désactive juste le type
    used = one(db.execute("SELECT COUNT(*) AS n FROM occupations WHERE type_id=?", (tid,)))
    if used and used["n"] > 0:
        db.execute("UPDATE occupation_types SET actif=0 WHERE id=?", (tid,))
    else:
        db.execute("DELETE FROM occupation_types WHERE id=?", (tid,))
    db.commit()
    return jsonify({"ok": True})

def _calc_heures(hd, hf):
    """Calcule les heures entre deux horaires HH:MM.
    Si heure_fin < heure_debut, considère que c'est le lendemain (passage minuit).
    Ex: 18:00 → 02:30 = 8h30"""
    if not hd or not hf: return 0
    try:
        p1 = hd.split(":"); p2 = hf.split(":")
        m1 = int(p1[0]) * 60 + int(p1[1])
        m2 = int(p2[0]) * 60 + int(p2[1])
        diff_min = m2 - m1
        # Passage minuit : ajouter 24h si fin < début
        if diff_min < 0:
            diff_min += 24 * 60
        return round((diff_min / 60) * 100) / 100
    except Exception:
        return 0

@app.route("/api/occupations")
@require_auth
def get_occupations():
    """Liste des occupations. Technicien = les siennes. Admin/Manager = toutes.
    Paramètres : ?debut=YYYY-MM-DD&fin=YYYY-MM-DD&technicien_id=X"""
    u = request.user
    db = get_db()
    sid = current_societe_id()
    sql = """SELECT o.*, u.nom AS technicien_nom, ot.nom AS type_nom, ot.couleur AS type_couleur,
                    ot.autorise_accompagnants AS type_autorise_accompagnants
             FROM occupations o
             LEFT JOIN utilisateurs u ON o.technicien_id = u.id
             LEFT JOIN occupation_types ot ON o.type_id = ot.id
             WHERE o.societe_id = ?"""
    params = [sid]
    # v218.52 : filtre via la table de liaison occupation_techniciens
    if u["role"] == "technicien":
        sql += " AND EXISTS (SELECT 1 FROM occupation_techniciens ot2 WHERE ot2.occupation_id=o.id AND ot2.technicien_id=?)"
        params.append(u["id"])
    elif request.args.get("technicien_id"):
        tid = to_int(request.args["technicien_id"])
        sql += " AND EXISTS (SELECT 1 FROM occupation_techniciens ot2 WHERE ot2.occupation_id=o.id AND ot2.technicien_id=?)"
        params.append(tid)
    if request.args.get("debut"):
        sql += " AND o.date >= ?"
        params.append(request.args["debut"])
    if request.args.get("fin"):
        sql += " AND o.date <= ?"
        params.append(request.args["fin"])
    sql += " ORDER BY o.date DESC, o.heure_debut"
    result = rows(db.execute(sql, params))
    # Résoudre les accompagnants en noms + ids pour chaque occupation
    _resolve_accompagnants(db, result)
    return jsonify(result)

def _resolve_accompagnants(db, occupations_list):
    """Ajoute techniciens_ids (liste d'ints) et techniciens_noms (liste de str)
    à chaque occupation. Mutation en place.
    v218.52 : utilise occupation_techniciens (table de liaison) au lieu de l'ancien champ JSON.
    Pour rétro-compat, on garde aussi accompagnants_ids et accompagnants_noms (= tous sauf le créateur)."""
    if not occupations_list:
        return
    occ_ids = [o["id"] if "id" in o else o.get("occupation_id") for o in occupations_list]
    occ_ids = [x for x in occ_ids if x]
    if not occ_ids:
        return
    # Charger en bulk : occupation_id → liste de (tech_id, nom)
    placeholders = ",".join(["?"] * len(occ_ids))
    rows_link = rows(db.execute(f"""
        SELECT ot.occupation_id, ot.technicien_id, u.nom
        FROM occupation_techniciens ot
        LEFT JOIN utilisateurs u ON u.id = ot.technicien_id
        WHERE ot.occupation_id IN ({placeholders})
        ORDER BY u.nom
    """, occ_ids))
    by_occ = {}
    for r in rows_link:
        oid = r["occupation_id"]
        if oid not in by_occ: by_occ[oid] = []
        by_occ[oid].append((r["technicien_id"], r.get("nom") or ""))
    for o in occupations_list:
        oid = o.get("id") or o.get("occupation_id")
        liens = by_occ.get(oid, [])
        ids = [t[0] for t in liens]
        noms = [t[1] for t in liens]
        o["techniciens_ids"] = ids
        o["techniciens_noms"] = noms
        # Rétro-compat : accompagnants = tous les techs sauf le créateur principal
        creator = o.get("technicien_id") or o.get("owner_tech_id")
        o["accompagnants_ids"] = [x for x in ids if x != creator]
        o["accompagnants_noms"] = [n for (i, n) in liens if i != creator]

@app.route("/api/occupations", methods=["POST"])
@require_auth
def create_occupation():
    d = request.json or {}
    u = request.user
    db = get_db()
    date = (d.get("date") or "").strip()
    if not date: return jsonify({"error":"Date requise"}), 400
    # Normaliser la liste des techniciens :
    # - techniciens_ids (v218.52, nouveau nom) ou technicien_ids (legacy) : tableau d'ids
    # - sinon technicien_id legacy (1 seule occupation)
    # - tech simple : force son propre id
    if u["role"] == "technicien":
        tech_ids = [u["id"]]
    else:
        raw_ids = d.get("techniciens_ids") or d.get("technicien_ids")
        if isinstance(raw_ids, list) and raw_ids:
            tech_ids = [to_int(t) for t in raw_ids if t]
        else:
            single = to_int(d.get("technicien_id")) or u["id"]
            tech_ids = [single] if single else []
        # Dédoublonner en préservant l'ordre
        seen = set(); clean = []
        for t in tech_ids:
            if t and t not in seen: seen.add(t); clean.append(t)
        tech_ids = clean
    if not tech_ids:
        return jsonify({"error":"Au moins un technicien requis"}), 400
    heure_debut = (d.get("heure_debut") or "").strip()
    heure_fin = (d.get("heure_fin") or "").strip()
    total_heures = d.get("total_heures")
    if total_heures is None:
        total_heures = _calc_heures(heure_debut, heure_fin)
    try:
        total_heures = float(total_heures or 0)
    except Exception:
        total_heures = 0
    # v218.104 : N° projet et nom chantier toujours optionnels (peu importe le type d'occupation).
    # Validation du format P+5 chiffres uniquement si saisi.
    type_id = to_int(d.get("type_id"))
    numero_projet = (d.get("numero_projet") or "").strip()
    nom_chantier = (d.get("nom_chantier") or "").strip()
    if numero_projet and not re.match(r"^P\d{5}$", numero_projet):
        return jsonify({"error":"Le numéro de projet doit être au format P suivi de 5 chiffres (ex: P00042)"}), 400
    # Accompagnants (purement informatif)
    acc_raw = d.get("accompagnants_ids") or []
    acc_clean = []
    if isinstance(acc_raw, list):
        for x in acc_raw:
            xi = to_int(x)
            if xi: acc_clean.append(xi)
    elif isinstance(acc_raw, str) and acc_raw.strip():
        try:
            parsed = json.loads(acc_raw)
            if isinstance(parsed, list):
                acc_clean = [to_int(x) for x in parsed if to_int(x)]
        except Exception:
            pass
    # Plage de dates : si date_fin fournie ET différente de date de début, itérer sur les jours OUVRÉS (lun-ven).
    # Sinon (date unique), créer l'occupation même si c'est un samedi ou dimanche (v218.103).
    date_fin = (d.get("date_fin") or "").strip() or date
    try:
        from datetime import datetime as _dt, timedelta as _td
        d0 = _dt.strptime(date, "%Y-%m-%d").date()
        d1 = _dt.strptime(date_fin, "%Y-%m-%d").date()
    except Exception:
        return jsonify({"error":"Format de date invalide (YYYY-MM-DD attendu)"}), 400
    if d1 < d0:
        return jsonify({"error":"date_fin doit être >= date"}), 400
    dates_to_create = []
    # v218.103 : si une seule date demandée, on l'accepte même un weekend.
    # Si plage (date != date_fin), on filtre les weekends pour ne créer que les jours ouvrés.
    if d0 == d1:
        dates_to_create.append(d0.strftime("%Y-%m-%d"))
    else:
        cur = d0
        while cur <= d1:
            # weekday(): 0=lundi ... 6=dimanche → garder 0..4 (lun-ven) pour les plages
            if cur.weekday() < 5:
                dates_to_create.append(cur.strftime("%Y-%m-%d"))
            cur += _td(days=1)
    if not dates_to_create:
        return jsonify({"error":"Aucun jour ouvré dans la plage (les samedis et dimanches sont ignorés sur les plages multi-jours)"}), 400

    # Créer une seule occupation par jour ouvré, avec tous les techs liés via occupation_techniciens
    # v218.52 : refonte - plus de duplication 1 occ/tech ; 1 occ liée à N techs.
    created_ids = []
    creator_tid = u["id"] if u["role"] != "technicien" else u["id"]  # technicien_id = créateur (ou premier des sélectionnés)
    if u["role"] != "technicien" and tech_ids:
        creator_tid = tech_ids[0]  # premier tech sélectionné = créateur principal
    for the_date in dates_to_create:
        db.execute("""INSERT INTO occupations
            (technicien_id, type_id, date, heure_debut, heure_fin, total_heures, notes, numero_projet, nom_chantier, accompagnants_ids)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (creator_tid, type_id, the_date, heure_debut, heure_fin,
             total_heures, d.get("notes",""), numero_projet, nom_chantier, ""))  # accompagnants_ids legacy = vide
        new_oid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        created_ids.append(new_oid)
        # Insérer tous les techs liés via la table de liaison
        for tid in tech_ids:
            db.execute("INSERT OR IGNORE INTO occupation_techniciens (occupation_id, technicien_id) VALUES (?, ?)",
                       (new_oid, tid))
    db.commit()
    # Log : une entrée par occupation créée
    ot_info = one(db.execute("SELECT nom FROM occupation_types WHERE id=?", (type_id,)))
    ot_nom = (ot_info or {}).get("nom", "Occupation")
    for i, oid in enumerate(created_ids):
        the_date = dates_to_create[i]
        techs_str = ",".join(str(t) for t in tech_ids)
        label = f"{ot_nom} - techs={techs_str} le {the_date}"
        log_action(request.user, "CREATE", "occupation", oid, label)
    # Notifications push
    for oid in created_ids:
        try:
            _notify_occupation(oid, "assigned")
        except Exception as e:
            logger.warning(f"[create_occupation] push échec oid={oid}: {e}")
    return jsonify({"id": created_ids[0], "ids": created_ids, "count": len(created_ids)}), 201

def _notify_occupation(occ_id, event_type, details="", extra_user_ids=None, explicit_user_ids=None):
    """Envoie une notification push pour une occupation (congé, formation, etc.).
    event_type : 'assigned' (nouveau), 'modified' (changement), 'removed' (supprimée)
    extra_user_ids : liste d'ids à notifier en plus du technicien principal
    explicit_user_ids : SI FOURNI, utilisé au lieu de charger depuis la base (ex: pour DELETE après suppression)
    """
    try:
        db = get_db()
        occ = None
        if not explicit_user_ids:
            occ = one(db.execute(
                """SELECT o.id, o.technicien_id, o.date, o.heure_debut, o.heure_fin,
                          o.accompagnants_ids, o.numero_projet, o.nom_chantier,
                          ot.nom AS type_nom, u.nom AS tech_nom
                   FROM occupations o
                   LEFT JOIN occupation_types ot ON o.type_id=ot.id
                   LEFT JOIN utilisateurs u ON o.technicien_id=u.id
                   WHERE o.id=?""", (occ_id,)
            ))
            if not occ: return
        user_ids = set()
        if explicit_user_ids:
            for uid in explicit_user_ids:
                if uid: user_ids.add(uid)
        else:
            if occ.get("technicien_id"): user_ids.add(occ["technicien_id"])
            acc = occ.get("accompagnants_ids") or ""
            if acc:
                try:
                    try:
                        lst = json.loads(acc)
                        if isinstance(lst, list):
                            for x in lst:
                                try: user_ids.add(int(x))
                                except: pass
                    except Exception:
                        for x in str(acc).split(","):
                            x = x.strip()
                            if x.isdigit(): user_ids.add(int(x))
                except Exception: pass
        if extra_user_ids:
            for uid in extra_user_ids:
                if uid: user_ids.add(uid)
        if not user_ids: return

        # ═══ Format de la notification ═══
        # Title : Type d'occurrence (Congé, Formation, etc.)
        # Body ligne 1 : N° projet · Nom chantier
        # Body ligne 2 : Date + horaires
        type_label = (occ or {}).get("type_nom") or "Occupation"
        projet_line = ""
        if (occ or {}).get("numero_projet") and (occ or {}).get("nom_chantier"):
            projet_line = f"{occ['numero_projet']} · {occ['nom_chantier']}"
        elif (occ or {}).get("numero_projet"):
            projet_line = occ["numero_projet"]
        elif (occ or {}).get("nom_chantier"):
            projet_line = occ["nom_chantier"]
        date_line = ""
        if (occ or {}).get("date"):
            try:
                parts = occ["date"][:10].split("-")
                if len(parts) == 3:
                    date_line = f"{parts[2]}/{parts[1]}/{parts[0]}"
                    if occ.get("heure_debut") and occ.get("heure_fin"):
                        date_line += f" {occ['heure_debut']}→{occ['heure_fin']}"
                    elif occ.get("heure_debut"):
                        date_line += f" à {occ['heure_debut']}"
            except Exception: pass
        body_lines = []
        if projet_line: body_lines.append(projet_line)
        if date_line: body_lines.append(date_line)
        if details: body_lines.append(details)
        payload = {
            "title": type_label,
            "body": "\n".join(body_lines),
            "icon": "/icons/socom-icon-192.png",
            "url": "/mobile",
            "tag": f"occ-{occ_id}"
        }
        send_push_to_users(list(user_ids), payload)
    except Exception as ex:
        logger.exception(f"[push notify_occupation] Erreur : {ex}")



@app.route("/api/occupations/<int:oid>", methods=["PATCH"])
@require_auth
def update_occupation(oid):
    d = request.json or {}
    u = request.user
    db = get_db()
    cur = one(db.execute("SELECT * FROM occupations WHERE id=?", (oid,)))
    if not cur: return jsonify({"error":"Introuvable"}), 404
    # v218.52 : un tech peut modifier l'occupation s'il est dans occupation_techniciens
    if u["role"] == "technicien":
        is_linked = one(db.execute(
            "SELECT 1 AS x FROM occupation_techniciens WHERE occupation_id=? AND technicien_id=?",
            (oid, u["id"])
        ))
        if not is_linked:
            return jsonify({"error":"Non autorisé"}), 403
    # v218.104 : N° projet et nom chantier toujours optionnels (peu importe le type).
    # Validation du format P+5 chiffres uniquement si saisi (non vide).
    if "numero_projet" in d:
        final_np = (d.get("numero_projet") or "").strip()
        if final_np and not re.match(r"^P\d{5}$", final_np):
            return jsonify({"error":"Le numéro de projet doit être au format P suivi de 5 chiffres (ex: P00042)"}), 400
        d["numero_projet"] = final_np
    if "nom_chantier" in d:
        d["nom_chantier"] = (d.get("nom_chantier") or "").strip()
    sets, params = [], []
    for f in ["technicien_id", "type_id", "date", "heure_debut", "heure_fin", "notes", "numero_projet", "nom_chantier"]:
        if f in d:
            sets.append(f"{f}=?")
            if f in ("technicien_id", "type_id"):
                params.append(to_int(d[f]))
            else:
                params.append(d[f])
    # Recalculer total_heures si horaires modifiés
    if "heure_debut" in d or "heure_fin" in d:
        hd = d.get("heure_debut", cur.get("heure_debut", ""))
        hf = d.get("heure_fin", cur.get("heure_fin", ""))
        sets.append("total_heures=?")
        params.append(_calc_heures(hd, hf))
    elif "total_heures" in d:
        sets.append("total_heures=?")
        try: params.append(float(d["total_heures"] or 0))
        except: params.append(0)
    if sets:
        params.append(oid)
        db.execute(f"UPDATE occupations SET {chr(44).join(sets)} WHERE id=?", params)
    # v218.52 : si techniciens_ids fourni, on remplace les liens
    if "techniciens_ids" in d:
        raw_tids = d.get("techniciens_ids") or []
        clean_tids = []
        if isinstance(raw_tids, list):
            seen = set()
            for x in raw_tids:
                xi = to_int(x)
                if xi and xi not in seen:
                    seen.add(xi); clean_tids.append(xi)
        if not clean_tids:
            return jsonify({"error":"Au moins un technicien requis"}), 400
        # Remplacer la liste des techs liés
        db.execute("DELETE FROM occupation_techniciens WHERE occupation_id=?", (oid,))
        for tid in clean_tids:
            db.execute("INSERT INTO occupation_techniciens (occupation_id, technicien_id) VALUES (?, ?)", (oid, tid))
        # Mettre aussi à jour le technicien_id (créateur principal) avec le premier de la liste
        # SAUF si déjà fourni dans d
        if "technicien_id" not in d:
            db.execute("UPDATE occupations SET technicien_id=? WHERE id=?", (clean_tids[0], oid))
    if not sets and "techniciens_ids" not in d:
        return jsonify({"error":"Rien"}), 400
    db.commit()
    log_action(request.user, "UPDATE", "occupation", oid, "")
    # Notification push
    try:
        detail_parts = []
        if "date" in d: detail_parts.append(f"date : {d['date']}")
        if "heure_debut" in d or "heure_fin" in d:
            hd = d.get("heure_debut", cur.get("heure_debut",""))
            hf = d.get("heure_fin", cur.get("heure_fin",""))
            detail_parts.append(f"{hd}→{hf}")
        if "techniciens_ids" in d: detail_parts.append("affectation modifiée")
        details = ", ".join(detail_parts)
        _notify_occupation(oid, "modified", details)
    except Exception as e:
        logger.warning(f"[update_occupation] push échec oid={oid}: {e}")
    return jsonify({"ok": True})

@app.route("/api/occupations/<int:oid>/duplicate", methods=["POST"])
@require_auth
def duplicate_occupation(oid):
    """Duplique une occupation existante sur une plage de dates (jours ouvrés lun-ven)."""
    u = request.user
    db = get_db()
    cur = one(db.execute("SELECT * FROM occupations WHERE id=?", (oid,)))
    if not cur: return jsonify({"error":"Introuvable"}), 404
    # Tech ne peut dupliquer que ses propres occupations
    if u["role"] == "technicien" and cur.get("technicien_id") != u["id"]:
        return jsonify({"error":"Non autorisé"}), 403
    body = request.json or {}
    ds = (body.get("date_debut") or "").strip()
    de = (body.get("date_fin") or "").strip()
    if not ds or not de:
        return jsonify({"error":"date_debut et date_fin requis"}), 400
    try:
        from datetime import datetime as _dt, timedelta as _td
        d0 = _dt.strptime(ds, "%Y-%m-%d").date()
        d1 = _dt.strptime(de, "%Y-%m-%d").date()
    except Exception:
        return jsonify({"error":"Format de date invalide (YYYY-MM-DD attendu)"}), 400
    if d1 < d0:
        return jsonify({"error":"date_fin doit être >= date_debut"}), 400

    # Calculer les jours ouvrés à dupliquer (en excluant la date d'origine si elle est dans la plage,
    # pour éviter un doublon parfait du même jour)
    src_date = (cur.get("date") or "")[:10]
    dates_to_create = []
    cur_d = d0
    while cur_d <= d1:
        if cur_d.weekday() < 5:
            ds_str = cur_d.strftime("%Y-%m-%d")
            if ds_str != src_date:
                dates_to_create.append(ds_str)
        cur_d += _td(days=1)
    if not dates_to_create:
        return jsonify({"error":"Aucun jour ouvré à dupliquer dans la plage"}), 400

    created_ids = []
    for the_date in dates_to_create:
        db.execute("""INSERT INTO occupations
            (technicien_id, type_id, date, heure_debut, heure_fin, total_heures, notes, numero_projet, nom_chantier, accompagnants_ids)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (cur.get("technicien_id"), cur.get("type_id"), the_date,
             cur.get("heure_debut") or "", cur.get("heure_fin") or "",
             cur.get("total_heures") or 0, cur.get("notes") or "",
             cur.get("numero_projet") or "", cur.get("nom_chantier") or "",
             cur.get("accompagnants_ids") or ""))
        created_ids.append(db.execute("SELECT last_insert_rowid()").fetchone()[0])
    db.commit()
    # Log
    ot_info = one(db.execute("SELECT nom FROM occupation_types WHERE id=?", (cur.get("type_id"),)))
    ot_nom = (ot_info or {}).get("nom", "Occupation")
    t_info = one(db.execute("SELECT nom FROM utilisateurs WHERE id=?", (cur.get("technicien_id"),)))
    t_nom = (t_info or {}).get("nom", "?")
    for i, new_oid in enumerate(created_ids):
        label = f"{ot_nom} - {t_nom} le {dates_to_create[i]} (dupliqué de #{oid})"
        log_action(request.user, "CREATE", "occupation", new_oid, label)
    # Notifications
    for new_oid in created_ids:
        try:
            _notify_occupation(new_oid, "assigned")
        except Exception as e:
            logger.warning(f"[duplicate_occupation] push échec oid={new_oid}: {e}")
    return jsonify({"ids": created_ids, "count": len(created_ids)}), 201

@app.route("/api/occupations/<int:oid>", methods=["DELETE"])
@require_auth
def delete_occupation(oid):
    u = request.user; db = get_db()
    cur = one(db.execute("SELECT * FROM occupations WHERE id=?", (oid,)))
    if not cur: return jsonify({"error":"Introuvable"}), 404
    # v218.52 : un tech peut supprimer s'il est lié à l'occupation
    if u["role"] == "technicien":
        is_linked = one(db.execute(
            "SELECT 1 AS x FROM occupation_techniciens WHERE occupation_id=? AND technicien_id=?",
            (oid, u["id"])
        ))
        if not is_linked:
            return jsonify({"error":"Non autorisé"}), 403
    # Info pour log
    t_info = one(db.execute("SELECT nom FROM utilisateurs WHERE id=?", (cur.get("technicien_id"),)))
    ot_info = one(db.execute("SELECT nom FROM occupation_types WHERE id=?", (cur.get("type_id"),)))
    label = f"{(ot_info or {}).get('nom','?')} - {(t_info or {}).get('nom','?')} le {cur.get('date','')}"
    # v218.52 : collecter les techs liés AVANT suppression via la table de liaison
    notify_users = set()
    for r in db.execute("SELECT technicien_id FROM occupation_techniciens WHERE occupation_id=?", (oid,)):
        if r["technicien_id"]:
            notify_users.add(r["technicien_id"])
    # ON DELETE CASCADE supprime aussi les liens automatiquement
    db.execute("DELETE FROM occupations WHERE id=?", (oid,))
    db.commit()
    log_action(u, "DELETE", "occupation", oid, label)
    # Notification (ligne déjà supprimée → envoi direct avec payload complet)
    try:
        ot_nom = (ot_info or {}).get("nom", "Occupation")
        if notify_users:
            # Même format que les autres notifs (4 lignes)
            projet_line = ""
            if cur.get("numero_projet") and cur.get("nom_chantier"):
                projet_line = f"{cur['numero_projet']} · {cur['nom_chantier']}"
            elif cur.get("numero_projet"):
                projet_line = cur["numero_projet"]
            elif cur.get("nom_chantier"):
                projet_line = cur["nom_chantier"]
            date_line = ""
            if cur.get("date"):
                try:
                    parts = cur["date"][:10].split("-")
                    if len(parts) == 3:
                        date_line = f"{parts[2]}/{parts[1]}/{parts[0]}"
                        if cur.get("heure_debut") and cur.get("heure_fin"):
                            date_line += f" {cur['heure_debut']}→{cur['heure_fin']}"
                        elif cur.get("heure_debut"):
                            date_line += f" à {cur['heure_debut']}"
                except Exception: pass
            body_lines = []
            if projet_line: body_lines.append(projet_line)
            if date_line: body_lines.append(date_line)
            send_push_to_users(list(notify_users), {
                "title": ot_nom + " (supprimée)",
                "body": "\n".join(body_lines),
                "icon": "/icons/socom-icon-192.png",
                "url": "/mobile",
                "tag": f"occ-{oid}"
            })
    except Exception as e:
        logger.warning(f"[delete_occupation] push échec oid={oid}: {e}")
    return jsonify({"ok": True})

# ══════════════════════════════════════════════════════════════════════
# DEMANDES DE CONGÉS — Workflow tech → manager
# Le technicien fait sa demande via mobile → en attente.
# Le manager (manager_id du tech) la valide ou refuse via la version Web.
# Si validée : création automatique d'occupations type "Congés" sur les jours ouvrés.
# ══════════════════════════════════════════════════════════════════════
def _dc_can_view(u, dc):
    """Vérifie si l'utilisateur peut voir une demande."""
    if u["role"] in ("admin", "superadmin"): return True
    if dc.get("technicien_id") == u["id"]: return True
    if u["role"] == "manager" and dc.get("manager_id") == u["id"]: return True
    return False

def _dc_can_decide(u, dc):
    """Vérifie si l'utilisateur peut valider/refuser une demande."""
    if dc.get("statut") != "EN_ATTENTE": return False
    if u["role"] in ("admin", "superadmin"): return True
    if u["role"] == "manager" and dc.get("manager_id") == u["id"]: return True
    return False

def _dc_resolve_manager_id(db, tech_id):
    """Renvoie l'id du manager d'un tech (manager_id), fallback sur le 1er admin si NULL."""
    t = one(db.execute("SELECT manager_id FROM utilisateurs WHERE id=?", (tech_id,)))
    if t and t.get("manager_id"):
        return t["manager_id"]
    a = one(db.execute("SELECT id FROM utilisateurs WHERE role='admin' AND actif=1 ORDER BY id LIMIT 1"))
    return a["id"] if a else None

def _dc_enrich(db, dc):
    """Ajoute les noms tech/manager/decideur pour l'affichage."""
    if dc.get("technicien_id"):
        t = one(db.execute("SELECT nom, matricule, email FROM utilisateurs WHERE id=?", (dc["technicien_id"],)))
        if t:
            dc["technicien_nom"] = t.get("nom","")
            dc["technicien_matricule"] = t.get("matricule","")
            dc["technicien_email"] = t.get("email","")
    if dc.get("manager_id"):
        m = one(db.execute("SELECT nom, email FROM utilisateurs WHERE id=?", (dc["manager_id"],)))
        if m:
            dc["manager_nom"] = m.get("nom","")
            dc["manager_email"] = m.get("email","")
    if dc.get("decision_par"):
        d = one(db.execute("SELECT nom FROM utilisateurs WHERE id=?", (dc["decision_par"],)))
        if d: dc["decision_par_nom"] = d.get("nom","")
    return dc

def _dc_fr_date(iso):
    """Convertit YYYY-MM-DD → DD/MM/YYYY."""
    if not iso: return ""
    parts = iso[:10].split("-")
    if len(parts) != 3: return iso
    return f"{parts[2]}/{parts[1]}/{parts[0]}"

def _dc_count_work_days(d0_iso, d1_iso):
    """Compte les jours ouvrés (lun-ven) entre 2 dates ISO incluses."""
    from datetime import datetime as _dt, timedelta as _td
    try:
        d0 = _dt.strptime(d0_iso, "%Y-%m-%d").date()
        d1 = _dt.strptime(d1_iso, "%Y-%m-%d").date()
    except Exception:
        return 0
    n = 0
    cur = d0
    while cur <= d1:
        if cur.weekday() < 5:
            n += 1
        cur += _td(days=1)
    return n

def _dc_build_pdf_payload(db, dc):
    """Construit le dict à passer à generate_demande_conge_pdf."""
    _dc_enrich(db, dc)
    # Nombre de jours
    if dc.get("demi_journee"):
        nb_jours = "0,5 jour"
    else:
        n = _dc_count_work_days(dc["date_debut"], dc["date_fin"])
        nb_jours = f"{n} jour" + ("s" if n > 1 else "")
    # Date de la demande (created_at est en SQLite datetime → on prend la date)
    date_demande_iso = (dc.get("created_at") or "")[:10]
    # Décision du responsable
    resp_decision = ""
    resp_date = ""
    if dc.get("statut") == "APPROUVEE":
        resp_decision = "ACCORD"
    elif dc.get("statut") == "REFUSEE":
        resp_decision = "REFUS"
    if dc.get("decision_at"):
        resp_date = _dc_fr_date(dc["decision_at"][:10])
    # Période avec demi-journée
    date_debut_str = _dc_fr_date(dc["date_debut"])
    date_fin_str = _dc_fr_date(dc["date_fin"])
    if dc.get("demi_journee") == "MATIN":
        date_debut_str += " (matin)"
        date_fin_str += " (matin)"
    elif dc.get("demi_journee") == "APRESMIDI":
        date_debut_str += " (après-midi)"
        date_fin_str += " (après-midi)"
    return {
        "nom_complet": dc.get("technicien_nom", ""),
        "matricule": dc.get("technicien_matricule", ""),
        "date_demande": _dc_fr_date(date_demande_iso),
        "date_debut": date_debut_str,
        "date_fin": date_fin_str,
        "nb_jours": nb_jours,
        "motif": dc.get("motif", "LEGAL"),
        "commentaire": dc.get("commentaire", ""),
        "responsable_nom": dc.get("decision_par_nom", "") or dc.get("manager_nom", ""),
        "responsable_decision": resp_decision,
        "responsable_date": resp_date,
        "directeur_nom": "",
        "directeur_decision": "",
        "directeur_date": "",
    }

def _dc_generate_pdf(db, dc):
    """Génère le PDF d'une demande. Renvoie (filename, bytes) ou (None, None) si échec."""
    try:
        from conges_pdf import generate_demande_conge_pdf
        payload = _dc_build_pdf_payload(db, dc)
        pdf_bytes = generate_demande_conge_pdf(payload)
        # Filename : Demande_de_conge_DD_MM_YY_NOM_Prenom.pdf
        nom = (dc.get("technicien_nom") or "").strip().replace(" ", "_")
        date_part = (dc.get("date_debut") or "").replace("-", "_")[2:]  # YY_MM_DD
        # Format final : DD_MM_YY
        try:
            d_iso = dc.get("date_debut", "")
            parts = d_iso.split("-")
            if len(parts) == 3:
                date_part = f"{parts[2]}_{parts[1]}_{parts[0][-2:]}"
        except Exception: pass
        fname = f"Demande_de_conge_{date_part}_{nom}.pdf"
        return fname, pdf_bytes
    except Exception as e:
        logger.error(f"[dc_pdf] échec génération PDF dc#{dc.get('id')}: {e}")
        return None, None


# ════════════════════════════════════════════════════════════════════════
# JOURS FÉRIÉS (v218.124)
# ════════════════════════════════════════════════════════════════════════

def _jf_compute_paques(year):
    """Calcule la date de Pâques pour une année donnée (algorithme de Butcher)."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19*a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2*e + 2*i - h - k) % 7
    m = (a + 11*h + 22*l) // 451
    month = (h + l - 7*m + 114) // 31
    day = ((h + l - 7*m + 114) % 31) + 1
    return date(year, month, day)


def _jf_legaux(year, pays):
    """Retourne la liste des jours fériés légaux pour un pays et une année.
    pays : 'FR' ou 'LU'."""
    paques = _jf_compute_paques(year)
    lundi_paques = paques + timedelta(days=1)
    ascension = paques + timedelta(days=39)
    lundi_pentecote = paques + timedelta(days=50)
    pentecote = paques + timedelta(days=49)
    out = [
        (f"{year}-01-01", "Jour de l'an"),
        (lundi_paques.strftime("%Y-%m-%d"), "Lundi de Pâques"),
        (f"{year}-05-01", "Fête du Travail"),
        (ascension.strftime("%Y-%m-%d"), "Ascension"),
        (lundi_pentecote.strftime("%Y-%m-%d"), "Lundi de Pentecôte"),
        (f"{year}-12-25", "Noël"),
        (f"{year}-12-26", "Saint-Étienne") if pays == "LU" else None,
    ]
    if pays == "FR":
        out += [
            (f"{year}-05-08", "Victoire 1945"),
            (f"{year}-07-14", "Fête nationale"),
            (f"{year}-08-15", "Assomption"),
            (f"{year}-11-01", "Toussaint"),
            (f"{year}-11-11", "Armistice 1918"),
        ]
    elif pays == "LU":
        out += [
            (f"{year}-05-09", "Journée de l'Europe"),
            (f"{year}-06-23", "Fête nationale luxembourgeoise"),
            (f"{year}-08-15", "Assomption"),
            (f"{year}-11-01", "Toussaint"),
        ]
    return [x for x in out if x is not None]


@app.route("/api/jours-feries", methods=["GET"])
@require_auth
def list_jours_feries():
    """Liste les jours fériés de la société courante. Optionnel : ?annee=2026."""
    db = get_db()
    sid = current_societe_id()
    annee = request.args.get("annee")
    where = ["societe_id=?"]; params = [sid]
    if annee:
        try:
            int(annee)
            where.append("date LIKE ?")
            params.append(f"{annee}-%")
        except: pass
    sql = "SELECT * FROM jours_feries WHERE " + " AND ".join(where) + " ORDER BY date"
    rs = rows(db.execute(sql, tuple(params)))
    # Récupérer aussi la config
    cfg = one(db.execute("SELECT * FROM jours_feries_config WHERE societe_id=?", (sid,)))
    if not cfg:
        db.execute("INSERT INTO jours_feries_config (societe_id) VALUES (?)", (sid,))
        db.commit()
        cfg = {"heure_debut_default": "08:00", "heure_fin_default": "16:00"}
    return jsonify({"jours_feries": rs, "config": cfg})


@app.route("/api/jours-feries", methods=["POST"])
@require_role("admin", "manager")
def create_jour_ferie():
    """Crée un jour férié. Body: { date, label, heure_debut?, heure_fin? }"""
    d = request.json or {}
    dt = (d.get("date") or "").strip()
    label = (d.get("label") or "").strip() or "Jour férié"
    if not dt:
        return jsonify({"error": "date requise"}), 400
    # Validation format YYYY-MM-DD
    try:
        datetime.strptime(dt, "%Y-%m-%d")
    except Exception:
        return jsonify({"error": "Format de date invalide (YYYY-MM-DD)"}), 400
    db = get_db()
    sid = current_societe_id()
    cfg = one(db.execute("SELECT * FROM jours_feries_config WHERE societe_id=?", (sid,))) or {}
    hd = (d.get("heure_debut") or cfg.get("heure_debut_default") or "08:00").strip()
    hf = (d.get("heure_fin") or cfg.get("heure_fin_default") or "16:00").strip()
    try:
        db.execute("""INSERT INTO jours_feries (societe_id, date, label, heure_debut, heure_fin)
                      VALUES (?,?,?,?,?)""", (sid, dt, label, hd, hf))
        db.commit()
        new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        log_action(request.user, "CREATE", "jour_ferie", new_id, f"{dt} {label}")
        return jsonify({"id": new_id, "date": dt, "label": label,
                        "heure_debut": hd, "heure_fin": hf}), 201
    except sqlite3.IntegrityError:
        return jsonify({"error": "Un jour férié existe déjà à cette date"}), 409


@app.route("/api/jours-feries/<int:jfid>", methods=["PATCH"])
@require_role("admin", "manager")
def update_jour_ferie(jfid):
    """Modifie un jour férié."""
    d = request.json or {}
    db = get_db()
    sid = current_societe_id()
    existing = one(db.execute("SELECT * FROM jours_feries WHERE id=? AND societe_id=?", (jfid, sid)))
    if not existing:
        return jsonify({"error": "Introuvable"}), 404
    sets = []; params = []
    if "label" in d:
        sets.append("label=?"); params.append((d["label"] or "Jour férié").strip())
    if "date" in d:
        dt = (d["date"] or "").strip()
        try: datetime.strptime(dt, "%Y-%m-%d")
        except: return jsonify({"error": "Format de date invalide"}), 400
        sets.append("date=?"); params.append(dt)
    if "heure_debut" in d:
        sets.append("heure_debut=?"); params.append((d["heure_debut"] or "08:00").strip())
    if "heure_fin" in d:
        sets.append("heure_fin=?"); params.append((d["heure_fin"] or "16:00").strip())
    if not sets:
        return jsonify({"error": "Rien à modifier"}), 400
    params.extend([jfid, sid])
    try:
        db.execute(f"UPDATE jours_feries SET {', '.join(sets)} WHERE id=? AND societe_id=?", params)
        db.commit()
        log_action(request.user, "UPDATE", "jour_ferie", jfid, f"sets={list(d.keys())}")
        return jsonify({"ok": True})
    except sqlite3.IntegrityError:
        return jsonify({"error": "Un jour férié existe déjà à cette date"}), 409


@app.route("/api/jours-feries/<int:jfid>", methods=["DELETE"])
@require_role("admin", "manager")
def delete_jour_ferie(jfid):
    """Supprime un jour férié."""
    db = get_db()
    sid = current_societe_id()
    existing = one(db.execute("SELECT * FROM jours_feries WHERE id=? AND societe_id=?", (jfid, sid)))
    if not existing:
        return jsonify({"error": "Introuvable"}), 404
    db.execute("DELETE FROM jours_feries WHERE id=? AND societe_id=?", (jfid, sid))
    db.commit()
    log_action(request.user, "DELETE", "jour_ferie", jfid, f"{existing['date']} {existing['label']}")
    return jsonify({"ok": True})


@app.route("/api/jours-feries/import-legaux", methods=["POST"])
@require_role("admin", "manager")
def import_jours_feries_legaux():
    """Importe automatiquement les fériés légaux pour une année + pays.
    Body : { annee: 2026, pays: 'FR' ou 'LU', remplacer?: bool }
    Si remplacer=False (défaut), ne touche pas aux fériés déjà saisis pour cette date."""
    d = request.json or {}
    try:
        annee = int(d.get("annee") or date.today().year)
    except:
        return jsonify({"error": "annee invalide"}), 400
    pays = (d.get("pays") or "FR").strip().upper()
    if pays not in ("FR", "LU"):
        return jsonify({"error": "pays doit être 'FR' ou 'LU'"}), 400
    remplacer = bool(d.get("remplacer"))
    db = get_db()
    sid = current_societe_id()
    cfg = one(db.execute("SELECT * FROM jours_feries_config WHERE societe_id=?", (sid,))) or {}
    hd = cfg.get("heure_debut_default") or "08:00"
    hf = cfg.get("heure_fin_default") or "16:00"
    feries = _jf_legaux(annee, pays)
    added = 0; skipped = 0; replaced = 0
    for dt, label in feries:
        existing = one(db.execute("SELECT id FROM jours_feries WHERE societe_id=? AND date=?", (sid, dt)))
        if existing:
            if remplacer:
                db.execute("UPDATE jours_feries SET label=? WHERE id=?", (label, existing["id"]))
                replaced += 1
            else:
                skipped += 1
        else:
            db.execute("""INSERT INTO jours_feries (societe_id, date, label, heure_debut, heure_fin)
                          VALUES (?,?,?,?,?)""", (sid, dt, label, hd, hf))
            added += 1
    db.commit()
    log_action(request.user, "IMPORT", "jours_feries", sid, f"{pays} {annee}",
               f"added={added} replaced={replaced} skipped={skipped}")
    return jsonify({"ok": True, "added": added, "replaced": replaced,
                    "skipped": skipped, "total": len(feries)})


@app.route("/api/jours-feries/config", methods=["PUT"])
@require_role("admin", "manager")
def update_jours_feries_config():
    """Met à jour les heures par défaut pour les nouveaux jours fériés."""
    d = request.json or {}
    hd = (d.get("heure_debut_default") or "08:00").strip()
    hf = (d.get("heure_fin_default") or "16:00").strip()
    db = get_db()
    sid = current_societe_id()
    existing = one(db.execute("SELECT societe_id FROM jours_feries_config WHERE societe_id=?", (sid,)))
    if existing:
        db.execute("""UPDATE jours_feries_config SET heure_debut_default=?, heure_fin_default=?,
                      updated_at=datetime('now') WHERE societe_id=?""", (hd, hf, sid))
    else:
        db.execute("""INSERT INTO jours_feries_config (societe_id, heure_debut_default, heure_fin_default)
                      VALUES (?,?,?)""", (sid, hd, hf))
    db.commit()
    log_action(request.user, "UPDATE", "jours_feries_config", sid,
               f"defaults={hd}-{hf}")
    return jsonify({"ok": True, "heure_debut_default": hd, "heure_fin_default": hf})


@app.route("/api/demandes_conges")
@require_auth
def list_demandes_conges():
    """Liste les demandes de congés.
    - tech : voit ses propres demandes
    - manager : voit les demandes de son équipe (technicien.manager_id = u.id)
    - admin : voit tout
    v218.75 : filtré par société active.
    Filtres optionnels : ?statut=EN_ATTENTE, ?technicien_id=X
    """
    u = request.user
    db = get_db()
    sid = current_societe_id()
    statut = request.args.get("statut")
    tech_filter = to_int(request.args.get("technicien_id"))
    sql = "SELECT * FROM demandes_conges WHERE societe_id=?"
    params = [sid]
    if u["role"] == "technicien":
        sql += " AND technicien_id=?"
        params.append(u["id"])
    elif u["role"] == "manager":
        # Demandes où il est le manager OU ses propres demandes
        sql += " AND (manager_id=? OR technicien_id=?)"
        params += [u["id"], u["id"]]
    # admin / superadmin : pas de filtre rôle (juste société)
    if statut:
        sql += " AND statut=?"
        params.append(statut)
    if tech_filter:
        sql += " AND technicien_id=?"
        params.append(tech_filter)
    sql += " ORDER BY created_at DESC"
    items = rows(db.execute(sql, params))
    for dc in items:
        _dc_enrich(db, dc)
    return jsonify(items)

@app.route("/api/demandes_conges/<int:dcid>")
@require_auth
def get_demande_conge(dcid):
    u = request.user; db = get_db()
    dc = one(db.execute("SELECT * FROM demandes_conges WHERE id=?", (dcid,)))
    if not dc: return jsonify({"error":"Introuvable"}), 404
    if not _dc_can_view(u, dc): return jsonify({"error":"Non autorisé"}), 403
    _dc_enrich(db, dc)
    return jsonify(dc)

@app.route("/api/demandes_conges", methods=["POST"])
@require_auth
def create_demande_conge():
    """Création d'une demande par un technicien (ou par admin pour un autre tech).
    Champs requis : date_debut, date_fin
    Optionnels : demi_journee ('', 'MATIN', 'APRESMIDI'), motif ('LEGAL'/'EXTRAORDINAIRE'),
                 commentaire
    """
    u = request.user
    d = request.json or {}
    db = get_db()
    # Tech cible : par défaut soi-même; admin peut créer pour quelqu'un d'autre
    tech_id = u["id"]
    if u["role"] in ("admin", "superadmin") and d.get("technicien_id"):
        tech_id = to_int(d["technicien_id"]) or u["id"]
    date_debut = (d.get("date_debut") or "").strip()
    date_fin = (d.get("date_fin") or "").strip()
    if not date_debut or not date_fin:
        return jsonify({"error":"Date de début et date de fin requises"}), 400
    try:
        from datetime import datetime as _dt
        d0 = _dt.strptime(date_debut, "%Y-%m-%d").date()
        d1 = _dt.strptime(date_fin, "%Y-%m-%d").date()
    except Exception:
        return jsonify({"error":"Format de date invalide (YYYY-MM-DD attendu)"}), 400
    if d1 < d0:
        return jsonify({"error":"date_fin doit être >= date_debut"}), 400
    demi = (d.get("demi_journee") or "").strip().upper()
    if demi not in ("", "MATIN", "APRESMIDI"):
        return jsonify({"error":"demi_journee doit être '', 'MATIN' ou 'APRESMIDI'"}), 400
    # Une demi-journée doit être sur un seul jour
    if demi and date_debut != date_fin:
        return jsonify({"error":"Une demi-journée ne peut couvrir qu'un seul jour"}), 400
    motif = (d.get("motif") or "LEGAL").strip().upper()
    if motif not in ("LEGAL", "EXTRAORDINAIRE"):
        motif = "LEGAL"
    commentaire = (d.get("commentaire") or "").strip()
    manager_id = _dc_resolve_manager_id(db, tech_id)
    sid = current_societe_id()  # v218.75
    db.execute("""INSERT INTO demandes_conges
        (technicien_id, date_debut, date_fin, demi_journee, motif, commentaire, statut, manager_id, societe_id)
        VALUES (?,?,?,?,?,?, 'EN_ATTENTE', ?, ?)""",
        (tech_id, date_debut, date_fin, demi, motif, commentaire, manager_id, sid))
    new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    db.commit()
    log_action(u, "CREATE", "demande_conge", new_id, f"Demande congé {date_debut} → {date_fin}")
    # Notification push au manager
    try:
        if manager_id:
            tech_info = one(db.execute("SELECT nom FROM utilisateurs WHERE id=?", (tech_id,)))
            tn = (tech_info or {}).get("nom","?")
            send_push_to_users([manager_id], {
                "title": "Nouvelle demande de congé",
                "body": f"{tn} : {date_debut} → {date_fin}",
                "icon": "/icons/socom-icon-192.png",
                "url": "/",
                "tag": f"dc-{new_id}"
            })
    except Exception as e:
        logger.warning(f"[create_demande_conge] push échec : {e}")
    # Envoi mail au manager
    try:
        if manager_id:
            mgr = one(db.execute("SELECT nom, email FROM utilisateurs WHERE id=?", (manager_id,)))
            tech_info = one(db.execute("SELECT nom, matricule FROM utilisateurs WHERE id=?", (tech_id,)))
            if mgr and mgr.get("email"):
                tn = (tech_info or {}).get("nom","?")
                tm = (tech_info or {}).get("matricule","")
                demi_label = ""
                if demi == "MATIN": demi_label = " (matin 08:00-12:00)"
                elif demi == "APRESMIDI": demi_label = " (après-midi 12:00-16:00)"
                motif_label = "Congé extraordinaire" if motif == "EXTRAORDINAIRE" else "Congé légal"
                period_str = (_dc_fr_date(date_debut) if date_debut == date_fin
                              else f"{_dc_fr_date(date_debut)} → {_dc_fr_date(date_fin)}")
                body_lines = [
                    f"Bonjour {(mgr.get('nom') or '').split()[0] if mgr.get('nom') else ''},",
                    "",
                    f"Une nouvelle demande de congé vient d'être déposée par {tn}" + (f" (matricule {tm})" if tm else "") + ".",
                    "",
                    f"Période : {period_str}{demi_label}",
                    f"Motif : {motif_label}",
                ]
                if commentaire:
                    body_lines += ["", f"Commentaire du demandeur : {commentaire}"]
                body_lines += [
                    "",
                    "Merci de vous connecter à la GMAO pour valider ou refuser cette demande :",
                    "https://socom-gmao.lu",
                    "",
                    "— SOCOM GMAO",
                ]
                send_mail(mgr["email"], f"[GMAO] Demande de congé — {tn}", "\n".join(body_lines))
    except Exception as e:
        logger.warning(f"[create_demande_conge] mail manager échec : {e}")
    return jsonify({"id": new_id, "ok": True}), 201

@app.route("/api/demandes_conges/<int:dcid>/valider", methods=["POST"])
@require_auth
def valider_demande_conge(dcid):
    """Validation par le manager. Crée automatiquement les occupations type 'Congés'
    sur les jours ouvrés (lun-ven) de la plage. Demi-journée supportée.
    """
    u = request.user; db = get_db()
    dc = one(db.execute("SELECT * FROM demandes_conges WHERE id=?", (dcid,)))
    if not dc: return jsonify({"error":"Introuvable"}), 404
    if not _dc_can_decide(u, dc): return jsonify({"error":"Non autorisé"}), 403
    d = request.json or {}
    decision_comm = (d.get("decision_commentaire") or "").strip()
    logger.info(f"[valider_dc] dcid={dcid} tech_id={dc.get('technicien_id')} debut={dc.get('date_debut')} fin={dc.get('date_fin')} demi={dc.get('demi_journee')}")
    # Type "Congés"
    type_conges = one(db.execute("SELECT id FROM occupation_types WHERE nom='Congés' AND actif=1"))
    if not type_conges:
        # Garantir l'existence (cas pathologique)
        logger.warning("[valider_dc] type Congés absent, création")
        db.execute("INSERT INTO occupation_types (nom, couleur) VALUES (?,?)", ("Congés","#10B981"))
        db.commit()
        type_conges = one(db.execute("SELECT id FROM occupation_types WHERE nom='Congés'"))
    type_id = type_conges["id"]
    logger.info(f"[valider_dc] type_id Congés = {type_id}")
    # Plage ouvrés
    from datetime import datetime as _dt, timedelta as _td
    d0 = _dt.strptime(dc["date_debut"], "%Y-%m-%d").date()
    d1 = _dt.strptime(dc["date_fin"], "%Y-%m-%d").date()
    demi = dc.get("demi_journee") or ""
    if demi == "MATIN":
        h_deb, h_fin, h_tot = "08:00", "12:00", 4.0
    elif demi == "APRESMIDI":
        h_deb, h_fin, h_tot = "12:00", "16:00", 4.0
    else:
        h_deb, h_fin, h_tot = "08:00", "16:00", 8.0
    notes = "Congé validé"
    if dc.get("motif") == "EXTRAORDINAIRE":
        notes = "Congé extraordinaire validé"
    if dc.get("commentaire"):
        notes += f" — {dc['commentaire']}"
    created_ids = []
    cur = d0
    while cur <= d1:
        if cur.weekday() < 5:  # lun-ven
            ds = cur.strftime("%Y-%m-%d")
            db.execute("""INSERT INTO occupations
                (technicien_id, type_id, date, heure_debut, heure_fin, total_heures, notes)
                VALUES (?,?,?,?,?,?,?)""",
                (dc["technicien_id"], type_id, ds, h_deb, h_fin, h_tot, notes))
            new_oid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            created_ids.append(new_oid)
            logger.info(f"[valider_dc] occ créée id={new_oid} tech={dc['technicien_id']} date={ds}")
        cur += _td(days=1)
    # Mise à jour de la demande
    occ_ids_json = json.dumps(created_ids)
    db.execute("""UPDATE demandes_conges SET
        statut='APPROUVEE', decision_par=?, decision_at=datetime('now'),
        decision_commentaire=?, occupation_ids=?
        WHERE id=?""",
        (u["id"], decision_comm, occ_ids_json, dcid))
    db.commit()
    logger.info(f"[valider_dc] commit OK, {len(created_ids)} occupation(s) créée(s)")
    log_action(u, "VALIDATE", "demande_conge", dcid, f"Validation demande #{dcid} ({len(created_ids)} occ)")
    # Notification push au tech
    try:
        send_push_to_users([dc["technicien_id"]], {
            "title": "Demande de congé approuvée ✓",
            "body": f"{dc['date_debut']} → {dc['date_fin']}",
            "icon": "/icons/socom-icon-192.png",
            "url": "/mobile",
            "tag": f"dc-{dcid}"
        })
    except Exception as e:
        logger.warning(f"[valider_demande_conge] push échec : {e}")
    # Mail au tech avec PDF en pièce jointe
    try:
        # Recharger dc avec les données mises à jour pour le PDF
        dc_fresh = one(db.execute("SELECT * FROM demandes_conges WHERE id=?", (dcid,)))
        if dc_fresh:
            _dc_enrich(db, dc_fresh)
            tech_email = dc_fresh.get("technicien_email", "")
            if tech_email:
                fname, pdf_bytes = _dc_generate_pdf(db, dc_fresh)
                period_str = (_dc_fr_date(dc_fresh["date_debut"]) if dc_fresh["date_debut"] == dc_fresh["date_fin"]
                              else f"{_dc_fr_date(dc_fresh['date_debut'])} → {_dc_fr_date(dc_fresh['date_fin'])}")
                first_name = (dc_fresh.get("technicien_nom") or "").split()
                # Le prénom est le 2e mot dans "SALMON David"
                prenom = first_name[-1] if len(first_name) > 1 else (first_name[0] if first_name else "")
                body_lines = [
                    f"Bonjour {prenom},",
                    "",
                    "Votre demande de congé a été APPROUVÉE.",
                    "",
                    f"Période : {period_str}",
                    f"Validée par : {dc_fresh.get('decision_par_nom','?')}",
                ]
                if decision_comm:
                    body_lines += ["", f"Commentaire : {decision_comm}"]
                body_lines += [
                    "",
                    "L'occupation a été créée automatiquement dans votre planning.",
                    "Vous trouverez ci-joint le formulaire SOCOM rempli pour archivage.",
                    "",
                    "— SOCOM GMAO",
                ]
                attachments = [(fname, pdf_bytes)] if pdf_bytes else None
                send_mail(tech_email, "[GMAO] Demande de congé approuvée ✓",
                          "\n".join(body_lines), attachments=attachments)
    except Exception as e:
        logger.warning(f"[valider_demande_conge] mail tech échec : {e}")
    return jsonify({"ok": True, "occupation_ids": created_ids, "count": len(created_ids)})

@app.route("/api/demandes_conges/<int:dcid>/refuser", methods=["POST"])
@require_auth
def refuser_demande_conge(dcid):
    """Refus par le manager."""
    u = request.user; db = get_db()
    dc = one(db.execute("SELECT * FROM demandes_conges WHERE id=?", (dcid,)))
    if not dc: return jsonify({"error":"Introuvable"}), 404
    if not _dc_can_decide(u, dc): return jsonify({"error":"Non autorisé"}), 403
    d = request.json or {}
    decision_comm = (d.get("decision_commentaire") or "").strip()
    db.execute("""UPDATE demandes_conges SET
        statut='REFUSEE', decision_par=?, decision_at=datetime('now'),
        decision_commentaire=?
        WHERE id=?""",
        (u["id"], decision_comm, dcid))
    db.commit()
    log_action(u, "REFUSE", "demande_conge", dcid, f"Refus demande #{dcid}")
    # Notification push au tech
    try:
        send_push_to_users([dc["technicien_id"]], {
            "title": "Demande de congé refusée",
            "body": f"{dc['date_debut']} → {dc['date_fin']}" + (f" — {decision_comm}" if decision_comm else ""),
            "icon": "/icons/socom-icon-192.png",
            "url": "/mobile",
            "tag": f"dc-{dcid}"
        })
    except Exception as e:
        logger.warning(f"[refuser_demande_conge] push échec : {e}")
    # Mail au tech pour information
    try:
        dc_fresh = one(db.execute("SELECT * FROM demandes_conges WHERE id=?", (dcid,)))
        if dc_fresh:
            _dc_enrich(db, dc_fresh)
            tech_email = dc_fresh.get("technicien_email", "")
            if tech_email:
                period_str = (_dc_fr_date(dc_fresh["date_debut"]) if dc_fresh["date_debut"] == dc_fresh["date_fin"]
                              else f"{_dc_fr_date(dc_fresh['date_debut'])} → {_dc_fr_date(dc_fresh['date_fin'])}")
                first_name = (dc_fresh.get("technicien_nom") or "").split()
                prenom = first_name[-1] if len(first_name) > 1 else (first_name[0] if first_name else "")
                body_lines = [
                    f"Bonjour {prenom},",
                    "",
                    "Votre demande de congé a été REFUSÉE.",
                    "",
                    f"Période : {period_str}",
                    f"Refusée par : {dc_fresh.get('decision_par_nom','?')}",
                ]
                if decision_comm:
                    body_lines += ["", f"Motif du refus : {decision_comm}"]
                body_lines += [
                    "",
                    "Pour toute question, veuillez vous rapprocher de votre manager.",
                    "",
                    "— SOCOM GMAO",
                ]
                send_mail(tech_email, "[GMAO] Demande de congé refusée",
                          "\n".join(body_lines))
    except Exception as e:
        logger.warning(f"[refuser_demande_conge] mail tech échec : {e}")
    return jsonify({"ok": True})

@app.route("/api/demandes_conges/<int:dcid>/annuler", methods=["POST"])
@require_auth
def annuler_demande_conge(dcid):
    """Annulation par le tech (uniquement si EN_ATTENTE) ou par admin (toujours)."""
    u = request.user; db = get_db()
    dc = one(db.execute("SELECT * FROM demandes_conges WHERE id=?", (dcid,)))
    if not dc: return jsonify({"error":"Introuvable"}), 404
    is_owner = (dc.get("technicien_id") == u["id"])
    is_admin = (u["role"] in ("admin", "superadmin"))
    if not (is_owner or is_admin):
        return jsonify({"error":"Non autorisé"}), 403
    if dc.get("statut") != "EN_ATTENTE" and not is_admin:
        return jsonify({"error":"Seules les demandes en attente peuvent être annulées"}), 400
    db.execute("UPDATE demandes_conges SET statut='ANNULEE' WHERE id=?", (dcid,))
    db.commit()
    log_action(u, "CANCEL", "demande_conge", dcid, f"Annulation demande #{dcid}")
    return jsonify({"ok": True})

@app.route("/api/demandes_conges/<int:dcid>", methods=["DELETE"])
@require_auth
def delete_demande_conge(dcid):
    """Suppression définitive d'une demande de congé.
    Si la demande est APPROUVEE, supprime aussi les occupations créées dans le planning.
    Autorisé : admin, manager du tech, ou le tech lui-même.
    """
    u = request.user; db = get_db()
    dc = one(db.execute("SELECT * FROM demandes_conges WHERE id=?", (dcid,)))
    if not dc: return jsonify({"error":"Introuvable"}), 404
    # Autorisation : admin, manager du tech, ou le tech lui-même
    is_admin = (u["role"] in ("admin", "superadmin"))
    is_owner = (dc.get("technicien_id") == u["id"])
    is_manager = (u["role"] == "manager" and dc.get("manager_id") == u["id"])
    if not (is_admin or is_owner or is_manager):
        return jsonify({"error":"Non autorisé"}), 403
    # Si demande approuvée → supprimer aussi les occupations associées
    occ_ids_deleted = []
    occ_ids_raw = dc.get("occupation_ids") or ""
    if dc.get("statut") == "APPROUVEE" and occ_ids_raw:
        try:
            occ_ids_list = json.loads(occ_ids_raw) if occ_ids_raw.startswith("[") else []
            if isinstance(occ_ids_list, list):
                for oid in occ_ids_list:
                    try:
                        oid_int = int(oid)
                        # Vérifier que l'occupation existe et appartient bien au tech demandeur
                        occ = one(db.execute("SELECT id, technicien_id FROM occupations WHERE id=?", (oid_int,)))
                        if occ and occ.get("technicien_id") == dc.get("technicien_id"):
                            db.execute("DELETE FROM occupations WHERE id=?", (oid_int,))
                            occ_ids_deleted.append(oid_int)
                    except Exception:
                        pass
        except Exception as e:
            logger.warning(f"[delete_dc] parse occupation_ids échec dc#{dcid}: {e}")
    # Suppression de la demande
    db.execute("DELETE FROM demandes_conges WHERE id=?", (dcid,))
    db.commit()
    logger.info(f"[delete_dc] dc#{dcid} supprimée par u#{u['id']} ({u['role']}) — {len(occ_ids_deleted)} occupation(s) supprimée(s)")
    log_action(u, "DELETE", "demande_conge", dcid,
               f"Suppression demande #{dcid}" + (f" + {len(occ_ids_deleted)} occ" if occ_ids_deleted else ""))
    # Notification push au tech (sauf s'il a supprimé lui-même)
    if not is_owner and dc.get("technicien_id"):
        try:
            send_push_to_users([dc["technicien_id"]], {
                "title": "Demande de congé supprimée",
                "body": f"{dc['date_debut']} → {dc['date_fin']}" + (f" ({len(occ_ids_deleted)} jour(s) retiré(s) du planning)" if occ_ids_deleted else ""),
                "icon": "/icons/socom-icon-192.png",
                "url": "/mobile",
                "tag": f"dc-{dcid}"
            })
        except Exception as e:
            logger.warning(f"[delete_dc] push échec : {e}")
    return jsonify({"ok": True, "occupations_supprimees": len(occ_ids_deleted)})

@app.route("/api/demandes_conges/<int:dcid>/pdf")
@require_auth
def download_demande_conge_pdf(dcid):
    """Télécharge le PDF du formulaire SOCOM rempli pour une demande.
    Accessible au tech demandeur, à son manager, et aux admins.
    """
    u = request.user; db = get_db()
    dc = one(db.execute("SELECT * FROM demandes_conges WHERE id=?", (dcid,)))
    if not dc: return jsonify({"error":"Introuvable"}), 404
    if not _dc_can_view(u, dc): return jsonify({"error":"Non autorisé"}), 403
    fname, pdf_bytes = _dc_generate_pdf(db, dc)
    if not pdf_bytes:
        return jsonify({"error":"Génération PDF impossible"}), 500
    from flask import Response
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"'}
    )

# ══ POINTAGE MENSUEL (depuis les CR intervenants) ══
@app.route("/api/pointage/<int:annee>/<int:mois>")
@require_auth
def pointage_mensuel(annee, mois):
    """Agrège les heures de cr_intervenants pour un mois donné.
    Filtre facultatif : ?technicien_id=X (sinon tous)."""
    u = request.user
    db = get_db()
    tech_filter = request.args.get("technicien_id")
    # Les techniciens ne peuvent voir que leur propre pointage
    if u["role"] == "technicien":
        tech_filter = u["id"]

    # Pour un manager : restreindre à son équipe (tech avec manager_id = son id) + lui-même
    # v218.73 : limiter à la société active
    sid_for_team = current_societe_id()
    allowed_ids = None  # None = pas de restriction (admin)
    if u["role"] == "manager":
        team_rows = rows(db.execute(
            """SELECT us.utilisateur_id AS id FROM utilisateur_societes us
               WHERE us.societe_id=? AND us.actif=1 AND (us.manager_id=? OR us.utilisateur_id=?)""",
            (sid_for_team, u["id"], u["id"])))
        allowed_ids = [r["id"] for r in team_rows]
        if tech_filter:
            tid = to_int(tech_filter)
            if tid not in allowed_ids:
                return jsonify({"error":"Technicien non autorisé"}), 403
    from calendar import monthrange
    nb_jours = monthrange(annee, mois)[1]
    date_debut = f"{annee:04d}-{mois:02d}-01"
    date_fin = f"{annee:04d}-{mois:02d}-{nb_jours:02d}"
    # Requête de base
    # v218.73 : filtre via i.societe_id (l'intervention porte la société)
    sid = current_societe_id()
    sql = """SELECT ci.date, ci.heure_debut, ci.heure_fin, ci.total_heures,
                    ci.utilisateur_id, COALESCE(u.nom, ci.nom, 'Inconnu') AS nom_tech,
                    cr.id AS cr_id, cr.intervention_id,
                    i.numero AS num_bon, i.type AS type_bon,
                    e.designation AS equipement, p.nom AS projet,
                    p.numero_projet AS num_projet
             FROM cr_intervenants ci
             JOIN comptes_rendus cr ON ci.cr_id = cr.id
             JOIN interventions i ON cr.intervention_id = i.id
             JOIN equipements e ON i.equipement_id = e.id
             JOIN projets p ON e.projet_id = p.id
             LEFT JOIN utilisateurs u ON ci.utilisateur_id = u.id
             WHERE i.societe_id = ? AND ci.date BETWEEN ? AND ?"""
    params = [sid, date_debut, date_fin]
    if tech_filter:
        sql += " AND ci.utilisateur_id = ?"
        params.append(to_int(tech_filter))
    elif allowed_ids is not None:
        if not allowed_ids:
            return jsonify({"totaux": [], "lignes": []})
        ph = ",".join(["?"] * len(allowed_ids))
        sql += f" AND ci.utilisateur_id IN ({ph})"
        params.extend(allowed_ids)
    sql += " ORDER BY nom_tech, ci.date, ci.heure_debut"
    lignes = rows(db.execute(sql, params))
    # Marquer comme lignes d'intervention
    for l in lignes:
        l["source"] = "intervention"

    # Récupérer les occupations sur la même période
    # v218.57 : multi-tech via occupation_techniciens. Chaque tech lié voit l'occupation dans son pointage.
    # v218.74 : filtre par société (occupations Omnis ne doivent pas remonter pour SOCOM et inversement)
    sql_occ = """SELECT o.date, o.heure_debut, o.heure_fin, o.total_heures,
                        link.technicien_id AS utilisateur_id,
                        COALESCE(u.nom, 'Inconnu') AS nom_tech,
                        o.notes, o.numero_projet AS num_projet_occ,
                        o.nom_chantier,
                        ot.nom AS type_occupation, ot.couleur AS type_couleur
                 FROM occupations o
                 JOIN occupation_techniciens link ON link.occupation_id = o.id
                 LEFT JOIN utilisateurs u ON link.technicien_id = u.id
                 LEFT JOIN occupation_types ot ON o.type_id = ot.id
                 WHERE o.societe_id = ? AND o.date BETWEEN ? AND ?"""
    params_occ = [sid, date_debut, date_fin]
    if tech_filter:
        sql_occ += " AND link.technicien_id = ?"
        params_occ.append(to_int(tech_filter))
    elif allowed_ids is not None:
        ph = ",".join(["?"] * len(allowed_ids))
        sql_occ += f" AND link.technicien_id IN ({ph})"
        params_occ.extend(allowed_ids)
    sql_occ += " ORDER BY nom_tech, o.date, o.heure_debut"
    occupations = rows(db.execute(sql_occ, params_occ))
    for o in occupations:
        o["source"] = "occupation"
        o["num_bon"] = ""
        o["type_bon"] = ""
        # Pour une occupation Offre → utiliser numero_projet saisi, sinon vide
        o["num_projet"] = o.get("num_projet_occ") or ""
        o["equipement"] = ""
        o["projet"] = o.get("type_occupation") or "Occupation"
        lignes.append(o)

    # Trier l'ensemble par tech/date/heure
    lignes.sort(key=lambda l: (l.get("nom_tech") or "", l.get("date") or "", l.get("heure_debut") or ""))

    # Totaux par technicien (incluant occupations)
    par_tech = {}
    for l in lignes:
        key = l.get("utilisateur_id") or l.get("nom_tech")
        if key not in par_tech:
            par_tech[key] = {
                "utilisateur_id": l.get("utilisateur_id"),
                "nom": l["nom_tech"],
                "total_heures": 0,
                "total_interventions_h": 0,
                "total_occupations_h": 0,
                "nb_jours": set(),
                "nb_interventions": set(),
            }
        h = float(l.get("total_heures") or 0)
        par_tech[key]["total_heures"] += h
        if l.get("source") == "occupation":
            par_tech[key]["total_occupations_h"] += h
        else:
            par_tech[key]["total_interventions_h"] += h
            if l.get("intervention_id"):
                par_tech[key]["nb_interventions"].add(l["intervention_id"])
        par_tech[key]["nb_jours"].add(l["date"])
    # Format JSON friendly
    totaux = [{
        "utilisateur_id": v["utilisateur_id"],
        "nom": v["nom"],
        "total_heures": round(v["total_heures"], 2),
        "total_interventions_h": round(v["total_interventions_h"], 2),
        "total_occupations_h": round(v["total_occupations_h"], 2),
        "nb_jours": len(v["nb_jours"]),
        "nb_interventions": len(v["nb_interventions"]),
    } for v in par_tech.values()]
    totaux.sort(key=lambda x: x["nom"])
    return jsonify({
        "annee": annee, "mois": mois,
        "date_debut": date_debut, "date_fin": date_fin,
        "lignes": lignes, "totaux": totaux,
    })

@app.route("/api/pointage/<int:annee>/<int:mois>/export")
@require_role("admin","manager")
def pointage_export(annee, mois):
    """Export Excel du pointage mensuel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import make_response
    db = get_db()
    sid = current_societe_id()  # v218.74
    from calendar import monthrange
    nb_jours = monthrange(annee, mois)[1]
    date_debut = f"{annee:04d}-{mois:02d}-01"
    date_fin = f"{annee:04d}-{mois:02d}-{nb_jours:02d}"
    lignes = rows(db.execute("""
        SELECT ci.date, ci.heure_debut, ci.heure_fin, ci.total_heures,
               COALESCE(u.nom, ci.nom, 'Inconnu') AS nom_tech,
               i.numero AS num_bon, i.type AS type_bon,
               e.designation AS equipement, p.nom AS projet,
               p.numero_projet AS num_projet
        FROM cr_intervenants ci
        JOIN comptes_rendus cr ON ci.cr_id = cr.id
        JOIN interventions i ON cr.intervention_id = i.id
        JOIN equipements e ON i.equipement_id = e.id
        JOIN projets p ON e.projet_id = p.id
        LEFT JOIN utilisateurs u ON ci.utilisateur_id = u.id
        WHERE i.societe_id = ? AND ci.date BETWEEN ? AND ?
        ORDER BY nom_tech, ci.date, ci.heure_debut
    """, [sid, date_debut, date_fin]))
    for l in lignes: l["source"] = "intervention"

    # Occupations sur la période
    occupations = rows(db.execute("""
        SELECT o.date, o.heure_debut, o.heure_fin, o.total_heures,
               COALESCE(u.nom, 'Inconnu') AS nom_tech,
               o.notes, o.numero_projet AS num_projet_occ,
               o.nom_chantier,
               ot.nom AS type_occupation
        FROM occupations o
        LEFT JOIN utilisateurs u ON o.technicien_id = u.id
        LEFT JOIN occupation_types ot ON o.type_id = ot.id
        WHERE o.societe_id = ? AND o.date BETWEEN ? AND ?
        ORDER BY nom_tech, o.date, o.heure_debut
    """, [sid, date_debut, date_fin]))
    for o in occupations:
        o["source"] = "occupation"
        o["num_bon"] = ""
        o["type_bon"] = o.get("type_occupation") or "Occupation"
        # Offre → numero_projet saisi à la place du N° Projet classique
        o["num_projet"] = o.get("num_projet_occ") or ""
        o["equipement"] = ""
        # Pour le projet, afficher le nom du chantier si présent (offre), sinon notes/type
        o["projet"] = o.get("nom_chantier") or o.get("notes") or o.get("type_occupation") or ""
    all_lignes = lignes + occupations

    # v218.167 : ajouter les jours fériés du mois pour chaque technicien actif
    # Heures = celles configurées sur chaque férié (par défaut 08:00-16:00 = 8h), catégorie "Férié"
    feries = rows(db.execute(
        "SELECT date, label, heure_debut, heure_fin FROM jours_feries WHERE societe_id=? AND date BETWEEN ? AND ? ORDER BY date",
        [sid, date_debut, date_fin]
    ))
    techniciens_actifs = rows(db.execute("""
        SELECT u.id, u.nom FROM utilisateurs u
        WHERE u.actif=1 AND EXISTS(
            SELECT 1 FROM utilisateur_societes us WHERE us.utilisateur_id=u.id AND us.societe_id=?
        )
        ORDER BY u.nom
    """, [sid]))
    logger.info(f"[pointage_export] {annee}-{mois}: {len(feries)} jour(s) férié(s), {len(techniciens_actifs)} technicien(s) actif(s)")
    feries_lignes = []
    for f in feries:
        libelle = f.get("label") or "Jour férié"
        hd = f.get("heure_debut") or "08:00"
        hf = f.get("heure_fin") or "16:00"
        # Calculer le total d'heures
        try:
            sh, sm = [int(x) for x in hd.split(":")]
            eh, em = [int(x) for x in hf.split(":")]
            total_h = (eh*60+em - sh*60-sm) / 60.0
            if total_h <= 0: total_h = 8.0
        except Exception:
            total_h = 8.0
        for tech in techniciens_actifs:
            feries_lignes.append({
                "source": "ferie",
                "nom_tech": tech["nom"],
                "date": f["date"],
                "heure_debut": hd,
                "heure_fin": hf,
                "total_heures": total_h,
                "num_bon": "",
                "type_bon": libelle,
                "num_projet": "",
                "equipement": "",
                "projet": libelle,
            })
    all_lignes += feries_lignes
    all_lignes.sort(key=lambda l: (l.get("nom_tech") or "", l.get("date") or "", l.get("heure_debut") or ""))

    wb = Workbook()
    # ── Feuille 1 : Détail ──
    ws = wb.active
    ws.title = "Détail"
    headers = ["Technicien", "Date", "Début", "Fin", "Total h", "Catégorie", "N° Bon", "Type", "N° Projet", "Équipement", "Projet / Notes"]
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="0F1E3D")
    occ_fill = PatternFill("solid", fgColor="FEF3C7")  # fond jaune pâle pour occupations
    ferie_fill = PatternFill("solid", fgColor="FECACA")  # v218.166 : fond rouge pâle pour fériés
    thin = Side(border_style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = head_font; c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    for idx, l in enumerate(all_lignes, 2):
        is_occ = l.get("source") == "occupation"
        is_ferie = l.get("source") == "ferie"
        if is_ferie: cat = "Férié"
        elif is_occ: cat = "Occupation"
        else: cat = "Intervention"
        row_data = [
            l["nom_tech"], l["date"], l.get("heure_debut",""), l.get("heure_fin",""),
            float(l.get("total_heures") or 0),
            cat,
            l.get("num_bon",""), l.get("type_bon",""),
            l.get("num_projet",""),
            l.get("equipement",""), l.get("projet","")
        ]
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(row=idx, column=ci, value=v)
            c.border = border
            if is_ferie: c.fill = ferie_fill
            elif is_occ: c.fill = occ_fill
    # Largeurs
    widths = [22, 12, 8, 8, 10, 12, 14, 14, 14, 30, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = "A2"
    # ── Feuille 2 : Totaux ──
    ws2 = wb.create_sheet("Totaux")
    # v218.166 : ajout colonne "dont Fériés"
    headers2 = ["Technicien", "Total heures", "dont Interventions", "dont Occupations", "dont Fériés", "Nb jours travaillés", "Nb interventions"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = head_font; c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    # Agrégation
    par_tech = {}
    for l in all_lignes:
        key = l["nom_tech"]
        if key not in par_tech:
            par_tech[key] = {"nom": key, "total": 0, "interv_h": 0, "occ_h": 0, "ferie_h": 0, "jours": set(), "bons": set()}
        h = float(l.get("total_heures") or 0)
        par_tech[key]["total"] += h
        if l.get("source") == "occupation":
            par_tech[key]["occ_h"] += h
        elif l.get("source") == "ferie":
            par_tech[key]["ferie_h"] += h
        else:
            par_tech[key]["interv_h"] += h
            if l.get("num_bon"): par_tech[key]["bons"].add(l["num_bon"])
        par_tech[key]["jours"].add(l["date"])
    for idx, (nom, v) in enumerate(sorted(par_tech.items()), 2):
        row_data = [v["nom"], round(v["total"],2), round(v["interv_h"],2), round(v["occ_h"],2),
                    round(v["ferie_h"],2),
                    len(v["jours"]), len(v["bons"])]
        for ci, val in enumerate(row_data, 1):
            c = ws2.cell(row=idx, column=ci, value=val)
            c.border = border
            if ci == 2: c.font = Font(bold=True)
    for i, w in enumerate([22, 14, 16, 16, 14, 20, 18], 1):
        ws2.column_dimensions[chr(64+i)].width = w
    ws2.freeze_panes = "A2"
    # ── Feuille 3 : Occupations ──
    ws3 = wb.create_sheet("Occupations")
    headers3 = ["Technicien", "Date", "Début", "Fin", "Total h", "Type", "N° Projet", "Nom du chantier", "Notes"]
    for i, h in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = head_font; c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    for idx, o in enumerate(occupations, 2):
        row_data = [
            o["nom_tech"], o["date"], o.get("heure_debut",""), o.get("heure_fin",""),
            float(o.get("total_heures") or 0),
            o.get("type_occupation",""),
            o.get("num_projet_occ","") or o.get("num_projet",""),
            o.get("nom_chantier",""),
            o.get("notes","")
        ]
        for ci, v in enumerate(row_data, 1):
            c = ws3.cell(row=idx, column=ci, value=v)
            c.border = border
    for i, w in enumerate([22, 12, 8, 8, 10, 18, 14, 30, 40], 1):
        ws3.column_dimensions[chr(64+i)].width = w
    ws3.freeze_panes = "A2"
    # Sortie
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"Pointage_{annee:04d}-{mois:02d}.xlsx"
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp

# ══ POINTAGE HEBDOMADAIRE ══

def _iso_week_dates(annee, semaine):
    """Retourne (lundi, dimanche) datetime.date de la semaine ISO donnée."""
    from datetime import date, timedelta
    # Le 4 janvier est toujours dans la semaine 1 (standard ISO)
    jan4 = date(annee, 1, 4)
    # Lundi de la semaine 1
    w1_mon = jan4 - timedelta(days=jan4.weekday())
    mon = w1_mon + timedelta(weeks=semaine - 1)
    sun = mon + timedelta(days=6)
    return mon, sun

@app.route("/api/pointage/week/<int:annee>/<int:semaine>")
@require_auth
def pointage_hebdo(annee, semaine):
    """Agrège les heures d'une semaine donnée. Par technicien et par projet/jour."""
    u = request.user
    db = get_db()
    try:
        mon, sun = _iso_week_dates(annee, semaine)
    except Exception:
        return jsonify({"error":"Semaine invalide"}), 400
    date_debut = mon.strftime("%Y-%m-%d")
    date_fin = sun.strftime("%Y-%m-%d")
    tech_filter = request.args.get("technicien_id")
    if u["role"] == "technicien":
        tech_filter = u["id"]

    # Pour un manager : restreindre à son équipe (tech avec manager_id = son id) + lui-même
    # v218.73 : limiter à la société active
    sid_for_team = current_societe_id()
    allowed_ids = None  # None = pas de restriction (admin)
    if u["role"] == "manager":
        team_rows = rows(db.execute(
            """SELECT us.utilisateur_id AS id FROM utilisateur_societes us
               WHERE us.societe_id=? AND us.actif=1 AND (us.manager_id=? OR us.utilisateur_id=?)""",
            (sid_for_team, u["id"], u["id"])))
        allowed_ids = [r["id"] for r in team_rows]
        # Si un filtre tech_id est demandé, vérifier qu'il est autorisé
        if tech_filter:
            tid = to_int(tech_filter)
            if tid not in allowed_ids:
                return jsonify({"error":"Technicien non autorisé"}), 403

    # Lignes interventions (via cr_intervenants)
    # v218.73 : filtre via i.societe_id
    sid = current_societe_id()
    sql = """SELECT ci.date, ci.heure_debut, ci.heure_fin, ci.total_heures,
                    ci.utilisateur_id, COALESCE(u.nom, ci.nom, 'Inconnu') AS nom_tech,
                    cr.id AS cr_id, cr.intervention_id,
                    i.numero AS num_bon, i.type AS type_bon,
                    e.designation AS equipement, p.nom AS projet,
                    p.numero_projet AS num_projet
             FROM cr_intervenants ci
             JOIN comptes_rendus cr ON ci.cr_id = cr.id
             JOIN interventions i ON cr.intervention_id = i.id
             JOIN equipements e ON i.equipement_id = e.id
             JOIN projets p ON e.projet_id = p.id
             LEFT JOIN utilisateurs u ON ci.utilisateur_id = u.id
             WHERE i.societe_id = ? AND ci.date BETWEEN ? AND ?"""
    params = [sid, date_debut, date_fin]
    if tech_filter:
        sql += " AND ci.utilisateur_id = ?"
        params.append(to_int(tech_filter))
    elif allowed_ids is not None:
        # Manager sans filtre : restreindre à son équipe
        if not allowed_ids:
            return jsonify({"totaux": [], "lignes": []})
        ph = ",".join(["?"] * len(allowed_ids))
        sql += f" AND ci.utilisateur_id IN ({ph})"
        params.extend(allowed_ids)
    sql += " ORDER BY nom_tech, ci.date, ci.heure_debut"
    lignes = rows(db.execute(sql, params))
    for l in lignes:
        l["source"] = "intervention"

    # Occupations sur la période
    # v218.57 : multi-tech via occupation_techniciens
    sql_occ = """SELECT o.date, o.heure_debut, o.heure_fin, o.total_heures,
                        link.technicien_id AS utilisateur_id,
                        COALESCE(u.nom, 'Inconnu') AS nom_tech,
                        o.notes, o.numero_projet AS num_projet_occ, o.nom_chantier,
                        ot.nom AS type_occupation, ot.couleur AS type_couleur
                 FROM occupations o
                 JOIN occupation_techniciens link ON link.occupation_id = o.id
                 LEFT JOIN utilisateurs u ON link.technicien_id = u.id
                 LEFT JOIN occupation_types ot ON o.type_id = ot.id
                 WHERE o.societe_id = ? AND o.date BETWEEN ? AND ?"""
    params_occ = [sid, date_debut, date_fin]
    if tech_filter:
        sql_occ += " AND link.technicien_id = ?"
        params_occ.append(to_int(tech_filter))
    elif allowed_ids is not None:
        # Manager sans filtre : restreindre à son équipe
        ph = ",".join(["?"] * len(allowed_ids))
        sql_occ += f" AND link.technicien_id IN ({ph})"
        params_occ.extend(allowed_ids)
    sql_occ += " ORDER BY nom_tech, o.date, o.heure_debut"
    occupations = rows(db.execute(sql_occ, params_occ))
    for o in occupations:
        o["source"] = "occupation"
        o["num_bon"] = ""
        o["type_bon"] = ""
        o["num_projet"] = o.get("num_projet_occ") or ""
        o["equipement"] = ""
        o["projet"] = o.get("nom_chantier") or o.get("type_occupation") or ""
        lignes.append(o)

    # v218.125 : Injecter les jours fériés comme lignes virtuelles pour chaque technicien
    # actif de la société (ils s'appliquent à TOUT LE MONDE).
    feries = rows(db.execute(
        """SELECT date, label, heure_debut, heure_fin FROM jours_feries
           WHERE societe_id=? AND date BETWEEN ? AND ?""",
        (sid, date_debut, date_fin)
    ))
    if feries:
        # Lister les techniciens concernés
        if tech_filter:
            tech_rows = rows(db.execute(
                """SELECT u.id, u.nom FROM utilisateurs u WHERE u.id=?""",
                (to_int(tech_filter),)))
        elif allowed_ids is not None:
            # Manager : son équipe
            ph_t = ",".join(["?"] * len(allowed_ids))
            tech_rows = rows(db.execute(
                f"""SELECT u.id, u.nom FROM utilisateurs u WHERE u.id IN ({ph_t})""",
                allowed_ids))
        else:
            # Admin : tous les techs actifs de la société courante
            tech_rows = rows(db.execute(
                """SELECT u.id, u.nom FROM utilisateurs u
                   JOIN utilisateur_societes us ON us.utilisateur_id=u.id
                   WHERE us.societe_id=? AND us.actif=1
                     AND u.role IN ('technicien','manager','admin')""",
                (sid,)))
        for jf in feries:
            try:
                hd = jf.get("heure_debut") or "08:00"
                hf = jf.get("heure_fin") or "16:00"
                h_total = _calc_heures(hd, hf)
            except Exception:
                h_total = 8.0
            for t in tech_rows:
                lignes.append({
                    "source": "ferie",
                    "date": jf["date"],
                    "heure_debut": hd,
                    "heure_fin": hf,
                    "total_heures": h_total,
                    "utilisateur_id": t["id"],
                    "nom_tech": t["nom"],
                    "type_occupation": jf.get("label") or "Jour férié",
                    "num_bon": "",
                    "type_bon": "",
                    "num_projet": "",
                    "equipement": "",
                    "projet": jf.get("label") or "Jour férié",
                })

    lignes.sort(key=lambda l: (l.get("nom_tech") or "", l.get("date") or "", l.get("heure_debut") or ""))

    # Totaux par technicien
    par_tech = {}
    for l in lignes:
        key = l.get("utilisateur_id") or l.get("nom_tech")
        if key not in par_tech:
            par_tech[key] = {
                "utilisateur_id": l.get("utilisateur_id"),
                "nom": l["nom_tech"],
                "total_heures": 0,
                "total_interventions_h": 0,
                "total_occupations_h": 0,
                "nb_jours": set(),
                "nb_interventions": set(),
            }
        h = float(l.get("total_heures") or 0)
        par_tech[key]["total_heures"] += h
        if l.get("source") in ("occupation", "ferie"):
            par_tech[key]["total_occupations_h"] += h
        else:
            par_tech[key]["total_interventions_h"] += h
            if l.get("intervention_id"):
                par_tech[key]["nb_interventions"].add(l["intervention_id"])
        par_tech[key]["nb_jours"].add(l["date"])
    totaux = [{
        "utilisateur_id": v["utilisateur_id"],
        "nom": v["nom"],
        "total_heures": round(v["total_heures"], 2),
        "total_interventions_h": round(v["total_interventions_h"], 2),
        "total_occupations_h": round(v["total_occupations_h"], 2),
        "nb_jours": len(v["nb_jours"]),
        "nb_interventions": len(v["nb_interventions"]),
    } for v in par_tech.values()]
    totaux.sort(key=lambda x: x["nom"])
    return jsonify({
        "annee": annee, "semaine": semaine,
        "date_debut": date_debut, "date_fin": date_fin,
        "lignes": lignes, "totaux": totaux,
    })


# ══ PARAMÈTRES HEURES JOUR/NUIT ══
def _get_plages_jour_nuit():
    """Retourne (debut_jour, fin_jour) en minutes depuis minuit.
    Par défaut 06:00→22:00 (jour), le reste = nuit."""
    db = get_db()
    rows_p = rows(db.execute(
        "SELECT cle, valeur FROM parametres_app WHERE cle IN ('plage_jour_debut','plage_jour_fin')"
    ))
    params = {r["cle"]: r["valeur"] for r in rows_p}
    def _hm_to_min(s, default):
        try:
            parts = (s or "").strip().split(":")
            if len(parts) >= 2:
                return int(parts[0])*60 + int(parts[1])
        except Exception:
            pass
        return default
    deb = _hm_to_min(params.get("plage_jour_debut"), 6*60)
    fin = _hm_to_min(params.get("plage_jour_fin"), 22*60)
    return deb, fin

def _split_jour_nuit(h_debut, h_fin):
    """Découpe un intervalle [h_debut, h_fin] (chaînes 'HH:MM') en heures jour et heures nuit
    selon les plages configurées. Retourne (h_jour, h_nuit) en heures décimales.
    Si h_fin <= h_debut, on considère que la fin est le lendemain (passage minuit).
    Si h_debut ou h_fin invalide, retourne (0, 0)."""
    def _parse(s):
        try:
            parts = (s or "").strip().split(":")
            if len(parts) >= 2:
                return int(parts[0])*60 + int(parts[1])
        except Exception:
            pass
        return None
    d = _parse(h_debut); f = _parse(h_fin)
    if d is None or f is None:
        return (0.0, 0.0)
    # Gestion passage minuit : on étale sur 0..2880 minutes (2 jours)
    if f <= d:
        f += 24*60
    deb_j, fin_j = _get_plages_jour_nuit()
    # Construire les segments JOUR sur 2 jours (jour 1 + jour 2 si on dépasse minuit)
    jour_segs = []
    if deb_j < fin_j:
        # Plage jour normale dans la journée
        jour_segs.append((deb_j, fin_j))
        jour_segs.append((deb_j + 24*60, fin_j + 24*60))
    elif deb_j > fin_j:
        # Plage jour qui chevauche minuit (cas exotique mais on gère)
        jour_segs.append((0, fin_j))
        jour_segs.append((deb_j, fin_j + 24*60))
        jour_segs.append((deb_j + 24*60, 24*60 + 24*60))
    else:
        # deb == fin → 0 heures de jour, tout en nuit
        pass
    # Calculer l'intersection [d, f] avec chaque segment jour
    minutes_jour = 0
    for (a, b) in jour_segs:
        inter = max(0, min(f, b) - max(d, a))
        minutes_jour += inter
    minutes_total = f - d
    minutes_nuit = max(0, minutes_total - minutes_jour)
    return (round(minutes_jour/60.0, 4), round(minutes_nuit/60.0, 4))

@app.route("/api/parametres/heures", methods=["GET"])
@require_auth
def get_parametres_heures():
    deb, fin = _get_plages_jour_nuit()
    def _to_hm(m):
        return f"{m//60:02d}:{m%60:02d}"
    return jsonify({
        "plage_jour_debut": _to_hm(deb),
        "plage_jour_fin": _to_hm(fin),
    })

@app.route("/api/parametres/heures", methods=["PATCH"])
@require_role("admin")
def update_parametres_heures():
    d = request.get_json(silent=True) or {}
    db = get_db()
    def _valid_hm(s):
        try:
            parts = (s or "").strip().split(":")
            if len(parts) != 2: return False
            h = int(parts[0]); m = int(parts[1])
            return 0 <= h <= 23 and 0 <= m <= 59
        except Exception:
            return False
    for cle in ("plage_jour_debut","plage_jour_fin"):
        if cle in d:
            v = (d.get(cle) or "").strip()
            if not _valid_hm(v):
                return jsonify({"error": f"Format invalide pour {cle} (attendu HH:MM)"}), 400
            db.execute(
                "INSERT INTO parametres_app (cle,valeur) VALUES (?,?) "
                "ON CONFLICT(cle) DO UPDATE SET valeur=excluded.valeur",
                (cle, v)
            )
    db.commit()
    deb, fin = _get_plages_jour_nuit()
    def _to_hm(m): return f"{m//60:02d}:{m%60:02d}"
    return jsonify({
        "plage_jour_debut": _to_hm(deb),
        "plage_jour_fin": _to_hm(fin),
    })


@app.route("/api/pointage/week/<int:annee>/<int:semaine>/<int:technicien_id>/fiche")
@require_role("admin","manager")
def pointage_fiche_hebdo(annee, semaine, technicien_id):
    """Génère une fiche de pointage Excel hebdomadaire pour un technicien,
    en utilisant le template pointage_template.xlsx."""
    from openpyxl import load_workbook
    from flask import make_response
    db = get_db()
    try:
        mon, sun = _iso_week_dates(annee, semaine)
    except Exception:
        return jsonify({"error":"Semaine invalide"}), 400
    # Récupérer le technicien + son manager
    tech = one(db.execute("SELECT * FROM utilisateurs WHERE id=?", (technicien_id,)))
    if not tech:
        return jsonify({"error":"Technicien introuvable"}), 404
    logger.info(f"[FICHE] tech_id={technicien_id} nom={tech.get('nom')} matricule={repr(tech.get('matricule'))} manager_id={tech.get('manager_id')}")
    date_debut = mon.strftime("%Y-%m-%d")
    date_fin = sun.strftime("%Y-%m-%d")
    # Localiser le template
    tpl_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pointage_template.xlsx")
    if not os.path.exists(tpl_path):
        return jsonify({"error":"Template introuvable"}), 500
    wb = load_workbook(tpl_path)
    ws = wb.active
    # Remplir en-tête
    ws["B5"] = semaine
    from datetime import datetime as _dt
    ws["G14"] = _dt(mon.year, mon.month, mon.day)  # formules Excel calculent les 6 autres jours
    ws["B7"] = tech.get("nom","")
    # Matricule : on tente d'abord l'entier (cohérent avec le template), sinon string
    mat_raw = (tech.get("matricule") or "").strip() if isinstance(tech.get("matricule"), str) else tech.get("matricule")
    if mat_raw:
        try:
            ws["I7"] = int(str(mat_raw).strip())
        except (ValueError, TypeError):
            ws["I7"] = str(mat_raw).strip()
    else:
        ws["I7"] = ""
    # Manager : récupérer depuis le lien manager_id du technicien
    mgr_nom = ""
    if tech.get("manager_id"):
        mgr = one(db.execute("SELECT nom FROM utilisateurs WHERE id=?", (tech.get("manager_id"),)))
        if mgr:
            mgr_nom = mgr.get("nom","") or ""
    logger.info(f"[FICHE] mgr_nom={repr(mgr_nom)} I7={repr(ws['I7'].value)}")
    ws["V7"] = mgr_nom
    # Bloc signature manager (B35)
    ws["B35"] = mgr_nom
    # Effacer les lignes exemples (16..32) pour les colonnes projet/données
    for row in range(16, 33):
        for col_letter in ['A','B','C','E','F','G','H','I','J','K','L','M','N',
                            'O','P','Q','R','S','T','U','V','W','X','Y','Z','AA']:
            coord = f'{col_letter}{row}'
            ws[coord] = None
    # Récupérer les CR+intervenants de la semaine pour ce tech
    cr_rows = rows(db.execute("""
        SELECT ci.date, ci.total_heures, ci.heure_debut, ci.heure_fin,
               i.numero AS num_bon, i.type AS type_bon,
               e.designation AS equipement, p.nom AS projet,
               p.numero_projet AS num_projet
        FROM cr_intervenants ci
        JOIN comptes_rendus cr ON ci.cr_id = cr.id
        JOIN interventions i ON cr.intervention_id = i.id
        JOIN equipements e ON i.equipement_id = e.id
        JOIN projets p ON e.projet_id = p.id
        WHERE ci.date BETWEEN ? AND ? AND ci.utilisateur_id = ?
        ORDER BY ci.date
    """, (date_debut, date_fin, technicien_id)))
    # Occupations de la semaine pour ce tech
    # v218.57 : multi-tech via occupation_techniciens
    occ_rows = rows(db.execute("""
        SELECT o.date, o.heure_debut, o.heure_fin, o.total_heures,
               o.notes, o.numero_projet, o.nom_chantier,
               ot.nom AS type_nom
        FROM occupations o
        LEFT JOIN occupation_types ot ON o.type_id = ot.id
        JOIN occupation_techniciens link ON link.occupation_id = o.id
        WHERE o.date BETWEEN ? AND ? AND link.technicien_id = ?
        ORDER BY o.date
    """, (date_debut, date_fin, technicien_id)))
    # Construire une structure : clé (num_projet, nom_projet, num_bon) -> {jour_idx: {heures, code}}
    # jours_idx : 0=lundi, 1=mardi, ..., 6=dimanche
    # Mapping vers colonnes Excel :
    #   L (lundi) = F (heures), G (nuit), H (code)
    #   M (mardi) = I, J, K
    #   M (mercr) = L, M, N
    #   J (jeudi) = O, P, Q
    #   V (vendr) = R, S, T
    #   S (samed) = U, V, W
    #   D (dim)   = X, Y, Z
    jour_cols = [
        ('F','G','H'),  # lundi
        ('I','J','K'),  # mardi
        ('L','M','N'),  # mercredi
        ('O','P','Q'),  # jeudi
        ('R','S','T'),  # vendredi
        ('U','V','W'),  # samedi
        ('X','Y','Z'),  # dimanche
    ]
    # Mapping code occupation → lettre code à utiliser dans la colonne Co
    # Selon la légende du modèle : C=Congé, CX=Congé exceptionnel, E=École, M=Maladie,
    # AT=Accident, F=Férié, AM=Absence motivée, A=Absence non motivée,
    # RF=Récup férié, RC=Récup compteur, FO=Formation
    code_map = {
        "congés": "C", "conges": "C", "congé": "C",
        "formation": "FO",
        "maladie": "M",
        "rtt": "RC",
        "réunion": None, "reunion": None,  # pas de code prévu pour réunion, laisser vide
        "école": "E", "ecole": "E",
        "férié": "F", "ferie": "F",
        "accident": "AT",
    }
    # Regrouper : {(num_projet, nom_projet, num_bon): {jour_idx: (h_jour, h_nuit, code)}}
    groupes = {}
    def _add(num_p, nom_p, num_b, jidx, hj, hn, code):
        k = (num_p or "", nom_p or "", num_b or "")
        if k not in groupes:
            groupes[k] = {}
        existing = groupes[k].get(jidx, (0.0, 0.0, ""))
        new_hj = (existing[0] or 0) + (hj or 0)
        new_hn = (existing[1] or 0) + (hn or 0)
        new_code = code or existing[2]
        groupes[k][jidx] = (new_hj, new_hn, new_code)
    # Interventions : split jour/nuit selon plages configurées
    for r in cr_rows:
        try:
            d = _dt.strptime(r["date"][:10], "%Y-%m-%d").date()
            jidx = (d - mon).days
            if 0 <= jidx <= 6:
                hd = r.get("heure_debut") or ""
                hf = r.get("heure_fin") or ""
                if hd and hf:
                    hj, hn = _split_jour_nuit(hd, hf)
                else:
                    # Pas d'horaires saisis : on met tout en jour par défaut
                    hj, hn = (r.get("total_heures") or 0, 0)
                _add(r.get("num_projet"), r.get("projet"), r.get("num_bon"), jidx, hj, hn, "")
        except Exception:
            continue
    # Occupations : split jour/nuit aussi quand horaires disponibles
    for o in occ_rows:
        try:
            d = _dt.strptime(o["date"][:10], "%Y-%m-%d").date()
            jidx = (d - mon).days
            if 0 <= jidx <= 6:
                t_nom = (o.get("type_nom") or "").strip().lower()
                code = code_map.get(t_nom, "")
                num_p = o.get("numero_projet") or ""
                nom_p = o.get("nom_chantier") or o.get("type_nom") or "Occupation"
                hd = o.get("heure_debut") or ""
                hf = o.get("heure_fin") or ""
                if hd and hf:
                    hj, hn = _split_jour_nuit(hd, hf)
                else:
                    hj, hn = (o.get("total_heures") or 0, 0)
                _add(num_p, nom_p, "/", jidx, hj, hn, code)
        except Exception:
            continue
    # v218.168 : Jours fériés de la semaine pour ce technicien (8h, code "F")
    sid = current_societe_id()
    feries_sem = rows(db.execute(
        "SELECT date, label, heure_debut, heure_fin FROM jours_feries WHERE societe_id=? AND date BETWEEN ? AND ?",
        (sid, date_debut, date_fin)
    ))
    logger.info(f"[FICHE] feries dans la semaine: {len(feries_sem)}")
    for f in feries_sem:
        try:
            d = _dt.strptime(f["date"][:10], "%Y-%m-%d").date()
            jidx = (d - mon).days
            if not (0 <= jidx <= 6): continue
            hd = f.get("heure_debut") or "08:00"
            hf = f.get("heure_fin") or "16:00"
            try:
                sh, sm = [int(x) for x in hd.split(":")]
                eh, em = [int(x) for x in hf.split(":")]
                total = (eh*60+em - sh*60-sm) / 60.0
                if total <= 0: total = 8.0
            except Exception:
                total = 8.0
            libelle = f.get("label") or "Férié"
            # Une ligne dédiée "Férié" : projet/numéro vides, code F, heures dans la colonne du jour
            _add("", libelle, "/", jidx, total, 0, "F")
        except Exception:
            continue
    # Écrire dans le fichier : une ligne par groupe, à partir de la ligne 16, max 17 lignes (16..32)
    current_row = 16
    max_row = 32
    # Matricule formaté pour colonne E (int si numérique, sinon string)
    mat_raw2 = tech.get("matricule") or ""
    try:
        mat_id = int(str(mat_raw2).strip()) if str(mat_raw2).strip() else ""
    except (ValueError, TypeError):
        mat_id = str(mat_raw2).strip()
    for (num_p, nom_p, num_b), jours in sorted(groupes.items()):
        if current_row > max_row:
            break
        ws[f"A{current_row}"] = num_p
        ws[f"B{current_row}"] = nom_p
        ws[f"C{current_row}"] = num_b
        ws[f"E{current_row}"] = mat_id
        for jidx, (hj, hn, code) in jours.items():
            hrs_col, hn_col, co_col = jour_cols[jidx]
            if hj:
                ws[f"{hrs_col}{current_row}"] = round(hj, 2)
            if hn:
                ws[f"{hn_col}{current_row}"] = round(hn, 2)
            if code:
                ws[f"{co_col}{current_row}"] = code
        current_row += 1
    # Les formules SUM(F16:F32) etc + AA33 = total général sont déjà dans le template.
    # Sortie
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    # Nom du fichier
    mois_noms = ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
    mois_nom = mois_noms[mon.month - 1]
    nom_safe = re.sub(r'[^A-Za-z0-9_-]', '_', tech.get("nom","Technicien"))
    fname = f"Pointage_{mois_nom}_S{semaine:02d}_{annee}_{nom_safe}.xlsx"
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp

# ══ PLANNING TECHNICIEN (vue unifiée créneaux + interventions) ══
@app.route("/api/planning/me")
@require_auth
def planning_me():
    """Retourne le planning d'un technicien sur une période.
    Combine les créneaux multi-jours et les interventions à date unique.
    Paramètres : ?debut=YYYY-MM-DD&fin=YYYY-MM-DD&technicien_id=X (admin/manager seulement)

    Un admin voit n'importe quel technicien. Un manager peut voir son équipe
    (les techniciens dont manager_id = son id) + lui-même.
    Un technicien ne peut voir que son propre planning."""
    u = request.user
    db = get_db()
    debut = request.args.get("debut")
    fin = request.args.get("fin")
    if not debut or not fin:
        return jsonify({"error":"debut et fin requis"}), 400
    # Déterminer le technicien cible
    target_id = u["id"]
    req_tid = request.args.get("technicien_id")
    if req_tid:
        req_tid = to_int(req_tid)
        if req_tid and req_tid != u["id"]:
            # Admin : accès à tout le monde
            if u["role"] in ("admin", "superadmin"):
                target_id = req_tid
            # Manager : accès seulement à son équipe
            elif u["role"] == "manager":
                target = one(db.execute("SELECT manager_id FROM utilisateurs WHERE id=?", (req_tid,)))
                if target and target.get("manager_id") == u["id"]:
                    target_id = req_tid
                else:
                    return jsonify({"error":"Technicien non autorisé (pas de votre équipe)"}), 403
            else:
                # Un technicien ne peut pas consulter un autre tech
                return jsonify({"error":"Non autorisé"}), 403
    items = []
    # Créneaux planifiés explicitement pour ce technicien
    # OU créneaux d'interventions dont ce technicien est multi-tech (pour voir le bon même si le créneau est assigné à un collègue)
    creneaux = rows(db.execute("""
        SELECT c.id AS creneau_id, c.date, c.heure_debut, c.heure_fin, c.notes,
               i.id AS intervention_id, i.numero, i.type, i.statut, i.description,
               e.designation AS equip_nom, e.type_technique,
               p.nom AS projet_nom
        FROM intervention_creneaux c
        JOIN interventions i ON c.intervention_id = i.id
        JOIN equipements e ON i.equipement_id = e.id
        JOIN projets p ON e.projet_id = p.id
        WHERE (
               c.technicien_id = ?
               OR (
                  (c.technicien_id IS NULL OR c.technicien_id = 0)
                  AND (i.technicien_id = ? OR EXISTS (
                       SELECT 1 FROM intervention_techniciens it WHERE it.intervention_id=i.id AND it.utilisateur_id=?))
               )
               OR EXISTS (
                  SELECT 1 FROM intervention_techniciens it WHERE it.intervention_id=i.id AND it.utilisateur_id=?
               )
        )
        AND c.date BETWEEN ? AND ?
        AND i.statut != 'ANNULEE'
        ORDER BY c.date, c.heure_debut
    """, (target_id, target_id, target_id, target_id, debut, fin)))
    for c in creneaux:
        c["source"] = "creneau"
        items.append(c)
    # Interventions sans aucun créneau, assignées au technicien sur la période
    ivs = rows(db.execute("""
        SELECT i.id AS intervention_id, i.numero, i.type, i.statut, i.description,
               i.date_prevue AS date, i.heure_prevue AS heure_debut,
               e.designation AS equip_nom, e.type_technique,
               p.nom AS projet_nom
        FROM interventions i
        JOIN equipements e ON i.equipement_id = e.id
        JOIN projets p ON e.projet_id = p.id
        WHERE (i.technicien_id = ? OR EXISTS (
              SELECT 1 FROM intervention_techniciens it WHERE it.intervention_id=i.id AND it.utilisateur_id=?))
          AND i.date_prevue BETWEEN ? AND ?
          AND i.statut != 'ANNULEE'
          AND NOT EXISTS (SELECT 1 FROM intervention_creneaux c WHERE c.intervention_id = i.id)
        ORDER BY i.date_prevue, i.heure_prevue
    """, (target_id, target_id, debut, fin)))
    for iv in ivs:
        iv["source"] = "intervention"
        iv["heure_fin"] = ""
        iv["creneau_id"] = None
        iv["notes"] = ""
        items.append(iv)
    # Si c'est un manager technique consultant SON propre planning :
    # ajouter les interventions liées aux techniques qu'il gère (même si non assignées à lui)
    if target_id == u["id"] and u["role"] == "manager":
        udb = one(db.execute("SELECT techniques FROM utilisateurs WHERE id=?", (u["id"],)))
        tech_csv = (udb and udb.get("techniques")) or ""
        tech_list = [t.strip() for t in tech_csv.split(",") if t.strip()]
        if tech_list:
            placeholders = ",".join(["?"] * len(tech_list))
            # Éviter les doublons : exclure les interventions déjà ajoutées
            existing_iv_ids = {it.get("intervention_id") for it in items if it.get("intervention_id")}
            # Créneaux liés aux techniques (tous, peu importe l'assignation)
            mgr_creneaux = rows(db.execute(f"""
                SELECT c.id AS creneau_id, c.date, c.heure_debut, c.heure_fin, c.notes,
                       i.id AS intervention_id, i.numero, i.type, i.statut, i.description,
                       e.designation AS equip_nom, e.type_technique,
                       p.nom AS projet_nom
                FROM intervention_creneaux c
                JOIN interventions i ON c.intervention_id = i.id
                JOIN equipements e ON i.equipement_id = e.id
                JOIN projets p ON e.projet_id = p.id
                WHERE e.type_technique IN ({placeholders})
                  AND c.date BETWEEN ? AND ?
                  AND i.statut != 'ANNULEE'
                ORDER BY c.date, c.heure_debut
            """, tech_list + [debut, fin]))
            for c in mgr_creneaux:
                # Éviter doublon avec les créneaux déjà ajoutés (via creneau_id)
                if not any(it.get("source")=="creneau" and it.get("creneau_id")==c["creneau_id"] for it in items):
                    c["source"] = "creneau"
                    items.append(c)
            # Interventions sans créneau, liées aux techniques
            mgr_ivs = rows(db.execute(f"""
                SELECT i.id AS intervention_id, i.numero, i.type, i.statut, i.description,
                       i.date_prevue AS date, i.heure_prevue AS heure_debut,
                       e.designation AS equip_nom, e.type_technique,
                       p.nom AS projet_nom
                FROM interventions i
                JOIN equipements e ON i.equipement_id = e.id
                JOIN projets p ON e.projet_id = p.id
                WHERE e.type_technique IN ({placeholders})
                  AND i.date_prevue BETWEEN ? AND ?
                  AND i.statut != 'ANNULEE'
                  AND NOT EXISTS (SELECT 1 FROM intervention_creneaux cc WHERE cc.intervention_id = i.id)
                ORDER BY i.date_prevue, i.heure_prevue
            """, tech_list + [debut, fin]))
            for iv in mgr_ivs:
                if iv["intervention_id"] not in existing_iv_ids:
                    iv["source"] = "intervention"
                    iv["heure_fin"] = ""
                    iv["creneau_id"] = None
                    iv["notes"] = ""
                    items.append(iv)
    # Occupations (congés, formation, etc.) du technicien cible
    # v218.52 : utilise la table de liaison occupation_techniciens (multi-tech)
    occs = rows(db.execute("""
        SELECT o.id AS occupation_id, o.date, o.heure_debut, o.heure_fin, o.notes,
               o.type_id, ot.nom AS type_occupation, ot.couleur AS type_couleur,
               o.total_heures, o.numero_projet, o.nom_chantier,
               o.technicien_id AS owner_tech_id, u_owner.nom AS owner_tech_nom,
               o.id AS id
        FROM occupations o
        LEFT JOIN occupation_types ot ON o.type_id = ot.id
        LEFT JOIN utilisateurs u_owner ON o.technicien_id = u_owner.id
        WHERE EXISTS (SELECT 1 FROM occupation_techniciens link WHERE link.occupation_id=o.id AND link.technicien_id=?)
        AND o.date BETWEEN ? AND ?
        ORDER BY o.date, o.heure_debut
    """, (target_id, debut, fin)))
    # Résoudre les accompagnants en noms pour l'affichage
    _resolve_accompagnants(db, occs)
    for o in occs:
        o["source"] = "occupation"
        # v218.49 : indiquer si l'utilisateur cible est accompagnant ou créateur
        o["is_accompagnant"] = (o.get("owner_tech_id") != target_id)
        # Pour cohérence avec les clés attendues par le mobile
        o["intervention_id"] = None
        o["numero"] = ""
        o["type"] = "OCCUPATION"
        o["statut"] = ""
        o["description"] = o.get("notes") or ""
        o["equip_nom"] = o.get("type_occupation") or "Occupation"
        o["projet_nom"] = ""
        items.append(o)

    # ═══════════════════════════════════════════════════════════════════
    # REMPLACEMENT DES BONS PAR LEURS CRs
    # Pour chaque intervention (MAINTENANCE ou DEPANNAGE) ayant au moins un CR,
    # on retire les items issus du bon et on les remplace par des pseudo-items
    # représentant les CRs (un par CR, à la date/heure du CR).
    # Règle :
    #   - MAINTENANCE sans CR → on garde le bon (affiché comme avant)
    #   - MAINTENANCE avec CR(s) → on remplace par les CR(s)
    #   - DEPANNAGE sans CR → on garde le bon
    #   - DEPANNAGE avec CR(s) → on remplace par les CR(s)
    # On charge les CRs pour TOUTES les interventions liées aux items du planning,
    # MAIS aussi pour toutes les interventions qui ont un CR dans la plage de dates affichée
    # (permet de capter les CRs dont la date diffère de la date_prevue du bon parent).
    # ═══════════════════════════════════════════════════════════════════
    try:
        # 1) Interventions déjà présentes dans les items
        intervention_ids = set(it.get("intervention_id") for it in items if it.get("intervention_id"))
        logger.info(f"[planning_me] target_id={target_id} debut={debut} fin={fin} items_before={len(items)} iv_ids_from_items={intervention_ids}")
        # 2) Interventions dont un CR tombe dans la fenêtre debut/fin
        # On regarde soit la date_intervention du CR, soit la date spécifique de l'intervenant
        # (qui peut différer en cas de CR multi-jours ou multi-intervenants)
        crs_in_range = rows(db.execute(
            """SELECT DISTINCT cr.intervention_id
               FROM comptes_rendus cr
               JOIN cr_intervenants ci ON ci.cr_id = cr.id
               WHERE ci.utilisateur_id = ?
                 AND (
                      (cr.date_intervention BETWEEN ? AND ?)
                   OR (NULLIF(ci.date,'') BETWEEN ? AND ?)
                 )""",
            (target_id, debut, fin, debut, fin)
        ))
        logger.info(f"[planning_me] crs_in_range={crs_in_range}")
        for row in crs_in_range:
            if row.get("intervention_id"):
                intervention_ids.add(row["intervention_id"])
        logger.info(f"[planning_me] intervention_ids_final={intervention_ids}")

        if intervention_ids:
            iid_list = list(intervention_ids)
            placeholders = ",".join(["?"] * len(iid_list))
            all_crs = rows(db.execute(
                f"SELECT id, intervention_id, numero, date_intervention, total_heures FROM comptes_rendus WHERE intervention_id IN ({placeholders}) ORDER BY date_intervention, id",
                iid_list
            ))
            crs_by_iv = {}
            for cr in all_crs:
                crs_by_iv.setdefault(cr["intervention_id"], []).append(cr)
            # Charger horaires + intervenants pour chaque CR
            # + date spécifique du technicien cible (table cr_intervenants a sa propre date)
            for cr in all_crs:
                try:
                    horaires = one(db.execute(
                        """SELECT MIN(NULLIF(heure_debut,'')) AS h_min, MAX(NULLIF(heure_fin,'')) AS h_max
                           FROM cr_intervenants WHERE cr_id=?""", (cr["id"],)))
                    cr["heure_debut"] = (horaires or {}).get("h_min") or ""
                    cr["heure_fin"] = (horaires or {}).get("h_max") or ""
                except Exception:
                    cr["heure_debut"] = ""
                    cr["heure_fin"] = ""
                try:
                    tech_rows = rows(db.execute(
                        "SELECT DISTINCT utilisateur_id FROM cr_intervenants WHERE cr_id=? AND utilisateur_id IS NOT NULL",
                        (cr["id"],)))
                    cr["technicien_ids"] = [t["utilisateur_id"] for t in tech_rows if t.get("utilisateur_id")]
                except Exception:
                    cr["technicien_ids"] = []
                # Date + horaires + heures personnelles spécifiques pour ce technicien cible
                # (part individuelle, pas le total du CR)
                try:
                    my_row = one(db.execute(
                        """SELECT date, heure_debut, heure_fin, total_heures FROM cr_intervenants
                           WHERE cr_id=? AND utilisateur_id=?
                           ORDER BY id LIMIT 1""",
                        (cr["id"], target_id)))
                    if my_row:
                        my_date = (my_row.get("date") or "").strip()
                        my_hd = (my_row.get("heure_debut") or "").strip()
                        my_hf = (my_row.get("heure_fin") or "").strip()
                        if my_date:
                            cr["date_for_target"] = my_date[:10]
                        if my_hd:
                            cr["heure_debut"] = my_hd
                        if my_hf:
                            cr["heure_fin"] = my_hf
                        # Part personnelle d'heures pour ce technicien
                        my_h = my_row.get("total_heures")
                        if my_h is not None:
                            try:
                                cr["heures_pour_target"] = float(my_h)
                            except Exception:
                                cr["heures_pour_target"] = None
                except Exception:
                    pass
                # Fallback : date_intervention du CR si pas de date intervenant spécifique
                if not cr.get("date_for_target"):
                    cr["date_for_target"] = (cr.get("date_intervention") or "")[:10]
                logger.info(f"[planning_me] CR {cr.get('numero')} (id={cr['id']}) iv={cr['intervention_id']} date_cr={cr.get('date_intervention')} date_target={cr.get('date_for_target')} hd={cr.get('heure_debut')} hf={cr.get('heure_fin')} techs={cr.get('technicien_ids')}")
            ivs_with_crs = set(crs_by_iv.keys())
            # Indexer métadonnées bons parents — d'abord via items, sinon requête directe
            iv_meta = {}
            for it in items:
                iid = it.get("intervention_id")
                if iid and iid in ivs_with_crs and iid not in iv_meta:
                    iv_meta[iid] = {
                        "numero": it.get("numero"),
                        "type": it.get("type"),
                        "statut": it.get("statut"),
                        "description": it.get("description"),
                        "equip_nom": it.get("equip_nom"),
                        "type_technique": it.get("type_technique"),
                        "projet_nom": it.get("projet_nom"),
                    }
            # Pour les interventions découvertes via les CRs mais pas dans items → requête directe
            missing_meta = [iid for iid in ivs_with_crs if iid not in iv_meta]
            if missing_meta:
                mm_ph = ",".join(["?"] * len(missing_meta))
                missing_rows = rows(db.execute(
                    f"""SELECT i.id AS intervention_id, i.numero, i.type, i.statut, i.description,
                               e.designation AS equip_nom, e.type_technique,
                               p.nom AS projet_nom
                       FROM interventions i
                       JOIN equipements e ON i.equipement_id = e.id
                       JOIN projets p ON e.projet_id = p.id
                       WHERE i.id IN ({mm_ph})""",
                    missing_meta
                ))
                for r in missing_rows:
                    iv_meta[r["intervention_id"]] = {
                        "numero": r.get("numero"),
                        "type": r.get("type"),
                        "statut": r.get("statut"),
                        "description": r.get("description"),
                        "equip_nom": r.get("equip_nom"),
                        "type_technique": r.get("type_technique"),
                        "projet_nom": r.get("projet_nom"),
                    }
            # Retirer tous les items dont l'intervention a au moins un CR
            items = [it for it in items if not (it.get("intervention_id") in ivs_with_crs)]
            # Ajouter les pseudo-items CR (uniquement ceux où le technicien cible est intervenant)
            for iid, crs in crs_by_iv.items():
                meta = iv_meta.get(iid, {})
                for cr in crs:
                    cr_date = cr.get("date_for_target") or (cr.get("date_intervention") or "")[:10]
                    # Filtrer : tech cible intervenant ET date dans la fenêtre
                    if target_id not in cr.get("technicien_ids", []):
                        continue
                    if cr_date < debut or cr_date > fin:
                        continue
                    # Part personnelle si disponible, sinon total CR (fallback ascendant)
                    cr_total_perso = cr.get("heures_pour_target")
                    cr_total_value = cr_total_perso if cr_total_perso is not None else (cr.get("total_heures") or 0)
                    items.append({
                        "source": "cr",
                        "cr_id": cr["id"],
                        "cr_numero": cr.get("numero") or "",
                        "intervention_id": iid,
                        "creneau_id": None,
                        "date": cr_date,
                        "heure_debut": cr.get("heure_debut") or "",
                        "heure_fin": cr.get("heure_fin") or "",
                        "total_heures": cr_total_value,
                        "notes": "",
                        "numero": meta.get("numero") or "",
                        "type": meta.get("type") or "",
                        "statut": meta.get("statut") or "",
                        "description": meta.get("description") or "",
                        "equip_nom": meta.get("equip_nom") or "",
                        "type_technique": meta.get("type_technique") or "",
                        "projet_nom": meta.get("projet_nom") or "",
                    })
    except Exception as ex:
        logger.exception(f"[planning_me] Erreur transformation CR : {ex}")

    items.sort(key=lambda x: (x.get("date") or "", x.get("heure_debut") or ""))
    return jsonify(items)

# ══ COMPTES RENDUS ══
@app.route("/api/comptes_rendus/<int:iid>")
@require_auth
def get_comptes_rendus(iid):
    db=get_db(); crs=rows(db.execute("SELECT * FROM comptes_rendus WHERE intervention_id=? ORDER BY created_at",(iid,)))
    for cr in crs:
        cr["releves"]=rows(db.execute("SELECT r.*,c.nom AS compteur_nom,c.numero,c.unite,c.localisation,ct.nom AS type_nom FROM releves_compteurs r JOIN compteurs c ON r.compteur_id=c.id LEFT JOIN compteur_types ct ON c.type_id=ct.id WHERE r.cr_id=? ORDER BY c.nom",(cr["id"],)))
    for cr in crs:
        try:
            cr["intervenants"]=rows(db.execute("""SELECT ci.*,COALESCE(u.nom,ci.nom,'') AS nom
                FROM cr_intervenants ci LEFT JOIN utilisateurs u ON ci.utilisateur_id=u.id WHERE ci.cr_id=?""",(cr["id"],)))
        except: cr["intervenants"]=[]
        try:
            cr["materiels"]=rows(db.execute("SELECT * FROM cr_materiels WHERE cr_id=? ORDER BY ordre,id",(cr["id"],)))
        except: cr["materiels"]=[]
        try:
            cr["photos"]=rows(db.execute("SELECT id,filename,mime_type,size,created_at FROM cr_photos WHERE cr_id=? ORDER BY id",(cr["id"],)))
        except: cr["photos"]=[]
    return jsonify(crs)

@app.route("/api/comptes_rendus/<int:iid>",methods=["POST"])
@require_auth
def create_compte_rendu(iid):
    d=request.json or {}; db=get_db()
    sid = current_societe_id()
    # v218.105 : normaliser les total_heures côté serveur (recalcul si heures saisies)
    _normalized_intervenants = []
    for iv in d.get("intervenants",[]):
        hd = iv.get("heure_debut","") or ""
        hf = iv.get("heure_fin","") or ""
        if hd and hf:
            th_iv = _calc_heures(hd, hf)
        else:
            try: th_iv = float(iv.get("total_heures") or 0)
            except: th_iv = 0
        _normalized_intervenants.append({
            "utilisateur_id": iv.get("utilisateur_id"),
            "nom": iv.get("nom",""),
            "date": iv.get("date",""),
            "heure_debut": hd,
            "heure_fin": hf,
            "total_heures": th_iv,
        })
    th = sum(float(iv.get("total_heures",0)) for iv in _normalized_intervenants)
    cr_num = next_numero(db,"CR","comptes_rendus","numero") if "numero" in [r[1] for r in db.execute("PRAGMA table_info(comptes_rendus)").fetchall()] else None
    if cr_num:
        db.execute("INSERT INTO comptes_rendus (intervention_id,date_intervention,observations,actions_realisees,mesures,recommandations,conclusion,total_heures,numero,societe_id) VALUES (?,?,?,?,?,?,?,?,?,?)",
                   (iid,d.get("date_intervention") or None,d.get("observations",""),d.get("actions_realisees",""),d.get("mesures",""),d.get("recommandations",""),d.get("conclusion",""),th,cr_num,sid))
    else:
        db.execute("INSERT INTO comptes_rendus (intervention_id,date_intervention,observations,actions_realisees,mesures,recommandations,conclusion,total_heures,societe_id) VALUES (?,?,?,?,?,?,?,?,?)",
                   (iid,d.get("date_intervention") or None,d.get("observations",""),d.get("actions_realisees",""),d.get("mesures",""),d.get("recommandations",""),d.get("conclusion",""),th,sid))
    db.commit(); cr_id=db.execute("SELECT last_insert_rowid()").fetchone()[0]
    # Mettre à jour nb_deplacements si fourni (colonne ajoutée via migration)
    try:
        nbd = int(d.get("nb_deplacements") or 0)
        if nbd > 0:
            db.execute("UPDATE comptes_rendus SET nb_deplacements=? WHERE id=?", (nbd, cr_id))
            db.commit()
    except Exception: pass
    # Stocker intervenants dans cr_intervenants si table existe
    try:
        for iv in _normalized_intervenants:
            db.execute("INSERT INTO cr_intervenants (cr_id,utilisateur_id,nom,date,heure_debut,heure_fin,total_heures) VALUES (?,?,?,?,?,?,?)",
                       (cr_id,to_int(iv.get("utilisateur_id")),iv.get("nom",""),iv.get("date",""),iv.get("heure_debut",""),iv.get("heure_fin",""),iv.get("total_heures",0)))
        db.commit()
    except Exception: pass
    # Stocker matériels
    try:
        for idx, m in enumerate(d.get("materiels",[])):
            des = (m.get("designation","") or "").strip()
            if not des: continue
            qte = float(m.get("quantite") or 0)
            db.execute("INSERT INTO cr_materiels (cr_id,designation,quantite,ordre) VALUES (?,?,?,?)",
                       (cr_id, des, qte, idx))
        db.commit()
    except Exception: pass
    return jsonify({"id":cr_id}),201

@app.route("/api/comptes_rendus/cr/<int:cr_id>",methods=["PATCH"])
@require_role("admin","manager")
def update_compte_rendu(cr_id):
    d=request.json or {}; db=get_db(); sets,params=[],[]
    for f in ["date_intervention","observations","actions_realisees","mesures","recommandations","conclusion","total_heures","nb_deplacements"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    if sets: params.append(cr_id); db.execute(f"UPDATE comptes_rendus SET {chr(44).join(sets)} WHERE id=?",params)
    if "intervenants" in d:
        db.execute("DELETE FROM cr_intervenants WHERE cr_id=?",(cr_id,))
        _total_cr = 0.0
        for iv in d["intervenants"]:
            hd = iv.get("heure_debut","") or ""
            hf = iv.get("heure_fin","") or ""
            # v218.105 : recalcul côté serveur du total_heures si début et fin sont remplis,
            # pour ne jamais avoir un total désynchronisé des horaires.
            if hd and hf:
                th = _calc_heures(hd, hf)
            else:
                try: th = float(iv.get("total_heures") or 0)
                except: th = 0
            _total_cr += th
            db.execute("INSERT INTO cr_intervenants (cr_id,utilisateur_id,nom,date,heure_debut,heure_fin,total_heures) VALUES (?,?,?,?,?,?,?)",
                       (cr_id,to_int(iv.get("utilisateur_id")),iv.get("nom",""),iv.get("date",""),hd,hf,th))
        # Mettre à jour aussi le total du CR (somme des intervenants)
        db.execute("UPDATE comptes_rendus SET total_heures=? WHERE id=?", (_total_cr, cr_id))
    if "materiels" in d:
        db.execute("DELETE FROM cr_materiels WHERE cr_id=?",(cr_id,))
        for idx, m in enumerate(d["materiels"]):
            des = (m.get("designation","") or "").strip()
            if not des: continue
            qte = float(m.get("quantite") or 0)
            db.execute("INSERT INTO cr_materiels (cr_id,designation,quantite,ordre) VALUES (?,?,?,?)",
                       (cr_id, des, qte, idx))
    db.commit(); return jsonify({"ok":True})

@app.route("/api/comptes_rendus/cr/<int:cr_id>",methods=["DELETE"])
@require_role("admin","manager")
def delete_compte_rendu(cr_id):
    db=get_db(); db.execute("DELETE FROM cr_intervenants WHERE cr_id=?",(cr_id,))
    db.execute("DELETE FROM comptes_rendus WHERE id=?",(cr_id,)); db.commit()
    return jsonify({"ok":True})

# ══ PHOTOS CR ══
@app.route("/api/comptes_rendus/cr/<int:cr_id>/photos", methods=["GET"])
@require_auth
def list_cr_photos(cr_id):
    db = get_db()
    photos = rows(db.execute("SELECT id,filename,mime_type,size,created_at FROM cr_photos WHERE cr_id=? ORDER BY id", (cr_id,)))
    return jsonify(photos)

@app.route("/api/comptes_rendus/cr/<int:cr_id>/photos", methods=["POST"])
@require_auth
def upload_cr_photo(cr_id):
    """Upload une photo au format base64. Body JSON: {filename, mime_type, data_base64}"""
    import base64
    d = request.json or {}
    data_b64 = d.get("data_base64", "")
    if not data_b64:
        return jsonify({"error": "data_base64 manquant"}), 400
    # Retirer préfixe data:image/...;base64, si présent
    if "," in data_b64:
        data_b64 = data_b64.split(",", 1)[1]
    try:
        blob = base64.b64decode(data_b64)
    except Exception as e:
        return jsonify({"error": f"base64 invalide: {e}"}), 400
    if len(blob) > 8 * 1024 * 1024:
        return jsonify({"error": "Photo trop volumineuse (max 8 Mo)"}), 400
    db = get_db()
    # Vérif CR existe
    if not db.execute("SELECT 1 FROM comptes_rendus WHERE id=?", (cr_id,)).fetchone():
        return jsonify({"error": "CR introuvable"}), 404
    fn = (d.get("filename") or "photo.jpg")[:200]
    mt = (d.get("mime_type") or "image/jpeg")[:80]
    db.execute("INSERT INTO cr_photos (cr_id,filename,mime_type,data,size) VALUES (?,?,?,?,?)",
               (cr_id, fn, mt, blob, len(blob)))
    db.commit()
    pid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    return jsonify({"id": pid, "size": len(blob)}), 201

@app.route("/api/cr_photos/<int:photo_id>", methods=["GET"])
@require_auth
def get_cr_photo(photo_id):
    from flask import make_response
    db = get_db()
    row = db.execute("SELECT mime_type,data FROM cr_photos WHERE id=?", (photo_id,)).fetchone()
    if not row:
        return jsonify({"error": "Photo introuvable"}), 404
    resp = make_response(bytes(row["data"]))
    resp.headers["Content-Type"] = row["mime_type"] or "image/jpeg"
    resp.headers["Cache-Control"] = "private, max-age=3600"
    return resp

@app.route("/api/cr_photos/<int:photo_id>", methods=["DELETE"])
@require_auth
def delete_cr_photo(photo_id):
    db = get_db()
    db.execute("DELETE FROM cr_photos WHERE id=?", (photo_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/comptes_rendus/cr/<int:cr_id>/debug")
@require_auth
def debug_cr(cr_id):
    """v218.101 - Debug endpoint : retourne tout ce qu'on sait sur un CR (intervenants détaillés).
    Utile pour diagnostiquer pourquoi un tech n'apparaît pas sur son planning."""
    db = get_db()
    cr = one(db.execute("SELECT id, intervention_id, date_intervention, total_heures FROM comptes_rendus WHERE id=?", (cr_id,)))
    if not cr:
        return jsonify({"error": "CR introuvable"}), 404
    # Tous les intervenants avec leurs détails
    intervenants = rows(db.execute(
        """SELECT ci.id, ci.utilisateur_id, ci.nom, ci.date, ci.heure_debut, ci.heure_fin, ci.total_heures,
                  u.nom AS user_nom, u.email AS user_email, u.role AS user_role
           FROM cr_intervenants ci
           LEFT JOIN utilisateurs u ON u.id = ci.utilisateur_id
           WHERE ci.cr_id = ?""", (cr_id,)
    ))
    # Le bon associé
    iv = one(db.execute("SELECT id, numero, type, technicien_id, equipe_id, societe_id, date_prevue FROM interventions WHERE id=?", (cr["intervention_id"],)))
    # Les techs liés à l'intervention via intervention_techniciens
    iv_techs = rows(db.execute(
        """SELECT it.utilisateur_id, u.nom FROM intervention_techniciens it
           LEFT JOIN utilisateurs u ON u.id = it.utilisateur_id
           WHERE it.intervention_id = ?""", (cr["intervention_id"],)
    ))
    return jsonify({
        "cr": cr,
        "intervention": iv,
        "intervenants": intervenants,  # ← chaque entrée doit avoir un utilisateur_id non-null
        "intervention_techniciens": iv_techs,
    })


# ══ TECHNIQUES ══
@app.route("/api/techniques")
@require_auth
def get_techniques():
    return jsonify(rows(get_db().execute("SELECT * FROM techniques ORDER BY nom")))

@app.route("/api/techniques",methods=["POST"])
@require_role("admin","manager")
def create_technique():
    d=request.json or {}
    if not d.get("nom"): return jsonify({"error":"nom requis"}),400
    db=get_db()
    try:
        db.execute("INSERT INTO techniques (nom,description) VALUES (?,?)",(d["nom"],d.get("description","")))
        db.commit(); return jsonify({"id":db.execute("SELECT last_insert_rowid()").fetchone()[0]}),201
    except Exception as e: return jsonify({"error":str(e)}),400

@app.route("/api/techniques/<int:tid>",methods=["PATCH"])
@require_role("admin","manager")
def update_technique(tid):
    d=request.json or {}; db=get_db(); sets,params=[],[]
    # v215.2 : récupérer l'ancien nom AVANT update pour propager dans equipements.type_technique
    old_row = one(db.execute("SELECT nom FROM techniques WHERE id=?", (tid,)))
    old_nom = (old_row or {}).get("nom") if old_row else None
    for f in ["nom","description"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    if not sets: return jsonify({"error":"Rien"}),400
    params.append(tid); db.execute(f"UPDATE techniques SET {chr(44).join(sets)} WHERE id=?",params)
    # v215.2 : si le nom change, propager dans tous les équipements qui utilisent cette technique
    new_nom = d.get("nom")
    if new_nom and old_nom and new_nom != old_nom:
        # Cas 1 : équipements liés via technique_id → on met à jour type_technique avec le nouveau nom
        db.execute("UPDATE equipements SET type_technique=? WHERE technique_id=?", (new_nom, tid))
        # Cas 2 : équipements legacy avec type_technique=ancien_nom mais technique_id NULL → on met à jour le nom et on fixe technique_id
        db.execute("UPDATE equipements SET type_technique=?, technique_id=? WHERE technique_id IS NULL AND type_technique=?",
                   (new_nom, tid, old_nom))
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/techniques/<int:tid>",methods=["DELETE"])
@require_role("admin","manager")
def delete_technique(tid):
    db=get_db(); db.execute("DELETE FROM techniques WHERE id=?",(tid,)); db.commit()
    return jsonify({"ok":True})

# ══════════════════════════════════════════════════════════════════════
# v217 : BIBLIOTHÈQUE D'IMAGES
# ══════════════════════════════════════════════════════════════════════
IMAGES_DIR = BASE_DIR / "uploads" / "images"
IMAGES_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_IMAGE_EXT = {"jpg", "jpeg", "png", "webp", "gif"}
MAX_IMAGE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB upload max

def _resize_and_save_image(file_storage, target_path):
    """Redimensionne (max 1600x1200) et sauvegarde en JPEG q=85 pour économiser de l'espace.
    Retourne le filename final ou None en cas d'erreur."""
    if not HAS_PILLOW:
        # Fallback : sauvegarder tel quel sans redim
        file_storage.save(str(target_path))
        return target_path.name
    try:
        img = PILImage.open(file_storage.stream)
        # Convertir RGBA/P/etc en RGB pour JPEG
        if img.mode in ("RGBA", "P", "LA"):
            background = PILImage.new("RGB", img.size, (255, 255, 255))
            if img.mode == "RGBA":
                background.paste(img, mask=img.split()[3])
            else:
                background.paste(img.convert("RGB"))
            img = background
        elif img.mode != "RGB":
            img = img.convert("RGB")
        # Redim si plus grand que 1600x1200
        img.thumbnail((1600, 1200), PILImage.LANCZOS)
        img.save(str(target_path), "JPEG", quality=85, optimize=True)
        return target_path.name
    except Exception as e:
        app.logger.error(f"Image resize failed: {e}")
        return None

@app.route("/api/image_library", methods=["GET"])
@require_auth
def list_image_library():
    db = get_db()
    rs = rows(db.execute("""
        SELECT il.id, il.nom, il.filename, il.created_at,
               u.nom AS created_by_nom
        FROM image_library il
        LEFT JOIN utilisateurs u ON il.created_by = u.id
        ORDER BY il.nom COLLATE NOCASE
    """))
    return jsonify(rs)

@app.route("/api/image_library", methods=["POST"])
@require_role("admin", "manager")
def create_image_library():
    """Upload nouvelle image. multipart/form-data avec champ 'image' + 'nom'."""
    nom = (request.form.get("nom") or "").strip()
    if not nom:
        return jsonify({"error": "Nom requis"}), 400
    if "image" not in request.files:
        return jsonify({"error": "Aucun fichier"}), 400
    f = request.files["image"]
    if not f.filename:
        return jsonify({"error": "Fichier vide"}), 400
    # Vérif extension
    ext = (f.filename.rsplit(".", 1)[-1] if "." in f.filename else "").lower()
    if ext not in ALLOWED_IMAGE_EXT:
        return jsonify({"error": f"Format non supporté ({', '.join(ALLOWED_IMAGE_EXT)})"}), 400
    # Vérif taille
    f.stream.seek(0, os.SEEK_END)
    sz = f.stream.tell()
    f.stream.seek(0)
    if sz > MAX_IMAGE_SIZE_BYTES:
        return jsonify({"error": f"Fichier trop gros ({sz/1024/1024:.1f} MB, max {MAX_IMAGE_SIZE_BYTES/1024/1024:.0f} MB)"}), 400

    db = get_db()
    # Insert d'abord pour obtenir l'id
    cur = db.execute("INSERT INTO image_library (nom, filename, created_by) VALUES (?, ?, ?)",
                     (nom, "", request.user["id"]))
    img_id = cur.lastrowid
    # Filename basé sur l'id (toujours .jpg car on convertit en JPEG)
    target_filename = f"{img_id}.jpg"
    target_path = IMAGES_DIR / target_filename
    saved = _resize_and_save_image(f, target_path)
    if not saved:
        # Rollback
        db.execute("DELETE FROM image_library WHERE id=?", (img_id,))
        db.commit()
        return jsonify({"error": "Échec traitement image"}), 500
    db.execute("UPDATE image_library SET filename=? WHERE id=?", (target_filename, img_id))
    db.commit()
    return jsonify({"ok": True, "id": img_id, "filename": target_filename})

@app.route("/api/image_library/<int:img_id>", methods=["PATCH"])
@require_role("admin", "manager")
def rename_image_library(img_id):
    d = request.json or {}
    nom = (d.get("nom") or "").strip()
    if not nom:
        return jsonify({"error": "Nom requis"}), 400
    db = get_db()
    db.execute("UPDATE image_library SET nom=? WHERE id=?", (nom, img_id))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/image_library/<int:img_id>", methods=["DELETE"])
@require_role("admin", "manager")
def delete_image_library(img_id):
    db = get_db()
    row = one(db.execute("SELECT filename FROM image_library WHERE id=?", (img_id,)))
    if not row:
        return jsonify({"error": "Image introuvable"}), 404
    # Supprimer fichier physique
    try:
        fpath = IMAGES_DIR / row["filename"]
        if fpath.exists():
            fpath.unlink()
    except Exception as e:
        app.logger.warning(f"Delete image file failed: {e}")
    # Supprimer entrée DB (les FK SET NULL sur equipements.image_id se feront automatiquement)
    db.execute("DELETE FROM image_library WHERE id=?", (img_id,))
    db.commit()
    return jsonify({"ok": True})

@app.route("/uploads/images/<path:filename>")
def serve_uploaded_image(filename):
    """Sert les images de la bibliothèque."""
    if ".." in filename or filename.startswith("/"):
        return "", 404
    f = IMAGES_DIR / filename
    if not f.exists() or not f.is_file():
        return "", 404
    resp = send_file(str(f))
    # Cache modéré (les images peuvent être remplacées)
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp

# ══════════════════════════════════════════════════════════════════════
# v218 : STATUTS D'ÉQUIPEMENT (configurables)
# ══════════════════════════════════════════════════════════════════════
@app.route("/api/equipement_statuts", methods=["GET"])
@require_auth
def list_equipement_statuts():
    db = get_db()
    rs = rows(db.execute("SELECT id, code, label, couleur, ordre FROM equipement_statuts ORDER BY ordre, label"))
    return jsonify(rs)

@app.route("/api/equipement_statuts", methods=["POST"])
@require_role("admin", "manager")
def create_equipement_statut():
    d = request.json or {}
    code = (d.get("code") or "").strip().upper().replace(" ", "_")
    label = (d.get("label") or "").strip()
    couleur = (d.get("couleur") or "#10B981").strip()
    ordre = int(d.get("ordre") or 0)
    if not code or not label:
        return jsonify({"error": "code et label requis"}), 400
    db = get_db()
    # Vérifier que le code n'existe pas déjà
    existing = one(db.execute("SELECT id FROM equipement_statuts WHERE code=?", (code,)))
    if existing:
        return jsonify({"error": "Ce code existe déjà"}), 400
    cur = db.execute("INSERT INTO equipement_statuts (code, label, couleur, ordre) VALUES (?, ?, ?, ?)",
                     (code, label, couleur, ordre))
    db.commit()
    log_action(request.user, "CREATE", "equipement_statut", cur.lastrowid, label)
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route("/api/equipement_statuts/<int:sid>", methods=["PATCH"])
@require_role("admin", "manager")
def update_equipement_statut(sid):
    d = request.json or {}
    db = get_db()
    existing = one(db.execute("SELECT * FROM equipement_statuts WHERE id=?", (sid,)))
    if not existing:
        return jsonify({"error": "Introuvable"}), 404
    fields = []
    params = []
    if "label" in d:
        fields.append("label=?"); params.append((d["label"] or "").strip())
    if "couleur" in d:
        fields.append("couleur=?"); params.append((d["couleur"] or "#10B981").strip())
    if "ordre" in d:
        fields.append("ordre=?"); params.append(int(d["ordre"] or 0))
    # Le code n'est pas modifiable (clé fonctionnelle utilisée dans equipements.statut)
    if not fields:
        return jsonify({"ok": True})
    params.append(sid)
    db.execute(f"UPDATE equipement_statuts SET {','.join(fields)} WHERE id=?", params)
    db.commit()
    log_action(request.user, "UPDATE", "equipement_statut", sid, existing["label"])
    return jsonify({"ok": True})

@app.route("/api/equipement_statuts/<int:sid>", methods=["DELETE"])
@require_role("admin")
def delete_equipement_statut(sid):
    db = get_db()
    existing = one(db.execute("SELECT * FROM equipement_statuts WHERE id=?", (sid,)))
    if not existing:
        return jsonify({"error": "Introuvable"}), 404
    # Vérifier qu'aucun équipement n'utilise ce statut
    cnt = one(db.execute("SELECT COUNT(*) AS n FROM equipements WHERE statut=?", (existing["code"],)))
    if cnt and cnt["n"]:
        return jsonify({"error": f"Impossible : {cnt['n']} équipement(s) utilisent ce statut. Modifie-les d'abord."}), 400
    db.execute("DELETE FROM equipement_statuts WHERE id=?", (sid,))
    db.commit()
    log_action(request.user, "DELETE", "equipement_statut", sid, existing["label"])
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════
# CHAMPS PERSONNALISÉS sur les équipements (v218.117)
# Catalogue géré par les admins. Chaque champ est soit "global" (s'applique à
# tous les équipements de la société), soit lié à une technique précise.
# Types supportés : text, number, date, boolean, select, textarea, file
# ══════════════════════════════════════════════════════════════════════

_CHAMP_TYPES_VALIDES = {"text", "number", "date", "boolean", "select", "textarea", "file"}

def _slugify_champ(label):
    """Convertit un label en slug technique (a-z0-9_)."""
    import unicodedata
    s = unicodedata.normalize('NFKD', str(label or '')).encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r'[^a-zA-Z0-9]+', '_', s).strip('_').lower()
    return s or 'champ'


# v218.121 : Champs système = champs en dur dans la table equipements, exposés
# via l'API equipement-champs pour s'afficher dans la section "Champs personnalisés"
# du formulaire. Pas de stockage dans equipement_valeurs_custom — ils restent
# dans les colonnes natives. is_system=True pour empêcher modif/suppression.
# label_technique = nom EXACT de la technique (sensible à la casse, attendu dans techniques.nom)
_EQ_SYSTEM_CHAMPS = [
    # ─── UPS ───
    {"col": "marque", "label": "Marque", "type": "text", "label_technique": "UPS", "ordre": 1000},
    {"col": "modele", "label": "Modèle", "type": "text", "label_technique": "UPS", "ordre": 1001},
    {"col": "numero_serie", "label": "N° de série", "type": "text", "label_technique": "UPS", "ordre": 1002},
    {"col": "puissance", "label": "Puissance", "type": "text", "label_technique": "UPS", "ordre": 1003},
    {"col": "in_out", "label": "In/Out", "type": "text", "label_technique": "UPS", "ordre": 1004},
    {"col": "date_mise_en_service", "label": "Mise en service", "type": "date", "label_technique": "UPS", "ordre": 1005},
    # ─── Transformateur (Haute tension) ───
    {"col": "trafo_marque", "label": "Marque (transfo)", "type": "text", "label_technique": "Haute tension", "ordre": 2000},
    {"col": "trafo_annee", "label": "Année (transfo)", "type": "text", "label_technique": "Haute tension", "ordre": 2001},
    {"col": "trafo_numero_serie", "label": "N° de série (transfo)", "type": "text", "label_technique": "Haute tension", "ordre": 2002},
    {"col": "trafo_puissance_kva", "label": "Puissance (kVA)", "type": "text", "label_technique": "Haute tension", "ordre": 2003},
    {"col": "trafo_refroidissement", "label": "Refroidissement", "type": "text", "label_technique": "Haute tension", "ordre": 2004},
    {"col": "trafo_poids_kg", "label": "Poids (kg)", "type": "text", "label_technique": "Haute tension", "ordre": 2005},
    {"col": "trafo_tension_entree_v", "label": "Tension d'entrée (V)", "type": "text", "label_technique": "Haute tension", "ordre": 2006},
    {"col": "trafo_courant_a", "label": "Courant (A)", "type": "text", "label_technique": "Haute tension", "ordre": 2007},
    {"col": "trafo_norme", "label": "Norme", "type": "text", "label_technique": "Haute tension", "ordre": 2008},
    {"col": "trafo_couplage", "label": "Couplage", "type": "text", "label_technique": "Haute tension", "ordre": 2009},
    {"col": "trafo_tension_service_v", "label": "Tension de service (V)", "type": "text", "label_technique": "Haute tension", "ordre": 2010},
    {"col": "trafo_reglage_tension_kv", "label": "Réglage tension (kV)", "type": "text", "label_technique": "Haute tension", "ordre": 2011},
]


def _system_champs_for_technique(technique_id, db):
    """Construit la liste des champs système applicables à une technique donnée.
    Retourne au format compatible avec equipement_champs (id=négatif pour différencier)."""
    if not technique_id:
        return []
    t = one(db.execute("SELECT id, nom FROM techniques WHERE id=?", (technique_id,)))
    if not t:
        return []
    tech_nom = t["nom"]
    out = []
    # ID négatif pour distinguer des champs custom (qui ont des IDs positifs)
    # On utilise un offset stable basé sur l'index dans la liste
    for idx, sys_c in enumerate(_EQ_SYSTEM_CHAMPS):
        if sys_c.get("label_technique") != tech_nom:
            continue
        out.append({
            "id": -(idx + 1),  # négatif pour système
            "is_system": True,
            "system_col": sys_c["col"],
            "societe_id": None,
            "technique_id": technique_id,
            "technique_nom": tech_nom,
            "slug": sys_c["col"],
            "label": sys_c["label"],
            "type": sys_c["type"],
            "options_json": "{}",
            "options": {},
            "ordre": sys_c.get("ordre", 0),
        })
    return out


@app.route("/api/equipement-champs", methods=["GET"])
@require_auth
def list_equipement_champs():
    """Liste les champs personnalisés. Optionnel : ?technique_id=X pour ne récupérer
    que les champs applicables à cette technique (= globaux + ceux de la technique).
    v218.121 : injecte aussi les champs système (en dur dans la table equipements)
    quand technique_id est fourni — pour qu'ils apparaissent dans la section
    'Champs personnalisés' du formulaire équipement."""
    db = get_db()
    sid = current_societe_id()
    logger.info(f"[equipement-champs GET] sid={sid} args={dict(request.args)}")
    technique_id = request.args.get("technique_id")
    where = ["c.societe_id=?"]; params = [sid]
    tid = None
    if technique_id:
        try:
            tid = int(technique_id)
            where.append("(c.technique_id IS NULL OR c.technique_id=?)")
            params.append(tid)
        except: pass
    sql = ("SELECT c.*, t.nom AS technique_nom "
           "FROM equipement_champs c "
           "LEFT JOIN techniques t ON c.technique_id=t.id "
           "WHERE " + " AND ".join(where) + " "
           "ORDER BY (c.technique_id IS NULL) DESC, c.ordre, c.label")
    try:
        champs = rows(db.execute(sql, tuple(params)))
        logger.info(f"[equipement-champs GET] {len(champs)} champ(s) custom trouvé(s) (sid={sid})")
    except Exception as e:
        logger.error(f"[equipement-champs GET] échec SQL : {e}", exc_info=True)
        return jsonify({"error": f"Erreur SQL : {e}"}), 500
    # Parser options_json
    for c in champs:
        try: c["options"] = json.loads(c.get("options_json") or "{}")
        except: c["options"] = {}
        c["is_system"] = False
    # v218.122 : injection des champs système DÉSACTIVÉE — elle causait une perte de données
    # quand le formulaire sauvegardait avant que la section "Champs personnalisés" ne soit peuplée.
    # Les champs en dur (marque, modele, trafo_*) restent gérés via leurs sections dédiées.
    return jsonify(champs)


@app.route("/api/equipement-champs", methods=["POST"])
@require_role("admin")
def create_equipement_champ():
    """Crée un champ personnalisé.
    Body : { label, type, technique_id?, options?, ordre? }"""
    d = request.json or {}
    label = (d.get("label") or "").strip()
    type_c = (d.get("type") or "text").strip().lower()
    if not label:
        return jsonify({"error": "label requis"}), 400
    if type_c not in _CHAMP_TYPES_VALIDES:
        return jsonify({"error": f"type invalide (valeurs : {', '.join(sorted(_CHAMP_TYPES_VALIDES))})"}), 400
    technique_id = to_int(d.get("technique_id")) if d.get("technique_id") else None
    options = d.get("options") or {}
    if not isinstance(options, dict): options = {}
    # Validation type select : choices doit être une liste non vide
    if type_c == "select":
        choices = options.get("choices") or []
        if not isinstance(choices, list) or not choices:
            return jsonify({"error": "type 'select' requiert options.choices = liste non vide"}), 400
    db = get_db()
    sid = current_societe_id()
    # Générer un slug unique
    base_slug = _slugify_champ(label)
    slug = base_slug
    n = 1
    while one(db.execute("SELECT id FROM equipement_champs WHERE societe_id=? AND slug=?", (sid, slug))):
        n += 1
        slug = f"{base_slug}_{n}"
    # ordre par défaut = max + 1 dans la portée
    max_o = one(db.execute("SELECT COALESCE(MAX(ordre),-1) AS m FROM equipement_champs WHERE societe_id=?", (sid,)))
    ordre = to_int(d.get("ordre")) if d.get("ordre") is not None else (max_o["m"] + 1)
    db.execute("""INSERT INTO equipement_champs
                  (societe_id, technique_id, slug, label, type, options_json, ordre)
                  VALUES (?,?,?,?,?,?,?)""",
               (sid, technique_id, slug, label, type_c, json.dumps(options, ensure_ascii=False), ordre))
    db.commit()
    new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    log_action(request.user, "CREATE", "equipement_champ", new_id, f"{label} ({type_c})")
    return jsonify({"id": new_id, "slug": slug, "label": label, "type": type_c,
                    "technique_id": technique_id, "ordre": ordre, "options": options}), 201


@app.route("/api/equipement-champs/<int:cid>", methods=["PATCH"])
@require_role("admin")
def update_equipement_champ(cid):
    """Met à jour un champ existant. Le slug N'est PAS modifiable (pour ne pas casser
    les valeurs déjà stockées). Pour renommer techniquement, supprime et recrée."""
    d = request.json or {}
    db = get_db()
    sid = current_societe_id()
    existing = one(db.execute("SELECT * FROM equipement_champs WHERE id=? AND societe_id=?", (cid, sid)))
    if not existing:
        return jsonify({"error": "Introuvable"}), 404
    sets = []; params = []
    if "label" in d:
        lbl = (d["label"] or "").strip()
        if not lbl: return jsonify({"error": "label requis"}), 400
        sets.append("label=?"); params.append(lbl)
    if "type" in d:
        t = (d["type"] or "").strip().lower()
        if t not in _CHAMP_TYPES_VALIDES:
            return jsonify({"error": "type invalide"}), 400
        sets.append("type=?"); params.append(t)
    if "technique_id" in d:
        tid = to_int(d["technique_id"]) if d["technique_id"] else None
        sets.append("technique_id=?"); params.append(tid)
    if "options" in d:
        opts = d["options"] or {}
        if not isinstance(opts, dict): opts = {}
        sets.append("options_json=?"); params.append(json.dumps(opts, ensure_ascii=False))
    if "ordre" in d:
        sets.append("ordre=?"); params.append(to_int(d["ordre"]) or 0)
    if not sets:
        return jsonify({"error": "Rien à modifier"}), 400
    params.append(cid); params.append(sid)
    db.execute(f"UPDATE equipement_champs SET {', '.join(sets)} WHERE id=? AND societe_id=?", params)
    db.commit()
    log_action(request.user, "UPDATE", "equipement_champ", cid, f"sets={list(d.keys())}")
    return jsonify({"ok": True})


@app.route("/api/equipement-champs/<int:cid>", methods=["DELETE"])
@require_role("admin")
def delete_equipement_champ(cid):
    """Supprime un champ et toutes ses valeurs."""
    db = get_db()
    sid = current_societe_id()
    existing = one(db.execute("SELECT label FROM equipement_champs WHERE id=? AND societe_id=?", (cid, sid)))
    if not existing:
        return jsonify({"error": "Introuvable"}), 404
    # CASCADE supprime aussi les valeurs et fichiers
    db.execute("DELETE FROM equipement_champs WHERE id=? AND societe_id=?", (cid, sid))
    db.commit()
    log_action(request.user, "DELETE", "equipement_champ", cid, existing.get("label", ""))
    return jsonify({"ok": True})


@app.route("/api/equipement-champs/reorder", methods=["POST"])
@require_role("admin")
def reorder_equipement_champs():
    """Réordonne les champs. Body : { ids: [3, 1, 5, 2] } → ordre 0, 1, 2, 3 dans cet ordre."""
    d = request.json or {}
    ids = d.get("ids") or []
    if not isinstance(ids, list):
        return jsonify({"error": "ids doit être une liste"}), 400
    db = get_db()
    sid = current_societe_id()
    for idx, cid in enumerate(ids):
        try:
            db.execute("UPDATE equipement_champs SET ordre=? WHERE id=? AND societe_id=?",
                       (idx, int(cid), sid))
        except: pass
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/equipements/<int:eid>/valeurs-custom", methods=["GET"])
@require_auth
def get_equipement_valeurs_custom(eid):
    """Retourne les valeurs des champs personnalisés d'un équipement.
    Retourne aussi le catalogue des champs applicables (globaux + de la technique de l'équip).
    Pour les fichiers : metadata (filename, size), pas le contenu binaire (à fetch séparément).
    v218.121 : inclut aussi les champs système (en dur dans la table equipements)
    avec leur valeur lue directement depuis la colonne native."""
    db = get_db()
    sid = current_societe_id()
    eq = one(db.execute("SELECT * FROM equipements WHERE id=?", (eid,)))
    if not eq: return jsonify({"error": "Équipement introuvable"}), 404
    tid = eq.get("technique_id")
    # v218.154 : fallback - si technique_id manque, déduire depuis type_technique (texte)
    if not tid and eq.get("type_technique"):
        tech_match = one(db.execute("SELECT id FROM techniques WHERE LOWER(nom)=LOWER(?) AND societe_id=?",
                                    (eq["type_technique"], sid)))
        if tech_match:
            tid = tech_match["id"]
            logger.info(f"[valeurs-custom GET] eid={eid} type_technique='{eq['type_technique']}' → technique_id résolu={tid}")
    # Champs applicables : globaux + de la technique
    champs = rows(db.execute("""SELECT c.*, t.nom AS technique_nom
                                FROM equipement_champs c
                                LEFT JOIN techniques t ON c.technique_id=t.id
                                WHERE c.societe_id=?
                                  AND (c.technique_id IS NULL OR c.technique_id=?)
                                ORDER BY (c.technique_id IS NULL) DESC, c.ordre, c.label""",
                              (sid, tid)))
    logger.info(f"[valeurs-custom GET] eid={eid} sid={sid} tid={tid} → {len(champs)} champ(s)")
    # Valeurs existantes
    valeurs = rows(db.execute("""SELECT v.*, f.id AS fichier_id, f.filename, f.size AS fichier_size
                                  FROM equipement_valeurs_custom v
                                  LEFT JOIN equipement_fichiers_custom f ON f.valeur_id=v.id
                                  WHERE v.equipement_id=?""", (eid,)))
    val_by_champ = {v["champ_id"]: v for v in valeurs}
    # Assembler les champs custom
    out = []
    for c in champs:
        try: opts = json.loads(c.get("options_json") or "{}")
        except: opts = {}
        v = val_by_champ.get(c["id"])
        item = {
            "champ_id": c["id"], "slug": c["slug"], "label": c["label"],
            "type": c["type"], "options": opts,
            "technique_id": c.get("technique_id"),
            "technique_nom": c.get("technique_nom"),
            "ordre": c.get("ordre", 0),
            "is_system": False,
            "valeur": (v or {}).get("valeur", ""),
        }
        if c["type"] == "file" and v and v.get("fichier_id"):
            item["fichier_id"] = v["fichier_id"]
            item["fichier_filename"] = v.get("filename")
            item["fichier_size"] = v.get("fichier_size")
        out.append(item)
    # v218.122 : injection des champs système DÉSACTIVÉE (cf. raison v218.122 plus haut)
    # Trier : globaux non-système d'abord, puis customs spécifiques
    out.sort(key=lambda c: (
        0 if c.get("technique_id") is None else 1,
        c.get("ordre", 0) or 0,
        c.get("label", "") or ""
    ))
    return jsonify(out)
    return jsonify(out)


@app.route("/api/equipements/<int:eid>/valeurs-custom", methods=["PUT"])
@require_auth
def set_equipement_valeurs_custom(eid):
    """Met à jour les valeurs des champs personnalisés.
    Body : { valeurs: [{champ_id, valeur}, ...] }
    Les fichiers ne passent PAS par ici (endpoint séparé /api/equipement-fichiers-custom)."""
    d = request.json or {}
    valeurs = d.get("valeurs") or []
    if not isinstance(valeurs, list):
        return jsonify({"error": "valeurs doit être une liste"}), 400
    db = get_db()
    sid = current_societe_id()
    logger.info(f"[valeurs-custom PUT] eid={eid} sid={sid} nb_valeurs={len(valeurs)}")
    eq = one(db.execute("SELECT id FROM equipements WHERE id=?", (eid,)))
    if not eq: return jsonify({"error": "Équipement introuvable"}), 404
    nb_inserted = 0
    nb_skipped = 0
    nb_protected = 0
    for item in valeurs:
        cid = to_int(item.get("champ_id"))
        if not cid:
            logger.info(f"[valeurs-custom PUT]   skip: cid invalide ({item})")
            nb_skipped += 1
            continue
        # Vérifier que le champ existe et appartient à la société
        c = one(db.execute("SELECT id, type, societe_id FROM equipement_champs WHERE id=? AND societe_id=?", (cid, sid)))
        if not c:
            # Diagnostic : le champ existe-t-il sous une autre société ?
            c_any = one(db.execute("SELECT id, societe_id FROM equipement_champs WHERE id=?", (cid,)))
            if c_any:
                logger.warning(f"[valeurs-custom PUT]   skip cid={cid}: champ existe mais societe_id={c_any.get('societe_id')} != requete sid={sid}")
            else:
                logger.warning(f"[valeurs-custom PUT]   skip cid={cid}: champ inexistant")
            nb_skipped += 1
            continue
        if c["type"] == "file":
            logger.info(f"[valeurs-custom PUT]   skip cid={cid}: type=file")
            nb_skipped += 1
            continue
        val = item.get("valeur")
        if val is None: val = ""
        if not isinstance(val, str): val = str(val)
        # v218.152 : PROTECTION CRITIQUE - ne pas écraser une valeur existante par chaîne vide
        # Empêche la perte de données quand _ccLoadValues n'a pas chargé les valeurs avant le save.
        if val == "":
            existing = one(db.execute("SELECT valeur FROM equipement_valeurs_custom WHERE equipement_id=? AND champ_id=?", (eid, cid)))
            if existing and (existing.get("valeur") or "") != "":
                logger.warning(f"[valeurs-custom PUT]   PROTECT cid={cid}: refuse écrasement de '{existing.get('valeur')}' par ''")
                nb_protected += 1
                continue
        try:
            db.execute("""INSERT INTO equipement_valeurs_custom (equipement_id, champ_id, valeur, updated_at)
                          VALUES (?, ?, ?, datetime('now'))
                          ON CONFLICT(equipement_id, champ_id) DO UPDATE SET
                            valeur=excluded.valeur, updated_at=excluded.updated_at""",
                       (eid, cid, val))
            nb_inserted += 1
        except Exception as e:
            logger.error(f"[valeurs-custom PUT]   ÉCHEC SQL cid={cid}: {e}")
            nb_skipped += 1
    db.commit()
    logger.info(f"[valeurs-custom PUT] eid={eid} inserted={nb_inserted} skipped={nb_skipped} protected={nb_protected}")
    return jsonify({"ok": True, "inserted": nb_inserted, "skipped": nb_skipped, "protected": nb_protected})


@app.route("/api/equipements/<int:eid>/fichier-custom/<int:cid>", methods=["POST"])
@require_auth
def upload_equipement_fichier_custom(eid, cid):
    """Téléverse un fichier pour un champ de type 'file' sur un équipement."""
    db = get_db()
    sid = current_societe_id()
    c = one(db.execute("SELECT id, type FROM equipement_champs WHERE id=? AND societe_id=?", (cid, sid)))
    if not c: return jsonify({"error": "Champ introuvable"}), 404
    if c["type"] != "file": return jsonify({"error": "Ce champ n'est pas de type fichier"}), 400
    if "file" not in request.files:
        return jsonify({"error": "Aucun fichier reçu"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Nom de fichier vide"}), 400
    data = f.read()
    if len(data) > 20 * 1024 * 1024:
        return jsonify({"error": "Fichier trop volumineux (max 20 Mo)"}), 400
    mime = f.mimetype or "application/octet-stream"
    # Créer la valeur si elle n'existe pas
    val_row = one(db.execute("SELECT id FROM equipement_valeurs_custom WHERE equipement_id=? AND champ_id=?", (eid, cid)))
    if val_row:
        valeur_id = val_row["id"]
        # Supprimer l'ancien fichier
        db.execute("DELETE FROM equipement_fichiers_custom WHERE valeur_id=?", (valeur_id,))
    else:
        db.execute("INSERT INTO equipement_valeurs_custom (equipement_id, champ_id, valeur) VALUES (?, ?, ?)",
                   (eid, cid, f.filename))
        valeur_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    db.execute("""INSERT INTO equipement_fichiers_custom (valeur_id, filename, mime_type, data, size)
                  VALUES (?, ?, ?, ?, ?)""",
               (valeur_id, f.filename, mime, data, len(data)))
    # Mettre à jour la valeur (= nom de fichier pour affichage)
    db.execute("UPDATE equipement_valeurs_custom SET valeur=?, updated_at=datetime('now') WHERE id=?",
               (f.filename, valeur_id))
    db.commit()
    fichier_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    return jsonify({"ok": True, "fichier_id": fichier_id, "filename": f.filename, "size": len(data)})


@app.route("/api/equipement-fichiers-custom/<int:fid>", methods=["GET"])
@require_auth
def get_equipement_fichier_custom(fid):
    """Télécharge le fichier."""
    db = get_db()
    row = one(db.execute("SELECT filename, mime_type, data FROM equipement_fichiers_custom WHERE id=?", (fid,)))
    if not row: return jsonify({"error": "Fichier introuvable"}), 404
    resp = make_response(bytes(row["data"]))
    resp.headers["Content-Type"] = row["mime_type"] or "application/octet-stream"
    resp.headers["Content-Disposition"] = f'attachment; filename="{row["filename"]}"'
    return resp


@app.route("/api/equipement-fichiers-custom/<int:fid>", methods=["DELETE"])
@require_auth
def delete_equipement_fichier_custom(fid):
    """Supprime un fichier custom (et la valeur associée)."""
    db = get_db()
    row = one(db.execute("SELECT valeur_id FROM equipement_fichiers_custom WHERE id=?", (fid,)))
    if not row: return jsonify({"error": "Introuvable"}), 404
    db.execute("DELETE FROM equipement_fichiers_custom WHERE id=?", (fid,))
    # Reset la valeur (ou la supprimer)
    db.execute("UPDATE equipement_valeurs_custom SET valeur='', updated_at=datetime('now') WHERE id=?", (row["valeur_id"],))
    db.commit()
    return jsonify({"ok": True})


# ══ GAMMES ══
# ══════════════════════════════════════════════════════════════════════
# SOUS-TYPES DE MAINTENANCE (Entretien / Visite / etc.)
# Configurables. Un sous-type s'attache à une gamme. Le sous-type d'un bon
# est calculé en agrégant les sous-types de ses gammes (cf. _bon_sous_type)
# ══════════════════════════════════════════════════════════════════════
@app.route("/api/maintenance_sous_types")
@require_auth
def get_maintenance_sous_types():
    """Liste tous les sous-types (publics)."""
    db = get_db()
    return jsonify(rows(db.execute(
        "SELECT * FROM maintenance_sous_types ORDER BY ordre, nom"
    )))

@app.route("/api/maintenance_sous_types", methods=["POST"])
@require_role("admin")
def create_maintenance_sous_type():
    """Crée un nouveau sous-type."""
    d = request.json or {}
    nom = (d.get("nom") or "").strip()
    if not nom: return jsonify({"error": "Nom requis"}), 400
    couleur = (d.get("couleur") or "#1E3A8A").strip()
    ordre = int(d.get("ordre") or 0)
    db = get_db()
    if one(db.execute("SELECT id FROM maintenance_sous_types WHERE nom=?", (nom,))):
        return jsonify({"error": "Un sous-type avec ce nom existe déjà"}), 400
    db.execute("INSERT INTO maintenance_sous_types (nom, couleur, ordre) VALUES (?, ?, ?)",
               (nom, couleur, ordre))
    db.commit()
    new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    log_action(request.user, "CREATE", "maintenance_sous_type", new_id, nom)
    return jsonify({"ok": True, "id": new_id}), 201

@app.route("/api/maintenance_sous_types/<int:sid>", methods=["PATCH"])
@require_role("admin")
def update_maintenance_sous_type(sid):
    """Modifie un sous-type."""
    d = request.json or {}
    db = get_db()
    if not one(db.execute("SELECT id FROM maintenance_sous_types WHERE id=?", (sid,))):
        return jsonify({"error": "Introuvable"}), 404
    fields = {}
    if "nom" in d:
        nom = (d["nom"] or "").strip()
        if not nom: return jsonify({"error": "Nom requis"}), 400
        # Vérif unicité
        existing = one(db.execute(
            "SELECT id FROM maintenance_sous_types WHERE nom=? AND id<>?", (nom, sid)
        ))
        if existing: return jsonify({"error": "Un autre sous-type a déjà ce nom"}), 400
        fields["nom"] = nom
    if "couleur" in d: fields["couleur"] = (d["couleur"] or "#1E3A8A").strip()
    if "ordre" in d: fields["ordre"] = int(d["ordre"] or 0)
    if not fields: return jsonify({"error": "Rien à modifier"}), 400
    placeholders = ", ".join(f"{k}=?" for k in fields)
    db.execute(f"UPDATE maintenance_sous_types SET {placeholders} WHERE id=?",
               (*fields.values(), sid))
    db.commit()
    log_action(request.user, "UPDATE", "maintenance_sous_type", sid, str(fields))
    return jsonify({"ok": True})

@app.route("/api/maintenance_sous_types/<int:sid>", methods=["DELETE"])
@require_role("admin")
def delete_maintenance_sous_type(sid):
    """Supprime un sous-type. Les gammes qui l'utilisaient repassent à 'aucun sous-type'."""
    db = get_db()
    st = one(db.execute("SELECT nom FROM maintenance_sous_types WHERE id=?", (sid,)))
    if not st: return jsonify({"error": "Introuvable"}), 404
    nb = one(db.execute("SELECT COUNT(*) AS n FROM gammes WHERE sous_type_id=?", (sid,)))
    nb_gammes = (nb or {}).get("n", 0)
    # On déréférence (les gammes passent à NULL automatiquement via update)
    db.execute("UPDATE gammes SET sous_type_id=NULL WHERE sous_type_id=?", (sid,))
    db.execute("DELETE FROM maintenance_sous_types WHERE id=?", (sid,))
    db.commit()
    log_action(request.user, "DELETE", "maintenance_sous_type", sid,
               f"{st['nom']} (déréférencé sur {nb_gammes} gamme(s))")
    return jsonify({"ok": True, "gammes_dereferencees": nb_gammes})


# ══════════════════════════════════════════════════════════════════════
# MESURES TECHNIQUES (par technique × sous-type)
# Configurables : ex. "UPS × Entretien" → 4 tableaux (entrée, by-pass, sortie, batterie)
# ══════════════════════════════════════════════════════════════════════
@app.route("/api/mesure_blocs")
@require_auth
def get_mesure_blocs():
    """Liste tous les blocs de mesures (avec leurs lignes).
    Optionnels : ?technique_id=X&sous_type_id=Y pour filtrer (mode BP).
    v218.109 : ?is_bc=1 pour récupérer les modules globaux des bons DÉPANNAGE (BC).
    En mode BC, technique_id et sous_type_id sont ignorés, on filtre seulement
    par societe_id (modules globaux par société)."""
    db = get_db()
    where = []; params = []
    is_bc_param = request.args.get("is_bc", "0")
    is_bc = is_bc_param in ("1", "true", "True", "yes")
    if is_bc:
        # Modules BC : globaux pour la société
        sid = current_societe_id()
        where.append("mb.is_bc=1"); params.append(0)  # placeholder removed below
        params = []  # reset
        where = ["mb.is_bc=1", "mb.societe_id=?"]
        params = [sid]
    else:
        # Modules BP : exclure les BC + filtrer par technique/sous_type si fourni
        where.append("(mb.is_bc=0 OR mb.is_bc IS NULL)")
        technique_id = request.args.get("technique_id")
        sous_type_id = request.args.get("sous_type_id")
        if technique_id:
            where.append("mb.technique_id=?"); params.append(int(technique_id))
        if sous_type_id:
            where.append("mb.sous_type_id=?"); params.append(int(sous_type_id))
    sql = """SELECT mb.*, t.nom AS technique_nom, st.nom AS sous_type_nom, st.couleur AS sous_type_couleur
             FROM mesure_blocs mb
             LEFT JOIN techniques t ON mb.technique_id = t.id
             LEFT JOIN maintenance_sous_types st ON mb.sous_type_id = st.id"""
    if where: sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY t.nom, st.ordre, mb.page_num, mb.ordre, mb.nom"
    blocs = rows(db.execute(sql, tuple(params)))
    # Charger les lignes de chaque bloc
    for b in blocs:
        b["lignes"] = rows(db.execute(
            "SELECT * FROM mesure_lignes WHERE bloc_id=? ORDER BY ordre, id",
            (b["id"],)
        ))
    return jsonify(blocs)

@app.route("/api/mesure_blocs", methods=["POST"])
@require_role("admin")
def create_mesure_bloc():
    """Crée un nouveau bloc de mesures (= module de structure rapport).
    Body : { technique_id, sous_type_id, nom, type, ordre, lignes: [{libelle, unite}, ...] }
    v218.109 : { is_bc: true } → module global BC (technique_id et sous_type_id ignorés)
    type ∈ {'tableau', 'texte', 'checklist'} (défaut: tableau)"""
    d = request.json or {}
    is_bc = bool(d.get("is_bc"))
    nom = (d.get("nom") or "").strip()
    if not nom:
        return jsonify({"error": "nom requis"}), 400
    if is_bc:
        # Mode BC : pas de technique/sous_type, on stocke avec societe_id
        technique_id = None
        sous_type_id = None
    else:
        technique_id = d.get("technique_id")
        sous_type_id = d.get("sous_type_id")
        if not technique_id or not sous_type_id:
            return jsonify({"error": "technique_id et sous_type_id requis (ou is_bc=true)"}), 400
    type_module = (d.get("type") or "tableau").strip().lower()
    if type_module not in ("tableau", "texte", "checklist", "graphique", "preconisations", "mesures_batteries", "image", "equipement", "image_checklist"):
        return jsonify({"error": "type invalide"}), 400
    ordre = int(d.get("ordre") or 0)
    # Config globale du module (utilisé pour graphique : titre, séries, unité Y…)
    bloc_fopts = d.get("field_options")
    if isinstance(bloc_fopts, dict):
        bloc_fopts = json.dumps(bloc_fopts, ensure_ascii=False)
    elif bloc_fopts is None:
        bloc_fopts = ""
    # v213 : largeur ('pleine' ou 'demi')
    largeur = (d.get("largeur") or "pleine").strip().lower()
    if largeur not in ("pleine", "demi"):
        largeur = "pleine"
    # v214 : numéro de page (default 1)
    try:
        page_num = max(1, int(d.get("page_num") or 1))
    except Exception:
        page_num = 1
    # v215 : icône (slug)
    icon = (d.get("icon") or "").strip()
    db = get_db()
    sid = current_societe_id()
    db.execute("""INSERT INTO mesure_blocs
                  (technique_id, sous_type_id, nom, type, ordre, field_options, largeur, page_num, icon, is_bc, societe_id)
                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
               (technique_id, sous_type_id, nom, type_module, ordre, bloc_fopts, largeur, page_num, icon,
                1 if is_bc else 0, sid))
    bid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    # Pour 'texte' : on crée une seule ligne sans libellé spécifique (la valeur sera la zone de texte)
    if type_module == "texte":
        db.execute("INSERT INTO mesure_lignes (bloc_id, libelle, unite, ordre, field_type, field_options) VALUES (?, ?, ?, ?, ?, ?)",
                   (bid, "_text_", "", 0, "text_long", ""))
    elif type_module == "preconisations":
        # Pas de lignes : le module lit dynamiquement les pièces de l'équipement.
        pass
    elif type_module == "mesures_batteries":
        # Pas de lignes : le module détecte les batteries depuis les pièces et stocke en JSON
        pass
    elif type_module == "image":
        # v217 : pas de lignes — le module récupère dynamiquement l'image principale de l'équipement
        pass
    elif type_module == "graphique":
        # Pour 'graphique' : chaque ligne = un point (libellé X + valeur ref dans field_options)
        for i, lg in enumerate(d.get("lignes") or []):
            lib = (lg.get("libelle") or "").strip()
            if not lib: continue
            fopts = lg.get("field_options")
            if isinstance(fopts, dict):
                fopts = json.dumps(fopts, ensure_ascii=False)
            elif fopts is None:
                fopts = ""
            db.execute("INSERT INTO mesure_lignes (bloc_id, libelle, unite, ordre, field_type, field_options) VALUES (?, ?, ?, ?, ?, ?)",
                       (bid, lib, "", i, "graph_point", fopts))
    else:
        for i, lg in enumerate(d.get("lignes") or []):
            lib = (lg.get("libelle") or "").strip()
            if not lib: continue
            un = (lg.get("unite") or "").strip()
            ftype = (lg.get("field_type") or "numeric").strip().lower()
            if ftype not in ("numeric", "text_short", "text_long", "binary", "select", "text_pair"):
                ftype = "numeric"
            fopts = lg.get("field_options")
            if isinstance(fopts, dict):
                fopts = json.dumps(fopts, ensure_ascii=False)
            elif fopts is None:
                fopts = ""
            db.execute("INSERT INTO mesure_lignes (bloc_id, libelle, unite, ordre, field_type, field_options) VALUES (?, ?, ?, ?, ?, ?)",
                       (bid, lib, un, i, ftype, fopts))
    db.commit()
    log_action(request.user, "CREATE", "mesure_bloc", bid, nom)
    return jsonify({"ok": True, "id": bid}), 201

@app.route("/api/mesure_blocs/<int:bid>", methods=["PATCH"])
@require_role("admin")
def update_mesure_bloc(bid):
    """Modifie un bloc et ses lignes (replace complet des lignes si fourni)."""
    d = request.json or {}
    db = get_db()
    if not one(db.execute("SELECT id FROM mesure_blocs WHERE id=?", (bid,))):
        return jsonify({"error": "Introuvable"}), 404
    fields = {}
    if "nom" in d:
        nom = (d["nom"] or "").strip()
        if not nom: return jsonify({"error": "Nom requis"}), 400
        fields["nom"] = nom
    if "ordre" in d: fields["ordre"] = int(d["ordre"] or 0)
    if "technique_id" in d: fields["technique_id"] = int(d["technique_id"])
    if "sous_type_id" in d: fields["sous_type_id"] = int(d["sous_type_id"])
    if "type" in d:
        type_module = (d["type"] or "tableau").strip().lower()
        if type_module not in ("tableau", "texte", "checklist", "graphique", "preconisations", "mesures_batteries", "image", "equipement", "image_checklist"):
            return jsonify({"error": "type invalide"}), 400
        fields["type"] = type_module
    if "field_options" in d:
        bfopts = d["field_options"]
        if isinstance(bfopts, dict):
            bfopts = json.dumps(bfopts, ensure_ascii=False)
        elif bfopts is None:
            bfopts = ""
        fields["field_options"] = bfopts
    if "largeur" in d:
        lg = (d["largeur"] or "pleine").strip().lower()
        if lg not in ("pleine", "demi"):
            lg = "pleine"
        fields["largeur"] = lg
    if "page_num" in d:
        try:
            fields["page_num"] = max(1, int(d["page_num"] or 1))
        except Exception:
            fields["page_num"] = 1
    if "icon" in d:
        fields["icon"] = (d["icon"] or "").strip()
    if fields:
        ph = ", ".join(f"{k}=?" for k in fields)
        db.execute(f"UPDATE mesure_blocs SET {ph} WHERE id=?", (*fields.values(), bid))
    if "lignes" in d:
        # Replace complet (les valeurs déjà saisies dans intervention_mesures référencent ligne_id ;
        # FK ON DELETE CASCADE → les saisies des anciennes lignes sont supprimées)
        # Pour 'texte' : on garde une ligne unique
        cur_type = (one(db.execute("SELECT type FROM mesure_blocs WHERE id=?", (bid,))) or {}).get("type", "tableau")
        if "type" in fields: cur_type = fields["type"]
        # v218.88 : pour les types sans lignes utilisateur, NE PAS toucher aux lignes
        # (sinon DELETE CASCADE supprime les valeurs saisies dans intervention_mesures)
        # 'texte' : la ligne unique _text_ contient la valeur saisie, donc à préserver aussi.
        if cur_type in ("equipement", "image", "preconisations", "mesures_batteries", "texte"):
            # Pour 'texte', s'assurer qu'il y a bien une ligne _text_ (au cas où elle n'existe pas)
            if cur_type == "texte":
                exists = one(db.execute("SELECT id FROM mesure_lignes WHERE bloc_id=? AND libelle='_text_'", (bid,)))
                if not exists:
                    db.execute("INSERT INTO mesure_lignes (bloc_id, libelle, unite, ordre, field_type, field_options) VALUES (?, ?, ?, ?, ?, ?)",
                               (bid, "_text_", "", 0, "text_long", ""))
            # Sinon (equipement/image/preco/batteries) : pas de lignes nécessaires
        else:
            # v218.88 : Upsert intelligent — préserver les ligne_id pour ne pas
            # supprimer les valeurs saisies dans intervention_mesures (FK CASCADE).
            # Les lignes envoyées avec un 'id' existant → UPDATE
            # Les lignes sans 'id' (ou id inconnu) → INSERT
            # Les lignes existantes non envoyées → DELETE (CASCADE des saisies)
            existing_ids = set(r["id"] for r in rows(db.execute(
                "SELECT id FROM mesure_lignes WHERE bloc_id=?", (bid,)
            )))
            sent_ids = set()
            for i, lg in enumerate(d["lignes"]):
                lib = (lg.get("libelle") or "").strip()
                if not lib: continue
                un = (lg.get("unite") or "").strip()
                if cur_type == "graphique":
                    ftype = "graph_point"
                    un = ""
                else:
                    ftype = (lg.get("field_type") or "numeric").strip().lower()
                    if ftype not in ("numeric", "text_short", "text_long", "binary", "select", "text_pair"):
                        ftype = "numeric"
                fopts = lg.get("field_options")
                if isinstance(fopts, dict):
                    fopts = json.dumps(fopts, ensure_ascii=False)
                elif fopts is None:
                    fopts = ""
                lid = lg.get("id")
                try:
                    lid = int(lid) if lid else None
                except Exception:
                    lid = None
                if lid and lid in existing_ids:
                    # UPDATE existant : préserve les saisies utilisateur
                    db.execute("""UPDATE mesure_lignes
                                  SET libelle=?, unite=?, ordre=?, field_type=?, field_options=?
                                  WHERE id=?""",
                               (lib, un, i, ftype, fopts, lid))
                    sent_ids.add(lid)
                else:
                    # INSERT nouvelle ligne
                    db.execute("""INSERT INTO mesure_lignes
                                  (bloc_id, libelle, unite, ordre, field_type, field_options)
                                  VALUES (?, ?, ?, ?, ?, ?)""",
                               (bid, lib, un, i, ftype, fopts))
            # DELETE des lignes qui n'ont pas été renvoyées
            to_delete = existing_ids - sent_ids
            for did in to_delete:
                db.execute("DELETE FROM mesure_lignes WHERE id=?", (did,))
    db.commit()
    log_action(request.user, "UPDATE", "mesure_bloc", bid, str(fields))
    return jsonify({"ok": True})

@app.route("/api/mesure_blocs/<int:bid>", methods=["DELETE"])
@require_role("admin")
def delete_mesure_bloc(bid):
    """Supprime un bloc (cascade sur mesure_lignes et intervention_mesures)."""
    db = get_db()
    b = one(db.execute("SELECT nom FROM mesure_blocs WHERE id=?", (bid,)))
    if not b: return jsonify({"error": "Introuvable"}), 404
    db.execute("DELETE FROM mesure_blocs WHERE id=?", (bid,))
    db.commit()
    log_action(request.user, "DELETE", "mesure_bloc", bid, b["nom"])
    return jsonify({"ok": True})

# v217.6 : actions sur les tuiles (couples technique × sous-type)
@app.route("/api/mesure_blocs/tile/move", methods=["PATCH"])
@require_role("admin", "manager")
def move_tile():
    """Réaffecte tous les blocs d'un couple (technique, sous-type) vers un autre couple.
    Body : {from_technique_id, from_sous_type_id, to_technique_id, to_sous_type_id}
    """
    d = request.json or {}
    try:
        from_t = int(d.get("from_technique_id") or 0)
        from_s = int(d.get("from_sous_type_id") or 0)
        to_t = int(d.get("to_technique_id") or 0)
        to_s = int(d.get("to_sous_type_id") or 0)
    except (ValueError, TypeError):
        return jsonify({"error": "IDs invalides"}), 400
    if not (from_t and from_s and to_t and to_s):
        return jsonify({"error": "Tous les IDs sont requis"}), 400
    if from_t == to_t and from_s == to_s:
        return jsonify({"error": "Aucun changement"}), 400
    db = get_db()
    # Vérifier que la cible n'a pas déjà des blocs (pour éviter les doublons)
    target = one(db.execute("SELECT COUNT(*) AS n FROM mesure_blocs WHERE technique_id=? AND sous_type_id=?", (to_t, to_s)))
    if target and target["n"]:
        return jsonify({"error": "La tuile cible existe déjà avec des modules. Supprime-la d'abord ou utilise Dupliquer."}), 400
    # Compter ce qu'on va déplacer
    src_count = one(db.execute("SELECT COUNT(*) AS n FROM mesure_blocs WHERE technique_id=? AND sous_type_id=?", (from_t, from_s)))
    if not src_count or not src_count["n"]:
        return jsonify({"error": "Tuile source introuvable"}), 404
    # Déplacer
    db.execute("UPDATE mesure_blocs SET technique_id=?, sous_type_id=? WHERE technique_id=? AND sous_type_id=?",
               (to_t, to_s, from_t, from_s))
    db.commit()
    log_action(request.user, "MOVE", "tile", None, f"{from_t}/{from_s} -> {to_t}/{to_s} ({src_count['n']} modules)")
    return jsonify({"ok": True, "moved": src_count["n"]})

@app.route("/api/mesure_blocs/tile/duplicate", methods=["POST"])
@require_role("admin", "manager")
def duplicate_tile():
    """Duplique tous les blocs (et leurs lignes) d'un couple vers un autre.
    Body : {from_technique_id, from_sous_type_id, to_technique_id, to_sous_type_id}
    """
    d = request.json or {}
    try:
        from_t = int(d.get("from_technique_id") or 0)
        from_s = int(d.get("from_sous_type_id") or 0)
        to_t = int(d.get("to_technique_id") or 0)
        to_s = int(d.get("to_sous_type_id") or 0)
    except (ValueError, TypeError):
        return jsonify({"error": "IDs invalides"}), 400
    if not (from_t and from_s and to_t and to_s):
        return jsonify({"error": "Tous les IDs sont requis"}), 400
    if from_t == to_t and from_s == to_s:
        return jsonify({"error": "La cible doit être différente de la source"}), 400
    db = get_db()
    # Vérifier que la cible est vide
    target = one(db.execute("SELECT COUNT(*) AS n FROM mesure_blocs WHERE technique_id=? AND sous_type_id=?", (to_t, to_s)))
    if target and target["n"]:
        return jsonify({"error": "La tuile cible existe déjà avec des modules. Supprime-la d'abord."}), 400
    # Récupérer tous les blocs source
    src_blocs = rows(db.execute("""
        SELECT id, nom, type, ordre, field_options, largeur, page_num, icon
        FROM mesure_blocs WHERE technique_id=? AND sous_type_id=?
        ORDER BY page_num, ordre, id
    """, (from_t, from_s)))
    if not src_blocs:
        return jsonify({"error": "Tuile source introuvable"}), 404
    # Pour chaque bloc, créer un duplicata + dupliquer ses lignes
    nb_blocs = 0
    nb_lignes = 0
    for b in src_blocs:
        cur = db.execute("""
            INSERT INTO mesure_blocs (technique_id, sous_type_id, nom, type, ordre, field_options, largeur, page_num, icon)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (to_t, to_s, b["nom"], b["type"], b["ordre"], b.get("field_options") or "",
              b.get("largeur") or "pleine", b.get("page_num") or 1, b.get("icon") or ""))
        new_bid = cur.lastrowid
        nb_blocs += 1
        # Dupliquer les lignes
        src_lignes = rows(db.execute("SELECT libelle, unite, ordre, field_type, field_options FROM mesure_lignes WHERE bloc_id=? ORDER BY ordre, id", (b["id"],)))
        for lg in src_lignes:
            db.execute("""
                INSERT INTO mesure_lignes (bloc_id, libelle, unite, ordre, field_type, field_options)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (new_bid, lg["libelle"], lg.get("unite") or "", lg.get("ordre") or 0,
                  lg.get("field_type") or "numeric", lg.get("field_options") or ""))
            nb_lignes += 1
    db.commit()
    log_action(request.user, "DUPLICATE", "tile", None, f"{from_t}/{from_s} -> {to_t}/{to_s} ({nb_blocs} modules, {nb_lignes} lignes)")
    return jsonify({"ok": True, "blocs": nb_blocs, "lignes": nb_lignes})

@app.route("/api/mesure_blocs/tile", methods=["DELETE"])
@require_role("admin")
def delete_tile():
    """Supprime tous les blocs d'un couple (technique, sous-type).
    Query params : ?technique_id=X&sous_type_id=Y
    """
    try:
        t_id = int(request.args.get("technique_id") or 0)
        s_id = int(request.args.get("sous_type_id") or 0)
    except (ValueError, TypeError):
        return jsonify({"error": "IDs invalides"}), 400
    if not (t_id and s_id):
        return jsonify({"error": "technique_id et sous_type_id requis"}), 400
    db = get_db()
    cnt = one(db.execute("SELECT COUNT(*) AS n FROM mesure_blocs WHERE technique_id=? AND sous_type_id=?", (t_id, s_id)))
    n = (cnt or {}).get("n", 0)
    if not n:
        return jsonify({"error": "Tuile introuvable"}), 404
    db.execute("DELETE FROM mesure_blocs WHERE technique_id=? AND sous_type_id=?", (t_id, s_id))
    db.commit()
    log_action(request.user, "DELETE", "tile", None, f"tuile {t_id}/{s_id} ({n} modules)")
    return jsonify({"ok": True, "deleted": n})

# v217.8 : Aperçu PDF d'une tuile avec valeurs factices
@app.route("/api/mesure_blocs/preview_pdf", methods=["POST"])
@require_role("admin", "manager")
def preview_pdf():
    """Génère un PDF d'aperçu pour une tuile (technique × sous-type) avec des
    valeurs factices, sans avoir besoin d'un bon réel.
    Body : {technique_id, sous_type_id}
    """
    d = request.json or {}
    try:
        tid = int(d.get("technique_id") or 0)
        sid = int(d.get("sous_type_id") or 0)
    except (ValueError, TypeError):
        return jsonify({"error": "IDs invalides"}), 400
    if not (tid and sid):
        return jsonify({"error": "technique_id et sous_type_id requis"}), 400
    db = get_db()
    technique = one(db.execute("SELECT nom FROM techniques WHERE id=?", (tid,)))
    if not technique:
        return jsonify({"error": "Technique introuvable"}), 404
    st = one(db.execute("SELECT nom FROM maintenance_sous_types WHERE id=?", (sid,)))
    if not st:
        return jsonify({"error": "Sous-type introuvable"}), 404

    # Récupérer les blocs de la tuile
    blocs = rows(db.execute("""
        SELECT id, nom, type, ordre, field_options, largeur, page_num, icon
        FROM mesure_blocs WHERE technique_id=? AND sous_type_id=?
        ORDER BY page_num, ordre, id
    """, (tid, sid)))

    # Construire les données factices pour chaque bloc
    mesures_techniques = []
    for b in blocs:
        bloc_data = {
            "nom": b["nom"],
            "type": b.get("type") or "tableau",
            "field_options": b.get("field_options") or "",
            "largeur": b.get("largeur") or "pleine",
            "page_num": int(b.get("page_num") or 1),
            "icon": b.get("icon") or "",
        }
        bloc_type = b.get("type") or "tableau"
        if bloc_type in ("tableau", "texte", "checklist"):
            # Récupérer les lignes et générer des valeurs factices
            lignes_db = rows(db.execute(
                "SELECT id, libelle, unite, ordre, field_type, field_options FROM mesure_lignes WHERE bloc_id=? ORDER BY ordre, id",
                (b["id"],)
            ))
            lignes_factices = []
            for i, lg in enumerate(lignes_db):
                ftype = lg.get("field_type") or "numeric"
                fopts = {}
                try:
                    fopts = json.loads(lg.get("field_options") or "{}") if isinstance(lg.get("field_options"), str) else (lg.get("field_options") or {})
                except Exception:
                    pass
                # Valeur factice selon le type
                if ftype == "numeric":
                    fake_val = str(round(230 + i * 12.3, 2))
                elif ftype == "binary":
                    fake_val = "1" if i % 2 == 0 else "0"
                elif ftype == "select":
                    opts = fopts.get("options") or ["Option 1", "Option 2"]
                    fake_val = opts[0] if opts else "—"
                elif ftype == "text_short":
                    fake_val = "Exemple " + str(i+1)
                elif ftype == "text_long":
                    fake_val = "Texte d'exemple sur plusieurs lignes pour montrer le rendu."
                elif ftype == "text_pair":
                    p1 = fopts.get("placeholder_1") or "Champ 1"
                    p2 = fopts.get("placeholder_2") or "Valeur exemple"
                    fake_val = p1 + "|||" + p2
                else:
                    fake_val = "—"
                ligne_data = {
                    "id": lg["id"],
                    "libelle": lg["libelle"],
                    "unite": lg.get("unite") or "",
                    "field_type": ftype,
                    "field_options": lg.get("field_options") or "",
                    "valeur": fake_val,
                }
                # Pour texte : utiliser _text_ comme libellé
                if bloc_type == "texte":
                    ligne_data["libelle"] = "_text_"
                    ligne_data["valeur"] = "Voici un exemple de texte libre saisi par le technicien lors de l'intervention. Il décrit l'état général de l'équipement, les opérations effectuées, et les recommandations."
                lignes_factices.append(ligne_data)
            bloc_data["lignes"] = lignes_factices
        elif bloc_type == "graphique":
            # Pas de lignes mais on génère un dataset factice (si supporté)
            bloc_data["lignes"] = []
            bloc_data["valeurs_graphique"] = {
                "labels": ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"],
                "datasets": [{"label": "Exemple", "data": [12, 19, 15, 25, 22, 30, 28]}]
            }
        elif bloc_type == "preconisations":
            # Pièces critiques factices
            bloc_data["lignes"] = []
            bloc_data["pieces"] = [
                {"type_piece": "Batterie", "reference": "NPC65", "quantite": 36,
                 "date_installation": "2020-01-01", "date_fin_de_vie": "2025-01-01",
                 "statut": "A_REMPLACER", "preconisation": "Fin de vie dépassée (2025) — remplacer", "preconisation_level": "critique"},
                {"type_piece": "Filtre", "reference": "FLT-001", "quantite": 4,
                 "date_installation": "2024-06-15", "date_fin_de_vie": "2027-06-15",
                 "statut": "OK", "preconisation": "OK — fin de vie prévue en 2027", "preconisation_level": "ok"},
            ]
        elif bloc_type == "mesures_batteries":
            bloc_data["lignes"] = []
            # Charger les options pour V_nom + tolerance
            try:
                fopts = json.loads(b.get("field_options") or "{}")
            except Exception:
                fopts = {}
            v_nom = float(fopts.get("tension_nominale") or 12.0)
            tol = float(fopts.get("tolerance_pct") or 10.0)
            # 12 batteries factices : 10 OK + 2 hors plage
            bloc_data["chains"] = [{"chain_idx": 0, "nb_batteries": 12, "piece_type": "Batterie 12V", "piece_id": 1}]
            bv = {}
            seuil_bas = v_nom * (1 - tol/100)
            seuil_haut = v_nom * (1 + tol/100)
            for i in range(1, 13):
                if i == 3:
                    bv[f"chain0_pos{i}"] = f"{seuil_bas - 0.5:.2f}"  # rouge bas
                elif i == 7:
                    bv[f"chain0_pos{i}"] = f"{seuil_haut + 0.3:.2f}"  # rouge haut
                else:
                    bv[f"chain0_pos{i}"] = f"{v_nom + 0.1:.2f}"  # vert
            bloc_data["bat_values"] = bv
        elif bloc_type == "image":
            # Module image — laisser data['equipement_image_path'] non renseigné
            # → le PDF affichera "Aucune image associée"
            bloc_data["lignes"] = []
        mesures_techniques.append(bloc_data)

    # Données globales factices
    fake_data = {
        "type_label": "MAINTENANCE",
        "sous_type": st["nom"],
        "numero_iv": "BC0000",
        "date": datetime.now().strftime("%d/%m/%Y"),
        "client": "Client de démo",
        "numero_projet": "P0000",
        "projet_nom": "Projet exemple",
        "projet_logo_path": "",
        "equipement": "Équipement exemple",
        "marque_modele": "Marque / Modèle",
        "localisation": "Local exemple",
        "technique": technique["nom"],
        "intervenants": "Technicien exemple",
        "is_maintenance": True,
        "comptes_rendus": [
            {"numero": "CR-DEMO", "date": datetime.now().strftime("%d/%m/%Y"),
             "observations": "Ceci est un exemple d'observation. Lors d'une vraie intervention, ce texte contiendra le compte-rendu rédigé par le technicien.",
             "intervenants": [{"nom": "Technicien Démo", "date": datetime.now().strftime("%Y-%m-%d"), "heure_debut": "08:00", "heure_fin": "16:00", "total_heures": 8.0}]}
        ],
        "gamme_operations": [
            {"gamme_nom": "Gamme exemple", "operations": [
                {"description": "Vérification visuelle de l'équipement", "done": True, "date_realisation": ""},
                {"description": "Mesure des tensions et courants", "done": True, "date_realisation": ""},
                {"description": "Test des sécurités", "done": False, "date_realisation": ""},
            ]}
        ],
        "mesures_techniques": mesures_techniques,
        "equipement_image_path": None,
    }
    try:
        from rapport_pdf import generate_rapport
        pdf_bytes = generate_rapport(fake_data)
    except Exception as e:
        return jsonify({"error": f"Erreur génération PDF : {e}"}), 500
    from flask import make_response
    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f'inline; filename="apercu-{tid}-{sid}.pdf"'
    return response

@app.route("/api/interventions/<int:iid>/mesures/debug")
def debug_intervention_mesures(iid):
    """Endpoint de debug : explique pourquoi un bon n'affiche pas les mesures.
    Pas d'auth requise, mais nécessite ?secret=socom-debug-2026 pour limiter l'accès."""
    if request.args.get("secret") != "socom-debug-2026":
        return jsonify({"error": "Ajoute ?secret=socom-debug-2026 à l'URL"}), 401
    db = get_db()
    diag = {"intervention_id": iid, "etapes": []}
    iv = one(db.execute("""
        SELECT i.id, i.type, i.equipement_id, e.designation AS equip_nom,
               e.technique_id, e.type_technique
        FROM interventions i
        JOIN equipements e ON i.equipement_id = e.id
        WHERE i.id = ?
    """, (iid,)))
    if not iv:
        diag["erreur"] = "Intervention introuvable"
        return jsonify(diag), 404
    diag["intervention"] = {
        "type": iv["type"],
        "equipement": iv["equip_nom"],
        "equipement_id": iv["equipement_id"],
        "technique_id": iv["technique_id"],
        "type_technique_legacy": iv["type_technique"],
    }
    # 1. Vérif technique_id (avec fallback sur type_technique)
    technique_id_resolved = _resolve_technique_id(iv["equipement_id"])
    if not technique_id_resolved:
        diag["etapes"].append("❌ Aucune technique trouvée pour cet équipement (technique_id=" + str(iv["technique_id"]) + ", type_technique='" + str(iv["type_technique"]) + "')")
        diag["solution"] = "Soit associer cet équipement à une technique dans sa fiche, soit créer une technique nommée '" + (iv["type_technique"] or "") + "' (insensible à la casse)."
        # Lister les techniques disponibles
        all_t = rows(db.execute("SELECT id, nom FROM techniques ORDER BY nom"))
        diag["techniques_existantes"] = [{"id": t["id"], "nom": t["nom"]} for t in all_t]
        return jsonify(diag)
    if iv["technique_id"]:
        diag["etapes"].append("✓ technique_id direct = " + str(technique_id_resolved))
    else:
        diag["etapes"].append("✓ technique résolue via fallback type_technique='" + str(iv["type_technique"]) + "' → technique_id=" + str(technique_id_resolved))
    # Nom de la technique
    t = one(db.execute("SELECT nom FROM techniques WHERE id=?", (technique_id_resolved,)))
    diag["technique_nom"] = t["nom"] if t else "(inconnu)"
    diag["etapes"].append("✓ technique nom = " + (t["nom"] if t else "?"))
    # 2. Sous-type via les gammes du bon
    iv_ops = rows(db.execute("""
        SELECT DISTINCT g.id AS gamme_id, g.nom AS gamme_nom, g.sous_type_id, st.nom AS sous_type_nom
        FROM intervention_operations io
        JOIN gamme_operations go ON io.gamme_operation_id = go.id
        JOIN gammes g ON go.gamme_id = g.id
        LEFT JOIN maintenance_sous_types st ON g.sous_type_id = st.id
        WHERE io.intervention_id = ?
    """, (iid,)))
    diag["gammes_du_bon"] = [{"nom": g["gamme_nom"], "sous_type": g["sous_type_nom"]} for g in iv_ops]
    if not iv_ops:
        diag["etapes"].append("❌ Aucune gamme trouvée via intervention_operations (le bon n'a pas d'opérations)")
        diag["solution"] = "Le bon doit avoir au moins une opération de gamme cochée. Le sous-type est déduit des gammes UTILISÉES (cochées dans le CR), pas seulement attribuées au bon."
        return jsonify(diag)
    sous_types_distincts = list(set(g["sous_type_nom"] for g in iv_ops if g["sous_type_nom"]))
    sans_sous_type = [g for g in iv_ops if not g["sous_type_id"]]
    if sans_sous_type:
        diag["etapes"].append("❌ " + str(len(sans_sous_type)) + " gamme(s) sans sous-type → fallback")
        diag["solution"] = "Toutes les gammes du bon doivent avoir le même sous-type pour que les mesures s'affichent."
        return jsonify(diag)
    if len(sous_types_distincts) > 1:
        diag["etapes"].append("❌ Mélange de sous-types (" + ", ".join(sous_types_distincts) + ") → fallback")
        return jsonify(diag)
    diag["etapes"].append("✓ sous-type unique = " + sous_types_distincts[0])
    diag["sous_type_label"] = sous_types_distincts[0]
    st_row = one(db.execute("SELECT id FROM maintenance_sous_types WHERE nom=?", (sous_types_distincts[0],)))
    if not st_row:
        diag["etapes"].append("❌ Sous-type non trouvé en table")
        return jsonify(diag)
    diag["sous_type_id"] = st_row["id"]
    # 3. Blocs configurés
    blocs = rows(db.execute("""
        SELECT id, nom FROM mesure_blocs
        WHERE technique_id=? AND sous_type_id=?
    """, (technique_id_resolved, st_row["id"])))
    diag["blocs_trouves"] = [b["nom"] for b in blocs]
    if not blocs:
        diag["etapes"].append("❌ Aucun bloc configuré pour technique_id=" + str(technique_id_resolved) + " × sous_type_id=" + str(st_row["id"]))
        diag["solution"] = "Va dans Paramètres → Mesures techniques et vérifie que les blocs sont bien attribués à la technique '" + diag["technique_nom"] + "' et au sous-type '" + diag["sous_type_label"] + "'."
        # Lister ce qui est configuré
        all_blocs = rows(db.execute("""
            SELECT mb.nom, t.nom AS tech_nom, st.nom AS st_nom
            FROM mesure_blocs mb
            LEFT JOIN techniques t ON mb.technique_id = t.id
            LEFT JOIN maintenance_sous_types st ON mb.sous_type_id = st.id
        """))
        diag["blocs_existants_dans_la_db"] = [
            {"nom": b["nom"], "technique": b["tech_nom"], "sous_type": b["st_nom"]}
            for b in all_blocs
        ]
        return jsonify(diag)
    diag["etapes"].append("✓ " + str(len(blocs)) + " bloc(s) trouvé(s) — les mesures DEVRAIENT s'afficher")
    return jsonify(diag)


@app.route("/api/interventions/<int:iid>/mesures")
@require_auth
def get_intervention_mesures(iid):
    """Retourne les blocs+lignes pertinents pour une intervention (selon sa technique
    et son sous-type), avec les valeurs déjà saisies."""
    db = get_db()
    iv = one(db.execute(
        "SELECT i.id, i.type, i.equipement_id FROM interventions i WHERE i.id = ?",
        (iid,)
    ))
    if not iv: return jsonify({"error": "Intervention introuvable"}), 404
    # Résoudre la technique avec fallback type_technique → techniques.nom
    technique_id = _resolve_technique_id(iv["equipement_id"])
    # Sous-type déduit des gammes
    sous_type_label = _bon_sous_type_label(iid)
    sous_type_id = None
    if sous_type_label:
        st = one(db.execute("SELECT id FROM maintenance_sous_types WHERE nom=?", (sous_type_label,)))
        if st: sous_type_id = st["id"]
    # Si pas de sous-type identifié (mélange ou aucune gamme) → aucun bloc applicable
    if not technique_id or not sous_type_id:
        return jsonify({"blocs": [], "valeurs": {},
                        "technique_id": technique_id,
                        "sous_type_id": sous_type_id,
                        "sous_type_label": sous_type_label})
    blocs = rows(db.execute("""
        SELECT mb.id, mb.nom, mb.ordre, mb.type, mb.field_options, mb.largeur, mb.page_num, mb.icon FROM mesure_blocs mb
        WHERE mb.technique_id = ? AND mb.sous_type_id = ?
        ORDER BY mb.page_num, mb.ordre, mb.nom
    """, (technique_id, sous_type_id)))
    for b in blocs:
        b["lignes"] = rows(db.execute(
            "SELECT * FROM mesure_lignes WHERE bloc_id=? ORDER BY ordre, id",
            (b["id"],)
        ))
        # v211 — Pour les modules 'preconisations', charger les pièces enrichies
        if (b.get("type") or "") == "preconisations":
            bcfg = {}
            try:
                bcfg = json.loads(b["field_options"]) if b.get("field_options") else {}
            except Exception:
                bcfg = {}
            seuil = int(bcfg.get("seuil_mois") or 6)
            pieces_rows = rows(db.execute(
                "SELECT * FROM pieces WHERE equipement_id=? ORDER BY date_fin_de_vie",
                (iv["equipement_id"],)
            ))
            pieces_data = []
            for p in pieces_rows:
                ns = statut_piece(p.get("date_fin_de_vie"))
                if ns != p.get("statut") and p.get("date_fin_de_vie"):
                    p["statut"] = ns
                prec_text, prec_level = _piece_preconisation(p, seuil_mois=seuil)
                p["preconisation"] = prec_text
                p["preconisation_level"] = prec_level
                pieces_data.append(p)
            b["pieces"] = pieces_data
        # v211 — Pour les modules 'mesures_batteries', charger les chaînes + valeurs saisies
        if (b.get("type") or "") == "mesures_batteries":
            chains = _detect_batteries_chaines(iv["equipement_id"])
            bat_values = {}
            bat_lignes = rows(db.execute(
                "SELECT id, libelle FROM mesure_lignes WHERE bloc_id=? AND libelle LIKE '_bat_%'",
                (b["id"],)
            ))
            if bat_lignes:
                line_ids = [l["id"] for l in bat_lignes]
                placeholders = ",".join("?" * len(line_ids))
                val_rows_bat = rows(db.execute(
                    f"SELECT ligne_id, valeur FROM intervention_mesures WHERE intervention_id=? AND ligne_id IN ({placeholders})",
                    (iid, *line_ids)
                ))
                val_by_lid = {v["ligne_id"]: v["valeur"] for v in val_rows_bat}
                for l in bat_lignes:
                    key = l["libelle"][len("_bat_"):]
                    if l["id"] in val_by_lid:
                        bat_values[key] = val_by_lid[l["id"]]
            b["chains"] = chains
            b["bat_values"] = bat_values
        # v218.88 : Pour les modules 'equipement', charger les données de l'équipement
        if (b.get("type") or "") == "equipement":
            eq_data = _bloc_equipement_data(iv["equipement_id"], iid, db)
            b["equipement_data"] = eq_data
        # v217.15 : Pour les modules 'image', charger le filename de l'image principale de l'équipement
        if (b.get("type") or "") == "image":
            iv2 = one(db.execute("""
                SELECT il.filename, il.nom AS image_nom
                FROM equipements e
                LEFT JOIN image_library il ON e.image_id = il.id
                WHERE e.id = ?
            """, (iv["equipement_id"],)))
            if iv2 and iv2.get("filename"):
                b["equipement_image_filename"] = iv2["filename"]
                b["equipement_image_nom"] = iv2.get("image_nom") or ""
    # Valeurs saisies
    val_rows = rows(db.execute(
        "SELECT ligne_id, valeur FROM intervention_mesures WHERE intervention_id=?",
        (iid,)
    ))
    valeurs = {v["ligne_id"]: v["valeur"] for v in val_rows}
    return jsonify({
        "blocs": blocs,
        "valeurs": valeurs,
        "technique_id": technique_id,
        "sous_type_id": sous_type_id,
        "sous_type_label": sous_type_label,
    })

@app.route("/api/interventions/<int:iid>/mesures", methods=["PUT"])
@require_auth
def save_intervention_mesures(iid):
    """Enregistre les valeurs des mesures pour un bon.
    Body : { valeurs: { ligne_id: valeur, ... }, valeurs_batteries: { bloc_id: { key: valeur, ... } } }
    Les valeurs_batteries créent des mesure_lignes à la volée (clé = libellé virtuel).
    """
    d = request.json or {}
    valeurs = d.get("valeurs") or {}
    valeurs_batteries = d.get("valeurs_batteries") or {}
    db = get_db()
    if not one(db.execute("SELECT id FROM interventions WHERE id=?", (iid,))):
        return jsonify({"error": "Intervention introuvable"}), 404
    n = 0
    # ─── Valeurs classiques (mesure_lignes existantes) ───
    for ligne_id, valeur in valeurs.items():
        try:
            lid = int(ligne_id)
        except (ValueError, TypeError):
            continue
        # Vérifier que la ligne existe
        if not one(db.execute("SELECT id FROM mesure_lignes WHERE id=?", (lid,))):
            continue
        v = (str(valeur) if valeur is not None else "").strip()
        if v:
            db.execute("""INSERT INTO intervention_mesures (intervention_id, ligne_id, valeur)
                          VALUES (?, ?, ?)
                          ON CONFLICT(intervention_id, ligne_id)
                          DO UPDATE SET valeur=excluded.valeur""",
                       (iid, lid, v))
        else:
            # Valeur vide → supprimer l'enregistrement
            db.execute("DELETE FROM intervention_mesures WHERE intervention_id=? AND ligne_id=?",
                       (iid, lid))
        n += 1
    # ─── Valeurs batteries (création de mesure_lignes à la volée) ───
    # Format reçu : { bloc_id: { "bat_chain0_pos0": "12.5", "bat_chain0_pos1": "12.4", ... } }
    for bloc_id_raw, vals_dict in valeurs_batteries.items():
        try:
            bloc_id = int(bloc_id_raw)
        except (ValueError, TypeError):
            continue
        # Vérifier que c'est bien un module mesures_batteries
        bloc_row = one(db.execute("SELECT type FROM mesure_blocs WHERE id=?", (bloc_id,)))
        if not bloc_row or bloc_row.get("type") != "mesures_batteries":
            continue
        # Indexer les lignes existantes du bloc par leur libellé
        existing_lignes = rows(db.execute("SELECT id, libelle FROM mesure_lignes WHERE bloc_id=?", (bloc_id,)))
        by_libelle = {l["libelle"]: l["id"] for l in existing_lignes}
        for key, valeur in (vals_dict or {}).items():
            # Le libellé virtuel : "bat_chain<idx>_pos<n>"
            libelle_virtual = "_bat_" + str(key)
            if libelle_virtual in by_libelle:
                lid = by_libelle[libelle_virtual]
            else:
                # Créer la ligne virtuelle
                db.execute(
                    "INSERT INTO mesure_lignes (bloc_id, libelle, unite, ordre, field_type, field_options) VALUES (?, ?, ?, ?, ?, ?)",
                    (bloc_id, libelle_virtual, "Vdc", 0, "battery", "")
                )
                lid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
                by_libelle[libelle_virtual] = lid
            v = (str(valeur) if valeur is not None else "").strip()
            if v:
                db.execute("""INSERT INTO intervention_mesures (intervention_id, ligne_id, valeur)
                              VALUES (?, ?, ?)
                              ON CONFLICT(intervention_id, ligne_id)
                              DO UPDATE SET valeur=excluded.valeur""",
                           (iid, lid, v))
            else:
                db.execute("DELETE FROM intervention_mesures WHERE intervention_id=? AND ligne_id=?",
                           (iid, lid))
            n += 1
    db.commit()
    return jsonify({"ok": True, "lignes_traitees": n})


def _resolve_technique_id(equipement_id):
    """Retourne le technique_id d'un équipement.
    Fallback : si la colonne technique_id est NULL, on cherche la technique par
    le nom dans equipements.type_technique (legacy).
    """
    db = get_db()
    e = one(db.execute(
        "SELECT technique_id, type_technique FROM equipements WHERE id=?",
        (equipement_id,)
    ))
    if not e:
        return None
    if e["technique_id"]:
        return e["technique_id"]
    # Fallback : matcher par nom
    type_tech = (e.get("type_technique") or "").strip()
    if not type_tech:
        return None
    t = one(db.execute(
        "SELECT id FROM techniques WHERE LOWER(nom) = LOWER(?)",
        (type_tech,)
    ))
    return t["id"] if t else None


def _detect_batteries_chaines(equipement_id):
    """Détecte les batteries d'un équipement et leur répartition en chaînes.
    Cherche les pièces dont type_piece contient 'batterie' (case-insensitive).

    Retour : liste de dicts { 'piece_id': X, 'chain_idx': N, 'nb_batteries': M, 'piece_type': 'Batterie 12V' }
    Une 'piece' avec quantite=72 et nbr_chaine=2 → produit 2 entrées (chain_idx=0 et 1) avec nb_batteries=36 chacune.
    """
    db = get_db()
    pieces = rows(db.execute(
        "SELECT id, type_piece, quantite, nbr_chaine FROM pieces "
        "WHERE equipement_id=? AND LOWER(type_piece) LIKE '%batterie%' "
        "ORDER BY id",
        (equipement_id,)
    ))
    logger.info(f"[BAT-DEBUG] equipement_id={equipement_id} pieces_found={len(pieces)} pieces={[dict(p) for p in pieces]}")
    chains = []
    chain_idx_counter = 0
    for p in pieces:
        nb_total = int(p.get("quantite") or 0)
        if nb_total <= 0:
            logger.info(f"[BAT-DEBUG] piece {p['id']} skipped: quantite={p.get('quantite')}")
            continue
        nb_chains = int(p.get("nbr_chaine") or 1)
        if nb_chains < 1:
            nb_chains = 1
        # Répartir équitablement nb_total sur nb_chains
        # Ex : 72 batteries / 2 chaînes → 36 + 36
        # Ex : 73 batteries / 2 chaînes → 37 + 36
        base = nb_total // nb_chains
        reste = nb_total - base * nb_chains
        for c in range(nb_chains):
            count = base + (1 if c < reste else 0)
            if count <= 0:
                continue
            chains.append({
                "piece_id": p["id"],
                "piece_type": p["type_piece"],
                "chain_idx": chain_idx_counter,
                "nb_batteries": count,
            })
            chain_idx_counter += 1
    logger.info(f"[BAT-DEBUG] equipement_id={equipement_id} → chains={chains}")
    return chains


def _piece_preconisation(piece, seuil_mois=6):
    """Calcule un texte de préconisation pour une pièce.
    Retourne (texte, niveau) où niveau ∈ {'critique','warning','info','ok'}.
    """
    statut = (piece.get("statut") or "").upper()
    dfv_str = piece.get("date_fin_de_vie") or ""
    today = datetime.now().date()
    dfv = None
    if dfv_str:
        try:
            dfv = datetime.strptime(dfv_str, "%Y-%m-%d").date()
        except Exception:
            dfv = None

    # Statut explicite : prioritaire
    if statut == "A_REMPLACER":
        return ("À remplacer immédiatement", "critique")

    # Date dépassée
    if dfv and dfv < today:
        return (f"Fin de vie dépassée ({dfv.strftime('%Y')}) — remplacer", "critique")

    # Date proche (selon seuil)
    if dfv:
        delta_days = (dfv - today).days
        seuil_days = int(seuil_mois) * 30
        if delta_days <= seuil_days:
            mois_restants = max(0, delta_days // 30)
            jours_restants = delta_days % 30
            if mois_restants == 0:
                return (f"Fin de vie dans {jours_restants} jour(s) — à anticiper", "warning")
            return (f"Fin de vie dans ~{mois_restants} mois — à anticiper", "warning")

    if statut == "A_SURVEILLER":
        return ("À surveiller (statut manuel)", "info")

    if dfv:
        return (f"OK — fin de vie prévue en {dfv.strftime('%Y')}", "ok")
    return ("OK", "ok")


@app.route("/api/interventions/<int:iid>/pieces")
@require_auth
def get_intervention_pieces(iid):
    """Retourne les pièces de l'équipement d'une intervention, enrichies de la préconisation.
    Optionnel : ?seuil_mois=N (défaut: 6)."""
    db = get_db()
    iv = one(db.execute("SELECT equipement_id FROM interventions WHERE id=?", (iid,)))
    if not iv:
        return jsonify({"error": "Intervention introuvable"}), 404
    seuil = int(request.args.get("seuil_mois") or 6)
    pieces = rows(db.execute(
        "SELECT * FROM pieces WHERE equipement_id=? ORDER BY date_fin_de_vie",
        (iv["equipement_id"],)
    ))
    out = []
    for p in pieces:
        # Recalcul du statut basé sur la date (cohérent avec get_pieces)
        ns = statut_piece(p.get("date_fin_de_vie"))
        if ns != p.get("statut") and p.get("date_fin_de_vie"):
            db.execute("UPDATE pieces SET statut=? WHERE id=?", (ns, p["id"]))
            p["statut"] = ns
        # Calcul préconisation
        prec_text, prec_level = _piece_preconisation(p, seuil_mois=seuil)
        p["preconisation"] = prec_text
        p["preconisation_level"] = prec_level
        out.append(p)
    db.commit()
    return jsonify(out)


def _generate_module_graphique_png(bloc):
    """Génère un PNG du graphique pour un bloc de type 'graphique'.
    Reçoit le dict du bloc avec 'field_options' (config) et 'lignes' (points avec valeur réf et mesurée).
    Retourne les bytes PNG, ou None si échec.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import io as _io
        # Config
        cfg = {}
        raw = bloc.get("field_options") or ""
        if raw:
            try: cfg = json.loads(raw) if isinstance(raw, str) else raw
            except Exception: cfg = {}
        title = cfg.get("title") or bloc.get("nom") or ""
        x_label = cfg.get("x_label") or ""
        y_label = cfg.get("y_label") or ""
        y_unit = cfg.get("y_unit") or ""
        ref_name = cfg.get("ref_series_name") or "Référence"
        mes_name = cfg.get("mes_series_name") or "Mesuré"
        ref_color = cfg.get("ref_color") or "#1E40AF"
        mes_color = cfg.get("mes_color") or "#DC2626"
        # Points : labels X, valeurs ref et mesurées
        labels = []
        vals_ref = []
        vals_mes = []
        for lg in bloc.get("lignes", []):
            labels.append(lg.get("libelle") or "")
            # valeur ref dans field_options de la ligne
            lg_opts = {}
            lr = lg.get("field_options") or ""
            if lr:
                try: lg_opts = json.loads(lr) if isinstance(lr, str) else lr
                except Exception: lg_opts = {}
            # v218.60 : helper qui accepte virgule décimale française
            def _to_float_fr(v):
                if v is None or v == "":
                    return None
                try:
                    return float(str(v).replace(",", ".").strip())
                except Exception:
                    return None
            vref = _to_float_fr(lg_opts.get("ref_value"))
            vals_ref.append(vref)
            vmes = _to_float_fr(lg.get("valeur"))
            vals_mes.append(vmes)
        if not labels:
            return None
        fig, ax = plt.subplots(figsize=(8, 4), dpi=120)
        x_idx = list(range(len(labels)))
        # Tracer chaque série en filtrant les None pour ne pas casser la ligne
        def _filter(values):
            xs = []; ys = []
            for i, v in enumerate(values):
                if v is not None:
                    xs.append(i); ys.append(v)
            return xs, ys
        xref, yref = _filter(vals_ref)
        xmes, ymes = _filter(vals_mes)
        if xref:
            ax.plot(xref, yref, marker='o', linewidth=2.0, markersize=6, color=ref_color, label=ref_name)
        if xmes:
            ax.plot(xmes, ymes, marker='s', linewidth=2.0, markersize=6, color=mes_color, label=mes_name)
        ax.set_xticks(x_idx)
        ax.set_xticklabels(labels, fontsize=9)
        ax.set_xlabel(x_label, fontsize=10)
        ylabel_full = y_label
        if y_unit:
            ylabel_full = f"{y_label} ({y_unit})" if y_label else f"({y_unit})"
        ax.set_ylabel(ylabel_full, fontsize=10)
        if title:
            ax.set_title(title, fontsize=12, fontweight='bold', color='#1E3A8A')
        ax.grid(True, alpha=0.3, linestyle='--')
        if xref or xmes:
            ax.legend(loc='best', fontsize=9, framealpha=0.95)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        plt.tight_layout()
        buf = _io.BytesIO()
        fig.savefig(buf, format='png', bbox_inches='tight', dpi=120)
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.warning(f"[graph] échec génération module graphique : {e}")
        return None


def _bloc_equipement_data(equipement_id, intervention_id, db):
    """v218.88 — Charge les données d'équipement nécessaires au module 'equipement'.
    v218.93 — Inclut maintenant les champs HT (transformateur), les cellules et tous les tableaux/bornes liés.
    Retourne un dict avec toutes les valeurs résolues."""
    if not equipement_id:
        return {}
    eq = one(db.execute("""
        SELECT e.id, e.designation, e.type_technique, e.localisation, e.marque, e.modele,
               e.puissance, e.numero_serie, e.in_out, e.date_mise_en_service, e.statut, e.notes,
               e.trafo_marque, e.trafo_annee, e.trafo_numero_serie, e.trafo_puissance_kva,
               e.trafo_refroidissement, e.trafo_poids_kg, e.trafo_tension_entree_v,
               e.trafo_courant_a, e.trafo_norme, e.trafo_couplage,
               e.trafo_tension_service_v, e.trafo_reglage_tension_kv,
               t.nom AS technique_nom
        FROM equipements e
        LEFT JOIN techniques t ON t.id = e.technique_id
        WHERE e.id = ?
    """, (equipement_id,)))
    if not eq:
        return {}
    # Marque + Modèle concaténés
    marque = (eq.get("marque") or "").strip()
    modele = (eq.get("modele") or "").strip()
    if marque and modele:
        marque_modele = f"{marque} — {modele}"
    elif marque:
        marque_modele = marque
    elif modele:
        marque_modele = modele
    else:
        marque_modele = ""
    # Tableau lié au bon (via intervention.tableau_id)
    tableau_label = "Tableau"
    tableau_value = ""
    if intervention_id:
        iv_row = one(db.execute("SELECT tableau_id, type_technique FROM interventions i LEFT JOIN equipements e ON e.id=i.equipement_id WHERE i.id=?", (intervention_id,)))
        if iv_row and iv_row.get("type_technique") and "borne" in (iv_row["type_technique"] or "").lower():
            tableau_label = "Borne"
        # Récupérer le tableau lié au bon
        tab_row = one(db.execute(
            "SELECT et.nom, et.localisation FROM interventions i LEFT JOIN equipement_tableaux et ON et.id=i.tableau_id WHERE i.id=? AND i.tableau_id IS NOT NULL",
            (intervention_id,)
        ))
        if tab_row and tab_row.get("nom"):
            tn = tab_row["nom"]
            tl = tab_row.get("localisation") or ""
            tableau_value = f"{tn} ({tl})" if tl else tn
    # Statut humain
    statut_labels = {
        'EN_SERVICE': 'En service', 'HORS_SERVICE': 'Hors service',
        'EN_PANNE': 'En panne', 'EN_MAINTENANCE': 'En maintenance', 'ARCHIVE': 'Archivé',
    }

    # v218.93 : Cellules HT (sous-équipements)
    cellules_list = []
    try:
        cellules_list = [dict(r) for r in rows(db.execute(
            "SELECT designation, marque, type FROM equipement_cellules WHERE equipement_id=? ORDER BY ordre, id",
            (equipement_id,)
        ))]
    except Exception:
        cellules_list = []

    # v218.93 : Tableaux/Bornes (sous-équipements) — tous ceux liés à cet équipement
    tableaux_list = []
    try:
        tableaux_list = [dict(r) for r in rows(db.execute(
            "SELECT nom, localisation FROM equipement_tableaux WHERE equipement_id=? ORDER BY ordre, id",
            (equipement_id,)
        ))]
    except Exception:
        tableaux_list = []

    # Type technique (label affichable selon la technique)
    type_technique = (eq.get("type_technique") or "").lower()
    is_borne = "borne" in type_technique
    sub_label = "Bornes" if is_borne else "Tableaux"

    return {
        # Champs basiques
        "designation": eq.get("designation") or "",
        "marque_modele": marque_modele,
        "puissance": eq.get("puissance") or "",
        "numero_serie": eq.get("numero_serie") or "",
        "in_out": eq.get("in_out") or "",
        "localisation": eq.get("localisation") or "",
        "tableau_label": tableau_label,
        "tableau": tableau_value,
        "technique": eq.get("technique_nom") or eq.get("type_technique") or "",
        "date_mise_service": eq.get("date_mise_en_service") or "",
        "statut": statut_labels.get(eq.get("statut") or "", eq.get("statut") or ""),
        "notes": eq.get("notes") or "",
        # v218.93 : Champs Transformateur (HT)
        "trafo_marque": eq.get("trafo_marque") or "",
        "trafo_annee": eq.get("trafo_annee") or "",
        "trafo_numero_serie": eq.get("trafo_numero_serie") or "",
        "trafo_puissance_kva": eq.get("trafo_puissance_kva") or "",
        "trafo_refroidissement": eq.get("trafo_refroidissement") or "",
        "trafo_poids_kg": eq.get("trafo_poids_kg") or "",
        "trafo_tension_entree_v": eq.get("trafo_tension_entree_v") or "",
        "trafo_courant_a": eq.get("trafo_courant_a") or "",
        "trafo_norme": eq.get("trafo_norme") or "",
        "trafo_couplage": eq.get("trafo_couplage") or "",
        "trafo_tension_service_v": eq.get("trafo_tension_service_v") or "",
        "trafo_reglage_tension_kv": eq.get("trafo_reglage_tension_kv") or "",
        # v218.93 : Sous-équipements
        "cellules_list": cellules_list,
        "tableaux_list": tableaux_list,
        "sub_label_tableaux": sub_label,  # "Tableaux" ou "Bornes"
        "type_technique": type_technique,
    }


def _intervention_mesures_data(intervention_id):
    """Construit la liste des blocs/lignes avec valeurs pour insertion dans le PDF.
    Retourne une liste : [{nom, lignes: [{libelle, unite, valeur}, ...]}, ...]
    v218.112 : pour les bons DÉPANNAGE, charge les modules globaux BC (is_bc=1)
    au lieu des modules technique × sous-type.
    Si aucun bloc applicable → liste vide.
    """
    db = get_db()
    iv = one(db.execute(
        "SELECT i.id, i.equipement_id, i.type, i.societe_id FROM interventions i WHERE i.id = ?",
        (intervention_id,)
    ))
    if not iv:
        return []
    # v218.112 : Pour un bon de DÉPANNAGE (BC), on utilise les modules globaux is_bc=1
    iv_type = (iv.get("type") or "").upper()
    if iv_type == "DEPANNAGE":
        sid = iv.get("societe_id") or current_societe_id()
        blocs = rows(db.execute("""
            SELECT id, nom, type, field_options, largeur, page_num, icon FROM mesure_blocs
            WHERE is_bc=1 AND societe_id=?
            ORDER BY page_num, ordre, nom
        """, (sid,)))
    else:
        technique_id = _resolve_technique_id(iv["equipement_id"])
        if not technique_id:
            return []
        sous_type_label = _bon_sous_type_label(intervention_id)
        if not sous_type_label:
            return []
        st = one(db.execute("SELECT id FROM maintenance_sous_types WHERE nom=?", (sous_type_label,)))
        if not st:
            return []
        blocs = rows(db.execute("""
            SELECT id, nom, type, field_options, largeur, page_num, icon FROM mesure_blocs
            WHERE technique_id=? AND sous_type_id=? AND (is_bc=0 OR is_bc IS NULL)
            ORDER BY page_num, ordre, nom
        """, (technique_id, st["id"])))
    if not blocs:
        return []
    val_rows = rows(db.execute(
        "SELECT ligne_id, valeur FROM intervention_mesures WHERE intervention_id=?",
        (intervention_id,)
    ))
    valeurs = {v["ligne_id"]: v["valeur"] for v in val_rows}
    out = []
    for b in blocs:
        lignes = rows(db.execute(
            "SELECT * FROM mesure_lignes WHERE bloc_id=? ORDER BY ordre, id",
            (b["id"],)
        ))
        bloc_data = {
            "nom": b["nom"],
            "type": b.get("type") or "tableau",
            "field_options": b.get("field_options") or "",
            "largeur": b.get("largeur") or "pleine",
            "page_num": int(b.get("page_num") or 1),
            "icon": b.get("icon") or "",
            "lignes": [{
                "libelle": lg["libelle"],
                "unite": lg["unite"] or "",
                "field_type": lg.get("field_type") or "numeric",
                "field_options": lg.get("field_options") or "",
                "valeur": valeurs.get(lg["id"], "") or "",
            } for lg in lignes]
        }
        # Pour les modules 'graphique' : pré-générer le PNG matplotlib
        if bloc_data["type"] == "graphique":
            png_bytes = _generate_module_graphique_png(bloc_data)
            if png_bytes:
                bloc_data["graph_png"] = png_bytes
        # Pour les modules 'preconisations' : charger les pièces enrichies
        if bloc_data["type"] == "preconisations":
            iv = one(db.execute("SELECT equipement_id FROM interventions WHERE id=?", (intervention_id,)))
            if iv:
                # Récupérer le seuil depuis field_options du bloc
                bcfg = {}
                try:
                    bcfg = json.loads(bloc_data["field_options"]) if bloc_data["field_options"] else {}
                except Exception:
                    bcfg = {}
                seuil = int(bcfg.get("seuil_mois") or 6)
                pieces_rows = rows(db.execute(
                    "SELECT * FROM pieces WHERE equipement_id=? ORDER BY date_fin_de_vie",
                    (iv["equipement_id"],)
                ))
                pieces_data = []
                for p in pieces_rows:
                    ns = statut_piece(p.get("date_fin_de_vie"))
                    if ns != p.get("statut") and p.get("date_fin_de_vie"):
                        p["statut"] = ns
                    prec_text, prec_level = _piece_preconisation(p, seuil_mois=seuil)
                    p["preconisation"] = prec_text
                    p["preconisation_level"] = prec_level
                    pieces_data.append(p)
                bloc_data["pieces"] = pieces_data
        # Pour les modules 'mesures_batteries' : charger les chaînes détectées + valeurs saisies
        if bloc_data["type"] == "mesures_batteries":
            iv = one(db.execute("SELECT equipement_id FROM interventions WHERE id=?", (intervention_id,)))
            if iv:
                chains = _detect_batteries_chaines(iv["equipement_id"])
                # Récupérer les valeurs saisies (lignes virtuelles _bat_*)
                bat_values = {}
                bat_lignes = rows(db.execute(
                    "SELECT id, libelle FROM mesure_lignes WHERE bloc_id=? AND libelle LIKE '_bat_%'",
                    (b["id"],)
                ))
                if bat_lignes:
                    line_ids = [str(l["id"]) for l in bat_lignes]
                    placeholders = ",".join("?" * len(line_ids))
                    val_rows = rows(db.execute(
                        f"SELECT ligne_id, valeur FROM intervention_mesures WHERE intervention_id=? AND ligne_id IN ({placeholders})",
                        (intervention_id, *line_ids)
                    ))
                    val_by_lid = {v["ligne_id"]: v["valeur"] for v in val_rows}
                    for l in bat_lignes:
                        # Le libellé est "_bat_chain<idx>_pos<n>", on extrait la clé
                        key = l["libelle"][len("_bat_"):]  # retirer le préfixe
                        if l["id"] in val_by_lid:
                            # v218.60 : normaliser virgule → point pour que le rendu PDF puisse parser en float
                            raw_v = val_by_lid[l["id"]]
                            if raw_v and isinstance(raw_v, str):
                                raw_v = raw_v.replace(",", ".").strip()
                            bat_values[key] = raw_v
                bloc_data["chains"] = chains
                bloc_data["bat_values"] = bat_values
        # v217 : Pour les modules 'image' : récupérer le filename de l'image principale de l'équipement
        if bloc_data["type"] == "image":
            iv2 = one(db.execute("""
                SELECT il.filename, il.nom AS image_nom
                FROM interventions i
                JOIN equipements e ON i.equipement_id = e.id
                LEFT JOIN image_library il ON e.image_id = il.id
                WHERE i.id=?
            """, (intervention_id,)))
            if iv2 and iv2.get("filename"):
                bloc_data["equipement_image_filename"] = iv2["filename"]
                bloc_data["equipement_image_nom"] = iv2.get("image_nom") or ""
        out.append(bloc_data)
    return out


def _bon_sous_type_label(intervention_id):
    """Retourne le sous-type d'une intervention.
    v218.108 : nouvelle logique multi-cas :
    1. Si intervention_operations existe (= au moins une op cochée par le tech) → on prend
       le sous-type des gammes effectivement réalisées (s'il est unique).
    2. Sinon, on regarde toutes les gammes liées à l'équipement :
       - Si toutes ont le même sous-type → on retourne ce sous-type
       - Si elles ont des sous-types différents → on prend celui de la gamme avec la
         périodicité la plus longue (ANNUEL > SEMESTRIEL > ... > QUOTIDIEN).
         Cette heuristique correspond à la gamme "dominante" pour la date prévue.
    3. Fallback : aucune gamme avec sous-type → None.
    """
    db = get_db()
    iv = one(db.execute("SELECT equipement_id FROM interventions WHERE id=?", (intervention_id,)))
    if not iv:
        return None
    eq_id = iv["equipement_id"]

    # Priorité 1 : sous-types via les opérations cochées
    ops_st = rows(db.execute("""
        SELECT DISTINCT g.sous_type_id, st.nom
        FROM intervention_operations io
        JOIN gamme_operations go ON go.id = io.gamme_operation_id
        JOIN gammes g ON g.id = go.gamme_id
        LEFT JOIN maintenance_sous_types st ON g.sous_type_id = st.id
        WHERE io.intervention_id = ? AND g.sous_type_id IS NOT NULL
    """, (intervention_id,)))
    if ops_st:
        names = set(s["nom"] for s in ops_st if s.get("nom"))
        if len(names) == 1:
            return next(iter(names))
        # Plusieurs sous-types cochés : fallback sur la priorité périodicité ci-dessous

    # Toutes les gammes liées à cet équipement
    gids_rows = rows(db.execute("SELECT gamme_id FROM equipement_gammes WHERE equipement_id=?", (eq_id,)))
    gids = [g["gamme_id"] for g in gids_rows]
    eq_row = one(db.execute("SELECT gamme_id FROM equipements WHERE id=?", (eq_id,)))
    if eq_row and eq_row.get("gamme_id") and eq_row["gamme_id"] not in gids:
        gids.append(eq_row["gamme_id"])
    if not gids:
        return None
    placeholders = ",".join(["?"] * len(gids))
    gammes_st = rows(db.execute(f"""
        SELECT g.id, g.nom, g.periodicite, g.sous_type_id, st.nom AS st_nom
        FROM gammes g
        LEFT JOIN maintenance_sous_types st ON g.sous_type_id = st.id
        WHERE g.id IN ({placeholders})
    """, gids))
    # Filtrer celles qui ont un sous-type
    gammes_with_st = [g for g in gammes_st if g.get("sous_type_id") and g.get("st_nom")]
    if not gammes_with_st:
        return None
    # Toutes le même sous-type ?
    distinct_names = set(g["st_nom"] for g in gammes_with_st)
    if len(distinct_names) == 1:
        return next(iter(distinct_names))
    # Sinon → priorité à la périodicité la plus longue
    # Ordre décroissant : QUINQUENNAL > QUADRIENNAL > TRIENNAL > BISANNUEL > ANNUEL > SEMESTRIEL > QUADRIMESTRIEL > TRIMESTRIEL > BIMESTRIEL > MENSUEL > BIMENSUEL > HEBDOMADAIRE > QUOTIDIEN
    PERIOD_PRIORITY = {
        "QUINQUENNAL": 130, "QUADRIENNAL": 120, "TRIENNAL": 110, "BISANNUEL": 100,
        "ANNUEL": 90, "ANNUELLE": 90,
        "SEMESTRIEL": 80, "SEMESTRIELLE": 80,
        "QUADRIMESTRIEL": 70,
        "TRIMESTRIEL": 60, "TRIMESTRIELLE": 60,
        "BIMESTRIEL": 50, "BIMESTRIELLE": 50,
        "MENSUEL": 40, "MENSUELLE": 40,
        "BIMENSUEL": 30,
        "HEBDOMADAIRE": 20,
        "QUOTIDIEN": 10,
    }
    def _prio(g):
        p = (g.get("periodicite") or "").strip().upper()
        return PERIOD_PRIORITY.get(p, 0)
    gammes_with_st.sort(key=_prio, reverse=True)
    return gammes_with_st[0]["st_nom"]


@app.route("/api/gammes")
@require_auth
def get_gammes():
    db=get_db(); gammes=rows(db.execute("""
        SELECT g.*, st.nom AS sous_type_nom, st.couleur AS sous_type_couleur
        FROM gammes g
        LEFT JOIN maintenance_sous_types st ON g.sous_type_id = st.id
        ORDER BY g.nom
    """))
    for g in gammes:
        g["operations"]=rows(db.execute("SELECT * FROM gamme_operations WHERE gamme_id=? ORDER BY ordre",(g["id"],)))
    return jsonify(gammes)

@app.route("/api/gammes/<int:gid>")
@require_auth
def get_gamme(gid):
    db=get_db(); g=one(db.execute("""
        SELECT g.*, st.nom AS sous_type_nom, st.couleur AS sous_type_couleur
        FROM gammes g
        LEFT JOIN maintenance_sous_types st ON g.sous_type_id = st.id
        WHERE g.id=?
    """,(gid,)))
    if not g: return jsonify({"error":"Non trouve"}),404
    g["operations"]=rows(db.execute("SELECT * FROM gamme_operations WHERE gamme_id=? ORDER BY ordre",(gid,)))
    return jsonify(g)

@app.route("/api/gammes",methods=["POST"])
@require_role("admin","manager")
def create_gamme():
    d=request.json or {}
    if not d.get("nom") or not d.get("periodicite"): return jsonify({"error":"nom et periodicite requis"}),400
    db=get_db()
    sous_type_id = d.get("sous_type_id") or None
    if sous_type_id == "" : sous_type_id = None
    db.execute("INSERT INTO gammes (nom,periodicite,temps,sous_type_id) VALUES (?,?,?,?)",
               (d["nom"],d["periodicite"],d.get("temps","00h00"),sous_type_id))
    db.commit(); gid=db.execute("SELECT last_insert_rowid()").fetchone()[0]
    for i,op in enumerate(d.get("operations",[])):
        if op.get("description"): db.execute("INSERT INTO gamme_operations (gamme_id,ordre,description) VALUES (?,?,?)",(gid,i,op["description"]))
    db.commit(); return jsonify({"id":gid}),201

@app.route("/api/gammes/<int:gid>",methods=["PATCH"])
@require_role("admin","manager")
def update_gamme(gid):
    d=request.json or {}; db=get_db(); sets,params=[],[]
    for f in ["nom","periodicite","temps"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    # sous_type_id : accepter NULL explicite via "" ou None
    if "sous_type_id" in d:
        v = d["sous_type_id"]
        if v == "" or v == 0: v = None
        sets.append("sous_type_id=?"); params.append(v)
    if sets: params.append(gid); db.execute(f"UPDATE gammes SET {chr(44).join(sets)} WHERE id=?",params)
    if "operations" in d:
        db.execute("DELETE FROM gamme_operations WHERE gamme_id=?",(gid,))
        for i,op in enumerate(d["operations"]):
            if op.get("description"): db.execute("INSERT INTO gamme_operations (gamme_id,ordre,description) VALUES (?,?,?)",(gid,i,op["description"]))
    db.commit(); return jsonify({"ok":True})

@app.route("/api/gammes/<int:gid>",methods=["DELETE"])
@require_role("admin","manager")
def delete_gamme(gid):
    db=get_db(); db.execute("DELETE FROM gammes WHERE id=?",(gid,)); db.commit()
    return jsonify({"ok":True})

# ══ SMTP ══
@app.route("/api/smtp")
@require_role("admin")
def get_smtp():
    r=one(get_db().execute("SELECT * FROM smtp_config WHERE id=1"))
    if r: r.pop("password",None)
    return jsonify(r or {})

@app.route("/api/smtp",methods=["PATCH"])
@require_role("admin")
def update_smtp():
    d=request.json or {}; db=get_db(); sets,params=[],[]
    for f in ["host","port","username","password","sender_email","sender_name","use_tls","enabled"]:
        if f in d: sets.append(f"{f}=?"); params.append(d[f])
    if not sets: return jsonify({"error":"Rien"}),400
    params.append(1); db.execute(f"UPDATE smtp_config SET {chr(44).join(sets)} WHERE id=?",params); db.commit()
    return jsonify({"ok":True})

@app.route("/api/smtp/test",methods=["POST"])
@require_role("admin")
def test_smtp():
    d=request.json or {}; to=d.get("to") or request.user.get("email")
    if not to: return jsonify({"error":"Destinataire requis"}),400
    try: send_mail(to,"[GMAO] Test SMTP","Test de configuration SMTP depuis SOCOM GMAO."); return jsonify({"ok":True})
    except Exception as e: return jsonify({"error":str(e)}),500

# ══ EXPORT/IMPORT EXCEL ══
@app.route("/api/export/excel")
@require_role("admin","manager")
def export_excel():
    try:
        import openpyxl; from openpyxl.styles import Font,PatternFill,Alignment
        db=get_db(); wb=openpyxl.Workbook()
        BF=PatternFill("solid",fgColor="244298"); WF=Font(bold=True,color="FFFFFF")
        def make_sheet(ws,headers,data_rows):
            ws.append(headers)
            for cell in ws[1]: cell.font=WF; cell.fill=BF; cell.alignment=Alignment(horizontal="center")
            for row in data_rows: ws.append([str(v) if v is not None else "" for v in row])
            for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=max(len(str(c.value or "")) for c in col)+4

        # Clients
        ws=wb.active; ws.title="Clients"
        make_sheet(ws,["ID","Société","Nom","Prénom","Email","Téléphone","Notes"],
            [(r["id"],r["societe"],r["nom"],r["prenom"],r["email"],r["telephone"],r["notes"])
             for r in rows(db.execute("SELECT * FROM clients ORDER BY societe"))])

        # Projets
        ws2=wb.create_sheet("Projets")
        make_sheet(ws2,["ID","Numéro","Nom","Client","Manager","Date début","Date fin","Statut"],
            [(r["id"],r["numero_projet"],r["nom"],r.get("client_nom",""),r.get("manager_nom",""),r["date_debut"],r["date_fin"],r["statut"])
             for r in rows(db.execute("""SELECT p.*,c.societe AS client_nom,u.nom AS manager_nom
                FROM projets p LEFT JOIN clients c ON p.client_id=c.id
                LEFT JOIN utilisateurs u ON p.manager_id=u.id ORDER BY p.nom"""))])

        # Equipements
        ws3=wb.create_sheet("Equipements")
        make_sheet(ws3,["ID","Désignation","Type","Projet","Localisation","Marque","Modèle","Puissance","N° Série","I/O","Mise en service","Statut"],
            [(e["id"],e["designation"],e["type_technique"],e.get("projet_nom",""),e.get("localisation",""),e.get("marque",""),e.get("modele",""),e.get("puissance",""),e.get("numero_serie",""),e.get("in_out",""),e.get("date_mise_en_service",""),e["statut"])
             for e in rows(db.execute("""SELECT e.*,p.nom AS projet_nom FROM equipements e
                JOIN projets p ON e.projet_id=p.id ORDER BY p.nom,e.designation"""))])

        # Pièces
        ws4=wb.create_sheet("Pieces")
        make_sheet(ws4,["ID","Type","Equipement","Projet","Qté","Référence","N° Série","Date install.","Durée vie (ans)","Fin de vie","Statut","Commentaire"],
            [(p["id"],p["type_piece"],p.get("equip_nom",""),p.get("projet_nom",""),p.get("quantite",1),p.get("reference",""),p.get("numero_serie",""),p.get("date_installation",""),p.get("duree_vie_estimee",""),p.get("date_fin_de_vie",""),p["statut"],p.get("commentaire",""))
             for p in rows(db.execute("""SELECT p.*,e.designation AS equip_nom,pr.nom AS projet_nom
                FROM pieces p JOIN equipements e ON p.equipement_id=e.id
                JOIN projets pr ON e.projet_id=pr.id ORDER BY pr.nom,e.designation"""))])

        # Interventions
        ws5=wb.create_sheet("Interventions")
        make_sheet(ws5,["ID","Numéro","Type","Statut","Equipement","Projet","Client","Technicien","Equipe","Date prévue","Date réalisation","Description"],
            [(i["id"],i["numero"],i["type"],i["statut"],i.get("equip_nom",""),i.get("projet_nom",""),i.get("client_nom",""),i.get("technicien_nom",""),i.get("equipe_nom",""),i.get("date_prevue",""),i.get("date_realisation",""),i.get("description",""))
             for i in rows(db.execute("""SELECT i.*,e.designation AS equip_nom,p.nom AS projet_nom,
                c.societe AS client_nom,u.nom AS technicien_nom,eq.nom AS equipe_nom
                FROM interventions i JOIN equipements e ON i.equipement_id=e.id
                JOIN projets p ON e.projet_id=p.id LEFT JOIN clients c ON p.client_id=c.id
                LEFT JOIN utilisateurs u ON i.technicien_id=u.id
                LEFT JOIN equipes eq ON i.equipe_id=eq.id ORDER BY i.created_at DESC"""))])

        # Comptes rendus
        ws6=wb.create_sheet("Comptes rendus")
        make_sheet(ws6,["ID","N° BT","Equipement","Date intervention","Observations","Actions réalisées","Mesures","Recommandations","Conclusion","Total heures"],
            [(r["id"],r.get("numero",""),r.get("equip_nom",""),r.get("date_intervention",""),r.get("observations",""),r.get("actions_realisees",""),r.get("mesures",""),r.get("recommandations",""),r.get("conclusion",""),r.get("total_heures",0))
             for r in rows(db.execute("""SELECT cr.*,i.numero,e.designation AS equip_nom
                FROM comptes_rendus cr JOIN interventions i ON cr.intervention_id=i.id
                JOIN equipements e ON i.equipement_id=e.id ORDER BY cr.created_at DESC"""))])

        # Gammes
        ws7=wb.create_sheet("Gammes")
        make_sheet(ws7,["ID","Nom","Périodicité","Temps","Opérations"],
            [(g["id"],g["nom"],g["periodicite"],g["temps"],
              " | ".join([o["description"] for o in rows(db.execute("SELECT * FROM gamme_operations WHERE gamme_id=? ORDER BY ordre",(g["id"],)))]))
             for g in rows(db.execute("SELECT * FROM gammes ORDER BY nom"))])

        # Techniques
        ws8=wb.create_sheet("Techniques")
        make_sheet(ws8,["ID","Nom","Description"],
            [(t["id"],t["nom"],t.get("description",""))
             for t in rows(db.execute("SELECT * FROM techniques ORDER BY nom"))])

        # Utilisateurs
        ws9=wb.create_sheet("Utilisateurs")
        make_sheet(ws9,["ID","Nom","Email","Rôle","Actif"],
            [(u["id"],u["nom"],u["email"],u["role"],u["actif"])
             for u in rows(db.execute("SELECT * FROM utilisateurs ORDER BY nom"))])

        # Equipes
        ws10=wb.create_sheet("Equipes")
        make_sheet(ws10,["ID","Nom","Manager","Membres"],
            [(eq["id"],eq["nom"],eq.get("manager_nom",""),
              " | ".join([m["nom"] for m in rows(db.execute("""SELECT u.nom FROM utilisateurs u
                JOIN equipe_membres em ON em.technicien_id=u.id WHERE em.equipe_id=?""",(eq["id"],)))]))
             for eq in rows(db.execute("""SELECT e.*,u.nom AS manager_nom FROM equipes e
                LEFT JOIN utilisateurs u ON e.manager_id=u.id ORDER BY e.nom"""))])

        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        as_attachment=True,download_name=f"gmao_export_{today()}.xlsx")
    except ImportError: return jsonify({"error":"openpyxl non installe"}),500
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/api/import/excel",methods=["POST"])
@require_role("admin")
def import_excel():
    try:
        import openpyxl; f=request.files.get("file")
        if not f: return jsonify({"error":"Fichier requis"}),400
        wb=openpyxl.load_workbook(f); db=get_db(); count=0
        if "Equipements" in wb.sheetnames:
            ws=wb["Equipements"]
            for row in ws.iter_rows(min_row=2,values_only=True):
                if not row[1]: continue
                try:
                    db.execute("INSERT OR IGNORE INTO equipements (designation,type_technique,localisation,marque,modele,statut,projet_id) VALUES (?,?,?,?,?,?,1)",
                               (row[1],row[2] or "UPS",row[4] or "",row[5] or "",row[6] or "",row[7] or "EN_SERVICE")); count+=1
                except Exception: pass
        db.commit(); return jsonify({"ok":True,"imported":count})
    except ImportError: return jsonify({"error":"openpyxl non installe"}),500
    except Exception as e: return jsonify({"error":str(e)}),500

# ══ PDF RAPPORT ══
@app.route("/api/interventions/<int:iid>/pdf")
@require_auth
def generate_pdf(iid):
    try: from rapport_pdf import generate_rapport
    except ImportError: return jsonify({"error":"Module rapport_pdf non disponible"}),500
    db=get_db()
    i=one(db.execute("""SELECT i.*,
           e.designation AS equip_nom, e.type_technique, e.localisation, e.marque, e.modele,
           e.puissance, e.numero_serie, e.in_out, e.date_mise_en_service AS eq_mise_service,
           e.statut AS eq_statut, e.notes AS eq_notes, e.technique_id AS eq_technique_id,
           e.image_id AS eq_image_id,
           e.trafo_marque, e.trafo_annee, e.trafo_numero_serie, e.trafo_puissance_kva,
           e.trafo_refroidissement, e.trafo_poids_kg, e.trafo_tension_entree_v,
           e.trafo_courant_a, e.trafo_norme, e.trafo_couplage,
           e.trafo_tension_service_v, e.trafo_reglage_tension_kv,
           p.id AS projet_id, p.nom AS projet_nom, p.numero_projet, p.logo_filename,
           c.societe AS client_nom,
           u.nom AS technicien_nom,
           et.nom AS tableau_nom, et.localisation AS tableau_localisation,
           il.filename AS eq_image_filename
           FROM interventions i
           JOIN equipements e ON i.equipement_id=e.id
           JOIN projets p ON e.projet_id=p.id
           LEFT JOIN clients c ON p.client_id=c.id
           LEFT JOIN utilisateurs u ON i.technicien_id=u.id
           LEFT JOIN equipement_tableaux et ON i.tableau_id=et.id
           LEFT JOIN image_library il ON e.image_id = il.id
           WHERE i.id=?""",(iid,)))
    if not i: return jsonify({"error":"Non trouve"}),404
    crs=rows(db.execute("SELECT * FROM comptes_rendus WHERE intervention_id=? ORDER BY created_at",(iid,)))
    for cr in crs:
        cr["intervenants"]=rows(db.execute("""
            SELECT ci.*,COALESCE(u.nom,ci.nom,'') AS nom
            FROM cr_intervenants ci LEFT JOIN utilisateurs u ON ci.utilisateur_id=u.id
            WHERE ci.cr_id=? ORDER BY ci.id""",(cr["id"],)))
        try:
            cr["materiels"]=rows(db.execute("SELECT * FROM cr_materiels WHERE cr_id=? ORDER BY ordre,id",(cr["id"],)))
        except: cr["materiels"]=[]
    # Construire la liste structurée des comptes-rendus pour le PDF
    def _calc_heures(hd, hf):
        """Calcule la différence d'heures entre deux horaires HH:MM.
        Si heure_fin < heure_debut, considère un passage minuit (+24h).
        Renvoie un float (ex: 2.25 pour 2h15) ou 0 si format invalide.
        """
        try:
            if not hd or not hf: return 0.0
            hd_parts = str(hd).split(":"); hf_parts = str(hf).split(":")
            if len(hd_parts)<2 or len(hf_parts)<2: return 0.0
            m_debut = int(hd_parts[0])*60 + int(hd_parts[1])
            m_fin = int(hf_parts[0])*60 + int(hf_parts[1])
            diff = m_fin - m_debut
            # Passage minuit : ajouter 24h si fin < début
            if diff < 0:
                diff += 24 * 60
            return round(diff / 60.0, 4)
        except Exception:
            return 0.0
    comptes_rendus_pdf = []
    for cr in crs:
        ivs = []
        for iv in cr.get("intervenants", []):
            nom = iv.get("nom", "") or ""
            if not nom and iv.get("utilisateur_id"):
                u = one(db.execute("SELECT nom FROM utilisateurs WHERE id=?", (iv["utilisateur_id"],)))
                if u: nom = u["nom"]
            # Total : priorité à la valeur stockée si >0, sinon calcul depuis horaires
            total_iv = 0.0
            try: total_iv = float(iv.get("total_heures", 0) or 0)
            except Exception: pass
            if total_iv <= 0:
                total_iv = _calc_heures(iv.get("heure_debut",""), iv.get("heure_fin",""))
            ivs.append({
                "nom": nom or "—",
                "date": iv.get("date", "") or cr.get("date_intervention", ""),
                "heure_debut": iv.get("heure_debut", ""),
                "heure_fin": iv.get("heure_fin", ""),
                "total_heures": total_iv
            })
        # Fallback : si pas d'intervenants CR, utiliser le technicien principal
        if not ivs and i.get("technicien_nom"):
            ivs.append({
                "nom": i["technicien_nom"],
                "date": cr.get("date_intervention", ""),
                "heure_debut": "",
                "heure_fin": "",
                "total_heures": cr.get("total_heures", 0)
            })
        comptes_rendus_pdf.append({
            "date": cr.get("date_intervention", "") or "—",
            "numero": cr.get("numero", "") or "",
            "observations": cr.get("observations", "") or "",
            "intervenants": ivs,
            "materiels": [
                {"designation": m.get("designation",""), "quantite": m.get("quantite",0)}
                for m in cr.get("materiels", []) if m.get("designation","")
            ],
            "photos": [bytes(p["data"]) for p in rows(db.execute("SELECT data FROM cr_photos WHERE cr_id=? ORDER BY id",(cr["id"],)))],
        })
    # Liste globale des matériels (toutes les lignes, même en doublon)
    materiels_globaux = []
    for cr in comptes_rendus_pdf:
        for m in cr.get("materiels", []):
            materiels_globaux.append({
                "designation": m["designation"],
                "quantite": m["quantite"],
                "cr_numero": cr.get("numero","") or ""
            })
    # Déplacement : km vient du projet, nombre vient du/des CR(s) (dépannage uniquement)
    depl_km = 0.0; depl_nb = 0
    # Récupérer km depuis le projet (via l'équipement)
    if i.get("equipement_id"):
        try:
            proj = one(db.execute("""SELECT p.deplacement_km
                FROM projets p JOIN equipements e ON e.projet_id=p.id WHERE e.id=?""",
                (i["equipement_id"],)))
            if proj:
                depl_km = float(proj.get("deplacement_km") or 0)
        except Exception: pass
    # Somme des nb_deplacements de tous les CRs
    try:
        for cr in crs:
            depl_nb += int(cr.get("nb_deplacements") or 0)
    except Exception: pass
    # Date du bon : première date de CR, sinon date_realisation ou date_prevue
    if comptes_rendus_pdf:
        date_iv = comptes_rendus_pdf[0].get("date", "") or i.get("date_realisation") or i.get("date_prevue") or today()
    else:
        date_iv = i.get("date_realisation") or i.get("date_prevue") or today()
    date_iv = (date_iv or "").split("\n")[0][:10]
    # Équipement affiché
    eq = i.get("equip_nom", "—")
    marque_modele = ""
    if i.get("marque") or i.get("modele"):
        marque_modele = f"{i.get('marque','')} {i.get('modele','')}".strip()
    # Intervenants (synthèse : noms uniques)
    noms_uniques = []
    seen_n = set()
    for cr in comptes_rendus_pdf:
        for iv in cr.get("intervenants", []):
            n = iv.get("nom", "")
            if n and n != "—" and n not in seen_n:
                seen_n.add(n); noms_uniques.append(n)
    if not noms_uniques and i.get("technicien_nom"):
        noms_uniques = [i["technicien_nom"]]
    intervenants_str = ", ".join(noms_uniques) if noms_uniques else "—"
    # Technique (nom lisible)
    technique = i.get("type_technique", "") or "—"
    # Récupérer le nom complet de la technique depuis la table techniques si lié
    try:
        tech_row = one(db.execute("SELECT t.nom FROM techniques t JOIN equipements e ON e.technique_id=t.id WHERE e.id=?", (i["equipement_id"],)))
        if tech_row and tech_row.get("nom"):
            technique = tech_row["nom"]
    except Exception:
        pass
    try:
        # Chemin absolu du logo projet s'il existe
        _projet_logo_path = ""
        _logo_fname = i.get("logo_filename") or ""
        if _logo_fname:
            _candidate = BASE_DIR / "uploads" / "projet_logos" / _logo_fname
            if _candidate.exists():
                _projet_logo_path = str(_candidate)
        # Sous-type maintenance déduit des gammes du bon
        _sous_type_label = _bon_sous_type_label(iid)
        # Mesures techniques (UPS×Entretien, etc.)
        _mesures_techniques = _intervention_mesures_data(iid)
        # v218.81 : charger la page de garde personnalisée de la société de l'intervention
        # v218.85 : charger aussi header_blocks et footer_blocks
        # v218.112 : pour les bons DÉPANNAGE, on charge la cover_pages_bc à la place
        _cover_blocks = None
        _header_blocks = None
        _footer_blocks = None
        _societe_nom = ""
        _societe_logo_path = ""
        try:
            _iv_sid = i.get("societe_id") or current_societe_id()
            _iv_type = (i.get("type") or "").upper()
            _cover_table = "cover_pages_bc" if _iv_type == "DEPANNAGE" else "cover_pages"
            _cover_row = one(db.execute(f"""SELECT blocks_json, header_blocks_json, footer_blocks_json
                                           FROM {_cover_table} WHERE societe_id=?""", (_iv_sid,)))
            if _cover_row:
                if _cover_row.get("blocks_json"):
                    _cover_blocks = json.loads(_cover_row["blocks_json"])
                    if not isinstance(_cover_blocks, list): _cover_blocks = None
                if _cover_row.get("header_blocks_json"):
                    _header_blocks = json.loads(_cover_row["header_blocks_json"])
                    if not isinstance(_header_blocks, list): _header_blocks = None
                if _cover_row.get("footer_blocks_json"):
                    _footer_blocks = json.loads(_cover_row["footer_blocks_json"])
                    if not isinstance(_footer_blocks, list): _footer_blocks = None
            # Récupérer nom + logo de la société
            _soc_row = one(db.execute("SELECT nom, logo_path FROM societes WHERE id=?", (_iv_sid,)))
            if _soc_row:
                _societe_nom = _soc_row.get("nom") or ""
                _logo_p = _soc_row.get("logo_path") or ""
                if _logo_p:
                    _abs = os.path.join(BASE_DIR, _logo_p)
                    if os.path.exists(_abs):
                        _societe_logo_path = _abs
        except Exception as _e:
            logger.warning(f"[cover-page] chargement échec pour intervention {iid}: {_e}")
        pdf = generate_rapport({
            "type_label": i.get("type", ""),
            "cover_page_blocks": _cover_blocks,
            "header_blocks": _header_blocks,
            "footer_blocks": _footer_blocks,
            "societe_nom": _societe_nom,
            "societe_logo_path": _societe_logo_path,
            "sous_type": _sous_type_label,
            "mesures_techniques": _mesures_techniques,
            "numero_iv": i.get("numero", "—"),
            "date": date_iv,
            "client": i.get("client_nom", "—"),
            "numero_projet": i.get("numero_projet", "") or "—",
            "projet_nom": i.get("projet_nom", "") or "",
            "projet_logo_path": _projet_logo_path,
            "equipement": eq,
            "marque_modele": marque_modele or "—",
            "localisation": i.get("localisation", "") or "—",
            # v217 : chemin absolu de l'image principale de l'équipement (None si pas d'image)
            "equipement_image_path": (str(IMAGES_DIR / i.get("eq_image_filename")) if i.get("eq_image_filename") else None),
            "tableau": (i.get("tableau_nom","") + (" (" + i.get("tableau_localisation","") + ")" if i.get("tableau_localisation") else "")) if i.get("tableau_nom") else "",
            "tableau_label": ("Borne" if (i.get("type_technique") or "").lower()=="borne de charge" else "Tableau"),
            "trafo": ({
                "marque": i.get("trafo_marque","") or "",
                "annee": i.get("trafo_annee","") or "",
                "numero_serie": i.get("trafo_numero_serie","") or "",
                "puissance_kva": i.get("trafo_puissance_kva","") or "",
                "refroidissement": i.get("trafo_refroidissement","") or "",
                "poids_kg": i.get("trafo_poids_kg","") or "",
                "tension_entree_v": i.get("trafo_tension_entree_v","") or "",
                "courant_a": i.get("trafo_courant_a","") or "",
                "norme": i.get("trafo_norme","") or "",
                "couplage": i.get("trafo_couplage","") or "",
                "tension_service_v": i.get("trafo_tension_service_v","") or "",
                "reglage_tension_kv": i.get("trafo_reglage_tension_kv","") or "",
            }) if (i.get("type_technique") or "").lower()=="haute tension" else None,
            "cellules": [],  # v218.160 : table supprimée
            # v218.158 : sous-équipements universels (toutes techniques)
            "sous_equipements": (rows(db.execute("SELECT designation, marque, type, nombre FROM equipement_sous_equipements WHERE equipement_id=? ORDER BY ordre, id", (i.get("equipement_id"),))) if i.get("equipement_id") else []),
            "securite_items": (rows(db.execute("""
                SELECT se.id, se.libelle, se.photo_data, se.photo_mime,
                       isec.present AS present,
                       isec.conforme AS conforme
                FROM securite_equipements se
                LEFT JOIN intervention_securite isec ON isec.securite_equipement_id=se.id AND isec.intervention_id=?
                WHERE se.actif=1
                ORDER BY se.ordre, se.id
            """, (iid,))) if (i.get("type_technique") or "").lower()=="haute tension" else []),
            "technique": technique,
            # Infos complémentaires de la fiche équipement
            "eq_puissance": i.get("puissance", "") or "",
            "eq_numero_serie": i.get("numero_serie", "") or "",
            "eq_in_out": i.get("in_out", "") or "",
            "eq_date_mise_service": i.get("eq_mise_service", "") or "",
            "eq_statut": i.get("eq_statut", "") or "",
            "eq_notes": i.get("eq_notes", "") or "",
            # v218.93 : liste de TOUS les tableaux/bornes liés à l'équipement (pour module 'equipement')
            "tableaux_eq": (lambda eq_id_: [dict(r) for r in rows(db.execute(
                "SELECT nom, localisation FROM equipement_tableaux WHERE equipement_id=? ORDER BY ordre, id",
                (eq_id_,)
            ))] if eq_id_ else [])(i.get("equipement_id")),
            "intervenants": intervenants_str,
            "description": i.get("description", "") or "",
            "comptes_rendus": comptes_rendus_pdf,
            "gamme_maintenance": (lambda iv_id, eq_id_, type_iv: (
                [
                    {
                        "nom": s["nom"],
                        "localisation": s.get("localisation","") or "",
                        "date_realisation": (one(db.execute("""SELECT cr.date_intervention FROM cr_tableaux ct
                            JOIN comptes_rendus cr ON ct.cr_id=cr.id
                            WHERE ct.tableau_id=? AND cr.intervention_id=?
                            ORDER BY cr.date_intervention, cr.id LIMIT 1""", (s["id"], iv_id))) or {}).get("date_intervention","") or "",
                        "cr_numero": (one(db.execute("""SELECT cr.numero FROM cr_tableaux ct
                            JOIN comptes_rendus cr ON ct.cr_id=cr.id
                            WHERE ct.tableau_id=? AND cr.intervention_id=?
                            ORDER BY cr.date_intervention, cr.id LIMIT 1""", (s["id"], iv_id))) or {}).get("numero","") or "",
                    }
                    for s in rows(db.execute("SELECT id, nom, localisation FROM equipement_tableaux WHERE equipement_id=? ORDER BY ordre, id", (eq_id_,)))
                ] if type_iv == "MAINTENANCE" else []
            ))(iid, i.get("equipement_id"), i.get("type")),
            "is_maintenance": (i.get("type") == "MAINTENANCE"),
            "gamme_operations": (lambda eq_id_, type_iv, iv_id: (
                (lambda gids: [
                    {
                        "gamme_nom": ((one(db.execute("SELECT nom FROM gammes WHERE id=?", (gid,))) or {}).get("nom","") or ""),
                        "operations": [
                            {
                                "description": op["description"],
                                "done": bool(_d),
                                "date_realisation": (_d.get("date_realisation") if _d else "") or "",
                                "technicien_nom": (_d.get("tech_nom") if _d else "") or "",
                                "cr_numero": (_d.get("cr_num") if _d else "") or "",
                            }
                            for op in rows(db.execute(
                                "SELECT id, description FROM gamme_operations WHERE gamme_id=? ORDER BY ordre, id", (gid,)
                            ))
                            for _d in [one(db.execute("""
                                SELECT io.date_realisation, u.nom AS tech_nom, cr.numero AS cr_num
                                FROM intervention_operations io
                                LEFT JOIN utilisateurs u ON io.technicien_id = u.id
                                LEFT JOIN comptes_rendus cr ON io.cr_id = cr.id
                                WHERE io.intervention_id=? AND io.gamme_operation_id=?
                            """, (iv_id, op["id"])))]
                        ]
                    }
                    for gid in gids
                    if rows(db.execute("SELECT 1 FROM gamme_operations WHERE gamme_id=? LIMIT 1", (gid,)))
                ])(
                    list(set(
                        [g["gamme_id"] for g in rows(db.execute("SELECT gamme_id FROM equipement_gammes WHERE equipement_id=?", (eq_id_,)))] +
                        ([(one(db.execute("SELECT gamme_id FROM equipements WHERE id=?", (eq_id_,))) or {}).get("gamme_id")]
                          if (one(db.execute("SELECT gamme_id FROM equipements WHERE id=?", (eq_id_,))) or {}).get("gamme_id") else [])
                    ))
                ) if type_iv == "MAINTENANCE" else []
            ))(i.get("equipement_id"), i.get("type"), iid),
            "materiels_globaux": materiels_globaux,
            "deplacement_km": depl_km,
            "nb_deplacements": depl_nb,
        })
    except Exception as e:
        import traceback
        logger.error(f"[generate_pdf] erreur iid={iid}: {e}\n{traceback.format_exc()}")
        return jsonify({"error": f"Erreur PDF : {str(e)}"}), 500
    fn = f"Rapport_{i.get('numero','BT00000')}_{(i.get('client_nom') or 'SOCOM').replace(' ','_')}.pdf"
    return send_file(io.BytesIO(pdf), mimetype="application/pdf", as_attachment=True, download_name=fn)
# ══ MIGRATIONS PLANNINGS ══
def migrate_plannings(db):
    # Migration plannings table
    pcols = [r[1] for r in db.execute("PRAGMA table_info(plannings)").fetchall()]
    for col, typ, default in [
        ("equipe_id", "INTEGER", "NULL"),
        ("heure_debut", "TEXT", "'08:00'"),
    ]:
        if col not in pcols:
            try: db.execute(f"ALTER TABLE plannings ADD COLUMN {col} {typ} DEFAULT {default}")
            except Exception: pass
    # Migration equipements
    cols = [r[1] for r in db.execute("PRAGMA table_info(equipements)").fetchall()]
    for col, typ, default in [
        ("planning_id", "INTEGER", "NULL"),
        ("semaine_planif", "INTEGER", "NULL"),
        ("jour_semaine_planif", "INTEGER", "NULL"),
        ("intervention_samedi", "INTEGER", "0"),
        ("intervention_dimanche", "INTEGER", "0"),
    ]:
        if col not in cols:
            db.execute(f"ALTER TABLE equipements ADD COLUMN {col} {typ} DEFAULT {default}")
    db.commit()

# ══ PLANNINGS ══
@app.route("/api/plannings")
@require_auth
def get_plannings():
    sid = current_societe_id()
    return jsonify(rows(get_db().execute("SELECT * FROM plannings WHERE societe_id=? ORDER BY nom", (sid,))))

@app.route("/api/plannings", methods=["POST"])
@require_role("admin","manager")
def create_planning():
    d = request.json or {}
    if not d.get("nom"): return jsonify({"error":"nom requis"}),400
    sid = current_societe_id()
    db = get_db()
    nb_sim = max(1, to_int(d.get("nb_simultane")) or 1)
    db.execute("INSERT INTO plannings (nom,equipe_id,heures_par_jour,heure_debut,nb_simultane,societe_id) VALUES (?,?,?,?,?,?)",
               (d["nom"],d.get("equipe_id"),d.get("heures_par_jour",8.0),d.get("heure_debut","08:00"),nb_sim,sid))
    db.commit()
    return jsonify({"id":db.execute("SELECT last_insert_rowid()").fetchone()[0]}),201

@app.route("/api/plannings/<int:pid>", methods=["PATCH"])
@require_role("admin","manager")
def update_planning(pid):
    d = request.json or {}; db = get_db(); sets,params=[],[]
    for f in ["nom","equipe_id","heures_par_jour","heure_debut","nb_simultane"]:
        if f in d:
            v = d[f]
            if f == "nb_simultane":
                v = max(1, to_int(v) or 1)
            sets.append(f"{f}=?"); params.append(v)
    if not sets: return jsonify({"error":"Rien"}),400
    params.append(pid)
    db.execute(f"UPDATE plannings SET {chr(44).join(sets)} WHERE id=?",params)
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/plannings/<int:pid>", methods=["DELETE"])
@require_role("admin","manager")
def delete_planning(pid):
    db = get_db()
    db.execute("UPDATE equipements SET planning_id=NULL WHERE planning_id=?",(pid,))
    db.execute("DELETE FROM plannings WHERE id=?",(pid,))
    db.commit()
    return jsonify({"ok":True})


